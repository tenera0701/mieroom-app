# ======================================================
# DBリセット手順:
#   PLRecordモデルに新カラムを追加したため、
#   既存のDBには新カラムが存在しない。
#   リセットするには instance/realestate.db を削除して再起動する。
#   例: del instance\realestate.db (Windows)
# ======================================================
import os
import json
import random
import tempfile
import re
import socket
import threading
import time
import hashlib
import imaplib
import smtplib
import email as emaillib
from email.header import decode_header as _decode_header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders as _email_encoders
from email.utils import parseaddr as _parseaddr
from functools import wraps
from flask import Flask, redirect, url_for, session, render_template, request, jsonify, make_response
from flask_sqlalchemy import SQLAlchemy
from dotenv import load_dotenv
import anthropic
from datetime import datetime, date, timedelta, timezone
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-key-change-in-production")

# Railway(PostgreSQL) or local(SQLite) の自動切り替え
_db_url = os.getenv("DATABASE_URL", "")
if _db_url.startswith("postgres://"):          # Railway が postgres:// を返す場合
    _db_url = _db_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = _db_url or "sqlite:///realestate.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0

_IS_POSTGRES = bool(_db_url)

db = SQLAlchemy(app)


@app.before_request
def sync_session_role():
    """セッションのロールをDBと同期（再ログイン不要でロール変更を反映）"""
    uid = session.get('app_user_id')
    if uid:
        user = AppUser.query.get(uid)
        if user and user.is_active:
            session['app_user_role'] = user.role
        elif user and not user.is_active:
            session.clear()


@app.after_request
def add_no_cache(response):
    """HTMLページとAPI(JSON)応答のブラウザキャッシュを無効化。
    APIをキャッシュすると編集後の再取得が古い値を返し、リロードするまで反映されない。"""
    ctype = response.content_type or ''
    is_api = request.path.startswith('/api/')
    if 'text/html' in ctype or 'application/json' in ctype or is_api:
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response


@app.context_processor
def inject_ui_context():
    """全テンプレートに共通変数を注入（is_premium はリクエスト時に評価）"""
    def _is_premium():
        uid = session.get('app_user_id')
        if not uid:
            return False
        u = AppUser.query.get(uid)
        if not u or u.role == 'super_admin' or not u.tenant_id:
            return False
        t = Tenant.query.get(u.tenant_id)
        return bool(t and t.plan == 'premium')

    def _sidebar_stores():
        """サイドバー用店舗リスト（店舗切替セレクト・本部リンク判定に使用）
        ignore_active=True で全店舗を返す（切替中の店舗に関わらず全店表示）"""
        uid = session.get('app_user_id')
        if not uid:
            return []
        u = AppUser.query.get(uid)
        if not u or u.role == 'super_admin':
            return []
        if u.role == 'owner':
            return Store.query.filter_by(tenant_id=u.tenant_id, is_active=True).all()
        # store_manager / staff: 自分の店舗のみ
        if u.store_id:
            s = Store.query.get(u.store_id)
            return [s] if s and s.is_active else []
        return []

    def _current_user_perms():
        uid = session.get('app_user_id')
        if not uid:
            return {}
        u = AppUser.query.get(uid)
        if not u:
            return {}
        return {
            'can_view_executive':   getattr(u, 'can_view_executive', True),
            'can_view_leads_page':  getattr(u, 'can_view_leads_page', True),
            'can_view_daily_report':getattr(u, 'can_view_daily_report', True),
            'can_view_leave':       getattr(u, 'can_view_leave', True),
            'can_view_accounting':  getattr(u, 'can_view_accounting', True),
        }

    def _is_chat_pro():
        uid = session.get('app_user_id')
        if not uid:
            return False
        u = AppUser.query.get(uid)
        if not u or u.role == 'super_admin' or not u.tenant_id:
            return False
        return current_has_option('chat_pro')

    def _has_floorplan():
        return current_has_option('floorplan')

    return {
        'is_premium': _is_premium(),
        'is_chat_pro': _is_chat_pro(),
        'has_floorplan': _has_floorplan(),
        'current_role': session.get('app_user_role', ''),
        'sidebar_stores': _sidebar_stores(),
        'user_perms': _current_user_perms(),
    }


anthropic_client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))


# ── 既存モデル ─────────────────────────────────────────────

# ── 幹部向け管理ツール：追加モデル ────────────────────────

class Tenant(db.Model):
    """テナント（契約会社）マスタ — マルチSaaS用"""
    __tablename__ = 'tenant'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)       # 会社名
    plan = db.Column(db.String(20), default='standard')   # standard / premium
    is_active = db.Column(db.Boolean, default=True)
    trial_ends_at = db.Column(db.DateTime, nullable=True)  # トライアル終了日時
    subscription_status = db.Column(db.String(20), default='trial')  # trial / active / locked / cancelled
    contract_start_date = db.Column(db.Date, nullable=True)  # 契約開始日
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class TenantOption(db.Model):
    """テナント／店舗に紐づくオプション（プランに追加するアドオン）。
    store_id=NULL はテナント全体（全店舗）、store_id 指定はその店舗のみ有効。
    option_key は PLAN_OPTION_DEFS のキー。"""
    __tablename__ = 'tenant_option'
    id          = db.Column(db.Integer, primary_key=True)
    tenant_id   = db.Column(db.Integer, db.ForeignKey('tenant.id'))
    store_id    = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=True)  # NULL=テナント全体
    option_key  = db.Column(db.String(40))
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)


# 利用可能なオプション一覧（今後ここに追加していく）
PLAN_OPTION_DEFS = [
    {'key': 'chat_pro', 'name': 'チャットPro',
     'desc': 'チャットで画像・PDFの添付が可能になり、メッセージを2年間保存（通常はテキストのみ・60日保存）'},
    {'key': 'floorplan', 'name': '間取り作成',
     'desc': '物件の間取り図をツール上で作成・保存・印刷できる「間取り作成」機能を利用できます'},
]


def _tenant_active_stores(tenant_id):
    return Store.query.filter_by(tenant_id=tenant_id, is_active=True).all() if tenant_id else []


def tenant_option_keys(tenant_id):
    """全店舗に共通で付与されているオプション（クライアント管理のバッジ表示用）"""
    stores = _tenant_active_stores(tenant_id)
    if not stores:
        return set()
    sets = [store_option_keys(s.id) for s in stores]
    return set.intersection(*sets) if sets else set()


def tenant_has_option(tenant_id, key):
    """全店舗に共通でオプションを持つか（後方互換・表示用）"""
    return key in tenant_option_keys(tenant_id)


def set_tenant_options(tenant_id, keys):
    """クライアント管理から: 指定オプションをテナントの全有効店舗に一括適用する
    （オプションは店舗単位で管理。ここで指定したものを全店舗に付与し、外したものを全店舗から外す）"""
    valid = {d['key'] for d in PLAN_OPTION_DEFS}
    want = {k for k in (keys or []) if k in valid}
    for s in _tenant_active_stores(tenant_id):
        cur = store_option_keys(s.id)
        for k in want - cur:
            db.session.add(TenantOption(tenant_id=tenant_id, store_id=s.id, option_key=k))
        for k in cur - want:
            TenantOption.query.filter_by(store_id=s.id, option_key=k).delete()


def store_option_keys(store_id):
    """店舗個別に付与されたオプションキー"""
    if not store_id:
        return set()
    return {o.option_key for o in TenantOption.query.filter_by(store_id=store_id).all()}


def set_store_options(store_id, keys):
    """店舗個別のオプションを keys で置き換える"""
    if not store_id:
        return
    valid = {d['key'] for d in PLAN_OPTION_DEFS}
    want = {k for k in (keys or []) if k in valid}
    cur = store_option_keys(store_id)
    store = Store.query.get(store_id)
    tid = store.tenant_id if store else None
    for k in want - cur:
        db.session.add(TenantOption(tenant_id=tid, store_id=store_id, option_key=k))
    for k in cur - want:
        TenantOption.query.filter_by(store_id=store_id, option_key=k).delete()


def store_has_option(tenant_id, store_id, key):
    """その店舗にオプションが付与されているか（店舗単位のみ。テナント全体の概念は廃止）"""
    if not store_id:
        return False
    return TenantOption.query.filter_by(
        store_id=store_id, option_key=key).first() is not None


def current_has_option(key):
    """ログイン中ユーザーの「現在の店舗」がオプションを持つか（super_adminは常に可）"""
    uid = session.get('app_user_id')
    if not uid:
        return False
    u = AppUser.query.get(uid)
    if not u:
        return False
    if u.role == 'super_admin':
        return True
    if not u.tenant_id:
        return False
    sid = session.get('active_store_id') or getattr(u, 'store_id', None)
    return store_has_option(u.tenant_id, sid, key)


def current_has_floorplan():
    return current_has_option('floorplan')


class Store(db.Model):
    """店舗マスタ"""
    __tablename__ = 'store'
    id = db.Column(db.Integer, primary_key=True)
    tenant_id = db.Column(db.Integer, db.ForeignKey('tenant.id'), nullable=True)
    name = db.Column(db.String(100), nullable=False)
    # 月次固定費（円）
    rent = db.Column(db.Float, default=0)           # 家賃
    parking_fee = db.Column(db.Float, default=0)    # 駐車場代
    copier_fee = db.Column(db.Float, default=0)     # コピー機リース
    internet_fee = db.Column(db.Float, default=0)   # インターネット
    consultant_fee = db.Column(db.Float, default=0) # コンサル費
    insurance_fee = db.Column(db.Float, default=0)  # 保険料
    cloud_fee = db.Column(db.Float, default=0)      # クラウド・SaaS
    is_active = db.Column(db.Boolean, default=True)
    is_locked = db.Column(db.Boolean, default=False)           # 店舗ロック
    contract_start_date = db.Column(db.Date, nullable=True)    # 店舗契約開始日
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Staff(db.Model):
    """スタッフマスタ"""
    __tablename__ = 'staff'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'))
    role = db.Column(db.String(50), default='営業')
    is_active = db.Column(db.Boolean, default=True)
    hired_at = db.Column(db.Date)


class SalesKPI(db.Model):
    """月次営業KPI"""
    __tablename__ = 'sales_kpi'
    id = db.Column(db.Integer, primary_key=True)
    staff_id = db.Column(db.Integer, db.ForeignKey('staff.id'), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    inquiries = db.Column(db.Integer, default=0)      # 反響数
    store_visits = db.Column(db.Integer, default=0)   # 来店数
    viewings = db.Column(db.Integer, default=0)       # 内見数
    applications = db.Column(db.Integer, default=0)   # 申込数
    contracts = db.Column(db.Integer, default=0)      # 契約数
    cancellations = db.Column(db.Integer, default=0)  # キャンセル数
    sales_amount = db.Column(db.Float, default=0)     # 売上（円）
    option_sales = db.Column(db.Float, default=0)     # オプション売上（円）
    estimated_sales = db.Column(db.Float, default=0)     # 売上見込み（円）
    target_sales = db.Column(db.Float, default=0)        # 月次目標売上（円）
    fire_insurance_count = db.Column(db.Integer, default=0)  # 火災保険件数
    lifeline_count = db.Column(db.Integer, default=0)        # ライフライン件数
    moving_count = db.Column(db.Integer, default=0)          # 引越し件数


class Lead(db.Model):
    """反響（リード）管理"""
    __tablename__ = 'lead'
    id = db.Column(db.Integer, primary_key=True)
    # 反響媒体：SUUMO, HOME'S, at home, Instagram, TikTok, Google, LINE, HP, 電話, MEO
    source = db.Column(db.String(50))
    received_at = db.Column(db.DateTime, default=datetime.utcnow)
    # ステータス：未対応, 対応中, 来店, 内見, 申込, 契約, 不成立
    status = db.Column(db.String(50), default='未対応')
    assigned_staff_id = db.Column(db.Integer, db.ForeignKey('staff.id'))
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'))
    customer_name = db.Column(db.String(100))
    note = db.Column(db.Text)
    line_added = db.Column(db.Boolean, default=False)


class LeadMediaStat(db.Model):
    """媒体別月次反響統計"""
    __tablename__ = 'lead_media_stat'
    id = db.Column(db.Integer, primary_key=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), default=1)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    media = db.Column(db.String(50), nullable=False)
    inquiries = db.Column(db.Integer, default=0)       # 反響数
    replies = db.Column(db.Integer, default=0)         # 返信
    line_added = db.Column(db.Integer, default=0)      # LINE追加
    visits = db.Column(db.Integer, default=0)          # 来店
    applications = db.Column(db.Integer, default=0)    # 申込
    contracts = db.Column(db.Integer, default=0)       # 契約
    cancellations = db.Column(db.Integer, default=0)   # キャンセル
    cancel_amount = db.Column(db.Float, default=0)     # キャンセル金額
    estimated_sales = db.Column(db.Float, default=0)   # 売上見込み
    actual_payment = db.Column(db.Float, default=0)    # 入金
    ad_cost = db.Column(db.Float, default=0)           # 広告費


class AdCost(db.Model):
    """月次広告費"""
    __tablename__ = 'ad_cost'
    id = db.Column(db.Integer, primary_key=True)
    source = db.Column(db.String(50), nullable=False)  # 媒体名
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    cost = db.Column(db.Float, default=0)              # 広告費（円）
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'))


class PLRecord(db.Model):
    """月次PL（損益計算）"""
    __tablename__ = 'pl_record'
    id = db.Column(db.Integer, primary_key=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    revenue = db.Column(db.Float, default=0)         # 売上
    gross_profit = db.Column(db.Float, default=0)    # 粗利
    ad_cost = db.Column(db.Float, default=0)         # 広告費
    labor_cost = db.Column(db.Float, default=0)      # 人件費
    other_fixed = db.Column(db.Float, default=0)     # その他固定費
    other_variable = db.Column(db.Float, default=0)  # その他変動費
    # 収入詳細
    brokerage_fee = db.Column(db.Float, default=0)        # 仲介手数料
    ad_income = db.Column(db.Float, default=0)            # AD収入
    lifeline_income = db.Column(db.Float, default=0)      # ライフライン収入
    moving_income = db.Column(db.Float, default=0)        # 引越収入
    fire_insurance_income = db.Column(db.Float, default=0) # 火災保険収入
    other_income = db.Column(db.Float, default=0)         # その他収入
    # 広告費詳細
    suumo_cost = db.Column(db.Float, default=0)
    homes_cost = db.Column(db.Float, default=0)
    athome_cost = db.Column(db.Float, default=0)
    instagram_cost = db.Column(db.Float, default=0)
    tiktok_cost = db.Column(db.Float, default=0)
    google_ads_cost = db.Column(db.Float, default=0)
    line_cost = db.Column(db.Float, default=0)
    hp_cost = db.Column(db.Float, default=0)
    meo_cost = db.Column(db.Float, default=0)
    other_ad_cost = db.Column(db.Float, default=0)
    # 人件費詳細（commission_payカラムを社会保険料として再利用）
    regular_salary = db.Column(db.Float, default=0)    # 正社員人件費
    parttime_salary = db.Column(db.Float, default=0)   # アルバイト人件費
    commission_pay = db.Column(db.Float, default=0)    # 社会保険料（旧:歩合給）
    # 固定費詳細
    pl_rent = db.Column(db.Float, default=0)           # 家賃
    pl_parking = db.Column(db.Float, default=0)        # 駐車場
    pl_copier = db.Column(db.Float, default=0)         # 複合機
    pl_internet = db.Column(db.Float, default=0)       # インターネット
    pl_consultant = db.Column(db.Float, default=0)     # 顧問料
    pl_insurance = db.Column(db.Float, default=0)      # 保険料
    pl_cloud = db.Column(db.Float, default=0)          # クラウド


class PLCustomItem(db.Model):
    """PLカスタム費用項目テンプレート（固定費・変動費・その他）"""
    __tablename__ = 'pl_custom_item'
    id = db.Column(db.Integer, primary_key=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), default=1)
    name = db.Column(db.String(100), nullable=False)
    item_type = db.Column(db.String(20), default='固定費')  # 固定費 / 変動費 / その他
    sort_order = db.Column(db.Integer, default=0)


class PLCustomValue(db.Model):
    """PLカスタム費用項目の月次金額"""
    __tablename__ = 'pl_custom_value'
    id = db.Column(db.Integer, primary_key=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), default=1)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    item_name = db.Column(db.String(100), nullable=False)
    item_type = db.Column(db.String(20), default='固定費')
    amount = db.Column(db.Float, default=0)


class UncollectedPayment(db.Model):
    """未入金（AD）管理"""
    __tablename__ = 'uncollected_payment'
    id = db.Column(db.Integer, primary_key=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), default=1)
    staff_id = db.Column(db.Integer, db.ForeignKey('staff.id'), nullable=True)
    property_name = db.Column(db.String(200))     # 物件名
    room_number   = db.Column(db.String(50))       # 合室
    application_date    = db.Column(db.Date)        # 申込日
    management_company  = db.Column(db.String(200)) # 管理会社名
    customer_name       = db.Column(db.String(100)) # お客様名
    expected_payment_date = db.Column(db.Date)      # 入金予定日
    amount = db.Column(db.Float, default=0)
    memo   = db.Column(db.Text)
    is_paid = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class AppUser(db.Model):
    """管理ツールログインユーザー"""
    __tablename__ = 'app_user'
    id = db.Column(db.Integer, primary_key=True)
    tenant_id = db.Column(db.Integer, db.ForeignKey('tenant.id'), nullable=True)  # super_admin=None
    username = db.Column(db.String(100), unique=True, nullable=False)
    email = db.Column(db.String(200), nullable=True)
    password_hash = db.Column(db.String(256))
    role = db.Column(db.String(20), default='staff')  # 'super_admin'/'owner'/'store_manager'/'staff'
    staff_id = db.Column(db.Integer, db.ForeignKey('staff.id'), nullable=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    # 権限フラグ（store_manager/staff向け）
    can_view_accounting = db.Column(db.Boolean, default=True)
    can_view_all_staff = db.Column(db.Boolean, default=True)
    can_edit_kpi = db.Column(db.Boolean, default=True)
    can_manage_uncollected = db.Column(db.Boolean, default=True)
    # ページアクセス権限（ナビ項目ごと）
    can_view_executive = db.Column(db.Boolean, default=True)      # 売上管理
    can_view_leads_page = db.Column(db.Boolean, default=True)     # 反響管理
    can_view_daily_report = db.Column(db.Boolean, default=True)   # 日報
    can_view_leave = db.Column(db.Boolean, default=True)          # 有給管理
    # クライアント管理画面の操作権限（sys_admin向け）
    admin_can_add_tenant    = db.Column(db.Boolean, default=False)  # 新規テナント追加
    admin_can_manage_stores = db.Column(db.Boolean, default=False)  # 店舗管理
    admin_can_delete_tenant = db.Column(db.Boolean, default=False)  # 削除
    admin_can_lock_tenant   = db.Column(db.Boolean, default=False)  # ロック・解除


class PasswordResetToken(db.Model):
    """パスワードリセットトークン"""
    __tablename__ = 'password_reset_token'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('app_user.id'), nullable=False)
    token = db.Column(db.String(256), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    used = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class LeaveRecord(db.Model):
    """有給・休暇管理"""
    __tablename__ = 'leave_record'
    id = db.Column(db.Integer, primary_key=True)
    store_id  = db.Column(db.Integer, db.ForeignKey('store.id'), default=1)
    staff_id  = db.Column(db.Integer, db.ForeignKey('staff.id'), nullable=False)
    leave_date = db.Column(db.Date, nullable=False)
    leave_type = db.Column(db.String(20), default='有給')  # 有給/半休/欠勤/遅刻/早退/その他
    days       = db.Column(db.Float, default=1.0)          # 日数（0.5=半日）
    memo       = db.Column(db.Text)
    status     = db.Column(db.String(20), default='承認済')  # 申請中/承認済/却下
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class LeaveBalance(db.Model):
    """年次有給日数管理"""
    __tablename__ = 'leave_balance'
    id = db.Column(db.Integer, primary_key=True)
    store_id      = db.Column(db.Integer, db.ForeignKey('store.id'), default=1)
    staff_id      = db.Column(db.Integer, db.ForeignKey('staff.id'), nullable=False)
    year          = db.Column(db.Integer, nullable=False)
    total_days    = db.Column(db.Float, default=10.0)    # 付与日数（法定またはカスタム）
    carryover_days= db.Column(db.Float, default=0.0)     # 前年繰越日数
    is_custom     = db.Column(db.Boolean, default=False) # True=手動設定 / False=法定自動計算
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)


class DailyTaskTemplate(db.Model):
    """日報タスクテンプレート（店舗ごとにカスタム可能）"""
    __tablename__ = 'daily_task_template'
    id = db.Column(db.Integer, primary_key=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), default=1)
    task_name = db.Column(db.String(200), nullable=False)
    is_default = db.Column(db.Boolean, default=False)  # デフォルトタスク
    is_active = db.Column(db.Boolean, default=True)
    sort_order = db.Column(db.Integer, default=0)


class DailyReport(db.Model):
    """日報"""
    __tablename__ = 'daily_report'
    id = db.Column(db.Integer, primary_key=True)
    staff_id = db.Column(db.Integer, db.ForeignKey('staff.id'), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), default=1)
    report_date = db.Column(db.Date, nullable=False)
    # タスクチェック（固定）
    prev_day_contact_done = db.Column(db.Boolean, default=False)   # 来店前日連絡
    same_day_contact_done = db.Column(db.Boolean, default=False)   # 来店当日連絡
    application_input_done = db.Column(db.Boolean, default=False)  # 申込管理入力
    # 申込数
    application_count = db.Column(db.Integer, default=0)
    # 明日の接客予定
    tomorrow_appointments = db.Column(db.Text)
    # メモ
    memo = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class DailyReportCustomer(db.Model):
    """日報の接客記録"""
    __tablename__ = 'daily_report_customer'
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey('daily_report.id'), nullable=False)
    customer_name = db.Column(db.String(100), nullable=False)
    applied = db.Column(db.Boolean, default=False)          # 申込になったか
    no_apply_reason = db.Column(db.Text)                    # 申込にならなかった理由
    improvement = db.Column(db.Text)                        # 改善案


class DailyTaskCheck(db.Model):
    """日報のカスタムタスクチェック"""
    __tablename__ = 'daily_task_check'
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey('daily_report.id'), nullable=False)
    task_id = db.Column(db.Integer, db.ForeignKey('daily_task_template.id'), nullable=False)
    checked = db.Column(db.Boolean, default=False)


class ContractRecord(db.Model):
    """申込・契約台帳（Excelインポートデータ）"""
    __tablename__ = 'contract_record'
    id = db.Column(db.Integer, primary_key=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'))
    staff_id = db.Column(db.Integer, db.ForeignKey('staff.id'))
    staff_name_raw = db.Column(db.String(100))       # Excelのシート名
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    seq_no = db.Column(db.Integer)
    status = db.Column(db.String(50))                # 契約/申込/キャンセル
    application_date = db.Column(db.Date)
    media = db.Column(db.String(100))
    property_name = db.Column(db.String(200))
    room_no = db.Column(db.String(50))
    customer_name = db.Column(db.String(100))
    phone = db.Column(db.String(50))
    rent = db.Column(db.Float)
    management_company = db.Column(db.String(200))
    review_status = db.Column(db.String(100))
    doc_arrival_date = db.Column(db.Date)
    contract_visit_date = db.Column(db.Date)
    settlement_date = db.Column(db.Date)
    contract_start_date = db.Column(db.Date)
    ad_income_date_raw = db.Column(db.String(100))   # テキスト含む
    ad_income_date = db.Column(db.Date)
    commission_pct = db.Column(db.Float)
    other_cost = db.Column(db.Float)
    ad_pct = db.Column(db.Float)
    ad_received = db.Column(db.String(10))           # ○/×
    lifeline = db.Column(db.String(10))
    moving = db.Column(db.String(10))
    fire_insurance = db.Column(db.String(10))
    application_amount = db.Column(db.Float, default=0)
    sales_amount = db.Column(db.Float, default=0)
    contract_amount = db.Column(db.Float, default=0)
    cancel_type = db.Column(db.String(50))
    source_file = db.Column(db.String(500))
    imported_at = db.Column(db.DateTime, default=datetime.utcnow)


class MediaType(db.Model):
    """媒体マスター（申込一覧のプルダウン用）"""
    __tablename__ = 'media_type'
    id = db.Column(db.Integer, primary_key=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), default=1)
    name = db.Column(db.String(100), nullable=False)
    sort_order = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True)


class StatusColor(db.Model):
    """申込ステータス別行カラー設定"""
    __tablename__ = 'status_color'
    id = db.Column(db.Integer, primary_key=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), default=1)
    status_key = db.Column(db.String(50), nullable=False)
    bg_color = db.Column(db.String(20), default='#ffffff')      # バッジ背景色
    text_color = db.Column(db.String(20), default='#111827')    # 文字色
    row_bg_color = db.Column(db.String(20), default='#ffffff')  # 行全体の背景色


class ApplicationRecord(db.Model):
    """申込一覧表"""
    __tablename__ = 'application_record'
    id = db.Column(db.Integer, primary_key=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    staff_id = db.Column(db.Integer, db.ForeignKey('staff.id'), nullable=True)
    application_date = db.Column(db.Date, nullable=False)
    media = db.Column(db.String(100))
    property_name = db.Column(db.String(200))
    room_number = db.Column(db.String(50))
    customer_name = db.Column(db.String(100))
    rent = db.Column(db.Float, default=0)
    contract_start_date = db.Column(db.Date)
    ad_payment_date = db.Column(db.Date)
    brokerage_fee = db.Column(db.Float, default=0)
    ancillary_services = db.Column(db.Text)  # 旧フィールド（廃止・後方互換のため保持）
    option_amount = db.Column(db.Float, default=0)
    ad_type = db.Column(db.String(10), default='amount')  # 'amount' or 'percent'
    ad_amount = db.Column(db.Float, default=0)
    lifeline = db.Column(db.Boolean, default=False)
    moving = db.Column(db.Boolean, default=False)
    fire_insurance = db.Column(db.Boolean, default=False)
    status = db.Column(db.String(50), default='申込')   # 申込/契約/キャンセル/キャンセル振替
    # 入金承認ワークフロー
    ad_settled = db.Column(db.Boolean, default=False)          # 営業がAD入金報告
    ad_approved = db.Column(db.Boolean, default=False)         # 店長がAD承認
    brokerage_settled = db.Column(db.Boolean, default=False)   # 営業が仲介入金報告
    brokerage_approved = db.Column(db.Boolean, default=False)  # 店長が仲介承認
    brokerage_payment_date = db.Column(db.Date, nullable=True) # 仲介入金日
    # その他費用（旧オプション）の入金承認ワークフロー（仲手とは別方向）
    option_settled = db.Column(db.Boolean, default=False)      # 営業がその他費用入金報告
    option_approved = db.Column(db.Boolean, default=False)     # 店長がその他費用承認
    option_payment_date = db.Column(db.Date, nullable=True)    # その他費用入金日
    management_company = db.Column(db.String(200))             # 管理会社名
    review_ng = db.Column(db.Boolean, default=False)          # 審査×（True=審査NG→キャンセル）旧フィールド
    review_status = db.Column(db.String(10), nullable=True)  # 審査状態: None=—, 'ok'=○, 'ng'=×
    past_customer = db.Column(db.Boolean, default=False)     # True=契約終了（顧客管理へ移動）
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class ContractDocument(db.Model):
    """契約書類（取引成立台帳など）の編集データ。申込1件につき1つ。data はJSON文字列。"""
    __tablename__ = 'contract_document'
    id = db.Column(db.Integer, primary_key=True)
    application_id = db.Column(db.Integer, db.ForeignKey('application_record.id'), unique=True, nullable=False)
    store_id = db.Column(db.Integer)
    data = db.Column(db.Text)   # JSON
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class FloorPlan(db.Model):
    """間取り（編集可能なキャンバスデータ。data は fabric.js のJSON）"""
    __tablename__ = 'floor_plan'
    id = db.Column(db.Integer, primary_key=True)
    store_id = db.Column(db.Integer)
    name = db.Column(db.String(200))
    data = db.Column(db.Text)   # fabric.js canvas JSON（背景画像のbase64含む）
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class DocCompanyInfo(db.Model):
    """会社情報（テナント単位で固定。帳票の会社情報タグへ自動差し込み）。data はJSON {ラベル: 値}"""
    __tablename__ = 'doc_company_info'
    id = db.Column(db.Integer, primary_key=True)
    tenant_id = db.Column(db.Integer, index=True, unique=True)
    data = db.Column(db.Text)   # JSON {"会社名":"...", "住所":"..."}
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class DocTemplate(db.Model):
    """クライアントのExcel帳票テンプレート。{{タグ}}で差し込み位置を表現する。"""
    __tablename__ = 'doc_template'
    id = db.Column(db.Integer, primary_key=True)
    tenant_id = db.Column(db.Integer, index=True)
    name = db.Column(db.String(160))            # 帳票名（例: 賃貸借契約書）
    filename = db.Column(db.String(255))        # 元ファイル名
    file_b64 = db.Column(db.Text)               # xlsx本体（base64）
    tags = db.Column(db.Text)                   # JSON: 検出タグ名のリスト
    mapping = db.Column(db.Text)                # JSON: {tag: {"scope":"company|case","company_key":..,"label":..}}
    is_active = db.Column(db.Boolean, default=True)
    created_by = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class EchoRecord(db.Model):
    """反響管理表（追客進捗管理）"""
    __tablename__ = 'echo_record'
    id            = db.Column(db.Integer, primary_key=True)
    store_id      = db.Column(db.Integer, db.ForeignKey('store.id'))
    staff_id      = db.Column(db.Integer, db.ForeignKey('staff.id'), nullable=True)
    list_name     = db.Column(db.String(200))   # リスト名
    echo_date     = db.Column(db.Date)           # 反響日
    media         = db.Column(db.String(100))    # 媒体
    method        = db.Column(db.String(100))    # 手段
    first_contact_date = db.Column(db.Date, nullable=True)  # 初回対応日
    followup_1    = db.Column(db.Date, nullable=True)
    followup_2    = db.Column(db.Date, nullable=True)
    followup_3    = db.Column(db.Date, nullable=True)
    followup_4    = db.Column(db.Date, nullable=True)
    followup_5    = db.Column(db.Date, nullable=True)
    followup_6    = db.Column(db.Date, nullable=True)
    followup_7    = db.Column(db.Date, nullable=True)
    followup_8    = db.Column(db.Date, nullable=True)
    followup_9    = db.Column(db.Date, nullable=True)
    followup_10   = db.Column(db.Date, nullable=True)
    followup_phone = db.Column(db.String(60), default='')  # 電話追客の追客番号（カンマ区切り 例:"1,3"）
    has_reply     = db.Column(db.Boolean, default=False)  # 返信有
    has_phone     = db.Column(db.Boolean, default=False)  # 電話対応有無
    has_line      = db.Column(db.Boolean, default=False)  # LINE追加
    memo          = db.Column(db.Text)
    external_id   = db.Column(db.String(160), nullable=True)  # 反響メール一意ID（重複取込防止）
    customer_email = db.Column(db.String(200), nullable=True) # お客様メール（送信先）
    has_unread_reply = db.Column(db.Boolean, default=False)   # 未読の返信あり
    has_phone_number = db.Column(db.Boolean, default=False)   # 電話番号の有無（〇/×）
    reply_dismissed  = db.Column(db.Boolean, default=False)   # 未返信アラートを「返信不要」として消したか
    status        = db.Column(db.String(40), nullable=True)   # 状況タグ（追客中/申込/終了 など）
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)


class MailMessage(db.Model):
    """反響ごとのメール会話（送信・受信）"""
    __tablename__ = 'mail_message'
    id          = db.Column(db.Integer, primary_key=True)
    store_id    = db.Column(db.Integer, db.ForeignKey('store.id'))
    echo_id     = db.Column(db.Integer, db.ForeignKey('echo_record.id'))
    direction   = db.Column(db.String(4))      # 'out'（送信）/ 'in'（受信）
    from_addr   = db.Column(db.String(300))
    to_addr     = db.Column(db.String(300))
    subject     = db.Column(db.String(500))
    body        = db.Column(db.Text)
    message_id  = db.Column(db.String(300))    # メールのMessage-ID
    in_reply_to = db.Column(db.String(300))
    is_read     = db.Column(db.Boolean, default=True)
    opened_at   = db.Column(db.DateTime, nullable=True)   # 送信メールを相手が開いた日時（既読）
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)


class MailAttachment(db.Model):
    """メールの添付ファイル（PDF/画像など）。本文と一緒に保存。"""
    __tablename__ = 'mail_attachment'
    id           = db.Column(db.Integer, primary_key=True)
    message_id   = db.Column(db.Integer, db.ForeignKey('mail_message.id'))
    store_id     = db.Column(db.Integer, db.ForeignKey('store.id'))
    filename     = db.Column(db.String(300))
    content_type = db.Column(db.String(120))
    size         = db.Column(db.Integer, default=0)
    data         = db.Column(db.LargeBinary)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)


class MailTemplate(db.Model):
    """返信テンプレート（テナント別）。"""
    __tablename__ = 'mail_template'
    id         = db.Column(db.Integer, primary_key=True)
    tenant_id  = db.Column(db.Integer, nullable=True)
    category   = db.Column(db.String(120), default='')   # フォルダ名（例：SUUMO / HOME'S）
    title      = db.Column(db.String(120))
    subject    = db.Column(db.String(300))
    body       = db.Column(db.Text)
    is_html    = db.Column(db.Boolean, default=False)   # 本文がHTMLか（リッチ編集）
    sort_order = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class CompanyProfile(db.Model):
    """会社情報（自動返信メールの差し込み用）。店舗ごと。"""
    __tablename__ = 'company_profile'
    id             = db.Column(db.Integer, primary_key=True)
    store_id       = db.Column(db.Integer, db.ForeignKey('store.id'), unique=True)
    company_name   = db.Column(db.String(200))
    store_name     = db.Column(db.String(200))        # 店舗名
    phone          = db.Column(db.String(60))
    fax            = db.Column(db.String(60))          # FAX番号
    email          = db.Column(db.String(200))
    address        = db.Column(db.String(300))
    representative = db.Column(db.String(100))         # 代表者氏名
    license_number = db.Column(db.String(120))         # 宅建業免許番号
    license_date   = db.Column(db.String(60))          # 宅建業免許取得日
    business_hours = db.Column(db.String(200))
    holidays       = db.Column(db.String(200))
    invoice_number = db.Column(db.String(60))          # インボイス登録番号
    line_url       = db.Column(db.String(300))
    logo_data      = db.Column(db.LargeBinary)        # 店舗ロゴ画像
    logo_type      = db.Column(db.String(80))         # ロゴのMIMEタイプ
    updated_at     = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class ChatChannel(db.Model):
    """社内チャットのチャンネル（全社 / 店舗別 / グループ）"""
    __tablename__ = 'chat_channel'
    id         = db.Column(db.Integer, primary_key=True)
    tenant_id  = db.Column(db.Integer, db.ForeignKey('tenant.id'))
    kind       = db.Column(db.String(10))   # 'company' / 'store' / 'group'
    name       = db.Column(db.String(120))
    store_id   = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=True)  # kind=store のとき
    created_by = db.Column(db.Integer, db.ForeignKey('app_user.id'), nullable=True)
    is_active  = db.Column(db.Boolean, default=True)
    pinned     = db.Column(db.Boolean, default=False)   # グループのピン止め（先頭に表示）
    sort_order = db.Column(db.Integer, default=0)       # グループの並び順
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class ChatMember(db.Model):
    """グループチャンネルのメンバー"""
    __tablename__ = 'chat_member'
    id         = db.Column(db.Integer, primary_key=True)
    channel_id = db.Column(db.Integer, db.ForeignKey('chat_channel.id'))
    user_id    = db.Column(db.Integer, db.ForeignKey('app_user.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class ChatMessage(db.Model):
    """社内チャットのメッセージ"""
    __tablename__ = 'chat_message'
    id         = db.Column(db.Integer, primary_key=True)
    channel_id = db.Column(db.Integer, db.ForeignKey('chat_channel.id'))
    tenant_id  = db.Column(db.Integer, db.ForeignKey('tenant.id'))
    user_id    = db.Column(db.Integer, db.ForeignKey('app_user.id'))
    user_name  = db.Column(db.String(120))   # 表示名スナップショット
    body       = db.Column(db.Text)
    has_attachments = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class ChatAttachment(db.Model):
    """チャットの添付ファイル（チャットProのみ）"""
    __tablename__ = 'chat_attachment'
    id           = db.Column(db.Integer, primary_key=True)
    message_id   = db.Column(db.Integer, db.ForeignKey('chat_message.id'))
    channel_id   = db.Column(db.Integer, db.ForeignKey('chat_channel.id'))
    tenant_id    = db.Column(db.Integer, db.ForeignKey('tenant.id'))
    filename     = db.Column(db.String(300))
    content_type = db.Column(db.String(120))
    size         = db.Column(db.Integer, default=0)
    data         = db.Column(db.LargeBinary)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)


class ChatRead(db.Model):
    """チャンネルごとのユーザー既読位置（last_read_id 以下は既読）"""
    __tablename__ = 'chat_read'
    id           = db.Column(db.Integer, primary_key=True)
    channel_id   = db.Column(db.Integer, db.ForeignKey('chat_channel.id'))
    user_id      = db.Column(db.Integer, db.ForeignKey('app_user.id'))
    last_read_id = db.Column(db.Integer, default=0)
    updated_at   = db.Column(db.DateTime, default=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint('channel_id', 'user_id', name='uq_chat_read_channel_user'),)


class MailSetting(db.Model):
    """店舗ごとの反響メール自動取込設定（IMAP）"""
    __tablename__ = 'mail_setting'
    id               = db.Column(db.Integer, primary_key=True)
    store_id         = db.Column(db.Integer, db.ForeignKey('store.id'), unique=True)
    imap_host        = db.Column(db.String(120), default='imap.gmail.com')
    imap_user        = db.Column(db.String(200))   # 連携するGmailアドレス
    imap_pass        = db.Column(db.String(200))   # アプリパスワード
    enabled          = db.Column(db.Boolean, default=False)  # 自動取込ON/OFF
    default_staff_id = db.Column(db.Integer, nullable=True)  # 取込時のデフォルト担当
    custom_keywords  = db.Column(db.Text)                    # 追加判定キーワード（1行=「語」or「語=媒体名」）
    import_after     = db.Column(db.DateTime, nullable=True) # この日時以降のメールのみ取込（過去分は取らない）
    last_fetch_at    = db.Column(db.DateTime, nullable=True)
    last_result      = db.Column(db.String(300))
    oauth_refresh_token = db.Column(db.Text, nullable=True)     # Google OAuth リフレッシュトークン
    oauth_email      = db.Column(db.String(200), nullable=True) # OAuthで連携したGmailアドレス
    auto_reply_enabled    = db.Column(db.Boolean, default=False)  # 新着反響に自動返信
    auto_reply_template_id = db.Column(db.Integer, nullable=True) # 自動返信に使うテンプレID
    created_at       = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at       = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class ProcessedReaction(db.Model):
    """取込済み（または削除済み）の反響メールID。削除後の再取込（復活）を防ぐ。"""
    __tablename__ = 'processed_reaction'
    id          = db.Column(db.Integer, primary_key=True)
    store_id    = db.Column(db.Integer, db.ForeignKey('store.id'))
    external_id = db.Column(db.String(160))
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)


class PortalSource(db.Model):
    """店舗ごとのポータル登録（差出人アドレス/ドメイン → 媒体名）"""
    __tablename__ = 'portal_source'
    id         = db.Column(db.Integer, primary_key=True)
    store_id   = db.Column(db.Integer, db.ForeignKey('store.id'))
    matcher    = db.Column(db.String(200))   # 差出人アドレス or @ドメイン（部分一致）
    media      = db.Column(db.String(100))   # 媒体名
    enabled    = db.Column(db.Boolean, default=True)
    auto_reply_template_id = db.Column(db.Integer, nullable=True)  # この媒体の自動返信テンプレ
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class CustomerServiceRecord(db.Model):
    """接客管理表"""
    __tablename__ = 'customer_service_record'
    id            = db.Column(db.Integer, primary_key=True)
    store_id      = db.Column(db.Integer, db.ForeignKey('store.id'))
    card_no       = db.Column(db.String(50))    # カードNo
    service_date  = db.Column(db.Date)          # 日付
    echo_media    = db.Column(db.String(100))   # 反響媒体
    staff_id      = db.Column(db.Integer, db.ForeignKey('staff.id'), nullable=True)
    customer_name = db.Column(db.String(100))   # お客様名
    service_type  = db.Column(db.String(50))    # 対応種別
    visit_count   = db.Column(db.Integer, default=0)  # 来店数
    status        = db.Column(db.String(20), default='追客中')  # 追客中/申込/他決/キャンセル
    memo          = db.Column(db.Text)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)


class DropdownOption(db.Model):
    """プルダウン選択肢マスタ（テナント別・カテゴリ別）"""
    __tablename__ = 'dropdown_option'
    id         = db.Column(db.Integer, primary_key=True)
    tenant_id  = db.Column(db.Integer, nullable=True)          # NULL=全テナント共通デフォルト
    category   = db.Column(db.String(50),  nullable=False)     # echo_media / echo_method / cs_media / cs_service_type / leads_media
    value      = db.Column(db.String(100), nullable=False)
    sort_order = db.Column(db.Integer, default=0)


class TrialApplication(db.Model):
    """トライアル申込フォーム送信履歴"""
    __tablename__ = 'trial_application'
    id         = db.Column(db.Integer, primary_key=True)
    company    = db.Column(db.String(200), nullable=False)
    name       = db.Column(db.String(100), nullable=False)
    email      = db.Column(db.String(200), nullable=False)
    phone      = db.Column(db.String(50),  nullable=False)
    stores     = db.Column(db.String(20),  nullable=False)
    message    = db.Column(db.Text,        nullable=True)
    status     = db.Column(db.String(20),  default='new')   # new / contacted / contracted / rejected
    memo       = db.Column(db.Text,        nullable=True)   # 管理者メモ
    created_at = db.Column(db.DateTime,    default=datetime.utcnow)


# ── Excel関連ヘルパー ─────────────────────────────────────

def excel_date_to_date(v):
    """ExcelシリアルまたはdatetimeをPython dateに変換"""
    if v is None:
        return None
    if isinstance(v, (int, float)) and v > 1:
        return (date(1899, 12, 30) + timedelta(days=int(v)))
    if hasattr(v, 'date'):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def to_float_safe(v):
    """安全にfloat変換、失敗時はNone"""
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def to_str_safe(v):
    """安全にstr変換"""
    if v is None:
        return None
    return str(v).strip() if str(v).strip() else None


def import_excel_file(file_path_or_stream, year, month, store_id=1):
    """
    Excelファイルを読んでContractRecordとSalesKPIを更新する。
    戻り値: {"imported": N, "staff_results": [...], "errors": [...]}
    """
    try:
        import openpyxl
    except ImportError:
        return {"imported": 0, "staff_results": [], "errors": ["openpyxl がインストールされていません"]}

    errors = []
    staff_results = []
    total_imported = 0

    try:
        if isinstance(file_path_or_stream, str):
            wb = openpyxl.load_workbook(file_path_or_stream, data_only=True)
        else:
            wb = openpyxl.load_workbook(file_path_or_stream, data_only=True)
    except Exception as e:
        return {"imported": 0, "staff_results": [], "errors": [f"Excelファイル読み込みエラー: {str(e)}"]}

    # テンプレート・集計シートは除外
    SKIP_SHEETS = {'原本', 'Sheet1', 'Sheet2', 'Sheet3', '集計', 'サマリー', '目標', 'テンプレート'}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        staff_name = sheet_name.strip()

        # テンプレートシートをスキップ
        if staff_name in SKIP_SHEETS or staff_name.startswith('Sheet'):
            continue

        # スタッフ検索・新規作成
        staff = Staff.query.filter_by(name=staff_name).first()
        if not staff:
            staff = Staff(name=staff_name, store_id=store_id, role='営業', is_active=True)
            db.session.add(staff)
            db.session.flush()

        # 同月の既存ContractRecordを削除（上書き）
        ContractRecord.query.filter_by(
            staff_id=staff.id, year=year, month=month
        ).delete()
        db.session.flush()

        sheet_imported = 0
        sheet_errors = []

        # 行12からデータ開始、2行1セット
        max_row = ws.max_row
        r = 12
        while r <= max_row:
            try:
                # 奇数行（メイン行）
                odd_row = r
                even_row = r + 1

                # 顧客名チェック（列7）
                customer_name = to_str_safe(ws.cell(odd_row, 7).value)
                if not customer_name:
                    r += 2
                    continue

                # ステータスチェック（列2）
                status_raw = to_str_safe(ws.cell(odd_row, 2).value)
                if not status_raw:
                    r += 2
                    continue

                # ---- 奇数行データ ----
                seq_no_v = ws.cell(odd_row, 1).value
                seq_no = int(seq_no_v) if seq_no_v and str(seq_no_v).isdigit() else None

                application_date = excel_date_to_date(ws.cell(odd_row, 4).value)
                property_name = to_str_safe(ws.cell(odd_row, 6).value)
                rent_v = to_float_safe(ws.cell(odd_row, 8).value)
                management_company = to_str_safe(ws.cell(odd_row, 9).value)
                commission_pct = to_float_safe(ws.cell(odd_row, 18).value)
                other_cost = to_float_safe(ws.cell(odd_row, 19).value)
                ad_pct = to_float_safe(ws.cell(odd_row, 20).value)
                ad_received = to_str_safe(ws.cell(odd_row, 21).value)
                lifeline = to_str_safe(ws.cell(odd_row, 22).value)
                moving = to_str_safe(ws.cell(odd_row, 23).value)
                fire_insurance = to_str_safe(ws.cell(odd_row, 24).value)
                application_amount = to_float_safe(ws.cell(odd_row, 34).value) or 0
                sales_amount_v = to_float_safe(ws.cell(odd_row, 35).value) or 0
                contract_amount = to_float_safe(ws.cell(odd_row, 36).value) or 0
                cancel_type = to_str_safe(ws.cell(odd_row, 48).value)

                # ---- 偶数行データ ----
                source_raw = to_str_safe(ws.cell(even_row, 4).value)   # 接客ソース
                media = to_str_safe(ws.cell(even_row, 5).value)
                room_no = to_str_safe(ws.cell(even_row, 6).value)
                phone = to_str_safe(ws.cell(even_row, 7).value)
                review_status = to_str_safe(ws.cell(even_row, 9).value)
                doc_arrival_date = excel_date_to_date(ws.cell(even_row, 10).value)
                contract_visit_date = excel_date_to_date(ws.cell(even_row, 11).value)
                settlement_date = excel_date_to_date(ws.cell(even_row, 12).value)
                contract_start_date = excel_date_to_date(ws.cell(even_row, 14).value)

                # AD入金日（テキスト含む場合あり）
                ad_date_raw_v = ws.cell(even_row, 15).value
                ad_income_date_raw = to_str_safe(ad_date_raw_v)
                ad_income_date = excel_date_to_date(ad_date_raw_v)

                record = ContractRecord(
                    store_id=store_id,
                    staff_id=staff.id,
                    staff_name_raw=staff_name,
                    year=year,
                    month=month,
                    seq_no=seq_no,
                    status=status_raw,
                    application_date=application_date,
                    media=media,
                    property_name=property_name,
                    room_no=room_no,
                    customer_name=customer_name,
                    phone=phone,
                    rent=rent_v,
                    management_company=management_company,
                    review_status=review_status,
                    doc_arrival_date=doc_arrival_date,
                    contract_visit_date=contract_visit_date,
                    settlement_date=settlement_date,
                    contract_start_date=contract_start_date,
                    ad_income_date_raw=ad_income_date_raw,
                    ad_income_date=ad_income_date,
                    commission_pct=commission_pct,
                    other_cost=other_cost,
                    ad_pct=ad_pct,
                    ad_received=ad_received,
                    lifeline=lifeline,
                    moving=moving,
                    fire_insurance=fire_insurance,
                    application_amount=application_amount,
                    sales_amount=sales_amount_v,
                    contract_amount=contract_amount,
                    cancel_type=cancel_type,
                )
                db.session.add(record)
                sheet_imported += 1

            except Exception as e:
                sheet_errors.append(f"行{r}: {str(e)}")

            r += 2

        db.session.flush()

        # シートごとのKPI集計
        records_this_sheet = ContractRecord.query.filter_by(
            staff_id=staff.id, year=year, month=month
        ).all()

        applications_cnt = sum(1 for rc in records_this_sheet if rc.status in ('申込', '契約'))
        contracts_cnt = sum(1 for rc in records_this_sheet if rc.status == '契約')
        cancellations_cnt = sum(1 for rc in records_this_sheet if rc.status == 'キャンセル')
        sales_amt = sum(
            (rc.contract_amount if rc.contract_amount else (rc.rent or 0))
            for rc in records_this_sheet if rc.status == '契約'
        )
        option_sales = (
            sum(1 for rc in records_this_sheet if rc.lifeline and '○' in rc.lifeline) * 5000 +
            sum(1 for rc in records_this_sheet if rc.moving and '○' in rc.moving) * 3000
        )

        # KPI更新
        kpi = SalesKPI.query.filter_by(staff_id=staff.id, store_id=store_id,
                                        year=year, month=month).first()
        if not kpi:
            kpi = SalesKPI(staff_id=staff.id, store_id=store_id, year=year, month=month)
            db.session.add(kpi)
        kpi.applications = applications_cnt
        kpi.contracts = contracts_cnt
        kpi.cancellations = cancellations_cnt
        kpi.sales_amount = sales_amt
        kpi.option_sales = float(option_sales)
        db.session.flush()

        total_imported += sheet_imported
        staff_results.append({
            'staff_name': staff_name,
            'staff_id': staff.id,
            'imported': sheet_imported,
            'contracts': contracts_cnt,
            'applications': applications_cnt,
            'cancellations': cancellations_cnt,
            'errors': sheet_errors,
        })

    db.session.commit()
    return {
        'imported': total_imported,
        'staff_results': staff_results,
        'errors': errors,
    }


# ── 初期店舗・スタッフ作成関数 ────────────────────────────

def init_store():
    """
    Storeテーブルが空のときだけデフォルトテナント・店舗・スタッフを作成する。
    """
    if Store.query.count() > 0:
        return

    print("初期テナント・店舗・スタッフを作成しています...")

    # デフォルトテナント
    tenant = Tenant.query.first()
    if not tenant:
        tenant = Tenant(name='ミエルーム', plan='standard', is_active=True)
        db.session.add(tenant)
        db.session.flush()

    store = Store(name='ミエルーム', is_active=True, tenant_id=tenant.id)
    db.session.add(store)
    db.session.flush()

    for i in range(1, 4):
        staff = Staff(name=f'スタッフ{i}', store_id=store.id, role='営業', is_active=True)
        db.session.add(staff)

    db.session.commit()
    init_default_media_types(store.id)
    print("初期テナント・店舗・スタッフの作成が完了しました。")

    # 初期オーナーアカウント作成
    if AppUser.query.count() == 0:
        owner = AppUser(
            username='owner',
            password_hash=generate_password_hash('roompick2024'),
            role='owner',
            tenant_id=tenant.id,
        )
        db.session.add(owner)
        db.session.commit()
        print("初期オーナーアカウントを作成しました。(username: owner)")
    return


def ensure_super_admin():
    """super_adminアカウントを確実に作成・維持する"""
    sa = AppUser.query.filter_by(username='super_admin').first()
    if not sa:
        sa = AppUser(
            username='super_admin',
            email='teneramente0701@gmail.com',
            password_hash=generate_password_hash('SuperAdmin2024!'),
            role='super_admin',
            tenant_id=None,
            is_active=True,
        )
        db.session.add(sa)
        db.session.commit()
        print("super_adminアカウントを作成しました (username: super_admin)")
    elif sa.role != 'super_admin':
        sa.role = 'super_admin'
        sa.tenant_id = None
        sa.is_active = True
        db.session.commit()
        print("super_adminのロールを修正しました")


def init_default_media_types(store_id):
    """デフォルト媒体マスターを初期化（新規店舗時）"""
    defaults = ['SUUMO', "HOME'S", 'at home', 'カナリー', 'スモッカ', 'HP', 'SNS', 'リピート', '紹介', '飛び込み', '直電話問合せ']
    for i, name in enumerate(defaults):
        if not MediaType.query.filter_by(store_id=store_id, name=name).first():
            db.session.add(MediaType(store_id=store_id, name=name, sort_order=i, is_active=True))
    db.session.commit()


def ensure_owner_account():
    """AppUserが存在しない場合にオーナーアカウントを作成する"""
    if AppUser.query.count() == 0:
        tenant = Tenant.query.first()
        owner = AppUser(
            username='owner',
            password_hash=generate_password_hash('roompick2024'),
            role='owner',
            tenant_id=tenant.id if tenant else None,
        )
        db.session.add(owner)
        db.session.commit()
        print("初期オーナーアカウントを作成しました。(username: owner)")


# ── DB初期化 & 初期データ作成 ──────────────────────────────

def migrate_db():
    """
    既存DBに新カラムを追加するマイグレーション。
    SQLiteはALTER TABLE ADD COLUMN をサポートしている。
    """
    import sqlite3
    db_path = os.path.join(app.instance_path, 'realestate.db')
    if not os.path.exists(db_path):
        return

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # pl_record テーブルの新カラムを追加
    pl_new_columns = [
        ('brokerage_fee', 'FLOAT DEFAULT 0'),
        ('ad_income', 'FLOAT DEFAULT 0'),
        ('lifeline_income', 'FLOAT DEFAULT 0'),
        ('moving_income', 'FLOAT DEFAULT 0'),
        ('fire_insurance_income', 'FLOAT DEFAULT 0'),
        ('other_income', 'FLOAT DEFAULT 0'),
        ('suumo_cost', 'FLOAT DEFAULT 0'),
        ('homes_cost', 'FLOAT DEFAULT 0'),
        ('athome_cost', 'FLOAT DEFAULT 0'),
        ('instagram_cost', 'FLOAT DEFAULT 0'),
        ('tiktok_cost', 'FLOAT DEFAULT 0'),
        ('google_ads_cost', 'FLOAT DEFAULT 0'),
        ('line_cost', 'FLOAT DEFAULT 0'),
        ('hp_cost', 'FLOAT DEFAULT 0'),
        ('meo_cost', 'FLOAT DEFAULT 0'),
        ('other_ad_cost', 'FLOAT DEFAULT 0'),
        ('regular_salary', 'FLOAT DEFAULT 0'),
        ('parttime_salary', 'FLOAT DEFAULT 0'),
        ('commission_pay', 'FLOAT DEFAULT 0'),
        ('pl_rent', 'FLOAT DEFAULT 0'),
        ('pl_parking', 'FLOAT DEFAULT 0'),
        ('pl_copier', 'FLOAT DEFAULT 0'),
        ('pl_internet', 'FLOAT DEFAULT 0'),
        ('pl_consultant', 'FLOAT DEFAULT 0'),
        ('pl_insurance', 'FLOAT DEFAULT 0'),
        ('pl_cloud', 'FLOAT DEFAULT 0'),
    ]

    # 既存カラムを取得
    cursor.execute("PRAGMA table_info(pl_record)")
    existing = {row[1] for row in cursor.fetchall()}

    for col_name, col_def in pl_new_columns:
        if col_name not in existing:
            try:
                cursor.execute(f"ALTER TABLE pl_record ADD COLUMN {col_name} {col_def}")
                print(f"  Added column pl_record.{col_name}")
            except Exception as e:
                print(f"  Skip {col_name}: {e}")

    # pl_custom_item / pl_custom_value の item_type カラム追加
    for tbl in ['pl_custom_item', 'pl_custom_value']:
        try:
            cursor.execute(f"PRAGMA table_info({tbl})")
            cols = {r[1] for r in cursor.fetchall()}
            if 'item_type' not in cols:
                cursor.execute(f"ALTER TABLE {tbl} ADD COLUMN item_type TEXT DEFAULT '固定費'")
                print(f"  Added column {tbl}.item_type")
        except Exception as e:
            print(f"  Skip {tbl}.item_type: {e}")

    # sales_kpi テーブルの新カラムを追加
    sales_kpi_new_cols = [
        ('estimated_sales',      'FLOAT DEFAULT 0'),
        ('target_sales',         'FLOAT DEFAULT 0'),
        ('fire_insurance_count', 'INTEGER DEFAULT 0'),
        ('lifeline_count',       'INTEGER DEFAULT 0'),
        ('moving_count',         'INTEGER DEFAULT 0'),
    ]
    cursor.execute("PRAGMA table_info(sales_kpi)")
    sk_existing = {row[1] for row in cursor.fetchall()}
    for col_name, col_def in sales_kpi_new_cols:
        if col_name not in sk_existing:
            try:
                cursor.execute(f"ALTER TABLE sales_kpi ADD COLUMN {col_name} {col_def}")
                print(f"  Added column sales_kpi.{col_name}")
            except Exception as e:
                print(f"  Skip sales_kpi.{col_name}: {e}")

    # app_user / store へ tenant_id カラムを追加
    for tbl in ['store', 'app_user']:
        cursor.execute(f"PRAGMA table_info({tbl})")
        cols = {r[1] for r in cursor.fetchall()}
        if 'tenant_id' not in cols:
            try:
                cursor.execute(f"ALTER TABLE {tbl} ADD COLUMN tenant_id INTEGER")
                print(f"  Added column {tbl}.tenant_id")
            except Exception as e:
                print(f"  Skip {tbl}.tenant_id: {e}")

    # store / tenant の後続カラム（migrate_postgres と整合させる）
    for tbl, extra_cols in [
        ('store', [
            ('is_locked', 'INTEGER DEFAULT 0'),
            ('contract_start_date', 'DATE'),
            ('created_at', 'TIMESTAMP'),
        ]),
        ('tenant', [
            ('trial_ends_at', 'TIMESTAMP'),
            ('subscription_status', "VARCHAR(20) DEFAULT 'trial'"),
            ('contract_start_date', 'DATE'),
        ]),
    ]:
        cursor.execute(f"PRAGMA table_info({tbl})")
        existing = {r[1] for r in cursor.fetchall()}
        for col_name, col_def in extra_cols:
            if col_name not in existing:
                try:
                    cursor.execute(f"ALTER TABLE {tbl} ADD COLUMN {col_name} {col_def}")
                    print(f"  Added column {tbl}.{col_name}")
                except Exception as e:
                    print(f"  Skip {tbl}.{col_name}: {e}")

    # app_user の権限フラグカラムを追加
    cursor.execute("PRAGMA table_info(app_user)")
    au_cols = {r[1] for r in cursor.fetchall()}
    for col_name, col_def in [
        ('can_view_accounting',    'INTEGER DEFAULT 1'),
        ('can_view_all_staff',     'INTEGER DEFAULT 1'),
        ('can_edit_kpi',           'INTEGER DEFAULT 1'),
        ('can_manage_uncollected', 'INTEGER DEFAULT 1'),
        ('can_view_executive',     'INTEGER DEFAULT 1'),
        ('can_view_leads_page',    'INTEGER DEFAULT 1'),
        ('can_view_daily_report',  'INTEGER DEFAULT 1'),
        ('can_view_leave',         'INTEGER DEFAULT 1'),
        ('admin_can_add_tenant',    'INTEGER DEFAULT 0'),
        ('admin_can_manage_stores', 'INTEGER DEFAULT 0'),
        ('admin_can_delete_tenant', 'INTEGER DEFAULT 0'),
        ('admin_can_lock_tenant',   'INTEGER DEFAULT 0'),
    ]:
        if col_name not in au_cols:
            try:
                cursor.execute(f"ALTER TABLE app_user ADD COLUMN {col_name} {col_def}")
                print(f"  Added column app_user.{col_name}")
            except Exception as e:
                print(f"  Skip app_user.{col_name}: {e}")

    # application_record の新カラムを追加
    cursor.execute("PRAGMA table_info(application_record)")
    ar_cols = {r[1] for r in cursor.fetchall()}
    for col_name, col_def in [
        ('option_amount', 'FLOAT DEFAULT 0'),
        ('option_settled', 'BOOLEAN DEFAULT 0'),
        ('option_approved', 'BOOLEAN DEFAULT 0'),
        ('option_payment_date', 'DATE'),
        ('brokerage_payment_date', 'DATE'),
        ('management_company', 'VARCHAR(200)'),
        ('review_ng', 'BOOLEAN DEFAULT 0'),
        ('review_status', 'VARCHAR(10)'),
        ('past_customer', 'BOOLEAN DEFAULT 0'),
    ]:
        if col_name not in ar_cols:
            try:
                cursor.execute(f"ALTER TABLE application_record ADD COLUMN {col_name} {col_def}")
                print(f"  Added column application_record.{col_name}")
            except Exception as e:
                print(f"  Skip application_record.{col_name}: {e}")

    # status_color に row_bg_color を追加
    cursor.execute("PRAGMA table_info(status_color)")
    scs_cols = {r[1] for r in cursor.fetchall()}
    if 'row_bg_color' not in scs_cols:
        try:
            cursor.execute("ALTER TABLE status_color ADD COLUMN row_bg_color VARCHAR(20) DEFAULT '#ffffff'")
            print("  Added column status_color.row_bg_color")
        except Exception as e:
            print(f"  Skip status_color.row_bg_color: {e}")

    # customer_service_record の status カラムを追加
    cursor.execute("PRAGMA table_info(customer_service_record)")
    csr_cols = {r[1] for r in cursor.fetchall()}
    if 'status' not in csr_cols:
        try:
            cursor.execute("ALTER TABLE customer_service_record ADD COLUMN status VARCHAR(20) DEFAULT '追客中'")
            print("  Added column customer_service_record.status")
        except Exception as e:
            print(f"  Skip customer_service_record.status: {e}")

    # daily_report の store_id カラムを追加
    cursor.execute("PRAGMA table_info(daily_report)")
    dr_cols = {r[1] for r in cursor.fetchall()}
    if 'store_id' not in dr_cols:
        try:
            cursor.execute("ALTER TABLE daily_report ADD COLUMN store_id INTEGER")
            print("  Added column daily_report.store_id")
        except Exception as e:
            print(f"  Skip daily_report.store_id: {e}")

    # echo_record の external_id カラムを追加（反響メール重複取込防止）
    cursor.execute("PRAGMA table_info(echo_record)")
    er_cols = {r[1] for r in cursor.fetchall()}
    if 'external_id' not in er_cols:
        try:
            cursor.execute("ALTER TABLE echo_record ADD COLUMN external_id VARCHAR(160)")
            print("  Added column echo_record.external_id")
        except Exception as e:
            print(f"  Skip echo_record.external_id: {e}")
    if 'customer_email' not in er_cols:
        try:
            cursor.execute("ALTER TABLE echo_record ADD COLUMN customer_email VARCHAR(200)")
            print("  Added column echo_record.customer_email")
        except Exception as e:
            print(f"  Skip echo_record.customer_email: {e}")
    if 'has_unread_reply' not in er_cols:
        try:
            cursor.execute("ALTER TABLE echo_record ADD COLUMN has_unread_reply BOOLEAN DEFAULT 0")
            print("  Added column echo_record.has_unread_reply")
        except Exception as e:
            print(f"  Skip echo_record.has_unread_reply: {e}")
    if 'has_phone_number' not in er_cols:
        try:
            cursor.execute("ALTER TABLE echo_record ADD COLUMN has_phone_number BOOLEAN DEFAULT 0")
            print("  Added column echo_record.has_phone_number")
        except Exception as e:
            print(f"  Skip echo_record.has_phone_number: {e}")
    if 'reply_dismissed' not in er_cols:
        try:
            cursor.execute("ALTER TABLE echo_record ADD COLUMN reply_dismissed BOOLEAN DEFAULT 0")
            print("  Added column echo_record.reply_dismissed")
        except Exception as e:
            print(f"  Skip echo_record.reply_dismissed: {e}")
    if 'status' not in er_cols:
        try:
            cursor.execute("ALTER TABLE echo_record ADD COLUMN status VARCHAR(40)")
            print("  Added column echo_record.status")
        except Exception as e:
            print(f"  Skip echo_record.status: {e}")

    # mail_message の opened_at カラムを追加（既読トラッキング）
    try:
        cursor.execute("PRAGMA table_info(mail_message)")
        mmcols = {r[1] for r in cursor.fetchall()}
        if mmcols and 'opened_at' not in mmcols:
            cursor.execute("ALTER TABLE mail_message ADD COLUMN opened_at DATETIME")
            print("  Added column mail_message.opened_at")
    except Exception as e:
        print(f"  Skip mail_message.opened_at: {e}")

    # mail_setting の custom_keywords カラムを追加
    try:
        cursor.execute("PRAGMA table_info(mail_setting)")
        mscols = {r[1] for r in cursor.fetchall()}
        if mscols and 'custom_keywords' not in mscols:
            cursor.execute("ALTER TABLE mail_setting ADD COLUMN custom_keywords TEXT")
            print("  Added column mail_setting.custom_keywords")
        if mscols and 'import_after' not in mscols:
            cursor.execute("ALTER TABLE mail_setting ADD COLUMN import_after DATETIME")
            print("  Added column mail_setting.import_after")
        if mscols and 'oauth_refresh_token' not in mscols:
            cursor.execute("ALTER TABLE mail_setting ADD COLUMN oauth_refresh_token TEXT")
            print("  Added column mail_setting.oauth_refresh_token")
        if mscols and 'oauth_email' not in mscols:
            cursor.execute("ALTER TABLE mail_setting ADD COLUMN oauth_email VARCHAR(200)")
            print("  Added column mail_setting.oauth_email")
        if mscols and 'auto_reply_enabled' not in mscols:
            cursor.execute("ALTER TABLE mail_setting ADD COLUMN auto_reply_enabled BOOLEAN DEFAULT 0")
            print("  Added column mail_setting.auto_reply_enabled")
        if mscols and 'auto_reply_template_id' not in mscols:
            cursor.execute("ALTER TABLE mail_setting ADD COLUMN auto_reply_template_id INTEGER")
            print("  Added column mail_setting.auto_reply_template_id")
    except Exception as e:
        print(f"  Skip mail_setting columns: {e}")

    # portal_source の auto_reply_template_id カラムを追加
    try:
        cursor.execute("PRAGMA table_info(portal_source)")
        pscols = {r[1] for r in cursor.fetchall()}
        if pscols and 'auto_reply_template_id' not in pscols:
            cursor.execute("ALTER TABLE portal_source ADD COLUMN auto_reply_template_id INTEGER")
            print("  Added column portal_source.auto_reply_template_id")
    except Exception as e:
        print(f"  Skip portal_source.auto_reply_template_id: {e}")

    # company_profile の logo カラムを追加
    try:
        cursor.execute("PRAGMA table_info(company_profile)")
        cpcols = {r[1] for r in cursor.fetchall()}
        if cpcols and 'logo_data' not in cpcols:
            cursor.execute("ALTER TABLE company_profile ADD COLUMN logo_data BLOB")
            print("  Added column company_profile.logo_data")
        if cpcols and 'logo_type' not in cpcols:
            cursor.execute("ALTER TABLE company_profile ADD COLUMN logo_type VARCHAR(80)")
            print("  Added column company_profile.logo_type")
        for col, typ in [('store_name', 'VARCHAR(200)'), ('fax', 'VARCHAR(60)'),
                         ('representative', 'VARCHAR(100)'), ('license_number', 'VARCHAR(120)'),
                         ('license_date', 'VARCHAR(60)'), ('invoice_number', 'VARCHAR(60)')]:
            if cpcols and col not in cpcols:
                cursor.execute(f"ALTER TABLE company_profile ADD COLUMN {col} {typ}")
                print(f"  Added column company_profile.{col}")
    except Exception as e:
        print(f"  Skip company_profile columns: {e}")

    # mail_template の is_html / category カラムを追加
    try:
        cursor.execute("PRAGMA table_info(mail_template)")
        mtcols = {r[1] for r in cursor.fetchall()}
        if mtcols and 'is_html' not in mtcols:
            cursor.execute("ALTER TABLE mail_template ADD COLUMN is_html BOOLEAN DEFAULT 0")
            print("  Added column mail_template.is_html")
        if mtcols and 'category' not in mtcols:
            cursor.execute("ALTER TABLE mail_template ADD COLUMN category VARCHAR(120) DEFAULT ''")
            print("  Added column mail_template.category")
    except Exception as e:
        print(f"  Skip mail_template columns: {e}")

    # mail_template の category（フォルダ）カラムを追加
    try:
        cursor.execute("PRAGMA table_info(mail_template)")
        mtcols = {r[1] for r in cursor.fetchall()}
        if mtcols and 'category' not in mtcols:
            cursor.execute("ALTER TABLE mail_template ADD COLUMN category VARCHAR(120) DEFAULT ''")
            print("  Added column mail_template.category")
    except Exception as e:
        print(f"  Skip mail_template.category: {e}")

    # echo_record の followup_phone（電話追客フラグ）カラムを追加
    try:
        cursor.execute("PRAGMA table_info(echo_record)")
        ercols = {r[1] for r in cursor.fetchall()}
        if ercols and 'followup_phone' not in ercols:
            cursor.execute("ALTER TABLE echo_record ADD COLUMN followup_phone VARCHAR(60) DEFAULT ''")
            print("  Added column echo_record.followup_phone")
    except Exception as e:
        print(f"  Skip echo_record.followup_phone: {e}")

    # tenant_option の store_id（店舗別オプション）カラムを追加
    try:
        cursor.execute("PRAGMA table_info(tenant_option)")
        tocols = {r[1] for r in cursor.fetchall()}
        if tocols and 'store_id' not in tocols:
            cursor.execute("ALTER TABLE tenant_option ADD COLUMN store_id INTEGER")
            print("  Added column tenant_option.store_id")
    except Exception as e:
        print(f"  Skip tenant_option.store_id: {e}")

    # chat_channel の pinned / sort_order カラムを追加
    try:
        cursor.execute("PRAGMA table_info(chat_channel)")
        ccols = {r[1] for r in cursor.fetchall()}
        if ccols and 'pinned' not in ccols:
            cursor.execute("ALTER TABLE chat_channel ADD COLUMN pinned BOOLEAN DEFAULT 0")
            print("  Added column chat_channel.pinned")
        if ccols and 'sort_order' not in ccols:
            cursor.execute("ALTER TABLE chat_channel ADD COLUMN sort_order INTEGER DEFAULT 0")
            print("  Added column chat_channel.sort_order")
    except Exception as e:
        print(f"  Skip chat_channel columns: {e}")

    conn.commit()
    conn.close()


def migrate_postgres():
    """PostgreSQL用: 新カラムが存在しない場合のみALTER TABLEで追加
    各カラムを独立したコネクションで処理し、1つの失敗が他に影響しないようにする"""
    new_cols = [
        ("sales_kpi",          "estimated_sales",         "FLOAT DEFAULT 0"),
        ("sales_kpi",          "target_sales",            "FLOAT DEFAULT 0"),
        ("sales_kpi",          "fire_insurance_count",    "INTEGER DEFAULT 0"),
        ("sales_kpi",          "lifeline_count",          "INTEGER DEFAULT 0"),
        ("sales_kpi",          "moving_count",            "INTEGER DEFAULT 0"),
        ("store",              "tenant_id",               "INTEGER"),
        ("app_user",           "tenant_id",               "INTEGER"),
        ("app_user",           "email",                   "VARCHAR(200)"),
        ("app_user",           "store_id",                "INTEGER"),
        ("app_user",           "last_login",              "TIMESTAMP"),
        ("app_user",           "can_view_accounting",     "BOOLEAN DEFAULT TRUE"),
        ("app_user",           "can_view_all_staff",      "BOOLEAN DEFAULT TRUE"),
        ("app_user",           "can_edit_kpi",            "BOOLEAN DEFAULT TRUE"),
        ("app_user",           "can_manage_uncollected",  "BOOLEAN DEFAULT TRUE"),
        ("app_user",           "can_view_executive",      "BOOLEAN DEFAULT TRUE"),
        ("app_user",           "can_view_leads_page",     "BOOLEAN DEFAULT TRUE"),
        ("app_user",           "can_view_daily_report",   "BOOLEAN DEFAULT TRUE"),
        ("app_user",           "can_view_leave",          "BOOLEAN DEFAULT TRUE"),
        ("application_record",        "option_amount", "FLOAT DEFAULT 0"),
        ("application_record", "brokerage_payment_date", "DATE"),
        ("application_record", "option_settled", "BOOLEAN DEFAULT FALSE"),
        ("application_record", "option_approved", "BOOLEAN DEFAULT FALSE"),
        ("application_record", "option_payment_date", "DATE"),
        ("application_record", "management_company", "VARCHAR(200)"),
        ("application_record", "review_ng", "BOOLEAN DEFAULT FALSE"),
        ("application_record", "review_status", "VARCHAR(10)"),
        ("application_record", "past_customer", "BOOLEAN DEFAULT FALSE"),
        ("status_color", "row_bg_color", "VARCHAR(20) DEFAULT '#ffffff'"),
        ("daily_report",             "store_id",     "INTEGER"),
        ("customer_service_record",  "status",       "VARCHAR(20) DEFAULT '追客中'"),
        ("echo_record",              "external_id",  "VARCHAR(160)"),
        ("echo_record",              "customer_email", "VARCHAR(200)"),
        ("echo_record",              "has_unread_reply", "BOOLEAN DEFAULT FALSE"),
        ("echo_record",              "has_phone_number", "BOOLEAN DEFAULT FALSE"),
        ("echo_record",              "reply_dismissed",  "BOOLEAN DEFAULT FALSE"),
        ("echo_record",              "status",          "VARCHAR(40)"),
        ("echo_record",              "followup_phone",  "VARCHAR(60) DEFAULT ''"),
        ("tenant_option",            "store_id",        "INTEGER"),
        ("chat_channel",             "pinned",          "BOOLEAN DEFAULT FALSE"),
        ("chat_channel",             "sort_order",      "INTEGER DEFAULT 0"),
        ("mail_message",             "opened_at",       "TIMESTAMP"),
        ("mail_setting",             "custom_keywords", "TEXT"),
        ("mail_setting",             "import_after",    "TIMESTAMP"),
        ("mail_setting",             "oauth_refresh_token", "TEXT"),
        ("mail_setting",             "oauth_email",     "VARCHAR(200)"),
        ("mail_setting",             "auto_reply_enabled", "BOOLEAN DEFAULT FALSE"),
        ("mail_setting",             "auto_reply_template_id", "INTEGER"),
        ("portal_source",            "auto_reply_template_id", "INTEGER"),
        ("mail_template",            "is_html",         "BOOLEAN DEFAULT FALSE"),
        ("mail_template",            "category",        "VARCHAR(120) DEFAULT ''"),
        ("company_profile",          "logo_data",       "BYTEA"),
        ("company_profile",          "logo_type",       "VARCHAR(80)"),
        ("company_profile",          "store_name",      "VARCHAR(200)"),
        ("company_profile",          "fax",             "VARCHAR(60)"),
        ("company_profile",          "representative",  "VARCHAR(100)"),
        ("company_profile",          "license_number",  "VARCHAR(120)"),
        ("company_profile",          "license_date",    "VARCHAR(60)"),
        ("company_profile",          "invoice_number",  "VARCHAR(60)"),
        ("mail_template",            "category",        "VARCHAR(120) DEFAULT ''"),
        ("tenant", "trial_ends_at",        "TIMESTAMP"),
        ("tenant", "subscription_status",  "VARCHAR(20) DEFAULT 'trial'"),
        ("tenant", "contract_start_date",     "DATE"),
        ("store",  "created_at",             "TIMESTAMP"),
        ("store",  "is_locked",              "BOOLEAN DEFAULT FALSE"),
        ("store",  "contract_start_date",    "DATE"),
        ("app_user", "admin_can_add_tenant",    "BOOLEAN DEFAULT FALSE"),
        ("app_user", "admin_can_manage_stores", "BOOLEAN DEFAULT FALSE"),
        ("app_user", "admin_can_delete_tenant", "BOOLEAN DEFAULT FALSE"),
        ("app_user", "admin_can_lock_tenant",   "BOOLEAN DEFAULT FALSE"),
    ]
    # 各カラムを独立した接続で追加（1つの失敗が他に波及しない）
    for tbl, col, typedef in new_cols:
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text(
                    f"ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS {col} {typedef}"
                ))
                conn.commit()
                print(f"PG migrate OK: {tbl}.{col}")
        except Exception as e:
            print(f"PG migrate skip {tbl}.{col}: {e}")


_DROPDOWN_DEFAULTS = {
    'echo_media':      ['SUUMO', "HOME'S", 'アットホーム', 'カナリー', 'Instagram', 'TikTok', '自社HP', '電話', 'SNS', '紹介', 'その他'],
    'echo_method':     ['メール', '電話', 'LINE', 'チャット', 'その他'],
    'echo_status':     ['追客中', '申込', '終了'],
    'cs_media':        ['SUUMO', "HOME'S", 'アットホーム', 'カナリー', 'Instagram', 'TikTok', '自社HP', '電話', 'SNS', '紹介', 'その他'],
    'cs_service_type': ['来店', '電話', 'メール', 'オンライン', 'LINE', 'その他'],
    'cs_status':       ['追客中', '申込', '他決', 'キャンセル'],
    'leads_media':     ['SUUMO', "HOME'S", 'アットホーム', 'カナリー', 'Instagram', 'TikTok', '自社HP', '電話', 'SNS', '紹介', 'その他'],
}


def dedupe_dropdown_options():
    """重複したプルダウン選択肢を除去（複数ワーカーの同時seedで二重挿入された分）。
    (tenant_id, category, value) が同じものは最小idのみ残す。冪等。"""
    try:
        seen = set()
        dup_ids = []
        for o in DropdownOption.query.order_by(DropdownOption.id).all():
            key = (o.tenant_id, o.category, o.value)
            if key in seen:
                dup_ids.append(o.id)
            else:
                seen.add(key)
        if dup_ids:
            DropdownOption.query.filter(DropdownOption.id.in_(dup_ids)).delete(synchronize_session=False)
            db.session.commit()
            print(f"dropdown dedupe: removed {len(dup_ids)} duplicate(s)")
    except Exception as e:
        db.session.rollback()
        print(f"dedupe_dropdown_options error: {e}")


def seed_dropdown_defaults():
    """各カテゴリにデフォルト選択肢が無い場合のみ挿入"""
    try:
        for cat, values in _DROPDOWN_DEFAULTS.items():
            if DropdownOption.query.filter_by(category=cat, tenant_id=None).count() == 0:
                for i, v in enumerate(values):
                    db.session.add(DropdownOption(tenant_id=None, category=cat, value=v, sort_order=i))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"seed_dropdown_defaults error: {e}")


def migrate_tenant_data():
    """既存データをデフォルトテナントに割り当てる（初回マイグレーション用）"""
    try:
        # テナントがなければ作成
        if Tenant.query.count() == 0:
            t = Tenant(name='ミエルーム', plan='standard', is_active=True)
            db.session.add(t)
            db.session.commit()

        # ORDER BY id で決定的に最初のテナントを取得（PostgreSQLでも安定）
        default_tenant = Tenant.query.order_by(Tenant.id).first()

        # tenant_idのないstoreを割り当て
        for s in Store.query.filter_by(tenant_id=None).all():
            s.tenant_id = default_tenant.id

        # super_admin以外でtenant_idのないユーザーを割り当て
        for u in AppUser.query.filter_by(tenant_id=None).all():
            if u.role != 'super_admin':
                u.tenant_id = default_tenant.id

        db.session.commit()

        # オーナーのテナントに店舗が無い場合は、そのテナント専用の店舗を新規作成する。
        # （旧実装は default_tenant の店舗を「移動」していたが、他テナント（特に既定テナント）の
        #   店舗を奪い、別会社の店舗が表示される重大なデータ混在を招くため廃止）
        for owner in AppUser.query.filter_by(role='owner').all():
            if owner.tenant_id:
                usable = Store.query.filter(
                    Store.tenant_id == owner.tenant_id,
                    db.or_(Store.is_active == True, Store.is_locked == True)).count()
                if usable == 0:
                    t = Tenant.query.get(owner.tenant_id)
                    db.session.add(Store(name=((t.name if t else None) or '本店'),
                                         tenant_id=owner.tenant_id, is_active=True))
                    db.session.commit()
                    print(f"オーナー(id={owner.id})のテナントに店舗を新規作成しました")

        # LeadMediaStat / Lead の store_id=NULL・0 を最初のアクティブ店舗に修復
        default_store = Store.query.filter_by(is_active=True).order_by(Store.id).first()
        if default_store:
            broken_stats = LeadMediaStat.query.filter(
                db.or_(LeadMediaStat.store_id == None, LeadMediaStat.store_id == 0)
            ).all()
            for s in broken_stats:
                s.store_id = default_store.id
            broken_leads = Lead.query.filter(
                db.or_(Lead.store_id == None, Lead.store_id == 0)
            ).all()
            for l in broken_leads:
                l.store_id = default_store.id
            if broken_stats or broken_leads:
                db.session.commit()
                print(f"store_id修復: LeadMediaStat={len(broken_stats)}件, Lead={len(broken_leads)}件")

        # 旧「テナント全体オプション」(store_id=NULL)を店舗単位に移行する。
        #  ・同一(テナント,オプション)で既に店舗別の設定があれば、店舗別を正とし NULL行は削除
        #  ・店舗別の設定が無ければ、全有効店舗に展開してから NULL行を削除
        try:
            null_opts = TenantOption.query.filter(TenantOption.store_id == None).all()
            if null_opts:
                from collections import defaultdict as _dd
                grouped = _dd(list)
                for o in null_opts:
                    grouped[(o.tenant_id, o.option_key)].append(o)
                migrated = 0
                for (otid, okey), rows in grouped.items():
                    has_store_level = TenantOption.query.filter(
                        TenantOption.tenant_id == otid,
                        TenantOption.option_key == okey,
                        TenantOption.store_id != None).first() is not None
                    if not has_store_level:
                        for s in Store.query.filter_by(tenant_id=otid, is_active=True).all():
                            db.session.add(TenantOption(tenant_id=otid, store_id=s.id, option_key=okey))
                    for o in rows:
                        db.session.delete(o)
                    migrated += 1
                db.session.commit()
                print(f"オプションを店舗単位に移行: {migrated}件")
        except Exception as _e:
            db.session.rollback()
            print(f"オプション店舗移行エラー: {_e}")

        print("テナントデータのマイグレーション完了")

        # 既存店舗に媒体マスターがなければ初期化
        for s in Store.query.all():
            if MediaType.query.filter_by(store_id=s.id).count() == 0:
                init_default_media_types(s.id)

        # 重複した媒体マスターを除去（同一店舗・同一名は最古の1件のみ残す）
        try:
            dup_removed = 0
            for s in Store.query.all():
                seen_names = {}
                rows = MediaType.query.filter_by(store_id=s.id, is_active=True)\
                    .order_by(MediaType.id.asc()).all()
                for m in rows:
                    if m.name in seen_names:
                        m.is_active = False
                        dup_removed += 1
                    else:
                        seen_names[m.name] = m.id
            if dup_removed:
                db.session.commit()
                print(f"媒体マスター重複除去: {dup_removed}件")
        except Exception as _e:
            db.session.rollback()
            print(f"媒体重複除去エラー: {_e}")
    except Exception as e:
        db.session.rollback()
        print(f"migrate_tenant_data error: {e}")


with app.app_context():
    db.create_all()
    if not _IS_POSTGRES:   # SQLite（ローカル）のみ
        migrate_db()
    else:
        migrate_postgres()
    init_store()
    ensure_owner_account()
    ensure_super_admin()
    migrate_tenant_data()
    seed_dropdown_defaults()
    dedupe_dropdown_options()   # 複数ワーカーの同時seedによる重複を除去
    # plan='chat_pro'（旧仕様）→ オプション 'chat_pro' に変換し、プランは standard に戻す
    try:
        legacy = Tenant.query.filter_by(plan='chat_pro').all()
        for t in legacy:
            if not TenantOption.query.filter_by(tenant_id=t.id, option_key='chat_pro').first():
                db.session.add(TenantOption(tenant_id=t.id, option_key='chat_pro'))
            t.plan = 'standard'
        if legacy:
            db.session.commit()
            print(f"chat_pro plan→option 変換: {len(legacy)}件")
    except Exception as e:
        db.session.rollback()
        print(f"chat_pro migrate error: {e}")
    # 無効なPLCustomItemテンプレートをクリーンアップ（数字のみの名前など）
    try:
        invalid_items = PLCustomItem.query.filter(
            db.or_(
                PLCustomItem.name.in_(['11', '1331', '13']),
                PLCustomItem.name == ''
            )
        ).all()
        for item in invalid_items:
            db.session.delete(item)
        if invalid_items:
            db.session.commit()
    except Exception:
        db.session.rollback()


# ── 認証デコレータ ────────────────────────────────────────

def is_tenant_locked(tenant):
    """テナントがロック/トライアル期限切れかどうかを判定"""
    if not tenant:
        return False
    if tenant.subscription_status == 'locked':
        return True
    if tenant.subscription_status == 'trial' and tenant.trial_ends_at:
        return datetime.utcnow() > tenant.trial_ends_at
    return False


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'app_user_id' not in session:
            return redirect(url_for('app_login'))
        # トライアル/ロックチェック（APIルートとロック通知ページ自身は除外）
        from flask import request as _req
        skip_paths = ('/trial-expired', '/app-login', '/forgot-password', '/reset-password')
        if not _req.path.startswith('/api/') and not any(_req.path.startswith(p) for p in skip_paths):
            user = AppUser.query.get(session['app_user_id'])
            if user and user.role != 'super_admin' and user.tenant_id:
                tenant = Tenant.query.get(user.tenant_id)
                if is_tenant_locked(tenant):
                    return redirect('/trial-expired')
        return f(*args, **kwargs)
    return decorated


def owner_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'app_user_id' not in session:
            return redirect(url_for('app_login'))
        user = AppUser.query.get(session['app_user_id'])
        if not user or user.role not in ('owner', 'super_admin'):
            return redirect(url_for('executive_dashboard'))
        return f(*args, **kwargs)
    return decorated


def owner_or_manager_required(f):
    """owner / store_manager / super_admin のみアクセス可（ログインアカウント管理など）"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'app_user_id' not in session:
            return redirect(url_for('app_login'))
        user = AppUser.query.get(session['app_user_id'])
        if not user or user.role not in ('owner', 'store_manager', 'super_admin'):
            return redirect(url_for('executive_dashboard'))
        return f(*args, **kwargs)
    return decorated


def super_admin_required(f):
    """super_admin または sys_admin がアクセス可"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'app_user_id' not in session:
            return redirect(url_for('app_login'))
        if session.get('app_user_role') not in ('super_admin', 'sys_admin'):
            return redirect(url_for('executive_dashboard'))
        return f(*args, **kwargs)
    return decorated


def _check_admin_perm(perm_field):
    """sys_admin の場合、指定したadmin権限フィールドを確認して403を返す（持てばOK）"""
    role = session.get('app_user_role')
    if role == 'super_admin':
        return None  # super_adminは常にOK
    if role == 'sys_admin':
        user = AppUser.query.get(session.get('app_user_id'))
        if user and getattr(user, perm_field, False):
            return None  # 権限あり
        from flask import jsonify as _j
        return _j({'error': 'この操作の権限がありません'}), 403
    from flask import jsonify as _j
    return _j({'error': 'この操作の権限がありません'}), 403


def super_admin_only(f):
    """破壊的操作：super_admin のみ（後方互換のため残す）"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'app_user_id' not in session:
            return redirect(url_for('app_login'))
        if session.get('app_user_role') != 'super_admin':
            from flask import jsonify as _jsonify
            return _jsonify({'error': 'この操作はスーパー管理者のみ実行できます'}), 403
        return f(*args, **kwargs)
    return decorated


def manager_or_above_required(f):
    """store_manager / owner のみアクセス可。staff と super_admin はブロック。"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'app_user_id' not in session:
            return redirect(url_for('app_login'))
        user = AppUser.query.get(session['app_user_id'])
        if not user:
            return redirect(url_for('app_login'))
        if user.role == 'super_admin':
            return redirect(url_for('admin_tenants'))
        if user.role == 'staff':
            return redirect(url_for('sales_management'))
        return f(*args, **kwargs)
    return decorated


def block_super_admin(f):
    """ビジネス系ページから super_admin を弾く（テナント管理のみ許可）"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('app_user_role') == 'super_admin':
            return redirect(url_for('admin_tenants'))
        return f(*args, **kwargs)
    return decorated


def is_premium_user():
    """現在のログインユーザーのテナントがプレミアプランかどうかを返す。
    super_admin はテナントを持たないため False。
    """
    uid = session.get('app_user_id')
    if not uid:
        return False
    user = AppUser.query.get(uid)
    if not user or user.role == 'super_admin' or not user.tenant_id:
        return False
    tenant = Tenant.query.get(user.tenant_id)
    return bool(tenant and tenant.plan == 'premium')


# ── ユーティリティ ─────────────────────────────────────────

def current_ym():
    """現在の年・月をタプルで返す"""
    now = datetime.now()
    return now.year, now.month


def get_allowed_store_ids(ignore_active=False):
    """ログインユーザーが参照できる店舗IDリストを返す（テナント分離）
    ignore_active=True の場合はセッションの active_store_id を無視して全店舗返す（本部ダッシュボード用）
    """
    uid = session.get('app_user_id')
    if not uid:
        return []
    user = AppUser.query.get(uid)
    if not user:
        return []
    if user.role == 'super_admin':
        return [s.id for s in Store.query.filter_by(is_active=True).order_by(Store.id).all()]
    elif user.role == 'owner':
        all_ids = [s.id for s in Store.query.filter_by(tenant_id=user.tenant_id, is_active=True).order_by(Store.id).all()]
        if not all_ids:
            null_stores = Store.query.filter(
                Store.tenant_id == None, Store.is_active == True
            ).order_by(Store.id).all()
            if null_stores:
                all_ids = [s.id for s in null_stores]
        if not ignore_active:
            active = session.get('active_store_id')
            if active and active in all_ids:
                return [active]
        return all_ids
    else:
        if user.store_id:
            return [user.store_id]
        if user.tenant_id:
            ids = [s.id for s in Store.query.filter_by(tenant_id=user.tenant_id, is_active=True).order_by(Store.id).all()]
            return ids[:1]
        return []


def get_allowed_stores(ignore_active=False):
    """ログインユーザーが参照できるStoreオブジェクトリストを返す"""
    uid = session.get('app_user_id')
    if not uid:
        return []
    user = AppUser.query.get(uid)
    if not user:
        return []
    if user.role == 'super_admin':
        return Store.query.filter_by(is_active=True).order_by(Store.id).all()
    elif user.role == 'owner':
        all_stores = Store.query.filter_by(tenant_id=user.tenant_id, is_active=True).order_by(Store.id).all()
        if not all_stores:
            all_stores = Store.query.filter(
                Store.tenant_id == None, Store.is_active == True
            ).order_by(Store.id).all()
        if not ignore_active:
            active = session.get('active_store_id')
            if active:
                filtered = [s for s in all_stores if s.id == active]
                if filtered:
                    return filtered
        return all_stores
    else:
        if user.store_id:
            s = Store.query.get(user.store_id)
            return [s] if s and s.is_active else []
        if user.tenant_id:
            return Store.query.filter_by(tenant_id=user.tenant_id, is_active=True).order_by(Store.id).limit(1).all()
        return []


def safe_store_id(requested_id=None):
    """テナント分離ヘルパー: リクエストの store_id を検証し安全な値を返す。
    - requested_id が許可範囲内なら そのまま返す
    - 無効 / 未指定なら active_store_id → 許可店舗の先頭 の順でフォールバック
    - 許可店舗がゼロなら None を返す（呼び出し側で 403 を返すこと）
    """
    allowed = get_allowed_store_ids()
    if not allowed:
        return None
    if requested_id and int(requested_id) in allowed:
        return int(requested_id)
    active = session.get('active_store_id')
    if active and active in allowed:
        return active
    return allowed[0]


# ── 既存ルーティング ──────────────────────────────────────

@app.route("/")
def index():
    # 管理ツールのログインページへリダイレクト
    return redirect(url_for('app_login'))


@app.route("/lp")
def lp():
    """ランディングページ（旧IeAI名義）"""
    return render_template("lp.html")


@app.route("/roompick-lp")
def roompick_lp():
    """ミエルーム ランディングページ"""
    return render_template("roompick_lp.html")


@app.route("/app-login", methods=["GET", "POST"])
def app_login():
    """管理ツール専用ログイン"""
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = AppUser.query.filter(
            db.or_(
                AppUser.username == username,
                db.func.lower(AppUser.email) == username.lower()
            ),
            AppUser.is_active == True
        ).first()
        if user and check_password_hash(user.password_hash, password):
            session.permanent = False  # ブラウザ閉じでセッション切れ
            session['app_user_id'] = user.id
            session['app_user_role'] = user.role
            session['app_username'] = user.username
            session['tenant_id'] = user.tenant_id
            user.last_login = datetime.utcnow()
            db.session.commit()
            # ログイン時に店舗を自動選択（複数店舗でも混在しないように）
            if user.role in ('owner', 'store_manager') and user.tenant_id:
                first_store = Store.query.filter_by(tenant_id=user.tenant_id, is_active=True).order_by(Store.id).first()
                if first_store:
                    session['active_store_id'] = first_store.id
            elif user.role == 'staff' and user.store_id:
                session['active_store_id'] = user.store_id
            # super_admin はテナント管理へ、それ以外は売上管理ダッシュボードへ
            if user.role == 'super_admin':
                dashboard_url = url_for('admin_tenants')
            elif user.role == 'staff':
                # 18: スタッフは自分のデータがデフォルト表示される顧客管理表へ
                dashboard_url = url_for('customer_management')
            else:
                dashboard_url = url_for('executive_dashboard')
            return make_response(f'''<!DOCTYPE html><html><head>
<meta charset="utf-8"><title>ログイン中...</title></head><body>
<script>
  sessionStorage.setItem("rp_auth","1");
  window.location.replace("{dashboard_url}");
</script>
</body></html>''')
        else:
            error = "ユーザー名またはパスワードが正しくありません。"
    return render_template("app_login.html", error=error)


@app.route("/app-logout")
def app_logout():
    """管理ツールログアウト"""
    session.pop('app_user_id', None)
    session.pop('app_user_role', None)
    session.pop('app_username', None)
    return redirect(url_for('app_login'))


# ── 幹部向け管理ツール：ページルート ─────────────────────

@app.route("/executive")
@login_required
@manager_or_above_required
def executive_dashboard():
    """売上管理ダッシュボード"""
    stores = get_allowed_stores(ignore_active=True)  # サイドバー用
    active_ids = get_allowed_store_ids()  # アクティブ店舗のみ
    allowed_ids = [s.id for s in stores]  # サイドバー用全店舗
    staff_list = Staff.query.filter(Staff.store_id.in_(active_ids), Staff.is_active == True).all()
    year, month = current_ym()
    store_id = active_ids[0] if active_ids else None
    return render_template("executive_dashboard.html",
                           stores=stores, staff_list=staff_list, year=year, month=month,
                           store_id=store_id, now=datetime.now())


def resolve_cur_staff_id(user):
    """ログインユーザーに対応する担当スタッフIDを返す。
    staff_id 未設定でも、同店舗で氏名が完全一致するスタッフが1人だけなら補完する。"""
    if not user:
        return None
    if getattr(user, 'staff_id', None):
        return user.staff_id
    if user.role == 'staff' and user.store_id and user.username:
        matches = Staff.query.filter_by(store_id=user.store_id, name=user.username,
                                        is_active=True).all()
        if len(matches) == 1:
            return matches[0].id
    return None


@app.route("/sales")
@login_required
@block_super_admin
def sales_management():
    """営業管理ページ：KPI入力・閲覧"""
    stores = get_allowed_stores()
    allowed_ids = [s.id for s in stores]
    staff_list = Staff.query.filter(Staff.store_id.in_(allowed_ids), Staff.is_active == True).all()
    year, month = current_ym()
    cur_user = AppUser.query.get(session.get('app_user_id'))
    cur_role = cur_user.role if cur_user else 'staff'
    cur_staff_id = resolve_cur_staff_id(cur_user)
    is_manager = cur_role in ('owner', 'store_manager', 'super_admin')
    store_id = allowed_ids[0] if allowed_ids else None
    media_types = MediaType.query.filter_by(store_id=store_id, is_active=True).order_by(MediaType.sort_order, MediaType.name).all() if store_id else []
    return render_template("sales_management.html", stores=stores, staff_list=staff_list,
                           year=year, month=month, now=datetime.now(),
                           cur_role=cur_role, cur_staff_id=cur_staff_id,
                           is_manager=is_manager, media_types=media_types,
                           store_id=store_id)


@app.route("/customer-management")
@login_required
@block_super_admin
def customer_management():
    """顧客管理表ページ：申込・入金管理に特化"""
    stores = get_allowed_stores(ignore_active=True)   # サイドバー用（全店舗）
    active_ids = get_allowed_store_ids()              # アクティブ店舗のみ
    staff_list = Staff.query.filter(Staff.store_id.in_(active_ids), Staff.is_active == True).all() if active_ids else []
    year, month = current_ym()
    cur_user = AppUser.query.get(session.get('app_user_id'))
    cur_role = cur_user.role if cur_user else 'staff'
    cur_staff_id = resolve_cur_staff_id(cur_user)
    is_manager = cur_role in ('owner', 'store_manager', 'super_admin')
    store_id = active_ids[0] if active_ids else None
    media_types = MediaType.query.filter_by(store_id=store_id, is_active=True).order_by(MediaType.sort_order, MediaType.name).all() if store_id else []
    return render_template("customer_management.html", stores=stores, staff_list=staff_list,
                           year=year, month=month, now=datetime.now(),
                           cur_role=cur_role, cur_staff_id=cur_staff_id,
                           is_manager=is_manager, media_types=media_types,
                           store_id=store_id)


@app.route("/contract-customers")
@login_required
@block_super_admin
def contract_customers():
    """契約中顧客管理ページ：審査〇で契約に進んだ顧客の契約書類を準備する"""
    stores = get_allowed_stores(ignore_active=True)
    active_ids = get_allowed_store_ids()
    staff_list = Staff.query.filter(Staff.store_id.in_(active_ids), Staff.is_active == True).all() if active_ids else []
    cur_user = AppUser.query.get(session.get('app_user_id'))
    cur_role = cur_user.role if cur_user else 'staff'
    cur_staff_id = resolve_cur_staff_id(cur_user)
    is_manager = cur_role in ('owner', 'store_manager', 'super_admin')
    store_id = active_ids[0] if active_ids else None
    return render_template("contract_customers.html", stores=stores, staff_list=staff_list,
                           now=datetime.now(), cur_role=cur_role, cur_staff_id=cur_staff_id,
                           is_manager=is_manager, store_id=store_id)


@app.route("/api/contract-customers")
@login_required
def api_contract_customers():
    """契約中顧客一覧：審査〇（review_status='ok'）になった顧客（＝契約に進んだ顧客）"""
    allowed_ids = get_allowed_store_ids()
    staff_id = request.args.get('staff_id', type=int)
    q = ApplicationRecord.query.filter(
        ApplicationRecord.store_id.in_(allowed_ids),
        ApplicationRecord.review_status == 'ok',
        ~ApplicationRecord.status.in_(['キャンセル', 'キャンセル振替']),
        db.or_(ApplicationRecord.past_customer == False, ApplicationRecord.past_customer == None),
    )
    if staff_id:
        q = q.filter(ApplicationRecord.staff_id == staff_id)
    recs = q.order_by(ApplicationRecord.application_date.asc(), ApplicationRecord.id.asc()).all()
    staff_map = {s.id: s.name for s in Staff.query.all()}
    return jsonify([_app_record_to_dict(r, staff_map) for r in recs])


@app.route("/past-customers")
@login_required
@block_super_admin
def past_customers():
    """顧客管理（契約終了）ページ"""
    stores = get_allowed_stores(ignore_active=True)
    active_ids = get_allowed_store_ids()
    staff_list = Staff.query.filter(Staff.store_id.in_(active_ids), Staff.is_active == True).all() if active_ids else []
    cur_user = AppUser.query.get(session.get('app_user_id'))
    cur_role = cur_user.role if cur_user else 'staff'
    cur_staff_id = resolve_cur_staff_id(cur_user)
    is_manager = cur_role in ('owner', 'store_manager', 'super_admin')
    store_id = active_ids[0] if active_ids else None
    return render_template("past_customers.html", stores=stores, staff_list=staff_list,
                           now=datetime.now(), cur_role=cur_role, cur_staff_id=cur_staff_id,
                           is_manager=is_manager, store_id=store_id)


@app.route("/api/past-customers")
@login_required
def api_past_customers():
    """顧客管理（契約終了）一覧"""
    allowed_ids = get_allowed_store_ids()
    staff_id = request.args.get('staff_id', type=int)
    q = request.args.get('q', '').strip()
    query = ApplicationRecord.query.filter(
        ApplicationRecord.store_id.in_(allowed_ids),
        ApplicationRecord.past_customer == True,
    )
    if staff_id:
        query = query.filter(ApplicationRecord.staff_id == staff_id)
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(
            ApplicationRecord.customer_name.ilike(like),
            ApplicationRecord.property_name.ilike(like),
            ApplicationRecord.room_number.ilike(like),
        ))
    recs = query.order_by(ApplicationRecord.application_date.desc()).all()
    staff_map = {s.id: s.name for s in Staff.query.all()}
    return jsonify([_app_record_to_dict(r, staff_map) for r in recs])


@app.route("/api/contract-customers/<int:rid>/past", methods=["POST"])
@login_required
def api_contract_customer_mark_past(rid):
    """契約終了フラグをトグル"""
    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403
    data = request.get_json() or {}
    rec.past_customer = bool(data.get('past', True))
    db.session.commit()
    return jsonify({'ok': True})


def _contract_doc_defaults(rec, staff):
    """申込データから契約書類（取引成立台帳）の初期値を作る"""
    fd = lambda d: d.isoformat() if d else ''
    return {
        'torihiki': '媒介',
        'keiyaku_date': fd(rec.contract_start_date),
        'hikiwatashi_date': fd(rec.contract_start_date),
        'kashinushi_name': '', 'kashinushi_addr': '', 'kashinushi_tel': '',
        'karinushi_name': rec.customer_name or '', 'karinushi_addr': '', 'karinushi_tel': '',
        'bukken_shozai': '',
        'kouzou': '', 'yane': '', 'youto': '居宅',
        'kaisuu_above': '', 'kaisuu_below': '', 'menseki': '', 'madori': '',
        'meisho': rec.property_name or '', 'goushitsu': rec.room_number or '',
        'setsubi': [], 'fuzoku': [], 'sonota': '',
        'chinryo': int(rec.rent or 0), 'kanrihi': '', 'chuusha': '',
        'reikin': '', 'shikikin': '', 'hosho': '', 'kazai_hoken': '',
        'cleaning': '', 'support24': '', 'chonaikai': '', 'catv': '', 'suidou': '',
        'keiyaku_shurui': '普通借家',
        'kanri_kaisha': rec.management_company or '',
        'tantou': staff.name if staff else '',
        'biko': '',
    }


@app.route("/contract-customers/<int:rid>/edit")
@login_required
@block_super_admin
def contract_document_edit(rid):
    """契約書類エディタページ（テンプレート方式）"""
    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return "権限がありません", 403
    return render_template("contract_doc_editor.html", rid=rid,
                           customer_name=rec.customer_name or '',
                           property_name=rec.property_name or '')


@app.route("/api/contract-customers/<int:rid>/document-data", methods=["GET"])
@login_required
def api_contract_document_get(rid):
    """契約書類の保存済み編集データと、申込からの初期値コンテキストを返す"""
    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403
    staff = Staff.query.get(rec.staff_id) if rec.staff_id else None
    store = Store.query.get(rec.store_id)
    fd = lambda d: d.isoformat() if d else ''
    ad_yen = round((rec.rent or 0) * (rec.ad_amount or 0) / 100) if (rec.ad_type or 'amount') == 'percent' else (rec.ad_amount or 0)
    ctx = {
        'customer_name': rec.customer_name or '',
        'property_name': rec.property_name or '',
        'room_number': rec.room_number or '',
        'rent': int(rec.rent or 0),
        'management_company': rec.management_company or '',
        'staff_name': staff.name if staff else '',
        'store_name': store.name if store else '',
        'contract_start_date': fd(rec.contract_start_date),
        'application_date': fd(rec.application_date),
        'brokerage_fee': int(rec.brokerage_fee or 0),
        'option_amount': int(rec.option_amount or 0),
        'ad_yen': int(ad_yen),
        'media': rec.media or '',
    }
    data = {}
    doc = ContractDocument.query.filter_by(application_id=rid).first()
    if doc and doc.data:
        try:
            saved = json.loads(doc.data)
            if isinstance(saved, dict):
                if saved.get('v') == 2:          # 新フォーマット {v:2, values:{tag:val}}
                    data = saved.get('values') or {}
                elif 'values' in saved and isinstance(saved.get('values'), dict):
                    data = saved['values']       # 旧1テンプレート形式 {template_id:.., values:{}}
                else:
                    data = saved                 # レガシー直フラット
        except Exception:
            pass
    return jsonify({'data': data, 'ctx': ctx, 'saved': bool(doc)})


@app.route("/api/contract-customers/<int:rid>/document-data", methods=["POST"])
@login_required
def api_contract_document_save(rid):
    """契約書類の編集データを保存"""
    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403
    cur_user = AppUser.query.get(session.get('app_user_id'))
    if cur_user and cur_user.role == 'staff' and rec.staff_id != cur_user.staff_id:
        return jsonify({'error': '権限がありません'}), 403
    payload = request.get_json() or {}
    # {v:2, values:{}} でも フラット{tag:val} でも受け付ける
    if isinstance(payload, dict) and payload.get('v') == 2:
        save_data = payload
    else:
        save_data = {'v': 2, 'values': payload}
    doc = ContractDocument.query.filter_by(application_id=rid).first()
    if not doc:
        doc = ContractDocument(application_id=rid, store_id=rec.store_id)
        db.session.add(doc)
    doc.data = json.dumps(save_data, ensure_ascii=False)
    doc.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'status': 'ok'})


# ── 契約書のAI読み取り（PDF/画像 → 項目自動抽出） ──────────────
# 抽出対象フィールド定義：key は契約書類エディタのフラットキー（docKey.fieldKey）と一致させる
CONTRACT_EXTRACT_FIELDS = [
    # 物件の表示
    {'k': 'keiyakusho.meisho',      'l': '物件名称',          't': 'text'},
    {'k': 'keiyakusho.goushitsu',   'l': '号室・部屋番号',     't': 'text'},
    {'k': 'keiyakusho.shozai',      'l': '物件所在地（住所）', 't': 'text'},
    {'k': 'keiyakusho.yukamenseki', 'l': '床面積（㎡・数値のみ）', 't': 'text'},
    {'k': 'keiyakusho.madori',      'l': '間取り（例:1LDK）',  't': 'text'},
    {'k': 'keiyakusho.kouzou',      'l': '構造（例:木造/RC造）', 't': 'text'},
    {'k': 'keiyakusho.kaisuu',      'l': '階数',              't': 'text'},
    # 契約期間
    {'k': 'keiyakusho.shiki', 'l': '契約期間の始期', 't': 'date'},
    {'k': 'keiyakusho.shuki', 'l': '契約期間の終期', 't': 'date'},
    {'k': 'keiyakusho.getsu', 'l': '契約月数',       't': 'num'},
    # 賃料等
    {'k': 'keiyakusho.chinryo',  'l': '賃料（月額・円）',     't': 'num'},
    {'k': 'keiyakusho.kyoekihi', 'l': '共益費・管理費（円）', 't': 'num'},
    {'k': 'keiyakusho.reikin',   'l': '礼金（円）',          't': 'num'},
    {'k': 'keiyakusho.shikikin', 'l': '敷金（円）',          't': 'num'},
    {'k': 'keiyakusho.chuusha',  'l': '駐車場料金（円）',     't': 'num'},
    {'k': 'keiyakusho.shokyaku', 'l': '償却・敷引',          't': 'text'},
    # 支払・口座
    {'k': 'keiyakusho.shiharaibi',   'l': '賃料支払日（毎月○日）', 't': 'text'},
    {'k': 'keiyakusho.kinyu',        'l': '金融機関名',        't': 'text'},
    {'k': 'keiyakusho.furikomi',     'l': '振込/引落（どちらか）', 't': 'sel', 'o': ['振込', '引落']},
    {'k': 'keiyakusho.kouza_no',     'l': '口座番号',          't': 'text'},
    {'k': 'keiyakusho.kouza_name',   'l': '口座名義人',        't': 'text'},
    {'k': 'keiyakusho.kanri_gyosha', 'l': '貸主及び管理業者',   't': 'text'},
    # 貸主・借主（取引成立台帳タブと共有）
    {'k': 'daicho.kashinushi_name', 'l': '貸主（オーナー）氏名・名称', 't': 'text'},
    {'k': 'daicho.kashinushi_addr', 'l': '貸主住所',          't': 'text'},
    {'k': 'daicho.kashinushi_tel',  'l': '貸主電話番号',       't': 'text'},
    {'k': 'daicho.karinushi_name',  'l': '借主（入居者）氏名',  't': 'text'},
    {'k': 'daicho.karinushi_addr',  'l': '借主（現）住所',     't': 'text'},
    {'k': 'daicho.karinushi_tel',   'l': '借主電話番号',       't': 'text'},
]

_ALLOWED_EXTRACT_MIME = {
    'application/pdf': 'pdf',
    'image/jpeg': 'image', 'image/jpg': 'image', 'image/png': 'image',
    'image/webp': 'image', 'image/gif': 'image',
}


@app.route("/api/contract-customers/<int:rid>/extract-document", methods=["POST"])
@login_required
@block_super_admin
def api_contract_extract_document(rid):
    """賃貸借契約書のPDF/画像をClaudeで読み取り、契約書類フィールドを自動抽出して返す。
    自動確定はせず、フロントの確認画面で人がチェック・修正してから反映する設計。"""
    import base64 as _b64
    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403

    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    raw = f.read()
    if len(raw) > 18 * 1024 * 1024:
        return jsonify({'error': 'ファイルが大きすぎます（18MBまで）'}), 400
    if len(raw) == 0:
        return jsonify({'error': 'ファイルが空です'}), 400

    mime = (f.mimetype or '').lower()
    # 拡張子からの補完
    if mime not in _ALLOWED_EXTRACT_MIME:
        ext = (f.filename.rsplit('.', 1)[-1] if '.' in f.filename else '').lower()
        mime = {'pdf': 'application/pdf', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                'png': 'image/png', 'webp': 'image/webp', 'gif': 'image/gif'}.get(ext, mime)
    kind = _ALLOWED_EXTRACT_MIME.get(mime)
    if not kind:
        return jsonify({'error': 'PDFまたは画像（JPEG/PNG）のみ対応しています'}), 400

    if not os.getenv("ANTHROPIC_API_KEY"):
        return jsonify({'error': 'AI読み取りが未設定です（管理者にお問い合わせください）'}), 503

    b64 = _b64.standard_b64encode(raw).decode()
    field_lines = []
    for fld in CONTRACT_EXTRACT_FIELDS:
        extra = ''
        if fld['t'] == 'date':
            extra = '（日付。YYYY-MM-DD形式。和暦は西暦へ変換）'
        elif fld['t'] == 'num':
            extra = '（半角数字のみ。カンマ・円記号・単位は付けない）'
        elif fld['t'] == 'sel':
            extra = f'（次のいずれか:{"/".join(fld.get("o", []))}）'
        field_lines.append(f'  "{fld["k"]}": "{fld["l"]}{extra}"')

    prompt = (
        "あなたは不動産の賃貸借契約書を読み取る専門アシスタントです。\n"
        "添付された賃貸借契約書（PDFまたは画像）から、以下のJSONキーに対応する値を抽出してください。\n\n"
        "対象フィールド（キー: 説明）:\n{\n" + ",\n".join(field_lines) + "\n}\n\n"
        "厳守ルール:\n"
        "1. 必ず上記キーだけを持つJSONオブジェクトを1つだけ返す。説明文やマークダウンは一切付けない。\n"
        "2. 書類に記載が無い・読み取れない項目は、そのキーを省略する（推測で埋めない）。\n"
        "3. 金額は半角数字のみ（例: 75000）。日付はYYYY-MM-DD。\n"
        "4. 自信が持てない値は省略する。誤った自動入力は事故につながるため、確実な箇所のみ返す。\n"
    )

    if kind == 'pdf':
        doc_block = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}}
    else:
        doc_block = {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}}

    try:
        msg = anthropic_client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1500,
            messages=[{"role": "user", "content": [doc_block, {"type": "text", "text": prompt}]}],
        )
        text = "".join(getattr(b, 'text', '') for b in msg.content).strip()
    except Exception as e:
        return jsonify({'error': f'AI読み取りに失敗しました: {e}'}), 502

    # JSON抽出（コードフェンスや前後テキストに耐性を持たせる）
    extracted = {}
    try:
        s, e = text.find('{'), text.rfind('}')
        if s >= 0 and e > s:
            parsed = json.loads(text[s:e + 1])
            if isinstance(parsed, dict):
                extracted = parsed
    except Exception:
        extracted = {}

    valid_keys = {fld['k'] for fld in CONTRACT_EXTRACT_FIELDS}
    out = []
    for fld in CONTRACT_EXTRACT_FIELDS:
        v = extracted.get(fld['k'])
        if v is None:
            continue
        v = str(v).strip()
        if v == '':
            continue
        if fld['t'] == 'num':
            v = re.sub(r'[^0-9]', '', v)
            if v == '':
                continue
        if fld['t'] == 'sel' and v not in fld.get('o', []):
            continue
        out.append({'key': fld['k'], 'label': fld['l'], 'type': fld['t'], 'value': v})

    return jsonify({'fields': out, 'found': len(out)})


# ── Excelテンプレート帳票（クライアント様式の取込→穴埋め→出力） ──────────
_DOC_TAG_RE = re.compile(r'\{\{\s*([^}]+?)\s*\}\}')


def _doc_tenant_id():
    u = AppUser.query.get(session.get('app_user_id'))
    return u.tenant_id if u else None


def _doc_can_manage():
    """テンプレート・会社情報の初期設定が可能なロールか"""
    return session.get('app_user_role') in ('owner', 'store_manager')


def _xlsx_extract_tags(xlsx_bytes):
    """xlsx内の全シート・全セルから {{タグ}} を抽出（重複除去・出現順）"""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    tags, seen = [], set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and '{{' in v:
                    for m in _DOC_TAG_RE.findall(v):
                        t = m.strip()
                        if t and t not in seen:
                            seen.add(t)
                            tags.append(t)
    return tags


def _xlsx_fill(xlsx_bytes, values):
    """{{タグ}} を values[タグ] で置換した xlsx バイト列を返す。
    セル全体が単一タグかつ値が数値ならば数値として書き込む（合計式などのため）。"""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and '{{' in v):
                    continue
                stripped = v.strip()
                m_full = _DOC_TAG_RE.fullmatch(stripped)
                if m_full:
                    key = m_full.group(1).strip()
                    rv = values.get(key, '')
                    rv = '' if rv is None else rv
                    sval = str(rv)
                    if sval != '' and re.fullmatch(r'-?\d+(\.\d+)?', sval):
                        cell.value = float(sval) if '.' in sval else int(sval)
                    else:
                        cell.value = sval
                else:
                    def _repl(mm):
                        k = mm.group(1).strip()
                        rv = values.get(k, '')
                        return '' if rv is None else str(rv)
                    cell.value = _DOC_TAG_RE.sub(_repl, v)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# 設定＞会社情報（CompanyProfile）を、契約フォーマットの固定差し込みタグ名に対応づける
_COMPANY_PROFILE_DOC_LABELS = [
    ('company_name',   '会社名'),
    ('store_name',     '店舗名'),
    ('phone',          '電話番号'),
    ('fax',            'FAX番号'),
    ('email',          'メールアドレス'),
    ('address',        '住所'),
    ('representative', '代表者氏名'),
    ('license_number', '宅建業免許番号'),
    ('license_date',   '宅建業免許取得日'),
    ('business_hours', '営業時間'),
    ('holidays',       '定休日'),
    ('invoice_number', 'インボイス登録番号'),
    ('line_url',       '公式LINEのURL'),
]


def _company_profile_doc_dict():
    """設定＞会社情報（CompanyProfile）を、契約フォーマットの固定差し込み辞書に変換。空欄は除外。"""
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    cp = CompanyProfile.query.filter_by(store_id=sid).first() if sid else None
    if not cp:
        return {}
    out = {}
    for attr, label in _COMPANY_PROFILE_DOC_LABELS:
        val = getattr(cp, attr, None)
        if val:
            out[label] = str(val)
    return out


def _doc_company_manual(tenant_id):
    """契約フォーマット側で手入力した追加項目（会社情報に無い項目）。"""
    rec = DocCompanyInfo.query.filter_by(tenant_id=tenant_id).first()
    if rec and rec.data:
        try:
            d = json.loads(rec.data)
            if isinstance(d, dict):
                return d
        except Exception:
            pass
    return {}


def _doc_company_data(tenant_id):
    """固定差し込み用の会社情報。設定＞会社情報を自動反映し、手入力の追加項目を重ねる
    （同名キーは設定＞会社情報を優先＝常に最新の内容が入る）。"""
    merged = dict(_doc_company_manual(tenant_id))
    merged.update(_company_profile_doc_dict())
    return merged


def _doc_default_mapping(tags, company_keys):
    """タグの初期割当：会社情報のキーと一致すれば company、それ以外は case"""
    mp = {}
    for t in tags:
        if t in company_keys:
            mp[t] = {'scope': 'company', 'company_key': t, 'label': t}
        else:
            mp[t] = {'scope': 'case', 'company_key': '', 'label': t}
    return mp


@app.route("/doc-templates")
@login_required
@block_super_admin
def doc_templates_page():
    return render_template("doc_templates.html", can_manage=_doc_can_manage())


@app.route("/doc-templates/embed")
@login_required
@block_super_admin
def doc_templates_embed():
    """契約管理ページのタブ内にiframeで埋め込む、サイドバー無しの設定UI"""
    return render_template("doc_templates_embed.html", can_manage=_doc_can_manage())


@app.route("/api/doc-company-info", methods=["GET"])
@login_required
@block_super_admin
def api_doc_company_get():
    tid = _doc_tenant_id()
    auto = _company_profile_doc_dict()                # 設定＞会社情報（自動）
    manual = _doc_company_manual(tid)                 # 手入力の保存内容
    extras = {k: v for k, v in manual.items() if k not in auto}   # 会社情報に無い追加項目のみ
    return jsonify({'data': _doc_company_data(tid), 'auto': auto, 'extras': extras})


@app.route("/api/doc-company-info", methods=["POST"])
@login_required
@block_super_admin
def api_doc_company_save():
    if not _doc_can_manage():
        return jsonify({'error': '権限がありません'}), 403
    tid = _doc_tenant_id()
    payload = request.get_json() or {}
    data = payload.get('data') if isinstance(payload, dict) else None
    if not isinstance(data, dict):
        return jsonify({'error': '不正なデータです'}), 400
    clean = {str(k).strip(): ('' if v is None else str(v)) for k, v in data.items() if str(k).strip()}
    rec = DocCompanyInfo.query.filter_by(tenant_id=tid).first()
    if not rec:
        rec = DocCompanyInfo(tenant_id=tid)
        db.session.add(rec)
    rec.data = json.dumps(clean, ensure_ascii=False)
    db.session.commit()
    return jsonify({'status': 'ok', 'data': clean})


@app.route("/api/doc-templates", methods=["GET"])
@login_required
@block_super_admin
def api_doc_templates_list():
    tid = _doc_tenant_id()
    rows = DocTemplate.query.filter_by(tenant_id=tid, is_active=True).order_by(DocTemplate.id.desc()).all()
    out = []
    for r in rows:
        try:
            tags = json.loads(r.tags) if r.tags else []
        except Exception:
            tags = []
        out.append({'id': r.id, 'name': r.name, 'filename': r.filename, 'tag_count': len(tags)})
    return jsonify({'templates': out})


@app.route("/api/doc-templates", methods=["POST"])
@login_required
@block_super_admin
def api_doc_template_upload():
    if not _doc_can_manage():
        return jsonify({'error': '権限がありません（オーナー・店長のみ設定できます）'}), 403
    import base64 as _b64
    tid = _doc_tenant_id()
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': 'Excelファイルを選択してください'}), 400
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Excel（.xlsx）形式のみ対応しています'}), 400
    raw = f.read()
    if len(raw) > 10 * 1024 * 1024:
        return jsonify({'error': 'ファイルが大きすぎます（10MBまで）'}), 400
    try:
        tags = _xlsx_extract_tags(raw)
    except Exception as e:
        return jsonify({'error': f'Excelの読み込みに失敗しました: {e}'}), 400
    name = (request.form.get('name') or '').strip() or f.filename.rsplit('.', 1)[0]
    company_keys = set(_doc_company_data(tid).keys())
    mapping = _doc_default_mapping(tags, company_keys)
    rec = DocTemplate(
        tenant_id=tid, name=name, filename=f.filename,
        file_b64=_b64.standard_b64encode(raw).decode(),
        tags=json.dumps(tags, ensure_ascii=False),
        mapping=json.dumps(mapping, ensure_ascii=False),
        created_by=session.get('app_user_id'),
    )
    db.session.add(rec)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': rec.id, 'tags': tags})


@app.route("/api/doc-templates/<int:tpl_id>", methods=["GET"])
@login_required
@block_super_admin
def api_doc_template_get(tpl_id):
    tid = _doc_tenant_id()
    r = DocTemplate.query.filter_by(id=tpl_id, tenant_id=tid, is_active=True).first_or_404()
    try:
        tags = json.loads(r.tags) if r.tags else []
    except Exception:
        tags = []
    try:
        mapping = json.loads(r.mapping) if r.mapping else {}
    except Exception:
        mapping = {}
    company = _doc_company_data(tid)
    # 未登録タグにデフォルト割当を補完
    for t in tags:
        if t not in mapping:
            mapping[t] = ({'scope': 'company', 'company_key': t, 'label': t}
                          if t in company else {'scope': 'case', 'company_key': '', 'label': t})
    return jsonify({
        'id': r.id, 'name': r.name, 'filename': r.filename,
        'tags': tags, 'mapping': mapping,
        'company_keys': list(company.keys()), 'company': company,
        'can_manage': _doc_can_manage(),
    })


@app.route("/api/doc-templates/<int:tpl_id>/mapping", methods=["POST"])
@login_required
@block_super_admin
def api_doc_template_mapping(tpl_id):
    if not _doc_can_manage():
        return jsonify({'error': '権限がありません'}), 403
    tid = _doc_tenant_id()
    r = DocTemplate.query.filter_by(id=tpl_id, tenant_id=tid, is_active=True).first_or_404()
    payload = request.get_json() or {}
    mapping = payload.get('mapping')
    if not isinstance(mapping, dict):
        return jsonify({'error': '不正なデータです'}), 400
    if 'name' in payload and str(payload['name']).strip():
        r.name = str(payload['name']).strip()
    r.mapping = json.dumps(mapping, ensure_ascii=False)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/doc-templates/<int:tpl_id>", methods=["DELETE"])
@login_required
@block_super_admin
def api_doc_template_delete(tpl_id):
    if not _doc_can_manage():
        return jsonify({'error': '権限がありません'}), 403
    tid = _doc_tenant_id()
    r = DocTemplate.query.filter_by(id=tpl_id, tenant_id=tid).first_or_404()
    r.is_active = False
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/doc-generate/<int:tpl_id>")
@login_required
@block_super_admin
def doc_generate_page(tpl_id):
    tid = _doc_tenant_id()
    r = DocTemplate.query.filter_by(id=tpl_id, tenant_id=tid, is_active=True).first_or_404()
    return render_template("doc_generate.html", tpl_id=tpl_id, tpl_name=r.name)


@app.route("/api/doc-templates/<int:tpl_id>/extract", methods=["POST"])
@login_required
@block_super_admin
def api_doc_template_extract(tpl_id):
    """添付資料（契約書/顧客/物件のPDF・画像）から案件タグの値をAIで抽出"""
    import base64 as _b64
    tid = _doc_tenant_id()
    r = DocTemplate.query.filter_by(id=tpl_id, tenant_id=tid, is_active=True).first_or_404()
    try:
        mapping = json.loads(r.mapping) if r.mapping else {}
    except Exception:
        mapping = {}
    # 案件タグ（AI抽出対象）
    case_tags = [t for t, m in mapping.items() if (m or {}).get('scope') != 'company']
    if not case_tags:
        return jsonify({'fields': [], 'found': 0})

    files = request.files.getlist('files') or ([request.files['file']] if 'file' in request.files else [])
    files = [f for f in files if f and f.filename]
    if not files:
        return jsonify({'error': '資料ファイルを選択してください'}), 400
    if not os.getenv("ANTHROPIC_API_KEY"):
        return jsonify({'error': 'AI読み取りが未設定です（管理者にお問い合わせください）'}), 503

    content = []
    total = 0
    for f in files[:5]:
        raw = f.read()
        total += len(raw)
        if total > 20 * 1024 * 1024:
            return jsonify({'error': '添付の合計が大きすぎます（20MBまで）'}), 400
        mime = (f.mimetype or '').lower()
        if mime not in _ALLOWED_EXTRACT_MIME:
            ext = (f.filename.rsplit('.', 1)[-1] if '.' in f.filename else '').lower()
            mime = {'pdf': 'application/pdf', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                    'png': 'image/png', 'webp': 'image/webp', 'gif': 'image/gif'}.get(ext, mime)
        kind = _ALLOWED_EXTRACT_MIME.get(mime)
        if not kind:
            continue
        b64 = _b64.standard_b64encode(raw).decode()
        if kind == 'pdf':
            content.append({"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}})
        else:
            content.append({"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}})
    if not content:
        return jsonify({'error': 'PDFまたは画像（JPEG/PNG）のみ対応しています'}), 400

    # 申込台帳の既知情報をコンテキストとして渡す（添付資料に無くても埋められるように）
    app_id = request.form.get('application_id')
    if app_id:
        try:
            rec = ApplicationRecord.query.get(int(app_id))
        except Exception:
            rec = None
        if rec and rec.store_id in get_allowed_store_ids():
            staff = Staff.query.get(rec.staff_id) if rec.staff_id else None
            fd = lambda d: d.isoformat() if d else ''
            known = {
                'お客様名・契約者名': rec.customer_name or '',
                '物件名': rec.property_name or '',
                '号室・部屋番号': rec.room_number or '',
                '賃料（月額）': int(rec.rent or 0),
                '管理会社': rec.management_company or '',
                '担当者': staff.name if staff else '',
                '契約開始日': fd(rec.contract_start_date),
                '申込日': fd(rec.application_date),
                '仲介手数料': int(rec.brokerage_fee or 0),
            }
            known = {k: v for k, v in known.items() if v not in ('', 0)}
            if known:
                ctx_text = "【この案件の既知情報（添付資料より優先して活用してよい）】\n" + \
                    "\n".join(f"・{k}: {v}" for k, v in known.items())
                content.append({"type": "text", "text": ctx_text})

    key_lines = ",\n".join(f'  "{t}": "{(mapping.get(t) or {}).get("label") or t}"' for t in case_tags)
    prompt = (
        "あなたは不動産書類を読み取る専門アシスタントです。\n"
        "添付資料（契約書・顧客情報・物件情報など）と既知情報から、以下のJSONキーに対応する値を抽出してください。\n\n"
        "対象フィールド（キー: 説明）:\n{\n" + key_lines + "\n}\n\n"
        "厳守ルール:\n"
        "1. 上記キーだけを持つJSONオブジェクトを1つだけ返す。説明文やマークダウンは付けない。\n"
        "2. 記載が無い・読み取れない・自信が無い項目はキーを省略する（推測で埋めない）。\n"
        "3. 金額は半角数字のみ、日付はYYYY-MM-DD（和暦は西暦へ）。\n"
        "4. 誤った自動入力は契約事故につながるため、確実な箇所のみ返す。\n"
    )
    content.append({"type": "text", "text": prompt})

    try:
        msg = anthropic_client.messages.create(
            model="claude-sonnet-4-6", max_tokens=2000,
            messages=[{"role": "user", "content": content}],
        )
        text = "".join(getattr(b, 'text', '') for b in msg.content).strip()
    except Exception as e:
        return jsonify({'error': f'AI読み取りに失敗しました: {e}'}), 502

    extracted = {}
    try:
        s, e = text.find('{'), text.rfind('}')
        if s >= 0 and e > s:
            parsed = json.loads(text[s:e + 1])
            if isinstance(parsed, dict):
                extracted = parsed
    except Exception:
        extracted = {}

    out = []
    for t in case_tags:
        v = extracted.get(t)
        if v is None:
            continue
        v = str(v).strip()
        if v == '':
            continue
        out.append({'key': t, 'label': (mapping.get(t) or {}).get('label') or t, 'value': v})
    return jsonify({'fields': out, 'found': len(out)})


@app.route("/api/doc-templates/<int:tpl_id>/render", methods=["POST"])
@login_required
@block_super_admin
def api_doc_template_render(tpl_id):
    """会社情報＋入力値でテンプレートを穴埋めし、xlsxをダウンロード返却"""
    import base64 as _b64
    from urllib.parse import quote as _urlquote
    tid = _doc_tenant_id()
    r = DocTemplate.query.filter_by(id=tpl_id, tenant_id=tid, is_active=True).first_or_404()
    try:
        mapping = json.loads(r.mapping) if r.mapping else {}
    except Exception:
        mapping = {}
    payload = request.get_json() or {}
    case_values = payload.get('values') or {}
    company = _doc_company_data(tid)

    values = {}
    for tag, m in mapping.items():
        m = m or {}
        if m.get('scope') == 'company':
            values[tag] = company.get(m.get('company_key') or tag, '')
        else:
            values[tag] = case_values.get(tag, '')
    # mappingに無いタグも case_values から補完
    for k, v in case_values.items():
        values.setdefault(k, v)

    try:
        raw = _b64.standard_b64decode(r.file_b64.encode())
        filled = _xlsx_fill(raw, values)
    except Exception as e:
        return jsonify({'error': f'出力に失敗しました: {e}'}), 500

    fname = f"{r.name or 'document'}.xlsx"
    resp = make_response(filled)
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = (
        "attachment; filename=\"document.xlsx\"; filename*=UTF-8''" + _urlquote(fname)
    )
    return resp


@app.route("/api/doc-templates/combined-schema", methods=["GET"])
@login_required
@block_super_admin
def api_doc_combined_schema():
    """全テンプレートの案件タグを集約して返す（書類一括編集画面用）"""
    try:
        tid = _doc_tenant_id()
        templates = DocTemplate.query.filter_by(tenant_id=tid, is_active=True).order_by(DocTemplate.id).all()
        tpl_list = []
        all_case_tags = {}   # 順序保持: tag -> label
        for tpl_row in templates:
            try:
                mapping = json.loads(tpl_row.mapping) if tpl_row.mapping else {}
                if not isinstance(mapping, dict):
                    mapping = {}
            except Exception:
                mapping = {}
            try:
                tags = json.loads(tpl_row.tags) if tpl_row.tags else []
                if not isinstance(tags, list):
                    tags = []
            except Exception:
                tags = []
            case_tags, company_tags = [], []
            for tag in tags:
                m = mapping.get(tag) or {}
                if not isinstance(m, dict):
                    m = {}
                if m.get('scope') == 'company':
                    company_tags.append(tag)
                else:
                    case_tags.append(tag)
                    if tag not in all_case_tags:
                        all_case_tags[tag] = m.get('label') or tag
            tpl_list.append({'id': tpl_row.id, 'name': tpl_row.name, 'filename': tpl_row.filename,
                             'case_tags': case_tags, 'company_tags': company_tags})
        return jsonify({
            'templates': tpl_list,
            'case_tags': [{'key': k, 'label': v} for k, v in all_case_tags.items()],
            'company': _doc_company_data(tid),
        })
    except Exception as _e:
        import traceback as _tb
        _tb.print_exc()
        return jsonify({'templates': [], 'case_tags': [], 'company': {}, 'error': str(_e)}), 500


@app.route("/api/doc-templates/extract-combined", methods=["POST"])
@login_required
@block_super_admin
def api_doc_extract_combined():
    """添付資料から全テンプレートの案件タグをまとめてAI抽出（1回の添付で全書類に反映）"""
    try:
        return _api_doc_extract_combined_impl()
    except Exception as _e:
        import traceback as _tb
        _tb.print_exc()
        return jsonify({'error': f'サーバーエラーが発生しました: {_e}'}), 500


def _api_doc_extract_combined_impl():
    import base64 as _b64
    tid = _doc_tenant_id()
    templates = DocTemplate.query.filter_by(tenant_id=tid, is_active=True).all()
    all_case_tags = {}
    for tpl_row in templates:
        try:
            mapping = json.loads(tpl_row.mapping) if tpl_row.mapping else {}
            if not isinstance(mapping, dict):
                mapping = {}
        except Exception:
            mapping = {}
        for tag, m in mapping.items():
            if not isinstance(m, dict):
                m = {}
            if m.get('scope') != 'company' and tag not in all_case_tags:
                all_case_tags[tag] = m.get('label') or tag
    if not all_case_tags:
        return jsonify({'values': {}, 'found': 0})

    # ファイルはJSONのbase64で受け取る（multipart/form-dataを避けてプロキシ問題を回避）
    payload = request.get_json(silent=True) or {}
    files_data = payload.get('files', [])   # [{name, type, data(base64)}]
    app_id = payload.get('application_id')

    if not files_data:
        return jsonify({'error': '資料ファイルを選択してください'}), 400
    if not os.getenv("ANTHROPIC_API_KEY"):
        return jsonify({'error': 'AI読み取りが未設定です（管理者にお問い合わせください）'}), 503

    content = []
    total = 0
    ext_map = {'pdf': 'application/pdf', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
               'png': 'image/png', 'webp': 'image/webp', 'gif': 'image/gif'}
    for fdata in files_data[:5]:
        if not isinstance(fdata, dict):
            continue
        b64 = fdata.get('data', '')
        if not b64:
            continue
        try:
            raw = _b64.standard_b64decode(b64)
        except Exception:
            continue
        total += len(raw)
        if total > 20 * 1024 * 1024:
            return jsonify({'error': '添付の合計が大きすぎます（20MBまで）'}), 400
        mime = (fdata.get('type') or '').lower()
        if mime not in _ALLOWED_EXTRACT_MIME:
            fname = fdata.get('name', '')
            ext = (fname.rsplit('.', 1)[-1] if '.' in fname else '').lower()
            mime = ext_map.get(ext, mime)
        kind = _ALLOWED_EXTRACT_MIME.get(mime)
        if not kind:
            continue
        if kind == 'pdf':
            content.append({"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}})
        else:
            content.append({"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}})
    if not content:
        return jsonify({'error': 'PDFまたは画像（JPEG/PNG）のみ対応しています'}), 400
    if app_id:
        try:
            rec = ApplicationRecord.query.get(int(app_id))
        except Exception:
            rec = None
        if rec and rec.store_id in get_allowed_store_ids():
            staff = Staff.query.get(rec.staff_id) if rec.staff_id else None
            _fd = lambda d: d.isoformat() if d else ''
            known = {k: v for k, v in {
                'お客様名・契約者名': rec.customer_name or '',
                '物件名': rec.property_name or '',
                '号室・部屋番号': rec.room_number or '',
                '賃料（月額）': int(rec.rent or 0),
                '管理会社': rec.management_company or '',
                '担当者': staff.name if staff else '',
                '契約開始日': _fd(rec.contract_start_date),
                '申込日': _fd(rec.application_date),
            }.items() if v not in ('', 0)}
            if known:
                content.append({"type": "text", "text": "【この案件の既知情報】\n" + "\n".join(f"・{k}: {v}" for k, v in known.items())})

    key_lines = ",\n".join(f'  "{t}": "{all_case_tags[t]}"' for t in all_case_tags)
    prompt = (
        "あなたは不動産書類を読み取る専門アシスタントです。\n"
        "添付資料と既知情報から、以下のJSONキーに対応する値を抽出してください。\n\n"
        "対象フィールド（キー: 説明）:\n{\n" + key_lines + "\n}\n\n"
        "厳守ルール:\n"
        "1. 上記キーだけを持つJSONオブジェクトを1つだけ返す。説明文は不要。\n"
        "2. 読み取れない・自信がない項目は省略する（推測しない）。\n"
        "3. 金額は半角数字のみ、日付はYYYY-MM-DD（和暦→西暦変換）。\n"
    )
    content.append({"type": "text", "text": prompt})

    try:
        msg = anthropic_client.messages.create(
            model="claude-sonnet-4-6", max_tokens=2000,
            messages=[{"role": "user", "content": content}],
        )
        text = "".join(getattr(b, 'text', '') for b in msg.content).strip()
    except Exception as e:
        return jsonify({'error': f'AI読み取りに失敗しました: {e}'}), 502

    extracted = {}
    try:
        s, e = text.find('{'), text.rfind('}')
        if s >= 0 and e > s:
            parsed = json.loads(text[s:e + 1])
            if isinstance(parsed, dict):
                extracted = parsed
    except Exception:
        extracted = {}

    out = {}
    for tag in all_case_tags:
        v = extracted.get(tag)
        if v is None:
            continue
        v = str(v).strip()
        if v:
            out[tag] = v
    return jsonify({'values': out, 'found': len(out)})


# ── 間取り作成 ──────────────────────────────────────────
@app.route("/floorplan")
@login_required
@block_super_admin
def floorplan_page():
    """間取り作成ページ（オプション契約が必要）"""
    if not current_has_floorplan():
        return redirect(url_for('customer_management'))
    return render_template("floorplan.html")


@app.route("/api/floorplans", methods=["GET"])
@login_required
def api_floorplans_list():
    if not current_has_floorplan():
        return jsonify({'error': '間取り作成はオプションプランです'}), 403
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    q = FloorPlan.query
    if sid:
        q = q.filter(FloorPlan.store_id == sid)
    items = q.order_by(FloorPlan.updated_at.desc()).all()
    return jsonify([{'id': f.id, 'name': f.name or '無題',
                     'updated_at': f.updated_at.strftime('%Y-%m-%d %H:%M') if f.updated_at else ''}
                    for f in items])


@app.route("/api/floorplans", methods=["POST"])
@login_required
def api_floorplans_create():
    if not current_has_floorplan():
        return jsonify({'error': '間取り作成はオプションプランです'}), 403
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    data = request.get_json() or {}
    fp_id = data.get('id')
    fp = FloorPlan.query.get(fp_id) if fp_id else None
    if fp:
        if fp.store_id != sid:
            return jsonify({'error': '権限がありません'}), 403
    else:
        fp = FloorPlan(store_id=sid)
        db.session.add(fp)
    fp.name = (data.get('name') or '無題')[:200]
    fp.data = data.get('data') or ''
    fp.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'status': 'ok', 'id': fp.id})


@app.route("/api/floorplans/<int:fid>", methods=["GET"])
@login_required
def api_floorplans_get(fid):
    if not current_has_floorplan():
        return jsonify({'error': '間取り作成はオプションプランです'}), 403
    allowed = get_allowed_store_ids()
    fp = FloorPlan.query.get_or_404(fid)
    if fp.store_id not in allowed:
        return jsonify({'error': '権限がありません'}), 403
    return jsonify({'id': fp.id, 'name': fp.name, 'data': fp.data or ''})


@app.route("/api/floorplans/<int:fid>", methods=["DELETE"])
@login_required
def api_floorplans_delete(fid):
    if not current_has_floorplan():
        return jsonify({'error': '間取り作成はオプションプランです'}), 403
    allowed = get_allowed_store_ids()
    fp = FloorPlan.query.get_or_404(fid)
    if fp.store_id not in allowed:
        return jsonify({'error': '権限がありません'}), 403
    db.session.delete(fp)
    db.session.commit()
    return jsonify({'status': 'ok'})


# ── 反響メール自動取込（IMAP） ─────────────────────────────

# ① 差出人ドメイン/アドレス → 媒体名（部分一致）
PORTAL_DOMAIN_MAP = [
    ('smocca',   'スモッカ'),
    ('suumo',    'SUUMO'),
    ('recruit',  'SUUMO'),
    ('homes',    "HOME'S"),
    ('lifull',   "HOME'S"),
    ('athome',   'アットホーム'),
    ('at-home',  'アットホーム'),
    ('chintai',  'CHINTAI'),
    ('apaman',   'アパマンショップ'),
    ('canary',   'カナリー'),
    ('cnary',    'カナリー'),
    ('ielove',   'いえらぶ'),
    ('tryell',   '住まい探しの窓口'),
    ('sumaisagashi', '住まい探しの窓口'),
    ('eheya',    'エイブル'),
    ('pittat',   'ピタットハウス'),
]

# 通知系（反響ではない）の件名キーワード → 取り込まない
NEG_SUBJECT_KEYWORDS = [
    '内見予約', '来店予約', '予約完了', '予約が完了', '予約確定', '予約受付', '予約のお知らせ',
    'ご予約', '申込完了', '申込が完了', 'お申込み完了', '申込受付', '申込のお知らせ',
    'キャンセル', '解約', '完了しました', '完了のお知らせ', '確定しました', '取消',
    # いい生活Square等の申込ステータス通知（「○○が △△ となりました」）は反響ではない
    'となりました', '申込者記入中', '下書き中', '審査中', '審査完了', 'ステータスを',
    'いい生活Square',
]

# ② 差出人名/件名/本文に含まれるキーワード → 媒体名（ドメインで判定できない場合）
PORTAL_KEYWORD_MAP = [
    # 実際のポータル名を優先（いえらぶCLOUD経由で複数ポータルの反響が届くため、いえらぶより前に置く）
    ('価格.com',     '価格.com'),
    ('価格コム',     '価格.com'),
    ('kakaku',       '価格.com'),
    ('住まい探しの窓口', '住まい探しの窓口'),
    ('住まい探し',   '住まい探しの窓口'),
    ('スモッカ',     'スモッカ'),
    ('SUUMO',        'SUUMO'),
    ('スーモ',       'SUUMO'),
    ("HOME'S",       "HOME'S"),
    ('ホームズ',     "HOME'S"),
    ('アットホーム', 'アットホーム'),
    ('カナリー',     'カナリー'),
    ('CHINTAI',      'CHINTAI'),
    ('アパマン',     'アパマンショップ'),
    ('いえらぶ',     'いえらぶ'),
    ('エイブル',     'エイブル'),
    ('ピタット',     'ピタットハウス'),
]

# 項目ラベルの同義語 → 正規キー（媒体差を吸収）
_FIELD_SYNONYMS = {
    'name':      ['氏名(漢字)', '氏名（漢字）', '氏名', 'お名前', 'お名前(漢字)', 'お名前（漢字）', '名前', 'ご氏名',
                  'お客様名', 'お客さま名', 'お客様氏名', 'ご担当者名', '申込者名', '申込者', 'お申込者', 'ご入居者名', '反響者名',
                  '顧客名', 'ご顧客名', 'ご入居者', '入居者名',
                  'ユーザ様氏名', 'ユーザー様氏名', 'ユーザ氏名', 'ユーザ様お名前', 'ユーザ様名'],
    'phone':     ['電話番号', 'tel', '電話', 'お電話番号', '携帯電話', '携帯番号', '連絡先電話番号', 'ご連絡先',
                  'ユーザ様電話番号', 'ユーザー様電話番号', 'ユーザ電話番号'],
    'email':     ['メールアドレス', 'email', 'e-mail', 'mail', 'メール', 'eメール',
                  'ユーザ様メールアドレス', 'ユーザー様メールアドレス', 'ユーザメールアドレス'],
    'property':  ['物件名', '建物名', 'マンション名', 'お問い合わせ物件名', 'お問合せ物件名', '問い合わせ物件名'],
    'room':      ['部屋番号', '号室'],
    'rent':      ['賃料', '家賃'],
    'address':   ['住所', '所在地', '物件所在地'],
    'madori':    ['間取り', '間取'],
    'menseki':   ['面積', '専有面積'],
    'station':   ['最寄駅', '最寄り駅'],
    'datetime':  ['お問合わせ日時', 'お問い合わせ日時', 'お問合せ日時', '問合せ日時', '問い合わせ日時', '受付日時', '反響日時', '受信日時'],
    'extid':     ['スモッカ反響id', '反響id', '問い合わせ番号', 'お問い合わせ番号', '受付番号', '反響番号', '問合せ番号'],
    'inquiry':   ['お問合せ内容', 'お問い合わせ内容', '物件に関するお問合せ内容', 'お問合せ', 'ご要望', '希望内容'],
    'bukken_no': ['物件管理番号', '物件番号', '管理番号', '自社管理番号'],
    'note':      ['備考', 'その他', 'ご質問', 'メッセージ', 'コメント'],
}


# 「問合せ／問い合わせ／問い合せ／問合わせ」等の表記ゆれを「問合せ」に統一
_INQUIRY_RE = re.compile(r'問\s*い?\s*合\s*わ?\s*せ')


def _normalize_inquiry_terms(s):
    return _INQUIRY_RE.sub('問合せ', s or '')


def _norm_label(s):
    s = (s or '').replace(' ', '').replace('　', '').strip().lower()
    return _normalize_inquiry_terms(s)


# 正規化済みラベル → キー
_LABEL_TO_KEY = {}
for _k, _syns in _FIELD_SYNONYMS.items():
    for _s in _syns:
        _LABEL_TO_KEY[_norm_label(_s)] = _k


def _decode_mime(s):
    if not s:
        return ''
    try:
        out = ''
        for txt, enc in _decode_header(s):
            if isinstance(txt, bytes):
                out += txt.decode(enc or 'utf-8', errors='replace')
            else:
                out += txt
        return out
    except Exception:
        return str(s)


def _email_plain_body(msg):
    """メールから本文(プレーンテキスト)を取得"""
    body = ''
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == 'text/plain' and 'attachment' not in str(part.get('Content-Disposition') or ''):
                payload = part.get_payload(decode=True)
                if payload:
                    charset = part.get_content_charset() or 'utf-8'
                    try:
                        body += payload.decode(charset, errors='replace')
                    except Exception:
                        body += payload.decode('utf-8', errors='replace')
        if not body:
            for part in msg.walk():
                if part.get_content_type() == 'text/html':
                    payload = part.get_payload(decode=True)
                    if payload:
                        charset = part.get_content_charset() or 'utf-8'
                        body += re.sub(r'<[^>]+>', ' ', payload.decode(charset, errors='replace'))
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            charset = msg.get_content_charset() or 'utf-8'
            body = payload.decode(charset, errors='replace')
    return body


def _email_attachments(msg):
    """受信メールから添付ファイル [(filename, content_type, bytes), ...] を取り出す。"""
    out = []
    if not msg.is_multipart():
        return out
    for part in msg.walk():
        if part.get_content_maintype() == 'multipart':
            continue
        cd = str(part.get('Content-Disposition') or '')
        fname = part.get_filename()
        if fname:
            try:
                fname = _decode_mime(fname)
            except Exception:
                pass
        # 添付 or インライン画像のみ（本文テキスト/HTMLは除外）
        is_attach = ('attachment' in cd.lower()) or bool(fname)
        if not is_attach:
            continue
        try:
            raw = part.get_payload(decode=True)
        except Exception:
            raw = None
        if not raw:
            continue
        ctype = part.get_content_type() or 'application/octet-stream'
        if not fname:
            ext = (ctype.split('/')[-1] or 'bin')[:6]
            fname = f'attachment.{ext}'
        out.append((fname[:300], ctype[:120], raw))
    return out


APP_BASE_URL = os.getenv('APP_BASE_URL', 'https://app.mieroom.cloud')


def _fmt_jst(dt, fmt='%Y-%m-%d %H:%M'):
    """UTC保存の日時を日本時間(JST=UTC+9)の文字列に変換"""
    if not dt:
        return ''
    return (dt + timedelta(hours=9)).strftime(fmt)
_IMG_EXTS = ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp')


def _text_to_html(text):
    """プレーンテキスト本文をHTMLに（改行→<br>、エスケープ）"""
    safe = (str(text or '')
            .replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))
    safe = safe.replace('\n', '<br>')
    return f'<div style="font-size:14px;line-height:1.7;color:#111827;white-space:normal;">{safe}</div>'


def parse_custom_keywords(text):
    """設定の追加キーワード文字列を [(キーワード, 媒体名), ...] に変換。
    1行＝「語」または「語=媒体名」「語,媒体名」「語：媒体名」。"""
    rules = []
    for raw in (text or '').replace('\r', '\n').split('\n'):
        line = raw.strip()
        if not line:
            continue
        m = re.split(r'[=＝,，：:]', line, 1)
        kw = m[0].strip()
        media = m[1].strip() if len(m) > 1 and m[1].strip() else kw
        if kw:
            rules.append((kw, media))
    return rules


def _detect_source(from_addr, subject='', body='', extra_map=None):
    """媒体名を ①ドメイン → ②追加キーワード → ③標準キーワード の順で判定。不明なら None。"""
    a = (from_addr or '').lower()
    for key, name in PORTAL_DOMAIN_MAP:
        if key in a:
            return name
    hay = f"{from_addr}\n{subject}\n{body[:800]}".upper()
    for key, name in (extra_map or []):
        if key and key.upper() in hay:
            return name
    for key, name in PORTAL_KEYWORD_MAP:
        if key.upper() in hay:
            return name
    return None


def _from_display_name(from_addr):
    """差出人名（または ドメイン名）を媒体名のフォールバックとして返す"""
    s = from_addr or ''
    m = re.match(r'\s*"?([^"<]+?)"?\s*<', s)
    if m:
        nm = m.group(1).strip()
        if nm and '@' not in nm:
            return nm[:40]
    md = re.search(r'@([\w.-]+)', s)
    if md:
        parts = md.group(1).split('.')
        if len(parts) >= 2:
            return parts[-2]
    return 'その他反響'


def parse_reaction_email(msg, extra_map=None, portal_map=None):
    """反響メールを解析。反響メールでなければ None を返す。
    登録ポータル（portal_map）からの受信は信頼扱いで確実に取り込む。
    未知の媒体でも、氏名/連絡先＋物件などの構造があれば取り込む。"""
    from_addr = _decode_mime(msg.get('From', ''))
    subject = _decode_mime(msg.get('Subject', ''))
    body = _email_plain_body(msg)
    if not body:
        return None

    fields = {}
    for raw_line in body.splitlines():
        line = raw_line.strip()
        if not line or ('：' not in line and ':' not in line):
            continue
        parts = re.split(r'[：:]', line, 1)
        if len(parts) != 2:
            continue
        label = _norm_label(parts[0])
        value = parts[1].strip()
        if not label or value in ('', '－', '-', 'ー', '−'):
            continue
        key = _LABEL_TO_KEY.get(label)
        if key and key not in fields:
            fields[key] = value

    # ＜ラベル＞/<ラベル> が単独行で、値が次の行にある形式（住まい探しの窓口 等）にも対応
    _blines = body.splitlines()
    for _i, _raw in enumerate(_blines):
        _ml = re.match(r'^\s*[<＜](.+?)[>＞]\s*$', _raw.strip())
        if not _ml:
            continue
        _key = _LABEL_TO_KEY.get(_norm_label(_ml.group(1)))
        if not _key or _key in fields:
            continue
        # 次の非空行を値として採用（次のラベル行に当たったら値なし）
        for _j in range(_i + 1, min(_i + 4, len(_blines))):
            _v = _blines[_j].strip()
            if not _v:
                continue
            if re.match(r'^[<＜].+[>＞]\s*$', _v):
                break
            if _v in ('－', '-', 'ー', '−'):
                break
            fields[_key] = _v
            break

    name = re.sub(r'[\s　]*(様|さま|殿)\s*$', '', fields.get('name', '')).strip()
    contact = name or fields.get('phone') or fields.get('email')
    prop = (fields.get('property') or fields.get('extid')
            or fields.get('bukken_no') or fields.get('inquiry'))

    # 媒体判定（優先度：登録ポータル ＞ ドメイン ＞ 追加KW ＞ 標準KW ＞ 差出人名）
    portal_media = None
    _fa = (from_addr or '').lower()
    for matcher, media in (portal_map or []):
        if matcher and matcher.lower() in _fa:
            portal_media = media
            break

    # 通知系（内見予約・申込完了・キャンセル等）は反響ではない → 除外
    if any(k in subject for k in NEG_SUBJECT_KEYWORDS):
        return None

    # 件名が「お問合せ受付」等の反響を示すか（表記ゆれを統一して判定）
    subj_norm = _normalize_inquiry_terms(subject)
    subj_is_inquiry = any(k in subj_norm for k in ('問合せ', '受付', '反響'))
    has_name = bool(name)

    if portal_media is not None:
        # 登録ポータルからの受信は信頼扱い：氏名か物件があれば取り込む
        if not (has_name or prop):
            return None
        source = portal_media
    else:
        # 未登録の差出人は精度重視：お客様の氏名が必須（＋物件 or 件名が反響）
        if not (has_name and (prop or subj_is_inquiry)):
            return None
        source = _detect_source(from_addr, subject, body, extra_map) or _from_display_name(from_addr)

    # 反響日時
    edate = None
    dt_raw = fields.get('datetime', '')
    mdt = re.search(r'(\d{4})[/\-年](\d{1,2})[/\-月](\d{1,2})', dt_raw)
    if mdt:
        try:
            edate = date(int(mdt.group(1)), int(mdt.group(2)), int(mdt.group(3)))
        except Exception:
            edate = None
    if edate is None:
        try:
            dh = emaillib.utils.parsedate_to_datetime(msg.get('Date'))
            if dh:
                edate = dh.date()
        except Exception:
            edate = None
    if edate is None:
        edate = date.today()

    # 一意ID（重複取込防止）
    raw_id = fields.get('extid', '')
    if raw_id:
        ext = f"{source}:{raw_id}"
    else:
        basis = f"{source}|{name}|{fields.get('property','')}|{fields.get('bukken_no','')}|{dt_raw}"
        ext = 'h:' + hashlib.md5(basis.encode('utf-8')).hexdigest()[:20]

    # 自由記述（お問合せ内容・備考）も拾う（表記ゆれを統一した本文で検索）
    nbody = _normalize_inquiry_terms(body)

    def _grab(markers):
        for mk in markers:
            nmk = _normalize_inquiry_terms(mk)
            i = nbody.find(nmk)
            if i < 0:
                continue
            collected = []
            for ln in nbody[i + len(nmk):].splitlines():
                s = ln.strip().lstrip('・').strip()
                if not s:
                    if collected:
                        break
                    continue
                # 区切り線/セクション見出し/ラベル行で終了
                if s.startswith('＜') or s.startswith('■') or set(s) <= set('-—─=＝_＿ 　'):
                    if collected:
                        break
                    continue
                if '：' in s or ':' in s:
                    if collected:
                        break
                    continue
                collected.append(s)
                if len(collected) >= 4:
                    break
            if collected:
                return ' '.join(collected)[:300]
        return ''

    inquiry_free = fields.get('inquiry') or _grab(['＜物件に関するお問合せ内容＞', '物件に関するお問合せ内容', '＜お問合せ内容＞', 'お問合せ内容'])
    biko_free = fields.get('note') or _grab(['■備考', '＜問い合わせ詳細＞', '備考'])

    # メモ生成（拾えた項目のみ）
    memo_lines = []
    for lbl, key in [('物件', 'property'), ('部屋番号', 'room'), ('賃料', 'rent'),
                     ('間取り', 'madori'), ('面積', 'menseki'), ('最寄駅', 'station'),
                     ('住所', 'address'), ('物件管理番号', 'bukken_no'),
                     ('電話', 'phone'), ('メール', 'email')]:
        v = fields.get(key)
        if v:
            memo_lines.append(f"{lbl}：{v}")
    if inquiry_free:
        memo_lines.append(f"お問合せ内容：{inquiry_free}")
    if biko_free:
        memo_lines.append(f"備考：{biko_free}")
    if raw_id:
        memo_lines.append(f"反響ID：{raw_id}")

    # メールアドレス抽出（値に紛れた場合も拾う）
    cust_email = fields.get('email', '')
    if cust_email:
        m_em = re.search(r'[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}', cust_email)
        cust_email = m_em.group(0) if m_em else ''

    return {
        'source': source,
        'name': name or '（氏名不明）',
        'date': edate,
        'memo': '\n'.join(memo_lines),
        'external_id': ext,
        'has_phone': bool(fields.get('phone')),
        'email': cust_email,
    }


def _ipv4_addr(host, port):
    """ホストのIPv4アドレス (ip, port) を返す。IPv6に経路が無い環境への対策。"""
    infos = socket.getaddrinfo(host, port, socket.AF_INET, socket.SOCK_STREAM)
    return infos[0][4]


class _IMAP4SSLIPv4(imaplib.IMAP4_SSL):
    """imap.gmail.com に必ずIPv4で接続するIMAP4_SSL（本番のIPv6経路なし対策）"""
    def _create_socket(self, timeout=None):
        sock = socket.create_connection(_ipv4_addr(self.host, self.port),
                                        timeout, getattr(self, 'source_address', None))
        return self.ssl_context.wrap_socket(sock, server_hostname=self.host)


# ── Google OAuth / Gmail API ─────────────────────────────────
# RailwayがSMTP送信ポートを塞ぐ環境向け：送信はGmail API(HTTPS)、受信はIMAP+XOAUTH2。
GOOGLE_CLIENT_ID = os.getenv('GOOGLE_CLIENT_ID', '')
GOOGLE_CLIENT_SECRET = os.getenv('GOOGLE_CLIENT_SECRET', '')
GOOGLE_OAUTH_SCOPE = 'https://mail.google.com/'   # IMAP + Gmail API send を包含
_google_token_cache = {}   # refresh_token -> (access_token, expiry_epoch)


def _http_post_form(url, data):
    import urllib.request, urllib.parse, json as _json
    body = urllib.parse.urlencode(data).encode()
    req = urllib.request.Request(url, data=body,
                                 headers={'Content-Type': 'application/x-www-form-urlencoded'},
                                 method='POST')
    with urllib.request.urlopen(req, timeout=20) as resp:
        return _json.loads(resp.read().decode('utf-8'))


def _google_access_token(refresh_token):
    """refresh_token から access_token を取得（キャッシュ付き）"""
    import time as _t
    if not refresh_token:
        raise RuntimeError('Google連携が未設定です')
    cached = _google_token_cache.get(refresh_token)
    if cached and cached[1] - 60 > _t.time():
        return cached[0]
    res = _http_post_form('https://oauth2.googleapis.com/token', {
        'client_id': GOOGLE_CLIENT_ID,
        'client_secret': GOOGLE_CLIENT_SECRET,
        'refresh_token': refresh_token,
        'grant_type': 'refresh_token',
    })
    at = res.get('access_token')
    if not at:
        raise RuntimeError('アクセストークンの取得に失敗しました')
    _google_token_cache[refresh_token] = (at, _t.time() + int(res.get('expires_in', 3600)))
    return at


def _gmail_profile_email(access_token):
    import urllib.request, json as _json
    req = urllib.request.Request('https://gmail.googleapis.com/gmail/v1/users/me/profile',
                                 headers={'Authorization': 'Bearer ' + access_token})
    with urllib.request.urlopen(req, timeout=20) as resp:
        return _json.loads(resp.read().decode('utf-8')).get('emailAddress')


def _gmail_api_send(refresh_token, raw_message_bytes):
    """Gmail API でメール送信（SMTPが使えない環境向け）"""
    import urllib.request, json as _json, base64
    token = _google_access_token(refresh_token)
    raw = base64.urlsafe_b64encode(raw_message_bytes).decode()
    body = _json.dumps({'raw': raw}).encode()
    req = urllib.request.Request('https://gmail.googleapis.com/gmail/v1/users/me/messages/send',
                                 data=body,
                                 headers={'Authorization': 'Bearer ' + token,
                                          'Content-Type': 'application/json'},
                                 method='POST')
    with urllib.request.urlopen(req, timeout=30) as resp:
        return _json.loads(resp.read().decode('utf-8'))


def _open_imap_raw(host, imap_user, imap_pass, oauth_refresh_token=None, oauth_email=None):
    """認証済みIMAP接続を返す（OAuth(XOAUTH2)優先・無ければapp-password）"""
    M = _IMAP4SSLIPv4(host or 'imap.gmail.com')
    if oauth_refresh_token:
        at = _google_access_token(oauth_refresh_token)
        user = oauth_email or imap_user
        auth_str = f'user={user}\x01auth=Bearer {at}\x01\x01'
        M.authenticate('XOAUTH2', lambda _x: auth_str.encode())
    else:
        M.login(imap_user, imap_pass)
    return M


def _open_imap(ms):
    return _open_imap_raw(ms.imap_host, ms.imap_user, ms.imap_pass,
                          ms.oauth_refresh_token, ms.oauth_email)


def test_imap_connection(host, user, password):
    try:
        M = _IMAP4SSLIPv4(host or 'imap.gmail.com')
        M.login(user, password)
        M.select('INBOX')
        try:
            M.logout()
        except Exception:
            pass
        return True, '接続に成功しました'
    except imaplib.IMAP4.error as e:
        return False, f'ログインに失敗しました（メールアドレス／アプリパスワードをご確認ください）: {e}'
    except Exception as e:
        return False, f'接続エラー：{e}'


def _msg_datetime(msg):
    """メールのDateヘッダを naive UTC datetime で返す。無ければ None。"""
    try:
        dh = emaillib.utils.parsedate_to_datetime(msg.get('Date'))
        if not dh:
            return None
        if dh.tzinfo:
            return dh.astimezone(timezone.utc).replace(tzinfo=None)
        return dh
    except Exception:
        return None


def _norm_name_key(s):
    """氏名の比較用キー（空白・全角空白を除去）"""
    return re.sub(r'[\s　]+', '', (s or '')).strip()


_UNKNOWN_NAME_KEYS = {'', '氏名不明', '不明', 'お客様', 'お名前不明', 'ー', '―'}


def _find_merge_target(store_id, name, email):
    """同一お客様の既存反響を探す（同名・メール不一致なら別人として除外）。無ければ None。"""
    nk = _norm_name_key(name)
    if not nk or nk in _UNKNOWN_NAME_KEYS:
        return None
    email_l = (email or '').strip().lower()
    cands = (EchoRecord.query.filter_by(store_id=store_id)
             .order_by(EchoRecord.id.desc()).limit(800).all())
    for c in cands:
        if _norm_name_key(c.list_name) != nk:
            continue
        cem = (c.customer_email or '').strip().lower()
        if email_l and cem and email_l != cem:
            continue   # メールが両方あって違う → 別人なのでまとめない
        return c
    return None


def _merge_into_echo(target, parsed):
    """既存反響に、追加反響の内容（物件名など）をメモへ追記して1件にまとめる。"""
    d = parsed.get('date')
    ds = f"{d.month}/{d.day}" if hasattr(d, 'month') else str(d or '')
    media = parsed.get('source') or ''
    info = (parsed.get('memo') or '').strip()
    head = f"【追加反響 {ds}" + (f" / {media}" if media else "") + "】"
    add = ("\n─────\n" + head + ("\n" + info if info else "")).rstrip()
    target.memo = ((target.memo or '').rstrip() + add).strip()
    # 不足情報を補完
    if not target.customer_email and parsed.get('email'):
        target.customer_email = parsed.get('email')
    if parsed.get('has_phone') and not target.has_phone_number:
        target.has_phone_number = True


def fetch_reactions_for_store(store_id, limit=120, since_days=30):
    """指定店舗のGmailから反響メールを取得し EchoRecord へ登録"""
    ms = MailSetting.query.filter_by(store_id=store_id).first()
    if not ms or not ((ms.imap_user and ms.imap_pass) or ms.oauth_refresh_token):
        return {'ok': False, 'error': '未設定', 'imported': 0, 'scanned': 0}
    extra_map = parse_custom_keywords(ms.custom_keywords)
    portal_map = [(p.matcher, p.media)
                  for p in PortalSource.query.filter_by(store_id=store_id, enabled=True).all()
                  if p.matcher and p.media]
    # 取込開始日時：未設定なら「今」にして、過去メールは取り込まない
    if ms.import_after is None:
        ms.import_after = datetime.utcnow()
        db.session.commit()
    import_after = ms.import_after

    imported = 0
    merged = 0
    scanned = 0
    new_exts = []
    try:
        M = _open_imap(ms)
        M.select('INBOX')
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        # IMAP検索は取込開始日以降に限定（過去分を遡らない）
        since = import_after.date()
        floor = date.today() - timedelta(days=since_days)
        if since < floor:
            since = floor
        since_str = f"{since.day:02d}-{months[since.month - 1]}-{since.year}"
        typ, data = M.search(None, f'(SINCE "{since_str}")')
        ids = data[0].split() if (data and data[0]) else []
        ids = ids[-limit:]
        for num in reversed(ids):
            try:
                typ, md = M.fetch(num, '(BODY.PEEK[])')
                if not md or not md[0]:
                    continue
                msg = emaillib.message_from_bytes(md[0][1])
            except Exception:
                continue
            scanned += 1
            # 取込開始日時より前のメールは無視（過去の反響を拾わない）
            mdt = _msg_datetime(msg)
            if mdt and import_after and mdt < import_after:
                continue
            parsed = parse_reaction_email(msg, extra_map, portal_map)
            if not parsed:
                # 反響でなければ、お客様からの返信かどうかを判定して会話に取り込む
                try:
                    _handle_incoming_reply(store_id, msg)
                except Exception as e:
                    print(f"reply handle error: {e}")
                continue
            ext = parsed['external_id']
            # 取込済み or 削除済み or 既存 → 取り込まない（削除後の復活を防ぐ）
            if (ProcessedReaction.query.filter_by(store_id=store_id, external_id=ext).first()
                    or EchoRecord.query.filter_by(store_id=store_id, external_id=ext).first()):
                continue
            # 同一お客様の既存反響があれば、新規行を作らずメモへ追記して1件にまとめる
            target = _find_merge_target(store_id, parsed['name'], parsed.get('email'))
            if target:
                _merge_into_echo(target, parsed)
                db.session.add(ProcessedReaction(store_id=store_id, external_id=ext))
                merged += 1
                continue
            db.session.add(EchoRecord(
                store_id=store_id,
                staff_id=ms.default_staff_id or None,
                list_name=parsed['name'],
                echo_date=parsed['date'],
                media=parsed['source'],
                method='メール',
                memo=parsed['memo'],
                has_phone=False,   # 電話対応は自動で〇にしない
                has_phone_number=bool(parsed.get('has_phone')),  # 電話番号の有無
                external_id=ext,
                customer_email=parsed.get('email') or None,
            ))
            db.session.add(ProcessedReaction(store_id=store_id, external_id=ext))
            new_exts.append(ext)
            imported += 1
        db.session.commit()
        try:
            M.close()
            M.logout()
        except Exception:
            pass
        # 新着反響への自動返信（設定ONかつお客様メールあり。媒体ごとにテンプレを切替）
        if new_exts and ms.auto_reply_enabled:
            # 媒体名 → テンプレID（ポータル登録ごとの指定）
            media_tpl = {}
            for p in PortalSource.query.filter_by(store_id=store_id).all():
                if p.media and p.auto_reply_template_id:
                    media_tpl.setdefault(p.media, p.auto_reply_template_id)
            default_tpl_id = ms.auto_reply_template_id
            _tpl_cache = {}
            def _get_tpl(tid):
                if not tid:
                    return None
                if tid not in _tpl_cache:
                    _tpl_cache[tid] = MailTemplate.query.get(tid)
                return _tpl_cache[tid]
            for ext in new_exts:
                rec = EchoRecord.query.filter_by(store_id=store_id, external_id=ext).first()
                if not rec or not rec.customer_email:
                    continue
                tpl = _get_tpl(media_tpl.get(rec.media)) or _get_tpl(default_tpl_id)
                if not tpl:
                    continue
                try:
                    send_mail_for_store(store_id, rec.id,
                                        tpl.subject or 'お問い合わせありがとうございます',
                                        tpl.body or '', base_url=APP_BASE_URL,
                                        is_html=bool(tpl.is_html))
                except Exception as e:
                    print(f"auto-reply send error (echo={rec.id}): {e}")
        ms.last_fetch_at = datetime.utcnow()
        mtxt = f" / {merged}件まとめ" if merged else ""
        ms.last_result = f"取得OK：{imported}件追加{mtxt} / {scanned}件確認（{_fmt_jst(datetime.utcnow(), '%m/%d %H:%M')}）"
        db.session.commit()
        return {'ok': True, 'imported': imported, 'merged': merged, 'scanned': scanned}
    except imaplib.IMAP4.error as e:
        db.session.rollback()
        try:
            ms.last_result = f"ログイン失敗：{e}"
            db.session.commit()
        except Exception:
            db.session.rollback()
        return {'ok': False, 'error': f'ログインに失敗しました（メールアドレス／アプリパスワードをご確認ください）', 'imported': imported, 'scanned': scanned}
    except Exception as e:
        db.session.rollback()
        try:
            ms.last_result = f"エラー：{e}"
            db.session.commit()
        except Exception:
            db.session.rollback()
        return {'ok': False, 'error': str(e), 'imported': imported, 'scanned': scanned}


# ── メール会話（送信／返信取り込み） ─────────────────────

def _addr_only(header_val):
    """ヘッダ値からメールアドレスのみ抽出"""
    try:
        return (_parseaddr(_decode_mime(header_val or ''))[1] or '').strip().lower()
    except Exception:
        return ''


def _handle_incoming_reply(store_id, msg):
    """反響でない受信メールが、既存反響への返信か判定して会話に取り込む。取り込んだら True。"""
    message_id = (msg.get('Message-ID') or '').strip()
    if message_id and MailMessage.query.filter_by(store_id=store_id, message_id=message_id).first():
        return False  # 取り込み済み

    refs = f"{msg.get('In-Reply-To','')} {msg.get('References','')}"
    echo_id = None
    m = re.search(r'mieroom-(\d+)-', refs)
    if m:
        eid = int(m.group(1))
        rec = EchoRecord.query.filter_by(id=eid, store_id=store_id).first()
        if rec:
            echo_id = rec.id
    if echo_id is None:
        from_email = _addr_only(msg.get('From'))
        if from_email:
            rec = (EchoRecord.query
                   .filter_by(store_id=store_id, customer_email=from_email)
                   .order_by(EchoRecord.id.desc()).first())
            if rec:
                echo_id = rec.id
    if echo_id is None:
        return False  # 既存反響に紐づかない → 取り込まない

    body = _email_plain_body(msg)
    mm = MailMessage(
        store_id=store_id, echo_id=echo_id, direction='in',
        from_addr=_decode_mime(msg.get('From', ''))[:300],
        to_addr=_decode_mime(msg.get('To', ''))[:300],
        subject=_decode_mime(msg.get('Subject', ''))[:500],
        body=body, message_id=message_id[:300] or None,
        in_reply_to=(msg.get('In-Reply-To') or '')[:300] or None,
        is_read=False,
    )
    db.session.add(mm)
    db.session.flush()  # mm.id 採番（添付の紐付け用）
    # 受信した添付ファイルを保存
    try:
        for (fname, ctype, raw) in _email_attachments(msg):
            if raw and len(raw) <= 20 * 1024 * 1024:   # 20MB上限
                db.session.add(MailAttachment(
                    message_id=mm.id, store_id=store_id,
                    filename=fname, content_type=ctype,
                    size=len(raw), data=raw))
    except Exception as e:
        print(f"  添付保存スキップ: {e}")
    rec = EchoRecord.query.get(echo_id)
    if rec:
        rec.has_unread_reply = True
        rec.reply_dismissed = False  # 新たな返信が来たら未返信アラートを再表示
    db.session.commit()
    return True


class _SMTPSSLIPv4(smtplib.SMTP_SSL):
    """smtp.gmail.com に必ずIPv4で接続するSMTP_SSL（証明書はホスト名で検証）"""
    def _get_socket(self, host, port, timeout):
        sock = socket.create_connection(_ipv4_addr(host, port), timeout, self.source_address)
        return self.context.wrap_socket(sock, server_hostname=self._host)


class _SMTPIPv4(smtplib.SMTP):
    """IPv4固定のSMTP（587 STARTTLS フォールバック用）"""
    def _get_socket(self, host, port, timeout):
        return socket.create_connection(_ipv4_addr(host, port), timeout, self.source_address)


def _smtp_deliver(host, user, pw, from_addr, to_addrs, msg_string):
    """IPv4でSMTP送信。465(SSL)失敗時は587(STARTTLS)へフォールバック。"""
    import ssl as _ssl
    ctx = _ssl.create_default_context()
    last_err = None
    # 1) 465 SMTPS（IPv4）
    try:
        s = _SMTPSSLIPv4(host, 465, local_hostname='mieroom.cloud', timeout=40, context=ctx)
        try:
            s.login(user, pw)
            s.sendmail(from_addr, to_addrs, msg_string)
        finally:
            try: s.quit()
            except Exception: pass
        return
    except Exception as e:
        last_err = e
    # 2) 587 STARTTLS（IPv4）フォールバック
    try:
        s = _SMTPIPv4(host, 587, local_hostname='mieroom.cloud', timeout=40)
        try:
            s.ehlo(); s.starttls(context=ctx); s.ehlo()
            s.login(user, pw)
            s.sendmail(from_addr, to_addrs, msg_string)
        finally:
            try: s.quit()
            except Exception: pass
        return
    except Exception as e2:
        last_err = e2
    raise last_err


def _split_name(full):
    parts = re.split(r'[\s　]+', (full or '').strip(), 1)
    last = parts[0] if parts and parts[0] else ''
    first = parts[1].strip() if len(parts) > 1 else ''
    return last, first


def _extract_property(memo):
    """メモから「物件：XXX」の最初の物件名を取り出す"""
    m = re.search(r'物件\s*[:：]\s*(.+)', memo or '')
    return m.group(1).strip() if m else ''


# 差し込み文字の一覧（UIの凡例と一致させる）
TEMPLATE_VARS = [
    ('#name#', 'お客様の氏名'),
    ('#lastName#', 'お客様の姓'),
    ('#firstName#', 'お客様の名'),
    ('#お問い合わせ物件名#', 'お問い合わせ物件名'),
    ('#物件名#', '物件名（同上）'),
    ('#会社名#', '会社名'),
    ('#店舗名#', '店舗名'),
    ('#会社電話番号#', '会社の電話番号'),
    ('#FAX番号#', 'FAX番号'),
    ('#会社メールアドレス#', '会社のメールアドレス'),
    ('#会社住所#', '会社の住所'),
    ('#代表者氏名#', '代表者氏名'),
    ('#宅建業免許番号#', '宅建業免許番号'),
    ('#宅建業免許取得日#', '宅建業免許取得日'),
    ('#会社営業時間#', '営業時間'),
    ('#会社定休日#', '定休日'),
    ('#インボイス番号#', 'インボイス登録番号'),
    ('#公式LINE#', '公式LINEのURL'),
]


def _apply_template_vars(text, rec, store_id):
    """テンプレ内の差し込み文字（#name# 等）を実データで置換"""
    if not text or '#' not in text:
        return text or ''
    cp = CompanyProfile.query.filter_by(store_id=store_id).first()
    last, first = _split_name(rec.list_name if rec else '')
    prop = _extract_property(rec.memo if rec else '')
    repl = {
        '#name#': (rec.list_name if rec else ''),
        '#lastName#': last, '#firstName#': first,
        '#お問い合わせ物件名#': prop, '#物件名#': prop,
        '#会社名#': (cp.company_name if cp else ''),
        '#店舗名#': (cp.store_name if cp else ''),
        '#会社電話番号#': (cp.phone if cp else ''),
        '#FAX番号#': (cp.fax if cp else ''),
        '#会社メールアドレス#': (cp.email if cp else ''),
        '#会社住所#': (cp.address if cp else ''),
        '#代表者氏名#': (cp.representative if cp else ''),
        '#宅建業免許番号#': (cp.license_number if cp else ''),
        '#宅建業免許取得日#': (cp.license_date if cp else ''),
        '#会社営業時間#': (cp.business_hours if cp else ''),
        '#会社定休日#': (cp.holidays if cp else ''),
        '#インボイス番号#': (cp.invoice_number if cp else ''),
        '#公式LINE#': (cp.line_url if cp else ''),
        '#LINE#': (cp.line_url if cp else ''),
    }
    for k, v in repl.items():
        text = text.replace(k, v or '')
    return text


def _html_to_text(html):
    """HTML本文をプレーンテキスト化（text/plain代替・チャット表示用）"""
    t = re.sub(r'(?i)<br\s*/?>', '\n', html or '')
    t = re.sub(r'(?i)</p>', '\n', t)
    t = re.sub(r'<[^>]+>', '', t)
    t = (t.replace('&nbsp;', ' ').replace('&amp;', '&')
         .replace('&lt;', '<').replace('&gt;', '>').replace('&quot;', '"'))
    return t.strip()


def send_mail_for_store(store_id, echo_id, subject, body, attachments=None, base_url=None, is_html=False):
    """店舗のGmail(SMTP)からお客様へ送信し、会話に保存。
    attachments: [(filename, content_type, bytes), ...]
    base_url: 既読トラッキング用ピクセルの絶対URLベース
    is_html: body がHTMLか（リッチテンプレ）"""
    ms = MailSetting.query.filter_by(store_id=store_id).first()
    use_oauth = bool(ms and ms.oauth_refresh_token)
    if not ms or not (use_oauth or (ms.imap_user and ms.imap_pass)):
        return {'ok': False, 'error': 'メール設定が未設定です'}
    sender = (ms.oauth_email or ms.imap_user or '').strip()
    rec = EchoRecord.query.filter_by(id=echo_id, store_id=store_id).first()
    if not rec:
        return {'ok': False, 'error': '対象の反響が見つかりません'}
    to_addr = (rec.customer_email or '').strip()
    if not to_addr:
        return {'ok': False, 'error': 'お客様のメールアドレスが未登録です'}

    # 差し込み文字（#name# 等）を実データに置換
    subject = _apply_template_vars(subject, rec, store_id)
    body = _apply_template_vars(body, rec, store_id)
    # チャット表示・保存用の本文（HTMLならテキスト化）
    display_body = _html_to_text(body) if is_html else body

    # スレッド情報（最後の受信メッセージに返信する形）
    last_in = (MailMessage.query
               .filter_by(store_id=store_id, echo_id=echo_id, direction='in')
               .order_by(MailMessage.id.desc()).first())
    msg_id = f"<mieroom-{echo_id}-{int(time.time()*1000)}@mieroom.cloud>"

    # 先に会話レコードを作成してIDを採番（既読ピクセルURLに使う）
    mm = MailMessage(
        store_id=store_id, echo_id=echo_id, direction='out',
        from_addr=sender, to_addr=to_addr,
        subject=subject or '', body=display_body or '',
        message_id=msg_id, is_read=True,
    )
    db.session.add(mm)
    db.session.flush()

    base = (base_url or APP_BASE_URL).rstrip('/')
    pixel = f'<img src="{base}/m/o/{mm.id}.gif" width="1" height="1" alt="" style="display:none">'
    if is_html:
        html_body = (body or '') + pixel
    else:
        html_body = _text_to_html(body or '') + pixel

    msg = MIMEMultipart('mixed')
    msg['Subject'] = subject or '（無題）'
    msg['From'] = sender
    msg['To'] = to_addr
    msg['Message-ID'] = msg_id
    if last_in and last_in.message_id:
        msg['In-Reply-To'] = last_in.message_id
        msg['References'] = last_in.message_id

    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(display_body or '', 'plain', 'utf-8'))
    alt.attach(MIMEText(html_body, 'html', 'utf-8'))
    msg.attach(alt)

    atts = attachments or []
    for (fname, ctype, raw) in atts:
        maintype, _, subtype = (ctype or 'application/octet-stream').partition('/')
        if not subtype:
            maintype, subtype = 'application', 'octet-stream'
        part = MIMEBase(maintype, subtype)
        part.set_payload(raw)
        _email_encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=str(fname))
        msg.attach(part)

    try:
        if use_oauth:
            _gmail_api_send(ms.oauth_refresh_token, msg.as_bytes())
        else:
            host = (ms.imap_host or 'imap.gmail.com').replace('imap.', 'smtp.')
            _smtp_deliver(host, ms.imap_user, ms.imap_pass, sender, [to_addr], msg.as_string())
    except Exception as e:
        db.session.rollback()
        return {'ok': False, 'error': f'送信に失敗しました：{e}'}

    # 添付を保存（送信成功後）
    for (fname, ctype, raw) in atts:
        db.session.add(MailAttachment(
            message_id=mm.id, store_id=store_id,
            filename=str(fname)[:300], content_type=(ctype or 'application/octet-stream')[:120],
            size=len(raw), data=raw))
    db.session.commit()
    return {'ok': True}


# ── 自動取込サービス：IMAP IDLE（リアルタイム push） ──
# gunicorn複数ワーカーでもソケットlockで1プロセスのみがIDLE接続を保持する。
_MAIL_SERVICE_STARTED = False
_IS_LEADER = False


def _imap_idle_wait(M, timeout, stop_event):
    """IMAP IDLE で新着を待つ。新着通知→True / タイムアウト→False。"""
    tag = M._new_tag()
    M.send(tag + b' IDLE\r\n')
    M.readline()  # '+ idling'
    got = False
    sock = M.socket()
    old_to = sock.gettimeout()
    end = time.time() + timeout
    try:
        sock.settimeout(5)
        while not stop_event.is_set() and time.time() < end:
            try:
                line = M.readline()
            except socket.timeout:
                continue
            except Exception:
                break
            if not line:
                break
            u = line.upper()
            if b'EXISTS' in u or b'RECENT' in u:
                got = True
                break
    finally:
        try:
            sock.settimeout(old_to)
        except Exception:
            pass
        try:
            M.send(b'DONE\r\n')
            for _ in range(10):
                l = M.readline()
                if not l or l.upper().startswith(tag.upper()):
                    break
        except Exception:
            pass
    return got


def _idle_worker(store_id, stop_event):
    """1店舗ぶんのIMAP接続を保持し、新着が来たら即フェッチ。切断時は自動再接続。"""
    backoff = 5
    while not stop_event.is_set():
        M = None
        try:
            with app.app_context():
                ms = MailSetting.query.filter_by(store_id=store_id).first()
                has_cred = ms and (ms.imap_pass or ms.oauth_refresh_token)
                if not ms or not ms.enabled or not has_cred:
                    return
                host = ms.imap_host or 'imap.gmail.com'
                user, pw = ms.imap_user, ms.imap_pass
                o_rt, o_em = ms.oauth_refresh_token, ms.oauth_email
            M = _open_imap_raw(host, user, pw, o_rt, o_em)
            M.select('INBOX')
            with app.app_context():   # 接続直後にキャッチアップ
                fetch_reactions_for_store(store_id)
            backoff = 5
            has_idle = b'IDLE' in (M.capabilities or ())
            while not stop_event.is_set():
                if has_idle:
                    _imap_idle_wait(M, 1500, stop_event)   # 最大25分でIDLE更新
                else:
                    stop_event.wait(60)                    # IDLE非対応→60秒間隔
                if stop_event.is_set():
                    break
                with app.app_context():
                    fetch_reactions_for_store(store_id)
        except Exception as e:
            print(f"idle worker store={store_id} reconnect: {e}")
            stop_event.wait(backoff)
            backoff = min(backoff * 2, 300)
        finally:
            try:
                if M:
                    M.logout()
            except Exception:
                pass


class _IdleManager:
    def __init__(self):
        self.workers = {}   # store_id -> {'thread','stop','sig'}
        self.lock = threading.Lock()

    @staticmethod
    def _sig(ms):
        return ((ms.imap_host or 'imap.gmail.com'), ms.imap_user, ms.imap_pass,
                bool(ms.oauth_refresh_token), ms.oauth_email)

    def sync(self):
        """有効店舗のIDLEワーカーを起動／無効・変更・停止したものを停止。"""
        try:
            with app.app_context():
                enabled = {ms.store_id: self._sig(ms)
                           for ms in MailSetting.query.filter_by(enabled=True).all()
                           if (ms.imap_user and ms.imap_pass) or ms.oauth_refresh_token}
        except Exception as e:
            print(f"idle sync query error: {e}")
            return
        with self.lock:
            for sid in list(self.workers.keys()):
                w = self.workers[sid]
                if sid not in enabled or enabled[sid] != w['sig'] or not w['thread'].is_alive():
                    w['stop'].set()
                    del self.workers[sid]
            for sid, sig in enabled.items():
                if sid not in self.workers:
                    ev = threading.Event()
                    t = threading.Thread(target=_idle_worker, args=(sid, ev), daemon=True)
                    self.workers[sid] = {'thread': t, 'stop': ev, 'sig': sig}
                    t.start()
                    print(f"idle worker started store={sid}")


idle_manager = _IdleManager()


def _mail_sync_loop():
    time.sleep(20)
    cnt = 0
    while True:
        try:
            idle_manager.sync()
            cnt += 1
            if cnt % 20 == 0:   # 約10分ごとに保険フェッチ（IDLE取りこぼし対策）
                with app.app_context():
                    for ms in MailSetting.query.filter_by(enabled=True).all():
                        if (ms.imap_user and ms.imap_pass) or ms.oauth_refresh_token:
                            try:
                                fetch_reactions_for_store(ms.store_id)
                            except Exception:
                                pass
        except Exception as e:
            print(f"mail sync loop error: {e}")
        time.sleep(30)


def start_mail_service():
    """ソケットlockでリーダーを1プロセスだけ選出し、IDLEサービスを開始。"""
    global _MAIL_SERVICE_STARTED, _IS_LEADER
    if _MAIL_SERVICE_STARTED:
        return
    try:
        lock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        lock.bind(('127.0.0.1', 47625))
        globals()['_MAIL_LOCK'] = lock   # GC防止のため保持
    except OSError:
        _MAIL_SERVICE_STARTED = True      # 非リーダーは以後起動しない
        return
    _MAIL_SERVICE_STARTED = True
    _IS_LEADER = True
    threading.Thread(target=_mail_sync_loop, daemon=True).start()
    print("mail IDLE service started (leader)")


def request_mail_sync():
    """設定保存時などに即時で有効店舗のワーカーを同期（リーダーのみ実行）。"""
    if _IS_LEADER:
        threading.Thread(target=idle_manager.sync, daemon=True).start()


# ── 反響メール取込 設定ページ・API ──
def _render_mail_settings(mode, title):
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    staff_list = Staff.query.filter(Staff.store_id == sid, Staff.is_active == True).all() if sid else []
    return render_template("mail_settings.html", staff_list=staff_list, mode=mode, page_title=title)


@app.route("/mail-settings")
@login_required
@block_super_admin
def mail_settings_page():
    # 旧URLは「メール自動取込設定」へ
    return redirect(url_for('settings_mail_import'))


@app.route("/settings/mail-templates")
@login_required
@block_super_admin
def settings_mail_templates():
    return _render_mail_settings('templates', 'メールテンプレート設定')


@app.route("/settings/mail-import")
@login_required
@block_super_admin
def settings_mail_import():
    return _render_mail_settings('import', 'メール自動取込設定')


@app.route("/settings/mail-automation")
@login_required
@block_super_admin
def settings_mail_automation():
    return _render_mail_settings('automation', 'メール自動返信設定')


@app.route("/api/mail-settings", methods=["GET"])
@login_required
def api_mail_settings_get():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    ms = MailSetting.query.filter_by(store_id=sid).first() if sid else None
    oauth_available = bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET)
    if not ms:
        return jsonify({'imap_user': '', 'imap_host': 'imap.gmail.com', 'enabled': False,
                        'default_staff_id': None, 'has_password': False, 'custom_keywords': '',
                        'last_result': '', 'last_fetch_at': None,
                        'oauth_available': oauth_available, 'oauth_connected': False, 'oauth_email': '',
                        'auto_reply_enabled': False, 'auto_reply_template_id': None})
    return jsonify({
        'imap_user': ms.imap_user or '',
        'imap_host': ms.imap_host or 'imap.gmail.com',
        'enabled': bool(ms.enabled),
        'default_staff_id': ms.default_staff_id,
        'has_password': bool(ms.imap_pass),
        'custom_keywords': ms.custom_keywords or '',
        'last_result': ms.last_result or '',
        'last_fetch_at': _fmt_jst(ms.last_fetch_at) if ms.last_fetch_at else None,
        'oauth_available': oauth_available,
        'oauth_connected': bool(ms.oauth_refresh_token),
        'oauth_email': ms.oauth_email or '',
        'auto_reply_enabled': bool(ms.auto_reply_enabled),
        'auto_reply_template_id': ms.auto_reply_template_id,
    })


@app.route("/api/mail-settings", methods=["POST"])
@login_required
def api_mail_settings_save():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    if not sid:
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json() or {}
    ms = MailSetting.query.filter_by(store_id=sid).first()
    if not ms:
        ms = MailSetting(store_id=sid)
        db.session.add(ms)
    # 部分更新：payloadに含まれるキーのみ更新（ページ分割で他項目を消さない）
    if 'imap_user' in data:
        ms.imap_user = (data.get('imap_user') or '').strip()[:200]
    if 'imap_host' in data:
        ms.imap_host = ((data.get('imap_host') or '').strip() or 'imap.gmail.com')[:120]
    pw = data.get('imap_pass')
    if pw:  # 入力があった時のみ更新（空欄なら既存パスワードを維持）
        ms.imap_pass = pw.replace(' ', '').strip()[:200]
    if 'enabled' in data:
        ms.enabled = bool(data.get('enabled'))
        if ms.enabled and ms.import_after is None:
            ms.import_after = datetime.utcnow()
    if 'default_staff_id' in data:
        dsid = data.get('default_staff_id')
        ms.default_staff_id = int(dsid) if dsid else None
    if 'custom_keywords' in data:
        ms.custom_keywords = (data.get('custom_keywords') or '')[:5000]
    if 'auto_reply_enabled' in data:
        ms.auto_reply_enabled = bool(data.get('auto_reply_enabled'))
    if 'auto_reply_template_id' in data:
        arid = data.get('auto_reply_template_id')
        ms.auto_reply_template_id = int(arid) if arid else None
    ms.updated_at = datetime.utcnow()
    db.session.commit()
    start_mail_service()   # 未起動なら起動
    request_mail_sync()    # 設定変更を即反映（リーダーのみ）
    return jsonify({'status': 'ok'})


# ── Google OAuth 連携（送信=Gmail API / 受信=IMAP XOAUTH2）──
def _oauth_redirect_uri():
    base = (request.url_root or APP_BASE_URL).rstrip('/')
    # 本番httpsを強制（Railwayはhttps終端でhostヘッダがhttpになることがある）
    if base.startswith('http://') and 'localhost' not in base and '127.0.0.1' not in base:
        base = 'https://' + base[len('http://'):]
    return base + '/oauth/google/callback'


@app.route("/oauth/google/start")
@login_required
@block_super_admin
def oauth_google_start():
    if not (GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET):
        return "Google連携が未設定です（管理者にGOOGLE_CLIENT_ID/SECRETの設定を依頼してください）", 400
    import urllib.parse, secrets as _secrets
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    if not sid:
        return "店舗が選択されていません", 400
    state = f"{sid}.{_secrets.token_urlsafe(16)}"
    session['oauth_state'] = state
    params = {
        'client_id': GOOGLE_CLIENT_ID,
        'redirect_uri': _oauth_redirect_uri(),
        'response_type': 'code',
        'scope': GOOGLE_OAUTH_SCOPE,
        'access_type': 'offline',
        'prompt': 'consent',           # 毎回 refresh_token を確実に得る
        'include_granted_scopes': 'true',
        'state': state,
        'login_hint': '',
    }
    url = 'https://accounts.google.com/o/oauth2/v2/auth?' + urllib.parse.urlencode(params)
    return redirect(url)


@app.route("/oauth/google/callback")
@login_required
@block_super_admin
def oauth_google_callback():
    err = request.args.get('error')
    if err:
        return redirect(url_for('mail_settings_page') + '?oauth=error')
    code = request.args.get('code')
    state = request.args.get('state') or ''
    if not code or not state or state != session.get('oauth_state'):
        return redirect(url_for('mail_settings_page') + '?oauth=state_error')
    try:
        sid = int(state.split('.')[0])
    except Exception:
        return redirect(url_for('mail_settings_page') + '?oauth=state_error')
    allowed = get_allowed_store_ids()
    if sid not in allowed:
        return redirect(url_for('mail_settings_page') + '?oauth=forbidden')
    try:
        tok = _http_post_form('https://oauth2.googleapis.com/token', {
            'code': code,
            'client_id': GOOGLE_CLIENT_ID,
            'client_secret': GOOGLE_CLIENT_SECRET,
            'redirect_uri': _oauth_redirect_uri(),
            'grant_type': 'authorization_code',
        })
        refresh = tok.get('refresh_token')
        access = tok.get('access_token')
        if not refresh:
            # 既に同意済みで refresh_token が返らないケース
            return redirect(url_for('mail_settings_page') + '?oauth=no_refresh')
        email = _gmail_profile_email(access) if access else ''
    except Exception as e:
        print(f"oauth callback error: {e}")
        return redirect(url_for('mail_settings_page') + '?oauth=token_error')

    ms = MailSetting.query.filter_by(store_id=sid).first()
    if not ms:
        ms = MailSetting(store_id=sid)
        db.session.add(ms)
    ms.oauth_refresh_token = refresh
    ms.oauth_email = email or ms.oauth_email
    if email:
        ms.imap_user = email   # 表示・XOAUTH2用
    ms.imap_host = 'imap.gmail.com'
    ms.updated_at = datetime.utcnow()
    db.session.commit()
    session.pop('oauth_state', None)
    start_mail_service()
    request_mail_sync()
    return redirect(url_for('mail_settings_page') + '?oauth=ok')


@app.route("/api/mail-settings/google-disconnect", methods=["POST"])
@login_required
@block_super_admin
def api_google_disconnect():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    ms = MailSetting.query.filter_by(store_id=sid).first() if sid else None
    if ms:
        ms.oauth_refresh_token = None
        ms.oauth_email = None
        db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/mail-settings/test", methods=["POST"])
@login_required
def api_mail_settings_test():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    data = request.get_json() or {}
    user = (data.get('imap_user') or '').strip()
    host = (data.get('imap_host') or 'imap.gmail.com').strip()
    pw = (data.get('imap_pass') or '')
    if not pw:  # 未入力なら保存済みパスワードでテスト
        ms = MailSetting.query.filter_by(store_id=sid).first() if sid else None
        pw = ms.imap_pass if ms else ''
    pw = (pw or '').replace(' ', '').strip()
    if not user or not pw:
        return jsonify({'ok': False, 'message': 'メールアドレスとアプリパスワードを入力してください'})
    ok, msg = test_imap_connection(host, user, pw)
    return jsonify({'ok': ok, 'message': msg})


@app.route("/api/mail-settings/fetch", methods=["POST"])
@login_required
def api_mail_settings_fetch():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    if not sid:
        return jsonify({'ok': False, 'error': 'unauthorized'}), 403
    res = fetch_reactions_for_store(sid)
    return jsonify(res)


# ── 反響取込 診断（読み取り専用：DBへ書き込まない） ──
def _diagnose_one_email(msg, extra_map, portal_map, store_id, import_after):
    """1通のメールを読み取り専用で解析し、取込可否と理由を返す（DB書き込みなし）。"""
    from_addr = _decode_mime(msg.get('From', '') or '')
    subject = _decode_mime(msg.get('Subject', '') or '')
    mdt = _msg_datetime(msg)
    out = {
        'from': from_addr[:160],
        'subject': subject[:160],
        'date': _fmt_jst(mdt, '%Y/%m/%d %H:%M') if mdt else (msg.get('Date') or ''),
    }
    if mdt and import_after and mdt < import_after:
        out['status'] = 'skip'
        out['reason'] = '取込開始日時より前のメール（過去分は取込対象外）'
        return out
    body = _email_plain_body(msg)
    if not body:
        out['status'] = 'skip'
        out['reason'] = '本文テキストを取得できない（HTMLのみ等で項目が読めない）'
        return out
    neg = next((k for k in NEG_SUBJECT_KEYWORDS if k in subject), None)
    if neg:
        out['status'] = 'skip'
        out['reason'] = f'件名の除外ワード「{neg}」に該当（通知系と判断）'
        return out
    parsed = parse_reaction_email(msg, extra_map, portal_map)
    if not parsed:
        fa = (from_addr or '').lower()
        matched = next((media for matcher, media in (portal_map or []) if matcher and matcher.lower() in fa), None)
        if matched:
            out['reason'] = f'ポータル「{matched}」一致だが氏名・物件のどちらも抽出できず（本文の項目表記が想定外）'
        else:
            out['reason'] = '差出人が未登録ポータルで、かつ「氏名＋(物件 or 反響件名)」を満たさず'
        out['status'] = 'skip'
        return out
    out['name'] = parsed.get('name')
    out['media'] = parsed.get('source')
    ext = parsed.get('external_id')
    if (ProcessedReaction.query.filter_by(store_id=store_id, external_id=ext).first()
            or EchoRecord.query.filter_by(store_id=store_id, external_id=ext).first()):
        out['status'] = 'dup'
        out['reason'] = '既に取込済み/既存（重複としてスキップ）'
        return out
    out['status'] = 'ok'
    out['reason'] = '反響として取込対象（この条件なら取り込まれる）'
    return out


def _is_mail_diag_manager():
    cur_user = AppUser.query.get(session.get('app_user_id'))
    return bool(cur_user and cur_user.role in ('owner', 'store_manager', 'super_admin'))


@app.route("/mail-diagnose")
@login_required
def mail_diagnose_page():
    if not _is_mail_diag_manager():
        return "この画面は管理者のみ利用できます。", 403
    stores = get_allowed_stores()
    return render_template("mail_diagnose.html", stores=stores)


@app.route("/api/mail-diagnose")
@login_required
def api_mail_diagnose():
    if not _is_mail_diag_manager():
        return jsonify({'error': '権限がありません'}), 403
    allowed = get_allowed_store_ids()
    sid = request.args.get('store_id', type=int) or (allowed[0] if allowed else None)
    if not sid or sid not in allowed:
        return jsonify({'error': '対象の店舗が見つかりません'}), 404
    ms = MailSetting.query.filter_by(store_id=sid).first()
    result = {'store_id': sid, 'setting': None, 'portals': [], 'emails': [], 'summary': {}}
    if not ms:
        result['error'] = 'この店舗のメール取込設定がありません（未設定）。'
        return jsonify(result)
    conn = 'OAuth(Google)連携' if ms.oauth_refresh_token else ('アプリパスワード' if (ms.imap_user and ms.imap_pass) else '未接続')
    result['setting'] = {
        'enabled': bool(ms.enabled),
        'connection': conn,
        'imap_user': ms.imap_user or ms.oauth_email or '',
        'import_after': _fmt_jst(ms.import_after, '%Y/%m/%d %H:%M') if ms.import_after else '（未設定）',
        'last_fetch_at': _fmt_jst(ms.last_fetch_at, '%Y/%m/%d %H:%M') if ms.last_fetch_at else '（なし）',
        'last_result': ms.last_result or '（なし）',
        'auto_reply_enabled': bool(ms.auto_reply_enabled),
    }
    portals = PortalSource.query.filter_by(store_id=sid).order_by(PortalSource.id.asc()).all()
    result['portals'] = [{'matcher': p.matcher or '', 'media': p.media or '', 'enabled': bool(p.enabled)} for p in portals]
    if conn == '未接続':
        result['error'] = 'メール接続が未設定です（IMAP/OAuthなし）。先に接続設定が必要です。'
        return jsonify(result)
    extra_map = parse_custom_keywords(ms.custom_keywords)
    portal_map = [(p.matcher, p.media) for p in portals if p.enabled and p.matcher and p.media]
    diag_days, limit = 21, 60
    try:
        M = _open_imap(ms)
        M.select('INBOX')
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        since = date.today() - timedelta(days=diag_days)
        since_str = f"{since.day:02d}-{months[since.month - 1]}-{since.year}"
        typ, data = M.search(None, f'(SINCE "{since_str}")')
        ids = data[0].split() if (data and data[0]) else []
        ids = ids[-limit:]
        emails = []
        for num in reversed(ids):
            try:
                typ, md = M.fetch(num, '(BODY.PEEK[])')
                if not md or not md[0]:
                    continue
                msg = emaillib.message_from_bytes(md[0][1])
            except Exception:
                continue
            emails.append(_diagnose_one_email(msg, extra_map, portal_map, sid, ms.import_after))
        try:
            M.close()
            M.logout()
        except Exception:
            pass
        result['emails'] = emails
        summary = {'total': len(emails), 'ok': 0, 'dup': 0, 'skip': 0}
        for e in emails:
            st = e.get('status', 'skip')
            summary[st] = summary.get(st, 0) + 1
        result['summary'] = summary
    except Exception as e:
        result['error'] = f'メール取得でエラー：{e}'
    return jsonify(result)


@app.route("/api/portal-sources", methods=["GET"])
@login_required
def api_portal_sources_get():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    items = (PortalSource.query.filter_by(store_id=sid).order_by(PortalSource.id.asc()).all()
             if sid else [])
    return jsonify([{'matcher': p.matcher or '', 'media': p.media or '', 'enabled': bool(p.enabled),
                     'auto_reply_template_id': p.auto_reply_template_id}
                    for p in items])


@app.route("/api/portal-sources", methods=["POST"])
@login_required
def api_portal_sources_save():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    if not sid:
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json() or {}
    rows = data.get('sources') or []
    PortalSource.query.filter_by(store_id=sid).delete()
    for r in rows:
        m = (r.get('matcher') or '').strip()[:200]
        md = (r.get('media') or '').strip()[:100]
        if m and md:
            arid = r.get('auto_reply_template_id')
            db.session.add(PortalSource(store_id=sid, matcher=m, media=md,
                                        enabled=bool(r.get('enabled', True)),
                                        auto_reply_template_id=(int(arid) if arid else None)))
    db.session.commit()
    request_mail_sync()
    return jsonify({'status': 'ok'})


@app.route("/api/echo-records/<int:rid>/messages", methods=["GET"])
@login_required
def api_echo_messages(rid):
    allowed = get_allowed_store_ids()
    rec = EchoRecord.query.get_or_404(rid)
    if rec.store_id not in allowed:
        return jsonify({'error': '権限がありません'}), 403
    msgs = (MailMessage.query.filter_by(echo_id=rid)
            .order_by(MailMessage.created_at.asc(), MailMessage.id.asc()).all())
    # 受信を既読化
    changed = False
    for m in msgs:
        if m.direction == 'in' and not m.is_read:
            m.is_read = True
            changed = True
    if rec.has_unread_reply:
        rec.has_unread_reply = False
        changed = True
    if changed:
        db.session.commit()
    # 添付ファイルをメッセージ単位で取得（1クエリ）
    msg_ids = [m.id for m in msgs]
    atts_by_msg = {}
    if msg_ids:
        for a in MailAttachment.query.filter(MailAttachment.message_id.in_(msg_ids)).all():
            fn = (a.filename or '').lower()
            atts_by_msg.setdefault(a.message_id, []).append({
                'id': a.id,
                'filename': a.filename or 'ファイル',
                'content_type': a.content_type or '',
                'size': a.size or 0,
                'is_image': (a.content_type or '').startswith('image/') or fn.endswith(_IMG_EXTS),
            })
    return jsonify({
        'customer_name': rec.list_name or '',
        'customer_email': rec.customer_email or '',
        'media': rec.media or '',
        'messages': [{
            'direction': m.direction,
            'subject': m.subject or '',
            'body': m.body or '',
            'from': m.from_addr or '',
            'at': _fmt_jst(m.created_at),
            'read': bool(m.opened_at) if m.direction == 'out' else None,
            'read_at': _fmt_jst(m.opened_at),
            'attachments': atts_by_msg.get(m.id, []),
        } for m in msgs],
    })


@app.route("/api/echo-records/<int:rid>/send", methods=["POST"])
@login_required
def api_echo_send(rid):
    allowed = get_allowed_store_ids()
    rec = EchoRecord.query.get_or_404(rid)
    if rec.store_id not in allowed:
        return jsonify({'error': '権限がありません'}), 403

    # JSON / multipart(添付あり) 両対応
    attachments = []
    if request.content_type and 'multipart/form-data' in request.content_type:
        subject = (request.form.get('subject') or '').strip()
        body = (request.form.get('body') or '').strip()
        for f in request.files.getlist('files'):
            if not f or not f.filename:
                continue
            raw = f.read()
            if not raw:
                continue
            if len(raw) > 20 * 1024 * 1024:
                return jsonify({'ok': False, 'error': f'添付「{f.filename}」が大きすぎます（20MBまで）'})
            attachments.append((f.filename, f.mimetype or 'application/octet-stream', raw))
    else:
        data = request.get_json() or {}
        subject = (data.get('subject') or '').strip()
        body = (data.get('body') or '').strip()

    if not body and not attachments:
        return jsonify({'ok': False, 'error': '本文を入力してください'})
    res = send_mail_for_store(rec.store_id, rid, subject, body,
                              attachments=attachments,
                              base_url=request.url_root)
    return jsonify(res)


@app.route("/api/mail-attachments/<int:aid>", methods=["GET"])
@login_required
def api_mail_attachment(aid):
    allowed = get_allowed_store_ids()
    a = MailAttachment.query.get_or_404(aid)
    if a.store_id not in allowed:
        return "権限がありません", 403
    from flask import Response as _Resp
    ctype = a.content_type or 'application/octet-stream'
    inline = ctype.startswith('image/') or ctype == 'application/pdf'
    disp = 'inline' if inline else 'attachment'
    fn = (a.filename or 'file').replace('"', '')
    resp = _Resp(a.data or b'', mimetype=ctype)
    try:
        from urllib.parse import quote as _q
        resp.headers['Content-Disposition'] = f"{disp}; filename*=UTF-8''{_q(fn)}"
    except Exception:
        resp.headers['Content-Disposition'] = f'{disp}; filename="{fn}"'
    return resp


@app.route("/m/o/<int:mid>.gif", methods=["GET"])
def mail_open_pixel(mid):
    """送信メール内のトラッキングピクセル。読み込まれたら既読(opened_at)を記録。"""
    try:
        m = MailMessage.query.get(mid)
        if m and m.direction == 'out' and not m.opened_at:
            m.opened_at = datetime.utcnow()
            db.session.commit()
    except Exception:
        db.session.rollback()
    # 1x1 透明GIF
    gif = (b'GIF89a\x01\x00\x01\x00\x80\x00\x00\xff\xff\xff\x00\x00\x00!\xf9'
           b'\x04\x01\x00\x00\x00\x00,\x00\x00\x00\x00\x01\x00\x01\x00\x00'
           b'\x02\x02D\x01\x00;')
    from flask import Response as _Resp
    resp = _Resp(gif, mimetype='image/gif')
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    return resp


# ── 返信テンプレート API ──────────────────────────────────
_MAIL_TEMPLATE_DEFAULTS = [
    ('お問い合わせ御礼', 'お問い合わせありがとうございます',
     'この度はお問い合わせいただき、誠にありがとうございます。\n担当させていただきます。\n\nご希望条件など改めて確認させていただきたく存じます。\n何卒よろしくお願いいたします。'),
    ('内見のご案内', '内見のご案内',
     'お世話になっております。\nご内見の日程について、下記のいずれかでご都合いかがでしょうか。\n\n・候補日①：\n・候補日②：\n\nご返信お待ちしております。'),
    ('追客フォロー', 'その後のご状況はいかがでしょうか',
     'お世話になっております。\nその後、お部屋探しのご状況はいかがでしょうか。\n新着のお部屋もございますので、ご希望条件を改めてお聞かせいただけますと幸いです。\n\nよろしくお願いいたします。'),
]


@app.route("/api/mail-templates", methods=["GET"])
@login_required
def api_mail_templates_get():
    tenant_id = _get_tenant_id()
    if MailTemplate.query.filter_by(tenant_id=tenant_id).count() == 0:
        for i, (t, s, b) in enumerate(_MAIL_TEMPLATE_DEFAULTS):
            db.session.add(MailTemplate(tenant_id=tenant_id, title=t, subject=s, body=b, sort_order=i))
        db.session.commit()
    items = (MailTemplate.query.filter_by(tenant_id=tenant_id)
             .order_by(MailTemplate.category, MailTemplate.sort_order, MailTemplate.id).all())
    return jsonify([{'id': t.id, 'category': t.category or '', 'title': t.title or '',
                     'subject': t.subject or '', 'body': t.body or '',
                     'is_html': bool(t.is_html)} for t in items])


@app.route("/api/mail-templates", methods=["POST"])
@login_required
def api_mail_templates_add():
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    body = (data.get('body') or '').strip()
    if not title or not body:
        return jsonify({'error': 'タイトルと本文を入力してください'}), 400
    tenant_id = _get_tenant_id()
    mx = db.session.query(db.func.max(MailTemplate.sort_order)).filter_by(tenant_id=tenant_id).scalar() or 0
    t = MailTemplate(tenant_id=tenant_id, category=(data.get('category') or '').strip()[:120],
                     title=title[:120],
                     subject=(data.get('subject') or '')[:300], body=body,
                     is_html=bool(data.get('is_html')), sort_order=mx + 1)
    db.session.add(t)
    db.session.commit()
    return jsonify({'id': t.id})


@app.route("/api/mail-templates/<int:tid>", methods=["PUT"])
@login_required
def api_mail_templates_update(tid):
    tenant_id = _get_tenant_id()
    t = MailTemplate.query.get_or_404(tid)
    if t.tenant_id != tenant_id:
        return jsonify({'error': 'forbidden'}), 403
    data = request.get_json() or {}
    if 'category' in data:
        t.category = (data.get('category') or '').strip()[:120]
    if 'title' in data:
        t.title = (data.get('title') or '')[:120]
    if 'subject' in data:
        t.subject = (data.get('subject') or '')[:300]
    if 'body' in data:
        t.body = data.get('body') or ''
    if 'is_html' in data:
        t.is_html = bool(data.get('is_html'))
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/company-profile", methods=["GET"])
@login_required
def api_company_profile_get():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    cp = CompanyProfile.query.filter_by(store_id=sid).first() if sid else None
    return jsonify({
        'company_name': cp.company_name if cp else '',
        'store_name': cp.store_name if cp else '',
        'phone': cp.phone if cp else '',
        'fax': cp.fax if cp else '',
        'email': cp.email if cp else '',
        'address': cp.address if cp else '',
        'representative': cp.representative if cp else '',
        'license_number': cp.license_number if cp else '',
        'license_date': cp.license_date if cp else '',
        'business_hours': cp.business_hours if cp else '',
        'holidays': cp.holidays if cp else '',
        'invoice_number': cp.invoice_number if cp else '',
        'line_url': cp.line_url if cp else '',
        'has_logo': bool(cp and cp.logo_data),
    })


@app.route("/api/company-profile", methods=["POST"])
@login_required
def api_company_profile_save():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    if not sid:
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json() or {}
    cp = CompanyProfile.query.filter_by(store_id=sid).first()
    if not cp:
        cp = CompanyProfile(store_id=sid)
        db.session.add(cp)
    cp.company_name   = (data.get('company_name') or '')[:200]
    cp.store_name     = (data.get('store_name') or '')[:200]
    cp.phone          = (data.get('phone') or '')[:60]
    cp.fax            = (data.get('fax') or '')[:60]
    cp.email          = (data.get('email') or '')[:200]
    cp.address        = (data.get('address') or '')[:300]
    cp.representative = (data.get('representative') or '')[:100]
    cp.license_number = (data.get('license_number') or '')[:120]
    cp.license_date   = (data.get('license_date') or '')[:60]
    cp.business_hours = (data.get('business_hours') or '')[:200]
    cp.holidays       = (data.get('holidays') or '')[:200]
    cp.invoice_number = (data.get('invoice_number') or '')[:60]
    cp.line_url       = (data.get('line_url') or '')[:300]
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/mail-templates/<int:tid>", methods=["DELETE"])
@login_required
def api_mail_templates_delete(tid):
    tenant_id = _get_tenant_id()
    t = MailTemplate.query.get_or_404(tid)
    if t.tenant_id != tenant_id:
        return jsonify({'error': 'forbidden'}), 403
    db.session.delete(t)
    db.session.commit()
    return jsonify({'status': 'ok'})


# ── 状況タグの行カラー API（反響管理表）──────────────────────
ECHO_STATUS_PALETTE = ['#CCFBF1', '#dbeafe', '#fef9c3', '#fde4cf', '#e9d5ff',
                       '#fecaca', '#d1fae5', '#fbcfe8', '#e0e7ff', '#cffafe']


@app.route("/api/echo-status-colors", methods=["GET"])
@login_required
def api_echo_status_colors_get():
    allowed = get_allowed_store_ids()
    store_id = allowed[0] if allowed else 1
    tenant_id = _get_tenant_id()
    opts = (DropdownOption.query.filter(
                DropdownOption.category == 'echo_status',
                db.or_(DropdownOption.tenant_id == tenant_id, DropdownOption.tenant_id == None))
            .order_by(DropdownOption.sort_order, DropdownOption.id).all())
    result = {}
    for i, o in enumerate(opts):
        sc = StatusColor.query.filter_by(store_id=store_id, status_key='echo:' + o.value).first()
        result[o.value] = (sc.row_bg_color if (sc and sc.row_bg_color)
                           else ECHO_STATUS_PALETTE[i % len(ECHO_STATUS_PALETTE)])
    return jsonify(result)


@app.route("/api/echo-status-colors", methods=["PUT"])
@login_required
def api_echo_status_colors_update():
    allowed = get_allowed_store_ids()
    store_id = allowed[0] if allowed else 1
    data = request.get_json() or {}
    colors = data.get('colors') if isinstance(data.get('colors'), dict) else data
    for key, color in (colors or {}).items():
        if not isinstance(color, str):
            continue
        sc = StatusColor.query.filter_by(store_id=store_id, status_key='echo:' + key).first()
        if not sc:
            sc = StatusColor(store_id=store_id, status_key='echo:' + key)
            db.session.add(sc)
        sc.row_bg_color = color or '#ffffff'
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/cs-status-colors", methods=["GET"])
@login_required
def api_cs_status_colors_get():
    """接客管理表の状況タグ別の色（badge背景色）。echo と同じ仕組み（status_key='cs:'）"""
    allowed = get_allowed_store_ids()
    store_id = allowed[0] if allowed else 1
    tenant_id = _get_tenant_id()
    opts = (DropdownOption.query.filter(
                DropdownOption.category == 'cs_status',
                db.or_(DropdownOption.tenant_id == tenant_id, DropdownOption.tenant_id == None))
            .order_by(DropdownOption.sort_order, DropdownOption.id).all())
    result = {}
    for i, o in enumerate(opts):
        sc = StatusColor.query.filter_by(store_id=store_id, status_key='cs:' + o.value).first()
        result[o.value] = (sc.row_bg_color if (sc and sc.row_bg_color)
                           else ECHO_STATUS_PALETTE[i % len(ECHO_STATUS_PALETTE)])
    return jsonify(result)


@app.route("/api/cs-status-colors", methods=["PUT"])
@login_required
def api_cs_status_colors_update():
    allowed = get_allowed_store_ids()
    store_id = allowed[0] if allowed else 1
    data = request.get_json() or {}
    colors = data.get('colors') if isinstance(data.get('colors'), dict) else data
    for key, color in (colors or {}).items():
        if not isinstance(color, str):
            continue
        sc = StatusColor.query.filter_by(store_id=store_id, status_key='cs:' + key).first()
        if not sc:
            sc = StatusColor(store_id=store_id, status_key='cs:' + key)
            db.session.add(sc)
        sc.row_bg_color = color or '#ffffff'
    db.session.commit()
    return jsonify({'status': 'ok'})


# ── 社内チャット ───────────────────────────────────────────
def _chat_user():
    uid = session.get('app_user_id')
    return AppUser.query.get(uid) if uid else None


def _chat_is_pro(tenant_id):
    # リクエスト中はログインユーザーの現在店舗で判定（店舗別オプション対応）、
    # それ以外（バックグラウンド処理等）はテナント全体で判定
    try:
        if session.get('app_user_id'):
            return current_has_option('chat_pro')
    except Exception:
        pass
    return tenant_has_option(tenant_id, 'chat_pro')


def _chat_retention_days(tenant_id):
    return 730 if _chat_is_pro(tenant_id) else 60


def _chat_display_name(u):
    if not u:
        return '不明'
    if u.staff_id:
        s = Staff.query.get(u.staff_id)
        if s and s.name:
            return s.name
    return u.username or '不明'


def _ensure_base_channels(tenant_id):
    """全社チャンネル＋全店舗チャンネルを整える。
    このテナントの有効店舗ぶんだけ作成し、テナントに属さない/無効店舗の
    店舗チャンネルは無効化する（他テナント店舗の混在掃除）。"""
    created = False
    if not ChatChannel.query.filter_by(tenant_id=tenant_id, kind='company').first():
        db.session.add(ChatChannel(tenant_id=tenant_id, kind='company', name='全社'))
        created = True
    valid_stores = Store.query.filter_by(tenant_id=tenant_id, is_active=True).all()
    valid_ids = {s.id for s in valid_stores}
    name_by_id = {s.id: (s.name or '店舗') for s in valid_stores}
    # 有効店舗のチャンネルを作成 or 再有効化
    for sid in valid_ids:
        ch = ChatChannel.query.filter_by(tenant_id=tenant_id, kind='store', store_id=sid).first()
        if not ch:
            db.session.add(ChatChannel(tenant_id=tenant_id, kind='store', store_id=sid, name=name_by_id[sid]))
            created = True
        elif not ch.is_active:
            ch.is_active = True
            created = True
    # このテナントの有効店舗以外を指す店舗チャンネルは無効化（他社店舗の混在を解消）
    for c in ChatChannel.query.filter_by(tenant_id=tenant_id, kind='store', is_active=True).all():
        if c.store_id not in valid_ids:
            c.is_active = False
            created = True
    if created:
        db.session.commit()


def _can_access_channel(c, user):
    if not c or not user or c.tenant_id != user.tenant_id or not c.is_active:
        return False
    if c.kind == 'company':
        return True
    if c.kind == 'store':
        return user.role == 'owner' or (user.store_id and c.store_id == user.store_id)
    if c.kind == 'group':
        if c.created_by == user.id:
            return True
        return ChatMember.query.filter_by(channel_id=c.id, user_id=user.id).first() is not None
    return False


def _visible_channels(tenant_id, user):
    _ensure_base_channels(tenant_id)
    chans = ChatChannel.query.filter_by(tenant_id=tenant_id, is_active=True).all()
    my_groups = {m.channel_id for m in ChatMember.query.filter_by(user_id=user.id).all()}
    out = []
    for c in chans:
        if c.kind == 'company':
            out.append(c)
        elif c.kind == 'store':
            if user.role == 'owner' or (user.store_id and c.store_id == user.store_id):
                out.append(c)
        elif c.kind == 'group':
            if c.id in my_groups or c.created_by == user.id:
                out.append(c)
    return out


def _chat_cleanup_channel(c):
    """チャンネルの保存期限切れメッセージ（と添付）を削除"""
    if not c:
        return
    cutoff = datetime.utcnow() - timedelta(days=_chat_retention_days(c.tenant_id))
    old = ChatMessage.query.filter(ChatMessage.channel_id == c.id,
                                   ChatMessage.created_at < cutoff).all()
    if not old:
        return
    ids = [m.id for m in old]
    ChatAttachment.query.filter(ChatAttachment.message_id.in_(ids)).delete(synchronize_session=False)
    ChatMessage.query.filter(ChatMessage.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()


@app.route("/chat")
@login_required
@block_super_admin
def chat_page():
    return render_template("chat.html")


@app.route("/api/chat/channels", methods=["GET"])
@login_required
def api_chat_channels():
    u = _chat_user()
    if not u or not u.tenant_id:
        return jsonify({'channels': [], 'is_pro': False, 'retention_days': 60, 'me': {}})
    chans = _visible_channels(u.tenant_id, u)
    kind_order = {'company': 0, 'store': 1, 'group': 2}
    # company → store → group。groupはピン優先→sort_order→名前
    chans.sort(key=lambda c: (
        kind_order.get(c.kind, 9),
        0 if (c.kind == 'group' and getattr(c, 'pinned', False)) else 1,
        getattr(c, 'sort_order', 0) or 0,
        c.name or '',
    ))
    # グループのメンバー数
    group_ids = [c.id for c in chans if c.kind == 'group']
    mcount = {}
    if group_ids:
        for cid_, cnt in db.session.query(ChatMember.channel_id, db.func.count(ChatMember.id))\
                .filter(ChatMember.channel_id.in_(group_ids)).group_by(ChatMember.channel_id).all():
            mcount[cid_] = cnt
    is_mgr = u.role in ('owner', 'store_manager')
    out = []
    for c in chans:
        item = {'id': c.id, 'kind': c.kind, 'name': c.name or '', 'store_id': c.store_id}
        if c.kind == 'group':
            item['pinned'] = bool(getattr(c, 'pinned', False))
            item['member_count'] = mcount.get(c.id, 0)
            item['can_manage'] = (c.created_by == u.id) or is_mgr   # 削除・並べ替え可
        out.append(item)
    return jsonify({
        'channels': out,
        'is_pro': _chat_is_pro(u.tenant_id),
        'retention_days': _chat_retention_days(u.tenant_id),
        'me': {'id': u.id, 'name': _chat_display_name(u)},
    })


@app.route("/api/chat/members", methods=["GET"])
@login_required
def api_chat_members():
    u = _chat_user()
    if not u or not u.tenant_id:
        return jsonify([])
    users = AppUser.query.filter(AppUser.tenant_id == u.tenant_id,
                                 AppUser.is_active == True,
                                 AppUser.role != 'super_admin').all()
    return jsonify([{'id': x.id, 'name': _chat_display_name(x)} for x in users if x.id != u.id])


@app.route("/api/chat/channels", methods=["POST"])
@login_required
def api_chat_create_group():
    u = _chat_user()
    if not u or not u.tenant_id:
        return jsonify({'error': 'unauthorized'}), 403
    if u.role not in ('owner', 'store_manager'):
        return jsonify({'error': 'グループの作成はオーナー・店長のみ可能です'}), 403
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'グループ名を入力してください'}), 400
    c = ChatChannel(tenant_id=u.tenant_id, kind='group', name=name[:120], created_by=u.id)
    db.session.add(c)
    db.session.flush()
    ids = set()
    for i in (data.get('member_ids') or []):
        try:
            ids.add(int(i))
        except (TypeError, ValueError):
            pass
    ids.add(u.id)  # 作成者は必ずメンバー
    valid = {x.id for x in AppUser.query.filter(AppUser.tenant_id == u.tenant_id,
                                                AppUser.id.in_(ids)).all()}
    for uid in valid:
        db.session.add(ChatMember(channel_id=c.id, user_id=uid))
    db.session.commit()
    return jsonify({'id': c.id})


def _can_manage_group(c, u):
    """グループの削除・ピン・並べ替えができるか（作成者 or オーナー/店長）"""
    if not c or c.kind != 'group':
        return False
    return c.created_by == u.id or u.role in ('owner', 'store_manager')


@app.route("/api/chat/channels/<int:cid>/members", methods=["GET"])
@login_required
def api_chat_channel_members(cid):
    """チャンネルのメンバー一覧（グループはメンバー、全社/店舗は対象ユーザー）"""
    u = _chat_user()
    c = ChatChannel.query.get_or_404(cid)
    if not _can_access_channel(c, u):
        return jsonify({'error': '権限がありません'}), 403
    if c.kind == 'group':
        uids = [m.user_id for m in ChatMember.query.filter_by(channel_id=cid).all()]
        users = AppUser.query.filter(AppUser.id.in_(uids)).all() if uids else []
    else:
        # 全社/店舗：テナントの該当ユーザー
        q = AppUser.query.filter(AppUser.tenant_id == u.tenant_id,
                                 AppUser.is_active == True, AppUser.role != 'super_admin')
        if c.kind == 'store' and c.store_id:
            q = q.filter(db.or_(AppUser.store_id == c.store_id, AppUser.role == 'owner'))
        users = q.all()
    return jsonify([{'id': x.id, 'name': _chat_display_name(x),
                     'is_me': x.id == u.id} for x in users])


@app.route("/api/chat/channels/<int:cid>", methods=["DELETE"])
@login_required
def api_chat_delete_group(cid):
    """グループを削除（作成者・オーナー・店長のみ）"""
    u = _chat_user()
    c = ChatChannel.query.get_or_404(cid)
    if c.tenant_id != u.tenant_id or c.kind != 'group':
        return jsonify({'error': 'グループのみ削除できます'}), 400
    if not _can_manage_group(c, u):
        return jsonify({'error': '削除する権限がありません'}), 403
    c.is_active = False
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/chat/channels/<int:cid>/pin", methods=["POST"])
@login_required
def api_chat_pin_group(cid):
    """グループのピン止めを切り替え"""
    u = _chat_user()
    c = ChatChannel.query.get_or_404(cid)
    if c.tenant_id != u.tenant_id or not _can_manage_group(c, u):
        return jsonify({'error': '権限がありません'}), 403
    c.pinned = not bool(getattr(c, 'pinned', False))
    db.session.commit()
    return jsonify({'status': 'ok', 'pinned': c.pinned})


@app.route("/api/chat/channels/<int:cid>/move", methods=["POST"])
@login_required
def api_chat_move_group(cid):
    """グループの並び順を1つ上/下に移動（隣のグループと並び順を入れ替え）"""
    u = _chat_user()
    c = ChatChannel.query.get_or_404(cid)
    if c.tenant_id != u.tenant_id or not _can_manage_group(c, u):
        return jsonify({'error': '権限がありません'}), 403
    direction = (request.get_json() or {}).get('direction', 'up')
    # 同じピン状態のグループ内で並べ替え
    groups = ChatChannel.query.filter_by(tenant_id=u.tenant_id, kind='group', is_active=True,
                                         pinned=bool(c.pinned)).all()
    groups.sort(key=lambda x: (x.sort_order or 0, x.name or '', x.id))
    idx = next((i for i, g in enumerate(groups) if g.id == cid), None)
    if idx is None:
        return jsonify({'error': 'not found'}), 404
    swap = idx - 1 if direction == 'up' else idx + 1
    if 0 <= swap < len(groups):
        # まず連番を振り直してから入れ替え（sort_orderが全部0でも動くように）
        for i, g in enumerate(groups):
            g.sort_order = i
        groups[idx].sort_order, groups[swap].sort_order = groups[swap].sort_order, groups[idx].sort_order
        db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/chat/channels/<int:cid>/messages", methods=["GET"])
@login_required
def api_chat_messages(cid):
    u = _chat_user()
    c = ChatChannel.query.get_or_404(cid)
    if not _can_access_channel(c, u):
        return jsonify({'error': '権限がありません'}), 403
    _chat_cleanup_channel(c)
    after = request.args.get('after', type=int) or 0
    q = ChatMessage.query.filter_by(channel_id=cid)
    if after:
        q = q.filter(ChatMessage.id > after)
    msgs = q.order_by(ChatMessage.id.asc()).limit(800).all()
    msg_ids = [m.id for m in msgs]
    atts_by = {}
    if msg_ids:
        for a in ChatAttachment.query.filter(ChatAttachment.message_id.in_(msg_ids)).all():
            fn = (a.filename or '').lower()
            atts_by.setdefault(a.message_id, []).append({
                'id': a.id, 'filename': a.filename or 'ファイル',
                'content_type': a.content_type or '',
                'is_image': (a.content_type or '').startswith('image/') or fn.endswith(_IMG_EXTS),
            })
    # 既読数：このチャンネルを読んだ各ユーザーの last_read_id を取得し、
    # 各メッセージについて「投稿者以外で last_read_id >= m.id の人数」を数える
    reads = [(r.user_id, r.last_read_id or 0)
             for r in ChatRead.query.filter_by(channel_id=cid).all()]
    def _read_count(m):
        return sum(1 for (uid_, lr) in reads if uid_ != m.user_id and lr >= m.id)
    return jsonify({'messages': [{
        'id': m.id, 'user_id': m.user_id, 'user_name': m.user_name or '',
        'body': m.body or '', 'mine': m.user_id == u.id,
        'at': _fmt_jst(m.created_at),
        'read_count': _read_count(m),
        'attachments': atts_by.get(m.id, []),
    } for m in msgs]})


@app.route("/api/chat/channels/<int:cid>/reads", methods=["GET"])
@login_required
def api_chat_reads(cid):
    """自分の投稿メッセージごとの既読人数を返す（既読表示の更新用・軽量）"""
    u = _chat_user()
    c = ChatChannel.query.get_or_404(cid)
    if not _can_access_channel(c, u):
        return jsonify({})
    reads = [(r.user_id, r.last_read_id or 0)
             for r in ChatRead.query.filter_by(channel_id=cid).all()]
    my_msgs = ChatMessage.query.filter_by(channel_id=cid, user_id=u.id).all()
    out = {}
    for m in my_msgs:
        out[str(m.id)] = sum(1 for (uid_, lr) in reads if uid_ != u.id and lr >= m.id)
    return jsonify(out)


@app.route("/api/chat/channels/<int:cid>/messages", methods=["POST"])
@login_required
def api_chat_send(cid):
    u = _chat_user()
    c = ChatChannel.query.get_or_404(cid)
    if not _can_access_channel(c, u):
        return jsonify({'error': '権限がありません'}), 403
    is_pro = _chat_is_pro(u.tenant_id)
    attachments = []
    if request.content_type and 'multipart/form-data' in request.content_type:
        body = (request.form.get('body') or '').strip()
        files = request.files.getlist('files')
        if files and not is_pro:
            return jsonify({'ok': False, 'error': 'ファイル添付はチャットProプランの機能です'})
        for f in files:
            if not f or not f.filename:
                continue
            raw = f.read()
            if not raw:
                continue
            if len(raw) > 20 * 1024 * 1024:
                return jsonify({'ok': False, 'error': f'「{f.filename}」が大きすぎます（20MBまで）'})
            attachments.append((f.filename, f.mimetype or 'application/octet-stream', raw))
    else:
        data = request.get_json() or {}
        body = (data.get('body') or '').strip()
    if not body and not attachments:
        return jsonify({'ok': False, 'error': 'メッセージを入力してください'})
    m = ChatMessage(channel_id=cid, tenant_id=u.tenant_id, user_id=u.id,
                    user_name=_chat_display_name(u), body=body, has_attachments=bool(attachments))
    db.session.add(m)
    db.session.flush()
    for (fn, ct, raw) in attachments:
        db.session.add(ChatAttachment(message_id=m.id, channel_id=cid, tenant_id=u.tenant_id,
                                      filename=fn[:300], content_type=(ct or 'application/octet-stream')[:120],
                                      size=len(raw), data=raw))
    db.session.commit()
    return jsonify({'ok': True, 'id': m.id})


def _chat_mark_read(u, cid):
    """指定チャンネルを最新メッセージまで既読にする"""
    last = db.session.query(db.func.max(ChatMessage.id)).filter(
        ChatMessage.channel_id == cid).scalar() or 0
    rec = ChatRead.query.filter_by(channel_id=cid, user_id=u.id).first()
    if not rec:
        rec = ChatRead(channel_id=cid, user_id=u.id, last_read_id=last)
        db.session.add(rec)
    elif (rec.last_read_id or 0) < last:
        rec.last_read_id = last
        rec.updated_at = datetime.utcnow()
    else:
        return last
    db.session.commit()
    return last


@app.route("/api/chat/channels/<int:cid>/read", methods=["POST"])
@login_required
def api_chat_mark_read(cid):
    u = _chat_user()
    c = ChatChannel.query.get_or_404(cid)
    if not _can_access_channel(c, u):
        return jsonify({'error': '権限がありません'}), 403
    last = _chat_mark_read(u, cid)
    return jsonify({'ok': True, 'last_read_id': last})


@app.route("/api/chat/unread", methods=["GET"])
@login_required
def api_chat_unread():
    """チャンネル別・合計の未読件数を返す（自分の発言は除外）"""
    u = _chat_user()
    if not u or not u.tenant_id:
        return jsonify({'total': 0, 'channels': {}})
    chans = _visible_channels(u.tenant_id, u)
    cids = [c.id for c in chans]
    if not cids:
        return jsonify({'total': 0, 'channels': {}})
    reads = {r.channel_id: (r.last_read_id or 0)
             for r in ChatRead.query.filter(ChatRead.user_id == u.id,
                                            ChatRead.channel_id.in_(cids)).all()}
    per = {}
    total = 0
    # チャンネルごとに last_read より新しい自分以外のメッセージ数を集計
    for cid in cids:
        lr = reads.get(cid, 0)
        cnt = db.session.query(db.func.count(ChatMessage.id)).filter(
            ChatMessage.channel_id == cid,
            ChatMessage.user_id != u.id,
            ChatMessage.id > lr).scalar() or 0
        if cnt:
            per[cid] = cnt
            total += cnt
    return jsonify({'total': total, 'channels': per})


@app.route("/api/chat/attachments/<int:aid>", methods=["GET"])
@login_required
def api_chat_attachment(aid):
    u = _chat_user()
    a = ChatAttachment.query.get_or_404(aid)
    if not u or a.tenant_id != u.tenant_id:
        return "権限がありません", 403
    c = ChatChannel.query.get(a.channel_id)
    if not _can_access_channel(c, u):
        return "権限がありません", 403
    from flask import Response as _Resp
    ctype = a.content_type or 'application/octet-stream'
    inline = ctype.startswith('image/') or ctype == 'application/pdf'
    resp = _Resp(a.data or b'', mimetype=ctype)
    fn = (a.filename or 'file').replace('"', '')
    try:
        from urllib.parse import quote as _q
        resp.headers['Content-Disposition'] = f"{'inline' if inline else 'attachment'}; filename*=UTF-8''{_q(fn)}"
    except Exception:
        resp.headers['Content-Disposition'] = f'{"inline" if inline else "attachment"}; filename="{fn}"'
    return resp


@app.route("/echo-management")
@login_required
@block_super_admin
def echo_management():
    """反響管理表ページ"""
    stores = get_allowed_stores(ignore_active=True)  # サイドバー用
    active_ids = get_allowed_store_ids()  # アクティブ店舗のみ
    allowed_ids = [s.id for s in stores]  # サイドバー用全店舗
    staff_list = Staff.query.filter(Staff.store_id.in_(active_ids), Staff.is_active == True).all() if active_ids else []
    year, month = current_ym()
    store_id = active_ids[0] if active_ids else None
    cur_user = AppUser.query.get(session.get('app_user_id'))
    cur_role = cur_user.role if cur_user else 'staff'
    cur_staff_id = resolve_cur_staff_id(cur_user)
    is_manager = cur_role in ('owner', 'store_manager', 'super_admin')
    return render_template("echo_management.html",
                           stores=stores, staff_list=staff_list,
                           year=year, month=month, store_id=store_id,
                           cur_staff_id=cur_staff_id, is_manager=is_manager,
                           now=datetime.now())


@app.route("/api/echo-records", methods=["GET"])
@login_required
def api_echo_records_list():
    allowed = get_allowed_store_ids()
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]
    staff_id = request.args.get('staff_id', type=int)

    from datetime import date as _date
    m_start = _date(year, month, 1)
    m_end   = _date(year + (month // 12), (month % 12) + 1, 1)

    q = EchoRecord.query.filter(
        EchoRecord.store_id.in_(allowed),
        EchoRecord.echo_date >= m_start,
        EchoRecord.echo_date <  m_end,
    )
    if staff_id:
        q = q.filter_by(staff_id=staff_id)
    records = q.order_by(EchoRecord.echo_date.asc(), EchoRecord.id.asc()).all()

    # 各反響の「最後のメッセージの向き」を取得（最後が受信＝未返信）
    rec_ids = [r.id for r in records]
    last_dir = {}
    if rec_ids:
        for m in (MailMessage.query
                  .filter(MailMessage.echo_id.in_(rec_ids))
                  .order_by(MailMessage.created_at.asc(), MailMessage.id.asc()).all()):
            last_dir[m.echo_id] = m.direction

    def fd(d): return d.strftime('%Y-%m-%d') if d else None
    def sname(sid): s = Staff.query.get(sid); return s.name if s else ''
    return jsonify([{
        'id': r.id, 'store_id': r.store_id, 'staff_id': r.staff_id,
        'staff_name': sname(r.staff_id),
        'list_name': r.list_name or '', 'echo_date': fd(r.echo_date),
        'media': r.media or '', 'method': r.method or '',
        'first_contact_date': fd(r.first_contact_date),
        **{f'followup_{i}': fd(getattr(r, f'followup_{i}')) for i in range(1, 11)},
        'followup_phone': r.followup_phone or '',
        'has_reply': r.has_reply, 'has_phone': r.has_phone, 'has_line': r.has_line,
        'memo': r.memo or '',
        'customer_email': r.customer_email or '',
        'has_unread_reply': bool(r.has_unread_reply),
        'needs_reply': last_dir.get(r.id) == 'in' and not r.reply_dismissed,
        'reply_dismissed': bool(r.reply_dismissed),
        'has_phone_number': bool(r.has_phone_number),
        'status': r.status or '',
    } for r in records])


@app.route("/api/echo-records", methods=["POST"])
@login_required
def api_echo_records_add():
    data = request.get_json() or {}
    allowed = get_allowed_store_ids()
    if not allowed:
        return jsonify({'error': 'unauthorized'}), 403
    from datetime import datetime as _dt
    def pd(s):
        if not s: return None
        try: return _dt.strptime(s, '%Y-%m-%d').date()
        except: return None

    sid = int(data.get('store_id') or 0) or allowed[0]
    if sid not in allowed:
        sid = allowed[0]
    r = EchoRecord(
        store_id=sid,
        staff_id=int(data.get('staff_id') or 0) or None,
        list_name=data.get('list_name', ''),
        echo_date=pd(data.get('echo_date')),
        media=data.get('media', ''),
        method=data.get('method', ''),
        first_contact_date=pd(data.get('first_contact_date')),
        **{f'followup_{i}': pd(data.get(f'followup_{i}')) for i in range(1, 11)},
        has_reply=bool(data.get('has_reply')),
        has_phone=bool(data.get('has_phone')),
        has_line=bool(data.get('has_line')),
        memo=data.get('memo', ''),
    )
    db.session.add(r)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': r.id})


@app.route("/api/echo-records/<int:rid>", methods=["PUT"])
@login_required
def api_echo_records_update(rid):
    r = EchoRecord.query.get_or_404(rid)
    data = request.get_json() or {}
    from datetime import datetime as _dt
    def pd(s):
        if not s: return None
        try: return _dt.strptime(s, '%Y-%m-%d').date()
        except: return None

    # 部分更新対応：payload に含まれるキーのみ更新（インライン編集で他項目を消さない）
    if 'staff_id' in data:
        r.staff_id = int(data.get('staff_id') or 0) or None
    if 'list_name' in data:
        r.list_name = data.get('list_name')
    if 'echo_date' in data:
        r.echo_date = pd(data.get('echo_date')) or r.echo_date
    if 'media' in data:
        r.media = data.get('media')
    if 'method' in data:
        r.method = data.get('method')
    if 'first_contact_date' in data:
        r.first_contact_date = pd(data.get('first_contact_date'))
    for i in range(1, 11):
        key = f'followup_{i}'
        if key in data:
            setattr(r, key, pd(data.get(key)))
    if 'followup_phone' in data:
        r.followup_phone = (data.get('followup_phone') or '')[:60]
    if 'has_reply' in data:
        r.has_reply = bool(data.get('has_reply'))
    if 'has_phone' in data:
        r.has_phone = bool(data.get('has_phone'))
    if 'has_line' in data:
        r.has_line = bool(data.get('has_line'))
    if 'has_phone_number' in data:
        r.has_phone_number = bool(data.get('has_phone_number'))
    if 'reply_dismissed' in data:
        r.reply_dismissed = bool(data.get('reply_dismissed'))
    if 'status' in data:
        r.status = (data.get('status') or '') or None
    if 'memo' in data:
        r.memo = data.get('memo')
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/echo-records/<int:rid>", methods=["DELETE"])
@login_required
def api_echo_records_delete(rid):
    r = EchoRecord.query.get_or_404(rid)
    # 自動取込分は「削除済み」として記録し、再取込での復活を防ぐ
    if r.external_id and not ProcessedReaction.query.filter_by(store_id=r.store_id, external_id=r.external_id).first():
        db.session.add(ProcessedReaction(store_id=r.store_id, external_id=r.external_id))
    db.session.delete(r)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/customer-service")
@login_required
@block_super_admin
def customer_service():
    """接客管理表ページ"""
    stores = get_allowed_stores(ignore_active=True)  # サイドバー用
    active_ids = get_allowed_store_ids()  # アクティブ店舗のみ
    allowed_ids = [s.id for s in stores]  # サイドバー用全店舗
    staff_list = Staff.query.filter(Staff.store_id.in_(active_ids), Staff.is_active == True).all() if active_ids else []
    year, month = current_ym()
    store_id = active_ids[0] if active_ids else None
    cur_user = AppUser.query.get(session.get('app_user_id'))
    cur_role = cur_user.role if cur_user else 'staff'
    cur_staff_id = resolve_cur_staff_id(cur_user)
    is_manager = cur_role in ('owner', 'store_manager', 'super_admin')
    return render_template("customer_service.html",
                           stores=stores, staff_list=staff_list,
                           year=year, month=month, store_id=store_id,
                           cur_staff_id=cur_staff_id, is_manager=is_manager,
                           now=datetime.now())


# ─── DropdownOption API ──────────────────────────────────────────────────

def _get_tenant_id():
    """現在ログイン中ユーザーのtenant_idを返す"""
    uid = session.get('app_user_id')
    if not uid:
        return None
    u = AppUser.query.get(uid)
    return u.tenant_id if u else None


@app.route("/api/dropdown/<category>", methods=["GET"])
@login_required
def api_dropdown_get(category):
    """カテゴリのプルダウン選択肢を返す（テナント固有 + 共通デフォルト）"""
    tenant_id = _get_tenant_id()
    opts = DropdownOption.query.filter(
        DropdownOption.category == category,
        db.or_(DropdownOption.tenant_id == tenant_id, DropdownOption.tenant_id == None)
    ).order_by(DropdownOption.sort_order, DropdownOption.id).all()
    return jsonify([{'id': o.id, 'value': o.value} for o in opts])


@app.route("/api/dropdown/<category>", methods=["POST"])
@login_required
def api_dropdown_add(category):
    """選択肢を追加する"""
    data = request.get_json() or {}
    value = (data.get('value') or '').strip()
    if not value:
        return jsonify({'error': 'value required'}), 400
    tenant_id = _get_tenant_id()
    max_order = db.session.query(db.func.max(DropdownOption.sort_order)).filter_by(category=category).scalar() or 0
    opt = DropdownOption(tenant_id=tenant_id, category=category, value=value, sort_order=max_order + 1)
    db.session.add(opt)
    db.session.commit()
    return jsonify({'id': opt.id, 'value': opt.value})


@app.route("/api/dropdown/item/<int:option_id>", methods=["DELETE"])
@login_required
def api_dropdown_delete(option_id):
    """選択肢を削除する"""
    opt = DropdownOption.query.get_or_404(option_id)
    db.session.delete(opt)
    db.session.commit()
    return jsonify({'ok': True})


# ─────────────────────────────────────────────────────────────────────────

@app.route("/api/customer-service-records", methods=["GET"])
@login_required
def api_cs_records_list():
    allowed = get_allowed_store_ids()
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]
    staff_id = request.args.get('staff_id', type=int)

    from datetime import date as _date
    m_start = _date(year, month, 1)
    m_end   = _date(year + (month // 12), (month % 12) + 1, 1)

    q = CustomerServiceRecord.query.filter(
        CustomerServiceRecord.store_id.in_(allowed),
        CustomerServiceRecord.service_date >= m_start,
        CustomerServiceRecord.service_date <  m_end,
    )
    if staff_id:
        q = q.filter_by(staff_id=staff_id)
    records = q.order_by(CustomerServiceRecord.service_date.asc(), CustomerServiceRecord.id.asc()).all()

    def fd(d): return d.strftime('%Y-%m-%d') if d else None
    def sname(sid): s = Staff.query.get(sid); return s.name if s else ''
    return jsonify([{
        'id': r.id, 'store_id': r.store_id,
        'card_no': r.card_no or '', 'service_date': fd(r.service_date),
        'echo_media': r.echo_media or '',
        'staff_id': r.staff_id, 'staff_name': sname(r.staff_id),
        'customer_name': r.customer_name or '',
        'service_type': r.service_type or '',
        'visit_count': r.visit_count or 0,
        'status': r.status or '追客中',
        'memo': r.memo or '',
    } for r in records])


@app.route("/api/customer-service-records", methods=["POST"])
@login_required
def api_cs_records_add():
    data = request.get_json() or {}
    allowed = get_allowed_store_ids()
    if not allowed:
        return jsonify({'error': 'unauthorized'}), 403
    from datetime import datetime as _dt
    def pd(s):
        if not s: return None
        try: return _dt.strptime(s, '%Y-%m-%d').date()
        except: return None

    sid = int(data.get('store_id') or 0) or allowed[0]
    if sid not in allowed:
        sid = allowed[0]
    r = CustomerServiceRecord(
        store_id=sid,
        card_no=data.get('card_no', ''),
        service_date=pd(data.get('service_date')),
        echo_media=data.get('echo_media', ''),
        staff_id=int(data.get('staff_id') or 0) or None,
        customer_name=data.get('customer_name', ''),
        service_type=data.get('service_type', ''),
        visit_count=int(data.get('visit_count') or 0),
        status=data.get('status', '追客中'),
        memo=data.get('memo', ''),
    )
    db.session.add(r)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': r.id})


@app.route("/api/customer-service-records/<int:rid>", methods=["PUT"])
@login_required
def api_cs_records_update(rid):
    r = CustomerServiceRecord.query.get_or_404(rid)
    data = request.get_json() or {}
    from datetime import datetime as _dt
    def pd(s):
        if not s: return None
        try: return _dt.strptime(s, '%Y-%m-%d').date()
        except: return None

    r.card_no      = data.get('card_no', r.card_no)
    r.service_date = pd(data.get('service_date')) or r.service_date
    r.echo_media   = data.get('echo_media', r.echo_media)
    r.staff_id     = int(data.get('staff_id') or 0) or None
    r.customer_name = data.get('customer_name', r.customer_name)
    r.service_type = data.get('service_type', r.service_type)
    r.visit_count  = int(data.get('visit_count') or 0)
    r.status       = data.get('status', r.status)
    r.memo         = data.get('memo', r.memo)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/customer-service-records/<int:rid>", methods=["DELETE"])
@login_required
def api_cs_records_delete(rid):
    r = CustomerServiceRecord.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/leads")
@login_required
@manager_or_above_required
def leads_management():
    """反響管理ページ：リード一覧・追加"""
    stores = get_allowed_stores()
    allowed_ids = [s.id for s in stores]
    staff_list = Staff.query.filter(Staff.store_id.in_(allowed_ids), Staff.is_active == True).all()
    leads = (Lead.query
             .filter(Lead.store_id.in_(allowed_ids))
             .order_by(Lead.received_at.desc())
             .limit(100)
             .all())
    year, month = current_ym()
    store_id = allowed_ids[0] if allowed_ids else None
    return render_template("leads_management.html", stores=stores, staff_list=staff_list,
                           year=year, month=month, store_id=store_id, now=datetime.now())


@app.route("/leads/add", methods=["POST"])
def leads_add():
    """反響追加（フォームPOST）"""
    data = request.form
    lead = Lead(
        source=data.get('source', ''),
        received_at=datetime.now(),
        status=data.get('status', '未対応'),
        assigned_staff_id=data.get('assigned_staff_id') or None,
        store_id=data.get('store_id') or None,
        customer_name=data.get('customer_name', ''),
        note=data.get('note', ''),
        line_added=data.get('line_added') == 'on',
    )
    db.session.add(lead)
    db.session.commit()
    return redirect(url_for('leads_management'))


@app.route("/api/leads/add", methods=["POST"])
def api_leads_add():
    """反響追加API（JSON/フォーム両対応）"""
    data = request.get_json() or request.form
    sid = safe_store_id(data.get('store_id'))
    if not sid:
        return jsonify({'error': 'unauthorized'}), 403
    lead = Lead(
        source=data.get('source') or data.get('media', ''),
        received_at=datetime.now(),
        status=data.get('status', '未対応'),
        assigned_staff_id=data.get('assigned_staff_id') or data.get('assignee_id') or None,
        store_id=sid,
        customer_name=data.get('customer_name', ''),
        note=data.get('note') or data.get('memo', ''),
        line_added=str(data.get('line_added', '0')) in ('1', 'true', 'True', 'on'),
    )
    db.session.add(lead)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': lead.id})


@app.route("/leads/status/<int:lead_id>")
def leads_status_update(lead_id):
    """ステータス更新（GETパラメータで受け取り）"""
    lead = Lead.query.get_or_404(lead_id)
    new_status = request.args.get('status', lead.status)
    lead.status = new_status
    db.session.commit()
    return redirect(url_for('leads_management'))


@app.route("/accounting")
@login_required
@manager_or_above_required
def accounting():
    """会計・PL管理ページ"""
    stores = get_allowed_stores()
    allowed_ids = [s.id for s in stores]
    year, month = current_ym()
    store_id = allowed_ids[0] if allowed_ids else None
    return render_template("accounting.html", stores=stores, year=year, month=month,
                           store_id=store_id, now=datetime.now())


@app.route("/staff-ranking")
@login_required
@block_super_admin
def staff_ranking():
    """スタッフランキングページ"""
    stores = get_allowed_stores()
    allowed_ids = [s.id for s in stores]
    staff_list = Staff.query.filter(Staff.store_id.in_(allowed_ids), Staff.is_active == True).all()
    year, month = current_ym()
    return render_template("staff_ranking.html", stores=stores, staff_list=staff_list,
                           year=year, month=month, now=datetime.now())


# ── 幹部向け管理ツール：データAPI（JSON） ────────────────

def _pl_profit_for(pls, y, m):
    """PLレコード群から利益合計を計算（PLCustomValue優先）"""
    total = 0
    for p in pls:
        cv = PLCustomValue.query.filter_by(store_id=p.store_id, year=y, month=m).all()
        ad_cvs    = [c for c in cv if c.item_type == '広告費']
        fixed_cvs = [c for c in cv if c.item_type == '固定費']
        var_cvs   = [c for c in cv if c.item_type == '変動費']
        if ad_cvs:
            ad_t = sum(c.amount for c in ad_cvs)
        else:
            ad_t = (p.ad_cost or 0) or sum(getattr(p, col, 0) or 0 for col in [
                'suumo_cost','homes_cost','athome_cost','instagram_cost','tiktok_cost',
                'google_ads_cost','line_cost','hp_cost','meo_cost','other_ad_cost'])
        lb_t = (p.labor_cost or 0) or ((p.regular_salary or 0) + (p.parttime_salary or 0) + (p.commission_pay or 0))
        ft = sum(c.amount for c in fixed_cvs)
        vt = sum(c.amount for c in var_cvs)
        total += (p.revenue or 0) - ad_t - lb_t - ft - vt
    return total


def _pl_ad_for(pls, y, m):
    """PLレコード群から広告費合計を計算（PLCustomValue優先）"""
    total = 0
    for p in pls:
        cv = PLCustomValue.query.filter_by(store_id=p.store_id, year=y, month=m).all()
        ad_cvs = [c for c in cv if c.item_type == '広告費']
        if ad_cvs:
            total += sum(c.amount for c in ad_cvs)
        else:
            total += (p.ad_cost or 0) or sum(getattr(p, col, 0) or 0 for col in [
                'suumo_cost','homes_cost','athome_cost','instagram_cost','tiktok_cost',
                'google_ads_cost','line_cost','hp_cost','meo_cost','other_ad_cost'])
    return total


@app.route("/api/kpi/by-store")
@login_required
def api_kpi_by_store():
    """店舗別KPIサマリー（売上/契約数/申込数/反響数/利益/経費）"""
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]
    stores = get_allowed_stores()
    result = []
    for store in stores:
        kpis = SalesKPI.query.filter_by(store_id=store.id, year=year, month=month).all()
        sales       = sum(k.sales_amount or 0 for k in kpis)
        contracts   = sum(k.contracts    or 0 for k in kpis)
        applications= sum(k.applications or 0 for k in kpis)
        inquiries   = sum(k.inquiries    or 0 for k in kpis)
        # 利益・経費はPLRecordから
        pl = PLRecord.query.filter_by(store_id=store.id, year=year, month=month).first()
        profit   = pl.net_profit if pl else None
        # 経費 = sales - profit (KPIベース)
        expenses = max(0, sales - (profit or 0)) if profit is not None else None
        # 入金済み申込件数
        from sqlalchemy import extract as _ex
        approved_cnt = ApplicationRecord.query.filter(
            ApplicationRecord.store_id == store.id,
            ~ApplicationRecord.status.in_(['キャンセル', 'キャンセル振替']),
            _ex('year',  ApplicationRecord.application_date) == year,
            _ex('month', ApplicationRecord.application_date) == month,
            db.or_(ApplicationRecord.ad_amount != 0, ApplicationRecord.brokerage_fee != 0),
            db.or_(ApplicationRecord.ad_amount <= 0, ApplicationRecord.ad_approved == True),
            db.or_(ApplicationRecord.brokerage_fee <= 0, ApplicationRecord.brokerage_approved == True),
        ).count()
        result.append({
            'store_id':    store.id,
            'store_name':  store.name,
            'sales':       sales,
            'contracts':   contracts,
            'applications': applications,
            'inquiries':   inquiries,
            'approved_apps': approved_cnt,
            'profit':      profit,
            'expenses':    expenses,
        })
    return jsonify(result)


@app.route("/api/kpi/summary")
def api_kpi_summary():
    """KPIサマリを返す（year/monthパラメータで任意月指定可）"""
    cy, cm = current_ym()
    year  = request.args.get('year',  type=int) or cy
    month = request.args.get('month', type=int) or cm

    allowed_ids = get_allowed_store_ids()

    def get_kpis(y, m):
        q = SalesKPI.query.filter_by(year=y, month=m)
        if allowed_ids:
            q = q.filter(SalesKPI.store_id.in_(allowed_ids))
        return q.all()

    def get_pl(y, m):
        q = PLRecord.query.filter_by(year=y, month=m)
        if allowed_ids:
            q = q.filter(PLRecord.store_id.in_(allowed_ids))
        return q.all()

    # 今月・前月
    kpis_now  = get_kpis(year, month)
    prev_m    = month - 1 if month > 1 else 12
    prev_y    = year if month > 1 else year - 1
    kpis_prev = get_kpis(prev_y, prev_m)
    pls_now   = get_pl(year, month)
    pls_prev  = get_pl(prev_y, prev_m)

    kpi_sales_now  = sum(k.sales_amount for k in kpis_now)
    kpi_sales_prev = sum(k.sales_amount for k in kpis_prev)
    contracts_now  = sum(k.contracts for k in kpis_now)
    contracts_prev = sum(k.contracts for k in kpis_prev)

    # 今月PL（経理データ優先）
    rev_now     = sum(p.revenue for p in pls_now)
    profit_now  = _pl_profit_for(pls_now, year, month)
    ad_now      = _pl_ad_for(pls_now, year, month)

    rev_prev    = sum(p.revenue for p in pls_prev)
    profit_prev = _pl_profit_for(pls_prev, prev_y, prev_m)
    ad_prev     = _pl_ad_for(pls_prev, prev_y, prev_m)

    # 経理データ優先で売上を決定（PLRecordなければKPIから取得）
    sales_now  = rev_now  if rev_now  > 0 else kpi_sales_now
    sales_prev = rev_prev if rev_prev > 0 else kpi_sales_prev

    # 広告ROI（今月）
    roi_now  = round(rev_now  / ad_now  * 100, 1) if ad_now  > 0 else 0
    roi_prev = round(rev_prev / ad_prev * 100, 1) if ad_prev > 0 else 0

    # 未対応反響（テナント分離：自分の店舗のみカウント）
    unhandled = Lead.query.filter(
        Lead.store_id.in_(allowed_ids),
        Lead.status == '未対応'
    ).count()

    # 着地予測（今月日割り）
    today_day = date.today().day
    days_in_month = 31 if month in [1,3,5,7,8,10,12] else 30 if month in [4,6,9,11] else 28
    forecast = round(sales_now / today_day * days_in_month) if today_day > 0 else sales_now

    # 前年同月
    kpis_yoy  = get_kpis(year - 1, month)
    pls_yoy   = get_pl(year - 1, month)
    kpi_sales_yoy = sum(k.sales_amount for k in kpis_yoy)
    contracts_yoy = sum(k.contracts for k in kpis_yoy)
    rev_yoy    = sum(p.revenue for p in pls_yoy)
    sales_yoy  = rev_yoy if rev_yoy > 0 else kpi_sales_yoy
    profit_yoy = _pl_profit_for(pls_yoy, year - 1, month)
    ad_yoy     = _pl_ad_for(pls_yoy, year - 1, month)
    roi_yoy    = round(rev_yoy / ad_yoy * 100, 1) if ad_yoy > 0 else 0

    def diff_pct(now, prev):
        if prev == 0:
            return 0
        return round((now - prev) / prev * 100, 1)

    # 反響管理データがあればそちらを優先（自動連携）
    lead_stats_now  = LeadMediaStat.query.filter(LeadMediaStat.store_id.in_(allowed_ids), LeadMediaStat.year==year,  LeadMediaStat.month==month).all()
    lead_stats_prev = LeadMediaStat.query.filter(LeadMediaStat.store_id.in_(allowed_ids), LeadMediaStat.year==prev_y, LeadMediaStat.month==prev_m).all()
    if lead_stats_now:
        total_inquiries   = sum(s.inquiries    for s in lead_stats_now)
        total_visits      = sum(s.visits       for s in lead_stats_now)
        total_applications= sum(s.applications for s in lead_stats_now)
        total_contracts_l = sum(s.contracts    for s in lead_stats_now)
        total_cancels     = sum(s.cancellations for s in lead_stats_now)
    else:
        total_inquiries   = sum(k.inquiries    for k in kpis_now)
        total_visits      = sum(k.store_visits for k in kpis_now)
        total_applications= sum(k.applications for k in kpis_now)
        total_contracts_l = contracts_now
        total_cancels     = sum(k.cancellations for k in kpis_now)

    return jsonify({
        'year': year, 'month': month,
        # 今月売上
        'month_sales':         sales_now,
        'prev_month_sales':    sales_prev,
        'prev_year_sales':     sales_yoy,
        'mom_sales':           diff_pct(sales_now, sales_prev),
        'yoy_sales':           diff_pct(sales_now, sales_yoy),
        # 今月利益
        'month_profit':        profit_now,
        'prev_month_profit':   profit_prev,
        'prev_year_profit':    profit_yoy,
        'mom_profit':          diff_pct(profit_now, profit_prev),
        'yoy_profit':          diff_pct(profit_now, profit_yoy),
        # 広告ROI
        'roi':                 roi_now,
        'prev_month_roi':      roi_prev,
        'prev_year_roi':       roi_yoy,
        'mom_roi':             diff_pct(roi_now, roi_prev),
        'yoy_roi':             diff_pct(roi_now, roi_yoy),
        # 経費合計（売上 - 利益）
        'total_expenses':      max(0, sales_now  - profit_now),
        'prev_month_expenses': max(0, sales_prev - profit_prev),
        # 契約数
        'month_contracts':     contracts_now,
        'prev_month_contracts':contracts_prev,
        'prev_year_contracts': contracts_yoy,
        'mom_contracts':       diff_pct(contracts_now, contracts_prev),
        'yoy_contracts':       diff_pct(contracts_now, contracts_yoy),
        # ファネル詳細（反響管理データ優先）
        'total_inquiries':     total_inquiries,
        'total_store_visits':  total_visits,
        'total_viewings':      sum(k.viewings for k in kpis_now),
        'total_applications':  total_applications,
        'total_contracts':     total_contracts_l,
        'total_cancellations': total_cancels,
        'total_option_sales':  sum(k.option_sales for k in kpis_now),
        'from_lead_stats':     bool(lead_stats_now),
        'staff_count':         len(set(k.staff_id for k in kpis_now)),
        'visit_rate':          round(total_visits / total_inquiries * 100, 1) if total_inquiries else 0,
        'contract_rate':       round(contracts_now / total_inquiries * 100, 1) if total_inquiries else 0,
    })


@app.route("/api/kpi/monthly")
def api_kpi_monthly():
    """月次KPIデータをグラフ用に返す（from/toパラメータで期間指定可）"""
    store_id   = request.args.get('store_id', type=int)
    staff_id   = request.args.get('staff_id', type=int)   # スタッフ別フィルタ
    from_param = request.args.get('from')   # YYYY-MM
    to_param   = request.args.get('to')     # YYYY-MM
    months_data = []

    today = date.today()

    # 期間リストを生成
    def ym_range(fy, fm, ty, tm):
        base = fy * 12 + fm - 1
        end  = ty  * 12 + tm  - 1
        result = []
        for t in range(base, end + 1):
            result.append((t // 12, t % 12 + 1))
        return result

    if from_param and to_param:
        try:
            fy, fm = int(from_param[:4]), int(from_param[5:7])
            ty, tm = int(to_param[:4]),   int(to_param[5:7])
            periods = ym_range(fy, fm, ty, tm)
        except Exception:
            periods = None
    else:
        periods = None

    if not periods:
        # デフォルト：直近12ヶ月（当月含む）
        base_total = today.year * 12 + today.month - 1
        periods = [(t // 12, t % 12 + 1) for t in range(base_total - 11, base_total + 1)]

    allowed_ids = get_allowed_store_ids()
    # store_idパラメータ指定時はさらに絞り込み
    if store_id and store_id in allowed_ids:
        filter_ids = [store_id]
    elif store_id:
        filter_ids = []
    else:
        filter_ids = allowed_ids

    for y, m in periods:

        query = SalesKPI.query.filter_by(year=y, month=m)
        if filter_ids:
            query = query.filter(SalesKPI.store_id.in_(filter_ids))
        if staff_id:
            query = query.filter(SalesKPI.staff_id == staff_id)
        kpis = query.all()

        # スタッフ別フィルタ時はSalesKPIのみ使用（PLRecordは店舗単位のため）
        if staff_id:
            kpi_sales = sum(k.sales_amount for k in kpis)
            months_data.append({
                'label':       f'{y}/{m:02d}',
                'year':        y,
                'month':       m,
                'inquiries':   sum(k.inquiries for k in kpis),
                'contracts':   sum(k.contracts for k in kpis),
                'sales':       kpi_sales,
                'gross_profit': 0,
                'ad_cost':     0,
            })
            continue

        pls = PLRecord.query.filter_by(year=y, month=m)
        if filter_ids:
            pls = pls.filter(PLRecord.store_id.in_(filter_ids))
        pls = pls.all()

        def _calc_gp(pl):
            cv = PLCustomValue.query.filter_by(store_id=pl.store_id, year=y, month=m).all()
            ad_cvs = [c for c in cv if c.item_type == '広告費']
            fixed_cvs = [c for c in cv if c.item_type == '固定費']
            var_cvs   = [c for c in cv if c.item_type == '変動費']
            if ad_cvs:
                ad_t = sum(c.amount for c in ad_cvs)
            else:
                ad_t = (pl.ad_cost or 0) or sum(getattr(pl, col, 0) or 0 for col in [
                    'suumo_cost','homes_cost','athome_cost','instagram_cost','tiktok_cost',
                    'google_ads_cost','line_cost','hp_cost','meo_cost','other_ad_cost'])
            lb_t = (pl.labor_cost or 0) or ((pl.regular_salary or 0)+(pl.parttime_salary or 0)+(pl.commission_pay or 0))
            ft = sum(c.amount for c in fixed_cvs)
            vt = sum(c.amount for c in var_cvs)
            return (pl.revenue or 0) - ad_t - lb_t - ft - vt

        pl_revenue = sum(p.revenue or 0 for p in pls)
        kpi_sales  = sum(k.sales_amount for k in kpis)
        months_data.append({
            'label':       f'{y}/{m:02d}',
            'year':        y,
            'month':       m,
            'inquiries':   sum(k.inquiries for k in kpis),
            'contracts':   sum(k.contracts for k in kpis),
            'sales':       pl_revenue if pl_revenue > 0 else kpi_sales,
            'gross_profit':sum(_calc_gp(p) for p in pls),
            'ad_cost':     sum(p.ad_cost or 0 for p in pls),
        })

    return jsonify(months_data)


@app.route("/api/kpi/staff")
def api_kpi_staff():
    """スタッフ別KPIデータを返す。
    実データから集計：反響数=反響管理表 / 接客数=接客管理表 / 申込数・契約数=申込一覧表 /
    売上=入金済み一覧（入金日が当月の承認済み金額）。
    目標(target_sales)と編集用のkpi_idのみKPI台帳(SalesKPI)から取得。"""
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]
    store_id    = request.args.get('store_id', type=int)
    allowed_ids = get_allowed_store_ids()
    filter_ids = [store_id] if (store_id and store_id in allowed_ids) else allowed_ids
    if not filter_ids:
        return jsonify([])

    import calendar as _cal
    m_start = date(year, month, 1)
    m_end   = date(year, month, _cal.monthrange(year, month)[1])

    # KPI台帳（目標・編集用のみ）
    kpi_map = {}
    for k in SalesKPI.query.filter_by(year=year, month=month).filter(SalesKPI.store_id.in_(filter_ids)).all():
        kpi_map[k.staff_id] = k

    # 反響数（反響管理表：当月の反響レコード件数）
    echo_cnt = {}
    for e in EchoRecord.query.filter(EchoRecord.store_id.in_(filter_ids),
                                     EchoRecord.echo_date >= m_start,
                                     EchoRecord.echo_date <= m_end).all():
        echo_cnt[e.staff_id] = echo_cnt.get(e.staff_id, 0) + 1

    # 接客数（接客管理表：当月の接客レコード件数）
    cs_cnt = {}
    for c in CustomerServiceRecord.query.filter(CustomerServiceRecord.store_id.in_(filter_ids),
                                                CustomerServiceRecord.service_date >= m_start,
                                                CustomerServiceRecord.service_date <= m_end).all():
        cs_cnt[c.staff_id] = cs_cnt.get(c.staff_id, 0) + 1

    # 申込一覧表（当月申込・キャンセル除く）：申込数 / 契約数 / 付帯件数
    app_stats = {}
    for a in ApplicationRecord.query.filter(
            ApplicationRecord.store_id.in_(filter_ids),
            ~ApplicationRecord.status.in_(['キャンセル', 'キャンセル振替']),
            db.extract('year',  ApplicationRecord.application_date) == year,
            db.extract('month', ApplicationRecord.application_date) == month).all():
        st = app_stats.setdefault(a.staff_id, {'app': 0, 'contracts': 0, 'll': 0, 'fire': 0, 'mv': 0})
        st['app'] += 1
        if a.status == '契約': st['contracts'] += 1
        if a.lifeline:         st['ll']   += 1
        if a.fire_insurance:   st['fire'] += 1
        if a.moving:           st['mv']   += 1

    # 売上（入金済み一覧：入金日が当月の承認済み金額をスタッフ別に集計）
    paid_rev = {}
    for r in ApplicationRecord.query.filter(
            ApplicationRecord.store_id.in_(filter_ids),
            ~ApplicationRecord.status.in_(['キャンセル', 'キャンセル振替'])).all():
        amt = 0
        if _approved_in_month(r, 'brokerage', year, month): amt += (r.brokerage_fee or 0)
        if _approved_in_month(r, 'option', year, month):    amt += (r.option_amount or 0)
        if _approved_in_month(r, 'ad', year, month):         amt += _ad_yen(r)
        if amt:
            paid_rev[r.staff_id] = paid_rev.get(r.staff_id, 0) + amt

    # 表示対象：実績があった or KPI台帳に行があるスタッフの和集合
    staff_ids = set(kpi_map) | set(echo_cnt) | set(cs_cnt) | set(app_stats) | set(paid_rev)
    staff_ids.discard(None)

    result = []
    for sid in staff_ids:
        staff = Staff.query.get(sid)
        if not staff or staff.store_id not in filter_ids:
            continue
        kpi = kpi_map.get(sid)
        st  = app_stats.get(sid, {'app': 0, 'contracts': 0, 'll': 0, 'fire': 0, 'mv': 0})
        app_ct = st['app']
        _r = lambda c: round(c / app_ct * 100, 1) if app_ct else 0
        store = Store.query.get(staff.store_id)
        result.append({
            'kpi_id':       kpi.id if kpi else None,
            'staff_id':     sid,
            'staff_name':   staff.name,
            'store_name':   store.name if store else '',
            'role':         staff.role or '',
            'inquiries':    echo_cnt.get(sid, 0),    # 反響数＝反響管理表
            'store_visits': cs_cnt.get(sid, 0),      # 接客数＝接客管理表
            'viewings':     0,
            'applications': app_ct,                  # 申込数＝申込一覧表
            'contracts':    st['contracts'],         # 契約数＝申込一覧表（ステータス=契約）
            'cancellations': kpi.cancellations if kpi else 0,
            'sales_amount': paid_rev.get(sid, 0),    # 売上＝入金済み一覧
            'option_sales': kpi.option_sales if kpi else 0,
            'estimated_sales': (kpi.estimated_sales or 0) if kpi else 0,
            'target_sales':    (kpi.target_sales or 0) if kpi else 0,
            'app_count_real':  app_ct,
            'll_contracts':    st['ll'],
            'fire_contracts':  st['fire'],
            'moving_contracts': st['mv'],
            'll_rate':         _r(st['ll']),
            'fire_rate':       _r(st['fire']),
            'moving_rate':     _r(st['mv']),
        })

    # 売上降順でソート
    result.sort(key=lambda x: x['sales_amount'], reverse=True)
    return jsonify(result)


@app.route("/api/leads/summary")
def api_leads_summary():
    """反響サマリを返す（媒体別・ステータス別）"""
    store_id  = request.args.get('store_id', type=int)
    year      = request.args.get('year',  type=int)
    month     = request.args.get('month', type=int)
    allowed_ids = get_allowed_store_ids()

    # テナント分離: 許可されたstore_idのみ
    query = Lead.query.filter(Lead.store_id.in_(allowed_ids))
    if store_id and store_id in allowed_ids:
        query = query.filter(Lead.store_id == store_id)
    if year and month:
        start = datetime(year, month, 1)
        if month == 12:
            end = datetime(year + 1, 1, 1)
        else:
            end = datetime(year, month + 1, 1)
        query = query.filter(Lead.received_at >= start, Lead.received_at < end)

    leads = query.all()

    # 媒体別集計
    by_source = {}
    for lead in leads:
        src = lead.source or '不明'
        by_source.setdefault(src, 0)
        by_source[src] += 1

    # ステータス別集計
    by_status = {}
    for lead in leads:
        st = lead.status or '不明'
        by_status.setdefault(st, 0)
        by_status[st] += 1

    # リード一覧
    leads_list = []
    for lead in leads:
        staff = Staff.query.get(lead.assigned_staff_id) if lead.assigned_staff_id else None
        leads_list.append({
            'id':            lead.id,
            'received_at':   lead.received_at.strftime('%Y-%m-%d %H:%M') if lead.received_at else '',
            'media':         lead.source or '',
            'source':        lead.source or '',
            'customer_name': lead.customer_name or '',
            'status':        lead.status or '未対応',
            'assignee':      staff.name if staff else '',
            'assigned_staff_id': lead.assigned_staff_id,
            'memo':          lead.note or '',
            'note':          lead.note or '',
            'line_added':    lead.line_added,
            'store_id':      lead.store_id,
        })

    # 媒体別サマリー（media_summary形式）
    media_summary_dict = {}
    for lead in leads:
        src = lead.source or '不明'
        if src not in media_summary_dict:
            media_summary_dict[src] = {'media': src, 'count': 0, 'cost': 0, 'visits': 0, 'contracts': 0}
        media_summary_dict[src]['count'] += 1
        if lead.status in ('来店', '申込', '契約', '内見'):
            media_summary_dict[src]['visits'] += 1
        if lead.status == '契約':
            media_summary_dict[src]['contracts'] += 1
    media_summary = list(media_summary_dict.values())

    # 月次推移（過去6ヶ月）
    monthly_labels = []
    monthly_datasets = {}
    today = date.today()
    for i in range(5, -1, -1):
        t = today.replace(day=1)
        for _ in range(i):
            t = (t - timedelta(days=1)).replace(day=1)
        y, m = t.year, t.month
        lbl = f'{y}/{m:02d}'
        monthly_labels.append(lbl)
        period_leads = Lead.query.filter(
            Lead.received_at >= datetime(y, m, 1),
            Lead.received_at < (datetime(y+1, 1, 1) if m == 12 else datetime(y, m+1, 1))
        ).all()
        for lead in period_leads:
            src = lead.source or '不明'
            if src not in monthly_datasets:
                monthly_datasets[src] = [0] * 6
            idx = 5 - i
            monthly_datasets[src][idx] += 1

    return jsonify({
        'total':            len(leads),
        'by_source':        by_source,
        'by_status':        by_status,
        'line_added_count': sum(1 for l in leads if l.line_added),
        'leads':            leads_list,
        'media_summary':    media_summary,
        'monthly':          {'labels': monthly_labels, 'datasets': monthly_datasets},
    })


_PL_AD_COL = {
    'SUUMO': 'suumo_cost', "HOME'S": 'homes_cost', 'アットホーム': 'athome_cost',
    'Instagram': 'instagram_cost', 'TikTok': 'tiktok_cost', '自社HP': 'hp_cost',
    'LINE': 'line_cost', 'MEO': 'meo_cost', 'Google広告': 'google_ads_cost',
}


def _auto_lead_media_stats(store_id, year, month):
    """媒体別の月次統計を各管理表から自動集計する（⑰）。
    反響数/返信/LINE追加→反響管理表, 接客数→接客管理表,
    申込/契約/キャンセル/売上見込み→顧客管理表, 広告費→経理PL"""
    from types import SimpleNamespace
    import calendar as _cal
    last_day = _cal.monthrange(year, month)[1]
    m_start = date(year, month, 1)
    m_end   = date(year, month, last_day)

    agg = {}  # media -> dict
    def slot(media):
        key = media or '不明'
        if key not in agg:
            agg[key] = dict(id=None, media=key, inquiries=0, replies=0, line_added=0, visits=0,
                            applications=0, contracts=0, cancellations=0, cancel_amount=0,
                            estimated_sales=0, ad_cost=0)
        return agg[key]

    # 反響管理表（EchoRecord）→ 反響数・返信・LINE追加
    echoes = EchoRecord.query.filter(
        EchoRecord.store_id == store_id,
        EchoRecord.echo_date >= m_start, EchoRecord.echo_date <= m_end,
    ).all()
    for e in echoes:
        s = slot(e.media)
        s['inquiries'] += 1
        if e.has_reply: s['replies'] += 1
        if e.has_line:  s['line_added'] += 1

    # 接客管理表（CustomerServiceRecord）→ 接客数
    css = CustomerServiceRecord.query.filter(
        CustomerServiceRecord.store_id == store_id,
        CustomerServiceRecord.service_date >= m_start,
        CustomerServiceRecord.service_date <= m_end,
    ).all()
    for c in css:
        slot(c.echo_media)['visits'] += 1

    # 顧客管理表（ApplicationRecord）→ 申込・契約・キャンセル・売上見込み
    apps = ApplicationRecord.query.filter(
        ApplicationRecord.store_id == store_id,
        db.extract('year',  ApplicationRecord.application_date) == year,
        db.extract('month', ApplicationRecord.application_date) == month,
    ).all()
    for a in apps:
        s = slot(a.media)
        if a.status in ('キャンセル', 'キャンセル振替'):
            s['cancellations'] += 1
            s['cancel_amount'] += _record_total_amount(a)
        else:
            s['applications'] += 1
            s['estimated_sales'] += _record_total_amount(a)
            if a.status == '契約':
                s['contracts'] += 1

    # 経理PL → 広告費（媒体別）
    pl = PLRecord.query.filter_by(store_id=store_id, year=year, month=month).first()
    if pl:
        for media, col in _PL_AD_COL.items():
            v = getattr(pl, col, 0) or 0
            if v:
                slot(media)['ad_cost'] += v
    for cv in PLCustomValue.query.filter_by(store_id=store_id, year=year, month=month).all():
        if cv.item_type == '広告費' and (cv.amount or 0):
            slot(cv.item_name)['ad_cost'] += cv.amount or 0

    # 手動入力・Excel取込（LeadMediaStat）をマージ（自動集計に加算）。手動分はidを持たせ編集可能に
    for ms in LeadMediaStat.query.filter_by(store_id=store_id, year=year, month=month).all():
        s = slot(ms.media)
        s['id'] = ms.id
        s['inquiries']       += ms.inquiries or 0
        s['replies']         += ms.replies or 0
        s['line_added']      += ms.line_added or 0
        s['visits']          += ms.visits or 0
        s['applications']    += ms.applications or 0
        s['contracts']       += ms.contracts or 0
        s['cancellations']   += ms.cancellations or 0
        s['cancel_amount']   += ms.cancel_amount or 0
        s['estimated_sales'] += ms.estimated_sales or 0
        s['ad_cost']         += ms.ad_cost or 0

    # 自動集計分はid=None（編集不可）、手動入力分はLeadMediaStatのid（編集可）
    return [SimpleNamespace(**v) for v in agg.values()]


@app.route("/api/leads/monthly-stats")
@login_required
def api_leads_monthly_stats():
    """媒体別月次反響統計を返す（各管理表から自動集計）"""
    cy, cm = current_ym()
    year  = request.args.get('year',  type=int) or cy
    month = request.args.get('month', type=int) or cm
    # ignore_active=True で保存時と同じ基準で店舗を解決する
    allowed = get_allowed_store_ids()
    if not allowed:
        return jsonify({'stats': [], 'totals': {}, 'trend': []}), 200
    req_sid = request.args.get('store_id', type=int) or 0
    store_id = req_sid if req_sid and req_sid in allowed else allowed[0]

    stats = _auto_lead_media_stats(store_id, year, month)

    # 前月比較
    prev_m = month - 1 if month > 1 else 12
    prev_y = year if month > 1 else year - 1
    prev_stats = _auto_lead_media_stats(store_id, prev_y, prev_m)
    prev_dict = {s.media: s for s in prev_stats}

    def safe_div(a, b, pct=False):
        if not b: return 0
        return round(a / b * (100 if pct else 1))

    def pct_diff(cur, prv):
        """前月比（%）を返す。Noneなら比較不可"""
        if prv is None or prv == 0: return None
        return round((cur - prv) / prv * 100, 1)

    result = []
    for s in stats:
        prev = prev_dict.get(s.media)
        # 売上見込みベースでROAS計算
        roas = round(s.estimated_sales / s.ad_cost * 100, 1) if s.ad_cost else 0
        result.append({
            'id': s.id,
            'media': s.media,
            'inquiries': s.inquiries,
            'replies': s.replies,
            'line_added': s.line_added,
            'visits': s.visits,
            'applications': s.applications,
            'contracts': s.contracts,
            'cancellations': s.cancellations,
            'cancel_amount': s.cancel_amount,
            'estimated_sales': s.estimated_sales,
            'ad_cost': s.ad_cost,
            'cvr': safe_div(s.contracts, s.inquiries, pct=True),
            'cpa': safe_div(s.ad_cost, s.inquiries) if s.inquiries else 0,
            'cpo': safe_div(s.ad_cost, s.contracts) if s.contracts else 0,
            'roas': roas,
            # 前月比
            'prev_inquiries':    prev.inquiries    if prev else None,
            'prev_applications': prev.applications if prev else None,
            'prev_contracts':    prev.contracts    if prev else None,
            'prev_estimated':    prev.estimated_sales if prev else None,
            'prev_ad_cost':      prev.ad_cost      if prev else None,
            'mom_inq':  pct_diff(s.inquiries,     prev.inquiries     if prev else None),
            'mom_app':  pct_diff(s.applications,  prev.applications  if prev else None),
            'mom_con':  pct_diff(s.contracts,     prev.contracts     if prev else None),
            'mom_est':  pct_diff(s.estimated_sales, prev.estimated_sales if prev else None),
        })

    # 合計
    def total(field):
        return sum(getattr(s, field, 0) or 0 for s in stats)

    def ptotal(field):
        return sum(getattr(s, field, 0) or 0 for s in prev_stats)

    tot_inq = total('inquiries')
    tot_app = total('applications')
    tot_con = total('contracts')
    tot_est = total('estimated_sales')
    tot_ad  = total('ad_cost')
    totals = {
        'media': '合計',
        'inquiries':      tot_inq,
        'replies':        total('replies'),
        'line_added':     total('line_added'),
        'visits':         total('visits'),
        'applications':   tot_app,
        'contracts':      tot_con,
        'cancellations':  total('cancellations'),
        'cancel_amount':  total('cancel_amount'),
        'estimated_sales': tot_est,
        'ad_cost':        tot_ad,
        'cvr':  safe_div(tot_con, tot_inq, pct=True),
        'cpa':  safe_div(tot_ad, tot_inq) if tot_inq else 0,
        'cpo':  safe_div(tot_ad, tot_con) if tot_con else 0,
        'roas': round(tot_est / tot_ad * 100, 1) if tot_ad else 0,
        # 前月合計
        'prev_inquiries':   ptotal('inquiries'),
        'prev_applications': ptotal('applications'),
        'prev_contracts':   ptotal('contracts'),
        'prev_estimated':   ptotal('estimated_sales'),
        'mom_inq': pct_diff(tot_inq, ptotal('inquiries')),
        'mom_app': pct_diff(tot_app, ptotal('applications')),
        'mom_con': pct_diff(tot_con, ptotal('contracts')),
        'mom_est': pct_diff(tot_est, ptotal('estimated_sales')),
    }

    # 月次トレンド（過去6ヶ月）
    trend = []
    for i in range(5, -1, -1):
        tm, ty = month - i, year
        while tm <= 0:
            tm += 12; ty -= 1
        ms = _auto_lead_media_stats(store_id, ty, tm)
        trend.append({
            'label': f'{ty}/{tm:02d}',
            'inquiries':      sum(s.inquiries for s in ms),
            'contracts':      sum(s.contracts for s in ms),
            'estimated_sales': sum(s.estimated_sales for s in ms),
        })

    return jsonify({'stats': result, 'totals': totals, 'trend': trend, 'year': year, 'month': month})


@app.route("/api/leads/trend")
def api_leads_trend():
    """反響月次トレンドを返す（from/toパラメータで期間指定可、デフォルト直近6ヶ月）"""
    store_id   = safe_store_id(request.args.get('store_id', type=int))
    from_param = request.args.get('from')
    to_param   = request.args.get('to')
    today = date.today()

    if from_param and to_param:
        try:
            fy, fm = int(from_param[:4]), int(from_param[5:7])
            ty, tm_e = int(to_param[:4]), int(to_param[5:7])
            base = fy * 12 + fm - 1
            end  = ty * 12 + tm_e - 1
            periods = [(t // 12, t % 12 + 1) for t in range(base, end + 1)]
        except Exception:
            periods = None
    else:
        periods = None

    if not periods:
        base_total = today.year * 12 + today.month - 1
        periods = [(t // 12, t % 12 + 1) for t in range(base_total - 5, base_total + 1)]

    trend = []
    for y, m in periods:
        ms = LeadMediaStat.query.filter_by(store_id=store_id, year=y, month=m).all()
        trend.append({
            'label':           f'{y}/{m:02d}',
            'inquiries':       sum(s.inquiries for s in ms),
            'contracts':       sum(s.contracts for s in ms),
            'estimated_sales': sum(s.estimated_sales for s in ms),
        })
    return jsonify(trend)


@app.route("/api/leads/monthly-stats", methods=["POST"])
@login_required
def api_leads_monthly_stats_input():
    """媒体別月次反響統計を手動入力"""
    data = request.get_json() or request.form
    year  = int(data.get('year', current_ym()[0]))
    month = int(data.get('month', current_ym()[1]))
    # ignore_active=True で保存・取得ともに同じ基準にする
    allowed = get_allowed_store_ids()
    if not allowed:
        return jsonify({'error': 'unauthorized'}), 403
    try:
        req_sid = int(data.get('store_id') or 0)
    except (TypeError, ValueError):
        req_sid = 0
    store_id = req_sid if req_sid and req_sid in allowed else allowed[0]
    media = data.get('media', '').strip()
    if not media:
        return jsonify({'error': '媒体名は必須です'}), 400

    stat = LeadMediaStat.query.filter_by(store_id=store_id, year=year, month=month, media=media).first()
    if not stat:
        stat = LeadMediaStat(store_id=store_id, year=year, month=month, media=media)
        db.session.add(stat)

    for field in ['inquiries','replies','line_added','visits','applications','contracts','cancellations']:
        v = data.get(field)
        if v is not None:
            setattr(stat, field, int(float(v) or 0))
    for field in ['cancel_amount','estimated_sales','actual_payment','ad_cost']:
        v = data.get(field)
        if v is not None:
            setattr(stat, field, float(v or 0))

    db.session.commit()
    return jsonify({'status': 'ok', 'id': stat.id, 'store_id': store_id, 'year': year, 'month': month})


@app.route("/api/leads/monthly-stats/<int:stat_id>", methods=["PUT"])
def api_leads_monthly_stats_update(stat_id):
    """媒体別月次反響統計を更新"""
    data = request.get_json() or request.form
    stat = LeadMediaStat.query.get_or_404(stat_id)

    if data.get('media'):
        stat.media = data.get('media').strip()
    for field in ['inquiries','replies','line_added','visits','applications','contracts','cancellations']:
        v = data.get(field)
        if v is not None:
            setattr(stat, field, int(float(v) or 0))
    for field in ['cancel_amount','estimated_sales','actual_payment','ad_cost']:
        v = data.get(field)
        if v is not None:
            setattr(stat, field, float(v or 0))

    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/leads/monthly-stats/<int:stat_id>", methods=["DELETE"])
def api_leads_monthly_stats_delete(stat_id):
    """媒体別月次反響統計を削除"""
    stat = LeadMediaStat.query.get_or_404(stat_id)
    db.session.delete(stat)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/leads/import-excel-stats", methods=["POST"])
def api_leads_import_excel_stats():
    """反響統計ExcelをインポートしてLeadMediaStatに保存"""
    year  = int(request.form.get('year',  current_ym()[0]))
    month = int(request.form.get('month', current_ym()[1]))
    store_id = safe_store_id(request.form.get('store_id', type=int))
    if not store_id:
        return jsonify({'error': 'unauthorized'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'ファイルが必要です'}), 400

    import openpyxl
    file = request.files['file']
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active

    SKIP = {'総費用対効果', '総反響数', None, ''}

    def safe_int(v):
        if v is None or str(v) in ('#DIV/0!', 'nan', ''): return 0
        try: return int(float(str(v)))
        except: return 0

    def safe_float(v):
        if v is None or str(v) in ('#DIV/0!', 'nan', ''): return 0.0
        try: return float(str(v))
        except: return 0.0

    imported = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 2: continue
        media = row[1]
        if not media or str(media).strip() in SKIP: continue
        media = str(media).strip()

        stat = LeadMediaStat.query.filter_by(store_id=store_id, year=year, month=month, media=media).first()
        if not stat:
            stat = LeadMediaStat(store_id=store_id, year=year, month=month, media=media)
            db.session.add(stat)

        stat.inquiries      = safe_int(row[2] if len(row) > 2 else None)
        stat.replies        = safe_int(row[3] if len(row) > 3 else None)
        stat.line_added     = safe_int(row[4] if len(row) > 4 else None)
        stat.visits         = safe_int(row[5] if len(row) > 5 else None)
        stat.applications   = safe_int(row[6] if len(row) > 6 else None)
        stat.contracts      = safe_int(row[7] if len(row) > 7 else None)
        stat.cancellations  = safe_int(row[8] if len(row) > 8 else None)
        stat.cancel_amount  = safe_float(row[9] if len(row) > 9 else None)
        stat.estimated_sales= safe_float(row[10] if len(row) > 10 else None)
        stat.actual_payment = safe_float(row[11] if len(row) > 11 else None)
        stat.ad_cost        = safe_float(row[12] if len(row) > 12 else None)
        imported += 1

    db.session.commit()
    return jsonify({'status': 'ok', 'imported': imported, 'year': year, 'month': month})


def _get_ad_items(store_id, y, m, pl=None):
    """広告費カスタム行を取得。なければ固定列からフォールバック"""
    ad_cvs = PLCustomValue.query.filter_by(store_id=store_id, year=y, month=m, item_type='広告費').all()
    if ad_cvs:
        return [{'name': v.item_name, 'amount': v.amount} for v in ad_cvs]
    if not pl:
        return []
    _AD_COLS = [('suumo_cost','SUUMO'),('homes_cost',"HOME'S"),('athome_cost','at home'),
                ('instagram_cost','Instagram'),('tiktok_cost','TikTok'),('google_ads_cost','Google広告'),
                ('line_cost','LINE'),('hp_cost','HP'),('meo_cost','MEO'),('other_ad_cost','その他')]
    return [{'name': n, 'amount': getattr(pl, col, 0) or 0}
            for col, n in _AD_COLS if (getattr(pl, col, 0) or 0) > 0]


@app.route("/api/pl/summary")
def api_pl_summary():
    """PLサマリを返す（店舗別・月次）"""
    store_id    = request.args.get('store_id', type=int)
    year        = request.args.get('year',  type=int) or current_ym()[0]
    month       = request.args.get('month', type=int) or current_ym()[1]
    allowed_ids = get_allowed_store_ids()

    # テナント分離: 許可されたstore_idのみ
    query = PLRecord.query.filter_by(year=year, month=month).filter(PLRecord.store_id.in_(allowed_ids))
    if store_id and store_id in allowed_ids:
        query = query.filter(PLRecord.store_id == store_id)
    pls = query.all()

    result = []
    for pl in pls:
        store = Store.query.get(pl.store_id)
        # カスタム費用（固定費・変動費・広告費）を取得
        custom_vals = PLCustomValue.query.filter_by(
            store_id=pl.store_id, year=year, month=month
        ).all()
        fixed_items    = [{'name': v.item_name, 'amount': v.amount} for v in custom_vals if (v.item_type or '固定費') == '固定費']
        variable_items = [{'name': v.item_name, 'amount': v.amount} for v in custom_vals if (v.item_type or '固定費') == '変動費']
        custom_items   = [{'name': v.item_name, 'amount': v.amount} for v in custom_vals if (v.item_type or '固定費') not in ('固定費','変動費','広告費')]
        ad_items       = _get_ad_items(pl.store_id, year, month, pl)

        # 広告費合計
        ad_cost_val = sum(i['amount'] for i in ad_items)

        # 人件費合計
        labor_total = (pl.regular_salary or 0) + (pl.parttime_salary or 0) + (pl.commission_pay or 0)
        labor_cost_val = pl.labor_cost if (pl.labor_cost and pl.labor_cost > 0) else labor_total

        # カスタム固定費・変動費合計
        fixed_total    = sum(item['amount'] for item in fixed_items)
        variable_total = sum(item['amount'] for item in variable_items)

        # 粗利・営業利益 = 売上 - 全経費
        total_cost = ad_cost_val + labor_cost_val + fixed_total + variable_total
        gross_profit_calc = pl.revenue - total_cost
        # DBに保存された gross_profit があればそちらを使用（0の場合は計算値）
        gross_profit_val = pl.gross_profit if pl.gross_profit != 0 else gross_profit_calc
        operating_profit = gross_profit_calc  # 営業利益 = 粗利（全経費控除後）

        result.append({
            'pl_id':          pl.id,
            'store_id':       pl.store_id,
            'store_name':     store.name if store else '不明',
            'revenue':        pl.revenue,
            'gross_profit':   gross_profit_calc,   # 自動計算値
            'gross_margin':   round(gross_profit_calc / pl.revenue * 100, 1) if pl.revenue else 0,
            'ad_cost':        ad_cost_val,
            'labor_cost':     labor_cost_val,
            'other_fixed':    fixed_total,
            'other_variable': variable_total,
            'operating_profit': operating_profit,
            'op_margin':      round(operating_profit / pl.revenue * 100, 1) if pl.revenue else 0,
            # 収入詳細
            'brokerage_fee':       pl.brokerage_fee or 0,
            'ad_income':           pl.ad_income or 0,
            'lifeline_income':     pl.lifeline_income or 0,
            'moving_income':       pl.moving_income or 0,
            'fire_insurance_income': pl.fire_insurance_income or 0,
            'other_income':        pl.other_income or 0,
            # 広告費詳細（動的行）
            'ad_items':       ad_items,
            # 人件費詳細
            'regular_salary':  pl.regular_salary or 0,
            'parttime_salary': pl.parttime_salary or 0,
            'social_insurance': pl.commission_pay or 0,  # commission_payを社会保険料として使用
            # 固定費詳細
            'pl_rent':       pl.pl_rent or 0,
            'pl_parking':    pl.pl_parking or 0,
            'pl_copier':     pl.pl_copier or 0,
            'pl_internet':   pl.pl_internet or 0,
            'pl_consultant': pl.pl_consultant or 0,
            'pl_insurance':  pl.pl_insurance or 0,
            'pl_cloud':      pl.pl_cloud or 0,
            # カスタム費用
            'fixed_items':    fixed_items,
            'variable_items': variable_items,
            'custom_items':   custom_items,
        })

    # テンプレート一覧（type別・テナント分離）
    _pl_sid = safe_store_id()
    all_items = PLCustomItem.query.filter_by(store_id=_pl_sid).order_by(PLCustomItem.sort_order).all() if _pl_sid else []
    template_fixed    = [i.name for i in all_items if (i.item_type or '固定費') == '固定費']
    template_variable = [i.name for i in all_items if (i.item_type or '固定費') == '変動費']
    template_items    = [i.name for i in all_items if (i.item_type or '固定費') not in ('固定費','変動費')]

    # prev_month / prev_year サポート
    def get_prev_summary(y, m):
        # 店舗フィルタ（選択中の店舗のみ）
        pq = PLRecord.query.filter_by(year=y, month=m)
        if allowed_ids:
            pq = pq.filter(PLRecord.store_id.in_(allowed_ids))
        if store_id and store_id in allowed_ids:
            pq = pq.filter(PLRecord.store_id == store_id)
        prev_pls = pq.all()
        if not prev_pls:
            return None
        pl = prev_pls[0]
        cv = PLCustomValue.query.filter_by(store_id=pl.store_id, year=y, month=m).all()
        fi = [{'name': v.item_name, 'amount': v.amount} for v in cv if (v.item_type or '固定費') == '固定費']
        vi = [{'name': v.item_name, 'amount': v.amount} for v in cv if (v.item_type or '固定費') == '変動費']
        ai = _get_ad_items(pl.store_id, y, m, pl)
        ad_v = sum(i['amount'] for i in ai) if ai else (pl.ad_cost or 0)
        lb_t = (pl.regular_salary or 0)+(pl.parttime_salary or 0)+(pl.commission_pay or 0)
        lb_v = pl.labor_cost if (pl.labor_cost and pl.labor_cost > 0) else lb_t
        ft = sum(i['amount'] for i in fi)
        vt = sum(i['amount'] for i in vi)
        gp = (pl.revenue or 0) - ad_v - lb_v - ft - vt
        return {
            'revenue': pl.revenue, 'gross_profit': gp, 'ad_cost': ad_v,
            'labor_cost': lb_v, 'other_fixed': ft, 'other_variable': vt,
            'operating_profit': gp,
            'ad_items': ai,
            'regular_salary': pl.regular_salary or 0, 'parttime_salary': pl.parttime_salary or 0,
            'social_insurance': pl.commission_pay or 0,
            'fixed_items': fi, 'variable_items': vi,
        }

    pm = month - 1 if month > 1 else 12
    py_m = year if month > 1 else year - 1
    prev_month = get_prev_summary(py_m, pm)
    prev_year  = get_prev_summary(year - 1, month)

    return jsonify({
        'year': year, 'month': month,
        'stores': result,
        'current': result[0] if result else None,
        'prev_month': prev_month,
        'prev_year':  prev_year,
        'template_fixed': template_fixed,
        'template_variable': template_variable,
        'template_items': template_items,
    })


@app.route("/api/pl/custom-items")
def api_pl_custom_items():
    """PLカスタム項目テンプレート一覧（type別・テナント分離）"""
    sid = safe_store_id()
    items = PLCustomItem.query.filter_by(store_id=sid).order_by(PLCustomItem.sort_order).all() if sid else []
    return jsonify({
        'fixed':    [{'id': i.id, 'name': i.name} for i in items if (i.item_type or '固定費') == '固定費'],
        'variable': [{'id': i.id, 'name': i.name} for i in items if (i.item_type or '固定費') == '変動費'],
        'other':    [{'id': i.id, 'name': i.name} for i in items if (i.item_type or '固定費') not in ('固定費','変動費')],
    })


@app.route("/api/pl/prev-month-data")
def api_pl_prev_month_data():
    """前月の固定費・PLデータを返す（新規入力時の自動入力用）"""
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]
    allowed_ids = get_allowed_store_ids()
    store_id = request.args.get('store_id', type=int)
    # テナント分離: store_idが未指定または許可外なら先頭の許可店舗を使用
    if not store_id or store_id not in allowed_ids:
        store_id = allowed_ids[0] if allowed_ids else None
    if not store_id:
        return jsonify({})

    # 前月
    pm = month - 1 if month > 1 else 12
    py = year if month > 1 else year - 1

    pl = PLRecord.query.filter_by(store_id=store_id, year=py, month=pm).first()
    custom_vals = PLCustomValue.query.filter_by(store_id=store_id, year=py, month=pm).all()

    fixed_items    = [{'name': v.item_name, 'amount': v.amount} for v in custom_vals if (v.item_type or '固定費') == '固定費']
    variable_items = [{'name': v.item_name, 'amount': v.amount} for v in custom_vals if (v.item_type or '固定費') == '変動費']
    ad_items       = _get_ad_items(store_id, py, pm, pl)

    result = {
        'fixed_items': fixed_items,
        'variable_items': variable_items,
        'ad_items': ad_items,
    }
    if pl:
        result.update({
            'regular_salary':  pl.regular_salary or 0,
            'parttime_salary': pl.parttime_salary or 0,
            'social_insurance': pl.commission_pay or 0,
        })
    return jsonify(result)


@app.route("/api/staff/ranking")
def api_staff_ranking():
    """スタッフランキングデータを返す"""
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]
    metric = request.args.get('metric', 'sales_amount')  # ランキング指標
    store_id = request.args.get('store_id', type=int)

    # 月間 or 累計（年間）の切り替え
    period = request.args.get('period', 'month')  # 'month' or 'year'

    if period == 'year':
        # 年間累計
        query = SalesKPI.query.filter_by(year=year)
    else:
        query = SalesKPI.query.filter_by(year=year, month=month)

    # 店舗フィルタ（テナント分離）
    _allowed = get_allowed_store_ids()
    if store_id and store_id in _allowed:
        query = query.filter_by(store_id=store_id)
    elif _allowed:
        query = query.filter(SalesKPI.store_id.in_(_allowed))
    kpis = query.all()

    # スタッフ別に集計
    staff_totals = {}
    for kpi in kpis:
        sid = kpi.staff_id
        if sid not in staff_totals:
            staff = Staff.query.get(sid)
            store = Store.query.get(kpi.store_id)
            staff_totals[sid] = {
                'staff_id':    sid,
                'staff_name':  staff.name if staff else '不明',
                'role':        staff.role if staff else '',
                'store_name':  store.name if store else '不明',
                'inquiries':   0,
                'store_visits':0,
                'viewings':    0,
                'applications':0,
                'contracts':   0,
                'sales_amount':0,
                'option_sales':0,
            }
        for field in ['inquiries', 'store_visits', 'viewings', 'applications',
                      'contracts', 'sales_amount', 'option_sales']:
            staff_totals[sid][field] += getattr(kpi, field)

    # 指定指標でソート
    ranking = sorted(staff_totals.values(), key=lambda x: x.get(metric, 0), reverse=True)
    # 順位付け
    for rank, item in enumerate(ranking, 1):
        item['rank'] = rank

    return jsonify({'year': year, 'month': month, 'period': period, 'ranking': ranking})


# ── 幹部向け管理ツール：データ入力API ────────────────────

@app.route("/api/kpi/input", methods=["POST"])
@login_required
def api_kpi_input():
    """KPIデータを入力・更新する"""
    import traceback
    try:
        data = request.get_json() or request.form
        staff_id_raw = data.get('staff_id', 0)
        try:
            staff_id = int(staff_id_raw)
        except (TypeError, ValueError):
            staff_id = 0
        store_id = safe_store_id(data.get('store_id'))
        if not store_id:
            return jsonify({'error': 'store not found', 'allowed': get_allowed_store_ids()}), 403
        if not staff_id or staff_id <= 0:
            return jsonify({'error': 'staff_id は必須です', 'got': staff_id_raw}), 400
        staff = db.session.get(Staff, staff_id) if hasattr(db.session, 'get') else Staff.query.get(staff_id)
        if not staff:
            return jsonify({'error': f'スタッフ(id={staff_id})が見つかりません'}), 400
        year  = int(data.get('year',  current_ym()[0]))
        month = int(data.get('month', current_ym()[1]))

        kpi = SalesKPI.query.filter_by(staff_id=staff_id, store_id=store_id,
                                        year=year, month=month).first()
        if not kpi:
            kpi = SalesKPI(staff_id=staff_id, store_id=store_id, year=year, month=month,
                           inquiries=0, store_visits=0, viewings=0, applications=0,
                           contracts=0, cancellations=0, sales_amount=0.0,
                           option_sales=0.0, estimated_sales=0.0, target_sales=0.0,
                           fire_insurance_count=0, lifeline_count=0, moving_count=0)
            db.session.add(kpi)
            db.session.flush()  # idを確定させてからフィールドを設定

        kpi.inquiries    = int(data.get('inquiries',     0) or 0)
        kpi.store_visits = int(data.get('store_visits',  0) or 0)
        kpi.viewings     = int(data.get('viewings',      0) or 0)
        kpi.applications = int(data.get('applications',  0) or 0)
        kpi.contracts    = int(data.get('contracts',     0) or 0)
        kpi.cancellations= int(data.get('cancellations', 0) or 0)
        kpi.sales_amount = float(data.get('sales_amount',  kpi.sales_amount  or 0) or 0)
        kpi.option_sales = float(data.get('option_sales',  kpi.option_sales  or 0) or 0)
        kpi.estimated_sales      = float(data.get('estimated_sales',      kpi.estimated_sales      or 0) or 0)
        kpi.target_sales         = float(data.get('target_sales',         kpi.target_sales         or 0) or 0)
        kpi.fire_insurance_count = int(data.get('fire_insurance_count',   kpi.fire_insurance_count or 0) or 0)
        kpi.lifeline_count       = int(data.get('lifeline_count',         kpi.lifeline_count       or 0) or 0)
        kpi.moving_count         = int(data.get('moving_count',           kpi.moving_count         or 0) or 0)
        db.session.commit()
        return jsonify({'status': 'ok', 'id': kpi.id})
    except Exception as e:
        db.session.rollback()
        tb = traceback.format_exc()
        print(f"api_kpi_input error: {e}\n{tb}")
        return jsonify({'error': str(e), 'detail': tb[-300:]}), 500


def _apply_pl_fields(pl, data):
    """PLレコードに詳細フィールドを適用する"""
    for field in ['revenue', 'gross_profit', 'ad_cost', 'labor_cost', 'other_fixed', 'other_variable',
                  'brokerage_fee', 'ad_income', 'lifeline_income', 'moving_income',
                  'fire_insurance_income', 'other_income',
                  'suumo_cost', 'homes_cost', 'athome_cost', 'instagram_cost', 'tiktok_cost',
                  'google_ads_cost', 'line_cost', 'hp_cost', 'meo_cost', 'other_ad_cost',
                  'regular_salary', 'parttime_salary', 'commission_pay',
                  'pl_rent', 'pl_parking', 'pl_copier', 'pl_internet',
                  'pl_consultant', 'pl_insurance', 'pl_cloud']:
        if field in data:
            setattr(pl, field, float(data.get(field, 0) or 0))
    # social_insurance → commission_payカラムにマッピング
    if 'social_insurance' in data:
        pl.commission_pay = float(data.get('social_insurance', 0) or 0)
    # 広告費合計：ad_itemsがある場合はそちらで計算（api_pl_inputで保存後に再計算）
    if 'ad_items' in data:
        ad_items_list = data.get('ad_items') or []
        if isinstance(ad_items_list, str):
            import json as _j
            try: ad_items_list = _j.loads(ad_items_list)
            except: ad_items_list = []
        ad_total = sum(float(i.get('amount', 0) or 0) for i in ad_items_list)
    else:
        ad_total = sum(float(data.get(k, 0) or 0) for k in [
            'suumo_cost','homes_cost','athome_cost','instagram_cost','tiktok_cost',
            'google_ads_cost','line_cost','hp_cost','meo_cost','other_ad_cost'
        ])
    if ad_total > 0:
        pl.ad_cost = ad_total
    # 人件費合計を自動計算して保存
    labor_total = (float(data.get('regular_salary', 0) or 0) +
                   float(data.get('parttime_salary', 0) or 0) +
                   float(data.get('social_insurance', data.get('commission_pay', 0)) or 0))
    if labor_total > 0:
        pl.labor_cost = labor_total


@app.route("/api/pl/monthly-chart")
@login_required
def api_pl_monthly_chart():
    """経理専用月次グラフ（PLRecordのみ・営業KPIを含まない）"""
    store_id   = request.args.get('store_id', type=int)
    from_param = request.args.get('from')
    to_param   = request.args.get('to')
    today      = date.today()
    allowed_ids = get_allowed_store_ids()
    filter_ids  = [store_id] if (store_id and store_id in allowed_ids) else allowed_ids

    if from_param and to_param:
        try:
            fy, fm = int(from_param[:4]), int(from_param[5:7])
            ty, tm = int(to_param[:4]),   int(to_param[5:7])
            base   = fy * 12 + fm - 1
            end    = ty  * 12 + tm  - 1
            periods = [(t // 12, t % 12 + 1) for t in range(base, end + 1)]
        except Exception:
            periods = None
    else:
        periods = None
    if not periods:
        base = today.year * 12 + today.month - 1
        periods = [(t // 12, t % 12 + 1) for t in range(base - 11, base + 1)]

    result = []
    for y, m in periods:
        pls = PLRecord.query.filter_by(year=y, month=m).filter(PLRecord.store_id.in_(filter_ids)).all()
        total_revenue  = sum(p.revenue or 0 for p in pls)
        total_expenses = 0
        for p in pls:
            cv = PLCustomValue.query.filter_by(store_id=p.store_id, year=y, month=m).all()
            ad_cvs    = [c for c in cv if c.item_type == '広告費']
            fixed_cvs = [c for c in cv if c.item_type == '固定費']
            var_cvs   = [c for c in cv if c.item_type == '変動費']
            ad_t = sum(c.amount for c in ad_cvs) if ad_cvs else (
                (p.ad_cost or 0) or sum(getattr(p, col, 0) or 0 for col in [
                    'suumo_cost','homes_cost','athome_cost','instagram_cost','tiktok_cost',
                    'google_ads_cost','line_cost','hp_cost','meo_cost','other_ad_cost']))
            lb_t = (p.labor_cost or 0) or ((p.regular_salary or 0)+(p.parttime_salary or 0)+(p.commission_pay or 0))
            ft   = sum(c.amount for c in fixed_cvs)
            vt   = sum(c.amount for c in var_cvs)
            total_expenses += ad_t + lb_t + ft + vt
        result.append({
            'label':    f'{y}/{m:02d}',
            'year':     y,
            'month':    m,
            'revenue':  total_revenue,
            'expenses': total_expenses,
            'profit':   total_revenue - total_expenses,
        })
    return jsonify(result)


@app.route("/api/pl/input", methods=["POST"])
def api_pl_input():
    """PLデータを入力・更新する"""
    data = request.get_json() or request.form
    store_id = safe_store_id(data.get('store_id'))
    if not store_id:
        return jsonify({'error': 'unauthorized'}), 403
    year     = int(data.get('year', current_ym()[0]))
    month    = int(data.get('month', current_ym()[1]))

    pl = PLRecord.query.filter_by(store_id=store_id, year=year, month=month).first()
    if not pl:
        pl = PLRecord(store_id=store_id, year=year, month=month)
        db.session.add(pl)

    _apply_pl_fields(pl, data)
    db.session.commit()

    # 固定費・変動費・カスタム項目を保存（type付き）
    def _save_typed_items(items_json_key, item_type):
        raw = None
        if request.is_json:
            raw = request.get_json().get(items_json_key)
        if raw is None:
            raw = data.get(items_json_key)
        if raw is None: return
        if isinstance(raw, str):
            import json as _json
            try: raw = _json.loads(raw)
            except: return
        # 既存エントリを全削除してから再挿入（ゴースト行防止）
        PLCustomValue.query.filter_by(store_id=store_id, year=year, month=month, item_type=item_type).delete()
        for item in raw:
            name   = str(item.get('name', '')).strip()
            amount = float(item.get('amount', 0) or 0)
            if not name: continue
            cv = PLCustomValue(store_id=store_id, year=year, month=month, item_name=name, item_type=item_type, amount=amount)
            db.session.add(cv)
            # テンプレート登録
            existing = PLCustomItem.query.filter_by(store_id=store_id, name=name, item_type=item_type).first()
            if not existing:
                max_order = db.session.query(db.func.max(PLCustomItem.sort_order)).filter_by(store_id=store_id).scalar() or 0
                db.session.add(PLCustomItem(store_id=store_id, name=name, item_type=item_type, sort_order=max_order + 1))

    _save_typed_items('fixed_items', '固定費')
    _save_typed_items('variable_items', '変動費')
    _save_typed_items('custom_items', 'その他')

    # 広告費カスタム行（送信されている場合は全置換）
    body = request.get_json() if request.is_json else {}
    if 'ad_items' in body:
        PLCustomValue.query.filter_by(
            store_id=store_id, year=year, month=month, item_type='広告費'
        ).delete()
        for item in (body['ad_items'] or []):
            name   = str(item.get('name', '')).strip()
            amount = float(item.get('amount', 0) or 0)
            if not name: continue
            cv = PLCustomValue(store_id=store_id, year=year, month=month,
                               item_name=name, item_type='広告費', amount=amount)
            db.session.add(cv)
        # 固定列をゼロクリア（重複カウント防止）
        for col in ['suumo_cost','homes_cost','athome_cost','instagram_cost','tiktok_cost',
                    'google_ads_cost','line_cost','hp_cost','meo_cost','other_ad_cost']:
            setattr(pl, col, 0)

    db.session.commit()

    return jsonify({'status': 'ok', 'id': pl.id})


@app.route("/api/ad/input", methods=["POST"])
def api_ad_input():
    """広告費を入力・更新する"""
    data = request.get_json() or request.form
    store_id = safe_store_id(data.get('store_id'))
    if not store_id:
        return jsonify({'error': 'unauthorized'}), 403
    source   = data.get('source', '')
    year     = int(data.get('year', current_ym()[0]))
    month    = int(data.get('month', current_ym()[1]))
    cost     = float(data.get('cost', 0))

    ac = AdCost.query.filter_by(store_id=store_id, source=source, year=year, month=month).first()
    if not ac:
        ac = AdCost(store_id=store_id, source=source, year=year, month=month)
        db.session.add(ac)
    ac.cost = cost
    db.session.commit()

    return jsonify({'status': 'ok', 'id': ac.id})


@app.route("/api/store/add", methods=["POST"])
@login_required
def api_store_add():
    """店舗を追加する（ログインユーザーのテナントに所属）"""
    uid = session.get('app_user_id')
    cur_user = AppUser.query.get(uid) if uid else None
    data = request.get_json() or request.form
    store = Store(
        name=data.get('name', ''),
        tenant_id=cur_user.tenant_id if cur_user else None,
        rent=float(data.get('rent', 0)),
        parking_fee=float(data.get('parking_fee', 0)),
        copier_fee=float(data.get('copier_fee', 0)),
        internet_fee=float(data.get('internet_fee', 0)),
        consultant_fee=float(data.get('consultant_fee', 0)),
        insurance_fee=float(data.get('insurance_fee', 0)),
        cloud_fee=float(data.get('cloud_fee', 0)),
        is_active=True,
    )
    db.session.add(store)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': store.id, 'name': store.name})


@app.route("/api/staff", methods=["GET"])
def api_staff_list():
    """スタッフ一覧（テナント分離）"""
    allowed_ids = get_allowed_store_ids()
    staff_list = Staff.query.filter(Staff.store_id.in_(allowed_ids), Staff.is_active == True).all()
    result = [{'id': s.id, 'name': s.name, 'role': s.role} for s in staff_list]
    return jsonify(result)


@app.route("/api/kpi/<int:kpi_id>", methods=["DELETE"])
def api_kpi_delete(kpi_id):
    """KPI削除"""
    kpi = SalesKPI.query.get_or_404(kpi_id)
    db.session.delete(kpi)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/kpi/<int:kpi_id>", methods=["PUT"])
def api_kpi_update(kpi_id):
    """KPI更新"""
    kpi = SalesKPI.query.get_or_404(kpi_id)
    data = request.get_json() or request.form
    kpi.inquiries     = int(data.get('inquiries',     kpi.inquiries))
    kpi.store_visits  = int(data.get('store_visits',  kpi.store_visits))
    kpi.viewings      = int(data.get('viewings',      kpi.viewings))
    kpi.applications  = int(data.get('applications',  kpi.applications))
    kpi.contracts     = int(data.get('contracts',     kpi.contracts))
    kpi.cancellations = int(data.get('cancellations', kpi.cancellations))
    kpi.sales_amount  = float(data.get('sales_amount', kpi.sales_amount))
    kpi.option_sales  = float(data.get('option_sales', kpi.option_sales))
    kpi.estimated_sales     = float(data.get('estimated_sales', kpi.estimated_sales or 0))
    kpi.fire_insurance_count= int(data.get('fire_insurance_count', kpi.fire_insurance_count or 0))
    kpi.lifeline_count      = int(data.get('lifeline_count', kpi.lifeline_count or 0))
    kpi.moving_count        = int(data.get('moving_count', kpi.moving_count or 0))
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/lead/<int:lead_id>", methods=["DELETE"])
def api_lead_delete(lead_id):
    """リード削除"""
    lead = Lead.query.get_or_404(lead_id)
    db.session.delete(lead)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/lead/<int:lead_id>", methods=["PUT"])
def api_lead_update(lead_id):
    """リード更新"""
    lead = Lead.query.get_or_404(lead_id)
    data = request.get_json() or request.form
    if 'source' in data or 'media' in data:
        lead.source = data.get('source') or data.get('media', lead.source)
    if 'status' in data:
        lead.status = data.get('status', lead.status)
    if 'customer_name' in data:
        lead.customer_name = data.get('customer_name', lead.customer_name)
    if 'assigned_staff_id' in data or 'assignee_id' in data:
        lead.assigned_staff_id = data.get('assigned_staff_id') or data.get('assignee_id') or lead.assigned_staff_id
    if 'note' in data or 'memo' in data:
        lead.note = data.get('note') or data.get('memo', lead.note)
    if 'line_added' in data:
        lead.line_added = str(data.get('line_added', '0')) in ('1', 'true', 'True', 'on')
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/pl/<int:pl_id>", methods=["DELETE"])
def api_pl_delete(pl_id):
    """PL削除"""
    pl = PLRecord.query.get_or_404(pl_id)
    PLCustomValue.query.filter_by(store_id=pl.store_id, year=pl.year, month=pl.month).delete()
    db.session.delete(pl)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/pl/<int:pl_id>", methods=["PUT"])
def api_pl_update(pl_id):
    """PL更新"""
    pl = PLRecord.query.get_or_404(pl_id)
    data = request.get_json() or request.form
    _apply_pl_fields(pl, data)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/staff/<int:staff_id>", methods=["DELETE"])
def api_staff_delete(staff_id):
    """スタッフ削除（論理削除）"""
    staff = Staff.query.get_or_404(staff_id)
    staff.is_active = False
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/staff/<int:staff_id>", methods=["PUT"])
def api_staff_update(staff_id):
    """スタッフ更新"""
    staff = Staff.query.get_or_404(staff_id)
    data = request.get_json() or request.form
    if 'name' in data:
        staff.name = data.get('name', staff.name)
    if 'role' in data:
        staff.role = data.get('role', staff.role)
    if 'store_id' in data:
        staff.store_id = int(data.get('store_id')) or staff.store_id
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/staff/add", methods=["POST"])
def api_staff_add():
    """スタッフを追加する"""
    data = request.get_json() or request.form
    hired_str = data.get('hired_at', '')
    hired_date = None
    if hired_str:
        try:
            hired_date = date.fromisoformat(hired_str)
        except ValueError:
            pass

    staff = Staff(
        name=data.get('name', ''),
        store_id=int(data.get('store_id', 0)) or None,
        role=data.get('role', '営業'),
        is_active=True,
        hired_at=hired_date,
    )
    db.session.add(staff)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': staff.id, 'name': staff.name})


# ── Excelインポート ──────────────────────────────────────

@app.route("/import-excel")
@login_required
def import_excel_page():
    """Excelインポートページ"""
    # インポート履歴（source_fileでグループ化）
    from sqlalchemy import func
    history = (db.session.query(
        ContractRecord.source_file,
        ContractRecord.year,
        ContractRecord.month,
        func.count(ContractRecord.id).label('count'),
        func.max(ContractRecord.imported_at).label('imported_at'),
    )
    .filter(ContractRecord.source_file != None)
    .group_by(ContractRecord.source_file, ContractRecord.year, ContractRecord.month)
    .order_by(func.max(ContractRecord.imported_at).desc())
    .limit(20)
    .all())

    year, month = current_ym()
    return render_template("import_excel.html",
                           history=history, year=year, month=month,
                           now=datetime.now())


@app.route("/api/import/excel", methods=["POST"])
def api_import_excel():
    """ExcelファイルをインポートしてJSONで結果を返す"""
    if 'file' not in request.files:
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'ファイル名が空です'}), 400

    year = request.form.get('year', type=int) or current_ym()[0]
    month = request.form.get('month', type=int) or current_ym()[1]
    store_id = safe_store_id(request.form.get('store_id', type=int))

    filename = secure_filename(f.filename)

    # 一時ファイルに保存してから処理
    suffix = os.path.splitext(filename)[1] or '.xlsx'
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        result = import_excel_file(tmp_path, year, month, store_id)
        # ソースファイル名を記録
        with app.app_context():
            ContractRecord.query.filter_by(year=year, month=month).filter(
                ContractRecord.source_file == None
            ).update({'source_file': filename})
            db.session.commit()
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    result['filename'] = filename
    result['year'] = year
    result['month'] = month
    return jsonify(result)


# ── 申込台帳 ─────────────────────────────────────────────

@app.route("/contracts")
@login_required
def contracts():
    """申込台帳ページ"""
    _active = get_allowed_store_ids()
    staff_list = Staff.query.filter(Staff.store_id.in_(_active), Staff.is_active == True).all() if _active else []
    year, month = current_ym()
    return render_template("contracts.html",
                           staff_list=staff_list, year=year, month=month,
                           now=datetime.now())


@app.route("/api/contracts")
def api_contracts():
    """ContractRecord一覧JSON (?year=&month=&staff_id=&status=)"""
    year = request.args.get('year', type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]
    staff_id = request.args.get('staff_id', type=int)
    status = request.args.get('status', '')

    query = ContractRecord.query.filter_by(year=year, month=month)
    if staff_id:
        query = query.filter_by(staff_id=staff_id)
    if status:
        query = query.filter_by(status=status)

    records = query.order_by(ContractRecord.seq_no).all()

    result = []
    for rc in records:
        result.append({
            'id': rc.id,
            'seq_no': rc.seq_no,
            'status': rc.status,
            'staff_name': rc.staff_name_raw,
            'application_date': rc.application_date.isoformat() if rc.application_date else None,
            'property_name': rc.property_name,
            'room_no': rc.room_no,
            'customer_name': rc.customer_name,
            'phone': rc.phone,
            'rent': rc.rent,
            'media': rc.media,
            'management_company': rc.management_company,
            'ad_pct': rc.ad_pct,
            'ad_received': rc.ad_received,
            'lifeline': rc.lifeline,
            'moving': rc.moving,
            'fire_insurance': rc.fire_insurance,
            'contract_amount': rc.contract_amount,
            'cancel_type': rc.cancel_type,
            'review_status': rc.review_status,
            'settlement_date': rc.settlement_date.isoformat() if rc.settlement_date else None,
            'ad_income_date': rc.ad_income_date.isoformat() if rc.ad_income_date else None,
            'ad_income_date_raw': rc.ad_income_date_raw,
        })

    return jsonify({
        'year': year,
        'month': month,
        'records': result,
        'total': len(result),
        'total_contract_amount': sum(r['contract_amount'] or 0 for r in result),
        'contracts': sum(1 for r in result if r['status'] == '契約'),
        'applications': sum(1 for r in result if r['status'] == '申込'),
        'cancellations': sum(1 for r in result if r['status'] == 'キャンセル'),
    })


@app.route("/api/contracts/<int:contract_id>", methods=["DELETE"])
def api_contract_delete(contract_id):
    """ContractRecord削除"""
    rc = ContractRecord.query.get_or_404(contract_id)
    db.session.delete(rc)
    db.session.commit()
    return jsonify({'status': 'ok'})


# ── 入金管理 ─────────────────────────────────────────────

@app.route("/payments")
@login_required
def payments():
    """入金管理ページ"""
    _allowed = get_allowed_store_ids()
    staff_list = Staff.query.filter(Staff.store_id.in_(_allowed), Staff.is_active == True).all() if _allowed else []
    year, month = current_ym()
    return render_template("payments.html",
                           staff_list=staff_list, year=year, month=month,
                           now=datetime.now())


@app.route("/api/payments/summary")
def api_payments_summary():
    """月次入金サマリーJSON (?year=&month=)"""
    year = request.args.get('year', type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]

    contracts = ContractRecord.query.filter_by(year=year, month=month, status='契約').all()

    # 仲介手数料予定額（賃料 × 手数料% または賃料1ヶ月分）
    commission_total = 0
    ad_income_total = 0
    ad_received_total = 0
    ad_unreceived_total = 0
    option_total = 0

    for rc in contracts:
        rent = rc.rent or 0
        commission_pct = rc.commission_pct or 100
        ad_pct = rc.ad_pct or 0

        commission = rent * (commission_pct / 100)
        commission_total += commission

        ad_amount = rent * (ad_pct / 100)
        ad_income_total += ad_amount

        if rc.ad_received and '○' in rc.ad_received:
            ad_received_total += ad_amount
        else:
            ad_unreceived_total += ad_amount

        if rc.lifeline and '○' in rc.lifeline:
            option_total += 5000
        if rc.moving and '○' in rc.moving:
            option_total += 3000

    # 入金状況詳細リスト
    details = []
    for rc in contracts:
        rent = rc.rent or 0
        ad_pct = rc.ad_pct or 0
        commission_pct = rc.commission_pct or 100
        ad_amount = rent * (ad_pct / 100)
        commission = rent * (commission_pct / 100)
        details.append({
            'id': rc.id,
            'customer_name': rc.customer_name,
            'property_name': rc.property_name,
            'room_no': rc.room_no,
            'staff_name': rc.staff_name_raw,
            'rent': rent,
            'ad_pct': ad_pct,
            'ad_amount': round(ad_amount),
            'ad_received': rc.ad_received,
            'ad_income_date': rc.ad_income_date.isoformat() if rc.ad_income_date else None,
            'ad_income_date_raw': rc.ad_income_date_raw,
            'commission': round(commission),
            'settlement_date': rc.settlement_date.isoformat() if rc.settlement_date else None,
            'contract_amount': rc.contract_amount or 0,
        })

    # 月別入金推移（過去6ヶ月）
    monthly_trend = []
    for i in range(5, -1, -1):
        t = date.today().replace(day=1)
        for _ in range(i):
            t = (t - timedelta(days=1)).replace(day=1)
        y, m = t.year, t.month
        recs = ContractRecord.query.filter_by(year=y, month=m, status='契約').all()
        received = 0
        unreceived = 0
        for rc in recs:
            rent = rc.rent or 0
            ad_pct_v = rc.ad_pct or 0
            ad_amt = rent * (ad_pct_v / 100)
            if rc.ad_received and '○' in rc.ad_received:
                received += ad_amt
            else:
                unreceived += ad_amt
        monthly_trend.append({
            'label': f'{y}/{m:02d}',
            'received': round(received),
            'unreceived': round(unreceived),
        })

    return jsonify({
        'year': year,
        'month': month,
        'commission_total': round(commission_total),
        'ad_income_total': round(ad_income_total),
        'ad_received_total': round(ad_received_total),
        'ad_unreceived_total': round(ad_unreceived_total),
        'option_total': round(option_total),
        'contract_count': len(contracts),
        'details': details,
        'monthly_trend': monthly_trend,
    })


# ── 設定ページ ───────────────────────────────────────────

def _render_settings(mode):
    stores     = get_allowed_stores(ignore_active=True)
    active_store_ids = get_allowed_store_ids(ignore_active=False)
    staff_list = Staff.query.filter(Staff.store_id.in_(active_store_ids), Staff.is_active == True).order_by(Staff.name).all() if active_store_ids else []
    store_name_map = {s.id: s.name for s in stores}
    user = AppUser.query.get(session['app_user_id'])
    is_owner = user and user.role == 'owner'
    is_manager = user and user.role == 'store_manager'
    if (is_owner or is_manager) and user.tenant_id:
        accounts = AppUser.query.filter_by(is_active=True, tenant_id=user.tenant_id).all()
    else:
        accounts = []
    titles = {'staff': 'スタッフ管理', 'accounts': 'ログイン管理',
              'company': '会社情報', 'profile': 'ID/パス管理'}
    return render_template("settings.html",
                           stores=stores, staff_list=staff_list, accounts=accounts,
                           store_name_map=store_name_map,
                           is_owner=is_owner, is_manager=is_manager,
                           current_user=user, mode=mode,
                           page_title=titles.get(mode, '設定'),
                           now=datetime.now())


@app.route("/settings")
@login_required
def settings():
    user = AppUser.query.get(session.get('app_user_id'))
    if user and user.role in ('owner', 'store_manager'):
        return redirect(url_for('settings_staff'))
    return redirect(url_for('settings_profile'))


@app.route("/settings/staff")
@login_required
def settings_staff():
    return _render_settings('staff')


@app.route("/settings/accounts")
@login_required
def settings_accounts():
    return _render_settings('accounts')


@app.route("/settings/company")
@login_required
def settings_company():
    return _render_settings('company')


@app.route("/settings/profile")
@login_required
def settings_profile():
    return _render_settings('profile')


@app.route("/api/company-logo", methods=["GET"])
@login_required
def api_company_logo_get():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    cp = CompanyProfile.query.filter_by(store_id=sid).first() if sid else None
    if not cp or not cp.logo_data:
        return "", 404
    from flask import Response as _Resp
    return _Resp(cp.logo_data, mimetype=cp.logo_type or 'image/png')


@app.route("/api/company-logo", methods=["POST"])
@login_required
def api_company_logo_upload():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    if not sid:
        return jsonify({'error': 'unauthorized'}), 403
    f = request.files.get('logo')
    if not f or not f.filename:
        return jsonify({'error': 'ファイルを選択してください'}), 400
    raw = f.read()
    if len(raw) > 5 * 1024 * 1024:
        return jsonify({'error': 'ロゴ画像は5MBまでです'}), 400
    cp = CompanyProfile.query.filter_by(store_id=sid).first()
    if not cp:
        cp = CompanyProfile(store_id=sid)
        db.session.add(cp)
    cp.logo_data = raw
    cp.logo_type = (f.mimetype or 'image/png')[:80]
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/company-logo", methods=["DELETE"])
@login_required
def api_company_logo_delete():
    allowed = get_allowed_store_ids()
    sid = allowed[0] if allowed else None
    cp = CompanyProfile.query.filter_by(store_id=sid).first() if sid else None
    if cp:
        cp.logo_data = None
        cp.logo_type = None
        db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/settings/staff/add", methods=["POST"])
@login_required
def api_settings_staff_add():
    """スタッフ追加"""
    data = request.get_json() or request.form
    # ignore_active=True で設定ページと同じ基準で店舗を解決する
    allowed = get_allowed_store_ids()
    if not allowed:
        return jsonify({'error': 'store not found'}), 403
    try:
        req_sid = int(data.get('store_id') or 0)
    except (TypeError, ValueError):
        req_sid = 0
    sid = req_sid if req_sid and req_sid in allowed else allowed[0]

    hired_str = data.get('hired_at', '')
    hired_date = None
    if hired_str:
        try:
            hired_date = date.fromisoformat(hired_str)
        except ValueError:
            pass
    staff = Staff(
        name=data.get('name', ''),
        store_id=sid,
        role=data.get('role', '営業'),
        is_active=True,
        hired_at=hired_date,
    )
    db.session.add(staff)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': staff.id, 'name': staff.name})


@app.route("/api/settings/staff/<int:staff_id>", methods=["PUT"])
@login_required
def api_settings_staff_update(staff_id):
    """スタッフ更新"""
    staff = Staff.query.get_or_404(staff_id)
    data = request.get_json() or request.form
    if 'name' in data:
        staff.name = data.get('name')
    if 'role' in data:
        staff.role = data.get('role')
    if 'hired_at' in data:
        try:
            staff.hired_at = date.fromisoformat(data['hired_at']) if data['hired_at'] else None
        except ValueError:
            pass
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/settings/staff/<int:staff_id>", methods=["DELETE"])
@login_required
def api_settings_staff_delete(staff_id):
    """スタッフ削除（論理削除）"""
    staff = Staff.query.get_or_404(staff_id)
    staff.is_active = False
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/settings/account/add", methods=["POST"])
@owner_or_manager_required
def api_settings_account_add():
    """アカウント追加（オーナーの自テナントに所属させる）"""
    data = request.get_json() or request.form
    cur = AppUser.query.get(session['app_user_id'])
    if not cur or not cur.tenant_id:
        return jsonify({'status': 'error', 'message': 'テナント情報が取得できません'}), 403
    if AppUser.query.filter_by(username=data.get('username', ''), is_active=True).first():
        return jsonify({'status': 'error', 'message': 'そのユーザー名は既に使用されています'}), 400
    # store_id: リクエストで指定されていればそれを使い、なければオーナーの最初の店舗
    sid = safe_store_id(data.get('store_id'))
    user = AppUser(
        tenant_id=cur.tenant_id,
        username=data.get('username', ''),
        password_hash=generate_password_hash(data.get('password', '')),
        role=data.get('role', 'staff'),
        staff_id=int(data.get('staff_id')) if data.get('staff_id') else None,
        store_id=sid,
        is_active=True,
        can_view_accounting=bool(data.get('can_view_accounting', True)),
        can_view_all_staff=bool(data.get('can_view_all_staff', True)),
        can_edit_kpi=bool(data.get('can_edit_kpi', True)),
        can_manage_uncollected=bool(data.get('can_manage_uncollected', True)),
    )
    db.session.add(user)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': user.id})


@app.route("/api/settings/account/<int:account_id>/password", methods=["PUT"])
@owner_or_manager_required
def api_settings_account_password(account_id):
    """パスワード変更（オーナーが他ユーザーのパスワードを変更）"""
    user = AppUser.query.get_or_404(account_id)
    data = request.get_json() or request.form
    new_password = data.get('password', '')
    if not new_password:
        return jsonify({'status': 'error', 'message': 'パスワードを入力してください'}), 400
    user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/settings/account/<int:account_id>", methods=["DELETE"])
@owner_or_manager_required
def api_settings_account_delete(account_id):
    """アカウント削除"""
    user = AppUser.query.get_or_404(account_id)
    if user.username == 'owner':
        return jsonify({'status': 'error', 'message': 'オーナーアカウントは削除できません'}), 400
    db.session.delete(user)   # 完全削除（ログなし）
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/settings/password", methods=["POST"])
@login_required
def api_settings_my_password():
    """自分のパスワード変更"""
    data = request.get_json() or request.form
    user = AppUser.query.get(session['app_user_id'])
    if not user:
        return jsonify({'status': 'error', 'message': 'ユーザーが見つかりません'}), 404
    current_password = data.get('current_password', '')
    new_password = data.get('new_password', '')
    if not check_password_hash(user.password_hash, current_password):
        return jsonify({'status': 'error', 'message': '現在のパスワードが正しくありません'}), 400
    if not new_password:
        return jsonify({'status': 'error', 'message': '新パスワードを入力してください'}), 400
    user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/settings/profile", methods=["PUT"])
@login_required
def api_settings_profile():
    """ログイン中ユーザーの表示名（username）変更"""
    user = AppUser.query.get_or_404(session['user_id'])
    data = request.get_json() or request.form
    new_name = (data.get('display_name') or '').strip()
    if not new_name:
        return jsonify({'status': 'error', 'message': '名前を入力してください'}), 400
    if AppUser.query.filter(AppUser.username == new_name, AppUser.id != user.id).first():
        return jsonify({'status': 'error', 'message': 'その名前は既に使用されています'}), 400
    user.username = new_name
    db.session.commit()
    return jsonify({'status': 'ok'})


# ── スタッフ別6ヶ月推移 API ──────────────────────────────────────
def _staff_real_stats(filter_ids, staff_id, y, m):
    """指定スタッフ・年月の実績を各管理表から集計（反響/接客/申込/売上）"""
    import calendar as _cal
    m_start = date(y, m, 1)
    m_end   = date(y, m, _cal.monthrange(y, m)[1])
    if not filter_ids:
        return {'inquiries': 0, 'visits': 0, 'applications': 0, 'revenue': 0}

    eq = EchoRecord.query.filter(EchoRecord.store_id.in_(filter_ids),
                                 EchoRecord.echo_date >= m_start, EchoRecord.echo_date <= m_end)
    cq = CustomerServiceRecord.query.filter(CustomerServiceRecord.store_id.in_(filter_ids),
                                            CustomerServiceRecord.service_date >= m_start,
                                            CustomerServiceRecord.service_date <= m_end)
    aq = ApplicationRecord.query.filter(ApplicationRecord.store_id.in_(filter_ids),
                                        ~ApplicationRecord.status.in_(['キャンセル', 'キャンセル振替']),
                                        db.extract('year',  ApplicationRecord.application_date) == y,
                                        db.extract('month', ApplicationRecord.application_date) == m)
    if staff_id:
        eq = eq.filter(EchoRecord.staff_id == staff_id)
        cq = cq.filter(CustomerServiceRecord.staff_id == staff_id)
        aq = aq.filter(ApplicationRecord.staff_id == staff_id)
    apps = aq.all()
    return {
        'inquiries':    eq.count(),
        'visits':       cq.count(),
        'applications': len(apps),
        'revenue':      sum(_record_approved_amount(a) for a in apps),
    }


@app.route("/api/kpi/staff-history")
def api_kpi_staff_history():
    """スタッフ別の6ヶ月推移データを返す"""
    staff_id = request.args.get('staff_id', type=int)
    store_id = request.args.get('store_id', type=int)
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]

    # 店舗フィルタ（テナント分離）
    allowed_ids = get_allowed_store_ids()
    filter_ids = [store_id] if (store_id and store_id in allowed_ids) else allowed_ids

    history = []
    yoy_data = {}
    for i in range(5, -1, -1):
        m = month - i
        y = year
        while m < 1:
            m += 12; y -= 1
        query = SalesKPI.query.filter_by(year=y, month=m)
        if filter_ids:
            query = query.filter(SalesKPI.store_id.in_(filter_ids))
        if staff_id:
            query = query.filter_by(staff_id=staff_id)
        kpis = query.all()
        history.append({
            'label':        f'{y}/{m:02d}',
            'contracts':    sum(k.contracts    or 0 for k in kpis),
            'applications': sum(k.applications or 0 for k in kpis),
            'revenue':      sum(k.sales_amount or 0 for k in kpis),
            'inquiries':    sum(k.inquiries    or 0 for k in kpis),
        })
        # 現在月のみ前年同月比を計算（実データから：反響=反響管理表/接客=接客管理表/申込・売上=顧客管理表）
        if i == 0:
            cur  = _staff_real_stats(filter_ids, staff_id, y, m)
            prev = _staff_real_stats(filter_ids, staff_id, y - 1, m)
            yoy_data = {
                'inquiries':    {'cur': cur['inquiries'],    'yoy': prev['inquiries']},
                'visits':       {'cur': cur['visits'],       'yoy': prev['visits']},
                'applications': {'cur': cur['applications'], 'yoy': prev['applications']},
                'revenue':      {'cur': cur['revenue'],      'yoy': prev['revenue']},
            }

    return jsonify({'history': history, 'yoy': yoy_data})


# ── 未入金（AD）管理 API ──────────────────────────────────────────
@app.route("/api/uncollected")
@login_required
def api_uncollected_list():
    store_id = request.args.get('store_id', type=int)
    staff_id = request.args.get('staff_id', type=int)
    include_paid = request.args.get('include_paid', 'false').lower() == 'true'

    q = UncollectedPayment.query
    if store_id:
        q = q.filter_by(store_id=store_id)
    if staff_id:
        q = q.filter_by(staff_id=staff_id)
    if not include_paid:
        q = q.filter_by(is_paid=False)

    items = q.order_by(UncollectedPayment.expected_payment_date.asc()).all()

    def fmt_date(d): return d.strftime('%Y-%m-%d') if d else None

    result = []
    for p in items:
        staff = Staff.query.get(p.staff_id) if p.staff_id else None
        result.append({
            'id':                   p.id,
            'store_id':             p.store_id,
            'staff_id':             p.staff_id,
            'staff_name':           staff.name if staff else '',
            'property_name':        p.property_name or '',
            'room_number':          p.room_number or '',
            'application_date':     fmt_date(p.application_date),
            'management_company':   p.management_company or '',
            'customer_name':        p.customer_name or '',
            'expected_payment_date':fmt_date(p.expected_payment_date),
            'amount':               p.amount or 0,
            'memo':                 p.memo or '',
            'is_paid':              p.is_paid,
        })
    return jsonify(result)


@app.route("/api/uncollected/paid-sum")
@login_required
def api_uncollected_paid_sum():
    """指定年月に入金済みとなった未入金データの合計額を返す"""
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]
    allowed = get_allowed_store_ids(ignore_active=True)

    from datetime import date as _date
    month_start = _date(year, month, 1)
    if month == 12:
        month_end = _date(year + 1, 1, 1)
    else:
        month_end = _date(year, month + 1, 1)

    q = UncollectedPayment.query.filter(
        UncollectedPayment.is_paid == True,
        UncollectedPayment.expected_payment_date >= month_start,
        UncollectedPayment.expected_payment_date <  month_end,
    )
    if allowed:
        q = q.filter(UncollectedPayment.store_id.in_(allowed))
    total = sum((p.amount or 0) for p in q.all())
    return jsonify({'total': total, 'year': year, 'month': month})


@app.route("/api/uncollected", methods=["POST"])
@login_required
def api_uncollected_add():
    data = request.get_json() or request.form
    from datetime import datetime as _dt
    def parse_date(s):
        if not s: return None
        try: return _dt.strptime(s, '%Y-%m-%d').date()
        except: return None

    p = UncollectedPayment(
        store_id=safe_store_id(data.get('store_id')) or (get_allowed_store_ids() or [1])[0],
        staff_id=int(data.get('staff_id') or 0) or None,
        property_name=data.get('property_name', ''),
        room_number=data.get('room_number', ''),
        application_date=parse_date(data.get('application_date')),
        management_company=data.get('management_company', ''),
        customer_name=data.get('customer_name', ''),
        expected_payment_date=parse_date(data.get('expected_payment_date')),
        amount=float(data.get('amount', 0) or 0),
        memo=data.get('memo', ''),
    )
    db.session.add(p)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': p.id})


@app.route("/api/uncollected/<int:pid>", methods=["PUT"])
@login_required
def api_uncollected_update(pid):
    p = UncollectedPayment.query.get_or_404(pid)
    data = request.get_json() or request.form
    from datetime import datetime as _dt
    def parse_date(s):
        if not s: return None
        try: return _dt.strptime(s, '%Y-%m-%d').date()
        except: return None

    if 'property_name'      in data: p.property_name      = data['property_name']
    if 'room_number'        in data: p.room_number        = data['room_number']
    if 'application_date'   in data: p.application_date   = parse_date(data['application_date'])
    if 'management_company' in data: p.management_company = data['management_company']
    if 'customer_name'      in data: p.customer_name      = data['customer_name']
    if 'expected_payment_date' in data: p.expected_payment_date = parse_date(data['expected_payment_date'])
    if 'amount'  in data: p.amount   = float(data['amount'] or 0)
    if 'memo'    in data: p.memo     = data['memo']
    if 'staff_id'in data: p.staff_id = int(data['staff_id'] or 0) or None
    if 'is_paid' in data: p.is_paid  = bool(data['is_paid'])
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/uncollected/<int:pid>", methods=["DELETE"])
@login_required
def api_uncollected_delete(pid):
    p = UncollectedPayment.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/uncollected/<int:pid>/paid", methods=["POST"])
@login_required
def api_uncollected_paid(pid):
    p = UncollectedPayment.query.get_or_404(pid)
    if not p.is_paid:
        p.is_paid = True
        # 対応スタッフのSalesKPIに売上を加算
        if p.staff_id and p.amount:
            data = request.get_json() or {}
            req_year  = data.get('year')
            req_month = data.get('month')
            if req_year and req_month:
                ref_year, ref_month = int(req_year), int(req_month)
            else:
                ref_date = p.application_date or p.expected_payment_date or date.today()
                ref_year, ref_month = ref_date.year, ref_date.month
            kpi = SalesKPI.query.filter_by(
                staff_id=p.staff_id, store_id=p.store_id,
                year=ref_year, month=ref_month
            ).first()
            if not kpi:
                kpi = SalesKPI(staff_id=p.staff_id, store_id=p.store_id,
                               year=ref_year, month=ref_month)
                db.session.add(kpi)
            kpi.sales_amount = (kpi.sales_amount or 0) + float(p.amount)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/uncollected/sync-from-applications", methods=["POST"])
@login_required
def api_uncollected_sync_from_applications():
    """前月の申込一覧で未承認の案件を未入金一覧に自動転記（重複スキップ）"""
    data = request.get_json() or {}
    year  = data.get('year')
    month = data.get('month')
    store_id_param = data.get('store_id')

    if not year or not month:
        return jsonify({'added': 0})

    year, month = int(year), int(month)
    allowed_ids = get_allowed_store_ids()

    if store_id_param and int(store_id_param) in allowed_ids:
        filter_ids = [int(store_id_param)]
    else:
        filter_ids = allowed_ids

    if not filter_ids:
        return jsonify({'added': 0})

    # 対象月の申込で未承認のもの（キャンセル除く）
    from sqlalchemy import extract
    apps = ApplicationRecord.query.filter(
        ApplicationRecord.store_id.in_(filter_ids),
        extract('year',  ApplicationRecord.application_date) == year,
        extract('month', ApplicationRecord.application_date) == month,
        ~ApplicationRecord.status.in_(['キャンセル', 'キャンセル振替']),
        db.or_(
            db.and_(ApplicationRecord.ad_amount != 0,       ApplicationRecord.ad_approved == False),
            db.and_(ApplicationRecord.brokerage_fee != 0,   ApplicationRecord.brokerage_approved == False),
        )
    ).all()

    added = 0
    for rec in apps:
        pending_amount = 0
        if (rec.ad_amount or 0) != 0 and not rec.ad_approved:
            pending_amount += rec.ad_amount or 0
        if (rec.brokerage_fee or 0) != 0 and not rec.brokerage_approved:
            pending_amount += rec.brokerage_fee or 0
        pending_amount += rec.option_amount or 0
        if pending_amount <= 0:
            continue

        # 同一案件の重複チェック（物件名＋顧客名＋申込日＋店舗）
        exists = UncollectedPayment.query.filter_by(
            store_id=rec.store_id,
            property_name=rec.property_name,
            customer_name=rec.customer_name,
            application_date=rec.application_date,
        ).first()
        if exists:
            continue

        up = UncollectedPayment(
            store_id=rec.store_id,
            staff_id=rec.staff_id,
            property_name=rec.property_name,
            room_number=rec.room_number,
            application_date=rec.application_date,
            customer_name=rec.customer_name,
            amount=pending_amount,
            memo='申込一覧から自動転記',
            is_paid=False,
        )
        db.session.add(up)
        added += 1

    db.session.commit()
    return jsonify({'added': added})


# ── 有給管理 ──────────────────────────────────────────────

@app.route("/leave-management")
@login_required
@manager_or_above_required
def leave_management():
    stores     = get_allowed_stores(ignore_active=True)   # サイドバー用
    active_ids = get_allowed_store_ids()                   # アクティブ店舗のみ
    staff_list = Staff.query.filter(
        Staff.store_id.in_(active_ids), Staff.is_active == True
    ).order_by(Staff.name).all()
    year = request.args.get('year', type=int) or date.today().year
    return render_template("leave_management.html", stores=stores, staff_list=staff_list, year=year)


@app.route("/api/set-active-store", methods=["POST"])
@login_required
def api_set_active_store():
    """プレミアオーナーがサイドバーで店舗を切り替えるためのAPI"""
    data = request.get_json() or {}
    store_id = data.get('store_id')
    allowed = get_allowed_store_ids(ignore_active=True)  # 全店舗で検証
    if store_id and int(store_id) in allowed:
        session['active_store_id'] = int(store_id)
    else:
        session.pop('active_store_id', None)
    return jsonify({'status': 'ok', 'active_store_id': session.get('active_store_id')})


@app.route("/api/leave", methods=["GET"])
@login_required
def api_leave_list():
    year     = request.args.get('year',     type=int) or date.today().year
    staff_id = request.args.get('staff_id', type=int)
    # アクティブ店舗のスタッフのみ対象
    _allowed_store_ids = get_allowed_store_ids()
    _allowed_staff_ids = [s.id for s in Staff.query.filter(
        Staff.store_id.in_(_allowed_store_ids), Staff.is_active == True
    ).all()] if _allowed_store_ids else []
    q = LeaveRecord.query.filter(
        db.extract('year', LeaveRecord.leave_date) == year,
        LeaveRecord.staff_id.in_(_allowed_staff_ids) if _allowed_staff_ids else db.false()
    )
    if staff_id:
        q = q.filter_by(staff_id=staff_id)
    records = q.order_by(LeaveRecord.leave_date.desc()).all()
    _allowed = get_allowed_store_ids()
    staff_map = {s.id: s.name for s in Staff.query.filter(Staff.store_id.in_(_allowed)).all() if _allowed}
    return jsonify([{
        'id':         r.id,
        'staff_id':   r.staff_id,
        'staff_name': staff_map.get(r.staff_id, '?'),
        'leave_date': r.leave_date.strftime('%Y-%m-%d'),
        'leave_type': r.leave_type,
        'days':       r.days,
        'memo':       r.memo or '',
        'status':     r.status,
    } for r in records])


@app.route("/api/leave", methods=["POST"])
@login_required
def api_leave_create():
    data = request.get_json() or {}
    try:
        ld = datetime.strptime(data['leave_date'], '%Y-%m-%d').date()
    except Exception:
        return jsonify({'error': '日付が不正です'}), 400
    r = LeaveRecord(
        staff_id   = data.get('staff_id'),
        leave_date = ld,
        leave_type = data.get('leave_type', '有給'),
        days       = float(data.get('days', 1.0)),
        memo       = data.get('memo', ''),
        status     = data.get('status', '承認済'),
    )
    db.session.add(r)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': r.id})


@app.route("/api/leave/<int:lid>", methods=["PUT"])
@login_required
def api_leave_update(lid):
    r = LeaveRecord.query.get_or_404(lid)
    data = request.get_json() or {}
    if 'leave_date' in data:
        r.leave_date = datetime.strptime(data['leave_date'], '%Y-%m-%d').date()
    if 'leave_type' in data: r.leave_type = data['leave_type']
    if 'days'       in data: r.days       = float(data['days'])
    if 'memo'       in data: r.memo       = data['memo']
    if 'status'     in data: r.status     = data['status']
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/leave/<int:lid>", methods=["DELETE"])
@login_required
def api_leave_delete(lid):
    r = LeaveRecord.query.get_or_404(lid)
    db.session.delete(r)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/leave/balance", methods=["GET"])
@login_required
def api_leave_balance():
    """スタッフ別有給残日数サマリー"""
    year = request.args.get('year', type=int) or date.today().year
    # 店舗フィルタ（テナント分離）
    _allowed = get_allowed_store_ids()
    staff_list = Staff.query.filter(
        Staff.store_id.in_(_allowed), Staff.is_active == True
    ).all() if _allowed else []
    result = []
    for s in staff_list:
        bal = LeaveBalance.query.filter_by(staff_id=s.id, year=year).first()
        total = bal.total_days if bal else 10.0
        used = db.session.query(db.func.sum(LeaveRecord.days)).filter(
            LeaveRecord.staff_id == s.id,
            db.extract('year', LeaveRecord.leave_date) == year,
            LeaveRecord.leave_type == '有給',
            LeaveRecord.status != '却下'
        ).scalar() or 0
        result.append({
            'staff_id':   s.id,
            'staff_name': s.name,
            'total_days': total,
            'used_days':  used,
            'remain_days': total - used,
        })
    return jsonify(result)


@app.route("/api/leave/balance", methods=["POST"])
@login_required
def api_leave_balance_update():
    """有給付与日数更新"""
    data = request.get_json() or {}
    staff_id = data.get('staff_id')
    year     = data.get('year') or date.today().year
    total    = float(data.get('total_days', 10))
    bal = LeaveBalance.query.filter_by(staff_id=staff_id, year=year).first()
    if bal:
        bal.total_days = total
    else:
        bal = LeaveBalance(staff_id=staff_id, year=year, total_days=total)
        db.session.add(bal)
    db.session.commit()
    return jsonify({'status': 'ok'})


# ── 日報 ──────────────────────────────────────────────────

@app.route("/daily-report")
@login_required
@block_super_admin
def daily_report():
    """日報ページ"""
    allowed_stores = get_allowed_stores()
    allowed_ids = [s.id for s in allowed_stores]
    stores = allowed_stores
    staff_list = Staff.query.filter(Staff.store_id.in_(allowed_ids), Staff.is_active == True).all()
    year, month = current_ym()
    today = date.today()
    # デフォルトタスクが未作成なら初期化（テナント分離）
    store_id = allowed_ids[0] if allowed_ids else None
    if store_id:
        default_tasks = [
            ("来店前日連絡", True, 1),
            ("来店当日連絡", True, 2),
            ("申込管理入力", True, 3),
        ]
        for task_name, is_def, order in default_tasks:
            exists = DailyTaskTemplate.query.filter_by(store_id=store_id, task_name=task_name).first()
            if not exists:
                db.session.add(DailyTaskTemplate(
                    store_id=store_id, task_name=task_name,
                    is_default=is_def, is_active=True, sort_order=order
                ))
        db.session.commit()
    tasks = DailyTaskTemplate.query.filter(
        DailyTaskTemplate.store_id.in_(allowed_ids),
        DailyTaskTemplate.is_active == True
    ).order_by(DailyTaskTemplate.sort_order).all()
    return render_template("daily_report.html",
                           stores=stores, staff_list=staff_list,
                           tasks=tasks, year=year, month=month,
                           store_id=store_id, today=today, now=datetime.now())


@app.route("/api/daily-report")
@login_required
def api_daily_report_list():
    """日報一覧取得"""
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]
    staff_id = request.args.get('staff_id', type=int)
    report_date_str = request.args.get('date')

    allowed_ids = get_allowed_store_ids()
    q = DailyReport.query
    if allowed_ids:
        q = q.filter(DailyReport.store_id.in_(allowed_ids))
    if report_date_str:
        try:
            rd = datetime.strptime(report_date_str, '%Y-%m-%d').date()
            q = q.filter_by(report_date=rd)
        except Exception:
            pass
    else:
        from calendar import monthrange
        first = date(year, month, 1)
        last  = date(year, month, monthrange(year, month)[1])
        q = q.filter(DailyReport.report_date >= first, DailyReport.report_date <= last)
    if staff_id:
        q = q.filter_by(staff_id=staff_id)
    reports = q.order_by(DailyReport.report_date.desc()).all()

    result = []
    for r in reports:
        staff = Staff.query.get(r.staff_id)
        customers = DailyReportCustomer.query.filter_by(report_id=r.id).all()
        task_checks = DailyTaskCheck.query.filter_by(report_id=r.id).all()
        result.append({
            'id': r.id,
            'staff_id': r.staff_id,
            'staff_name': staff.name if staff else '不明',
            'report_date': r.report_date.isoformat(),
            'prev_day_contact_done': r.prev_day_contact_done,
            'same_day_contact_done': r.same_day_contact_done,
            'application_input_done': r.application_input_done,
            'application_count': r.application_count or 0,
            'tomorrow_appointments': r.tomorrow_appointments or '',
            'memo': r.memo or '',
            'customers': [
                {
                    'id': c.id,
                    'customer_name': c.customer_name,
                    'applied': c.applied,
                    'no_apply_reason': c.no_apply_reason or '',
                    'improvement': c.improvement or '',
                } for c in customers
            ],
            'task_checks': {tc.task_id: tc.checked for tc in task_checks},
        })
    return jsonify(result)


@app.route("/api/daily-report", methods=["POST"])
@login_required
def api_daily_report_save():
    """日報保存（新規 or 更新）"""
    data = request.get_json() or {}
    staff_id = int(data.get('staff_id') or 0) or None
    store_id = safe_store_id(data.get('store_id'))
    if not store_id:
        return jsonify({'error': 'unauthorized'}), 403
    report_date_str = data.get('report_date') or date.today().isoformat()
    try:
        report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
    except Exception:
        report_date = date.today()

    # 同日・同スタッフの日報があれば更新、なければ作成
    report = DailyReport.query.filter_by(staff_id=staff_id, report_date=report_date).first()
    if not report:
        report = DailyReport(staff_id=staff_id, store_id=store_id, report_date=report_date)
        db.session.add(report)

    report.prev_day_contact_done  = bool(data.get('prev_day_contact_done'))
    report.same_day_contact_done  = bool(data.get('same_day_contact_done'))
    report.application_input_done = bool(data.get('application_input_done'))
    report.application_count      = int(data.get('application_count') or 0)
    report.tomorrow_appointments  = data.get('tomorrow_appointments', '')
    report.memo                   = data.get('memo', '')
    report.updated_at             = datetime.utcnow()
    db.session.flush()

    # 接客記録: 全削除 → 再登録
    DailyReportCustomer.query.filter_by(report_id=report.id).delete()
    for c in (data.get('customers') or []):
        if c.get('customer_name', '').strip():
            db.session.add(DailyReportCustomer(
                report_id=report.id,
                customer_name=c['customer_name'].strip(),
                applied=bool(c.get('applied')),
                no_apply_reason=c.get('no_apply_reason', ''),
                improvement=c.get('improvement', ''),
            ))

    # カスタムタスクチェック
    DailyTaskCheck.query.filter_by(report_id=report.id).delete()
    for task_id_str, checked in (data.get('task_checks') or {}).items():
        try:
            db.session.add(DailyTaskCheck(
                report_id=report.id,
                task_id=int(task_id_str),
                checked=bool(checked),
            ))
        except Exception:
            pass

    db.session.commit()
    return jsonify({'status': 'ok', 'id': report.id})


@app.route("/api/daily-report/<int:rid>", methods=["DELETE"])
@login_required
def api_daily_report_delete(rid):
    r = DailyReport.query.get_or_404(rid)
    DailyReportCustomer.query.filter_by(report_id=rid).delete()
    DailyTaskCheck.query.filter_by(report_id=rid).delete()
    db.session.delete(r)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/daily-task-template")
@login_required
def api_task_template_list():
    """タスクテンプレート一覧"""
    store_id = safe_store_id(request.args.get('store_id', type=int))
    if not store_id:
        return jsonify([])
    allowed_ids = get_allowed_store_ids()
    tasks = DailyTaskTemplate.query.filter(
        DailyTaskTemplate.store_id.in_(allowed_ids),
        DailyTaskTemplate.is_active == True
    ).order_by(DailyTaskTemplate.sort_order).all()
    return jsonify([{'id': t.id, 'task_name': t.task_name, 'is_default': t.is_default} for t in tasks])


@app.route("/api/daily-task-template", methods=["POST"])
@login_required
def api_task_template_add():
    """タスクテンプレート追加"""
    data = request.get_json() or {}
    task_name = (data.get('task_name') or '').strip()
    if not task_name:
        return jsonify({'error': 'task_name required'}), 400
    store_id = safe_store_id(data.get('store_id'))
    if not store_id:
        return jsonify({'error': 'unauthorized'}), 403
    max_order = db.session.query(db.func.max(DailyTaskTemplate.sort_order)).filter_by(
        store_id=store_id).scalar() or 0
    t = DailyTaskTemplate(store_id=store_id, task_name=task_name, sort_order=max_order + 1)
    db.session.add(t)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': t.id})


@app.route("/api/daily-task-template/<int:tid>", methods=["PUT"])
@login_required
def api_task_template_update(tid):
    """タスクテンプレート名変更（デフォルト含む）"""
    t = DailyTaskTemplate.query.get_or_404(tid)
    data = request.get_json() or {}
    new_name = (data.get('task_name') or '').strip()
    if not new_name:
        return jsonify({'error': 'task_name required'}), 400
    t.task_name = new_name
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/daily-task-template/<int:tid>", methods=["DELETE"])
@login_required
def api_task_template_delete(tid):
    """タスクテンプレート削除（デフォルトは削除不可）"""
    t = DailyTaskTemplate.query.get_or_404(tid)
    if t.is_default:
        return jsonify({'error': 'デフォルトタスクは削除できません'}), 400
    t.is_active = False
    db.session.commit()
    return jsonify({'status': 'ok'})


# ── パスワードリセット ──────────────────────────────────────

import secrets
from itsdangerous import URLSafeTimedSerializer

def _reset_serializer():
    return URLSafeTimedSerializer(app.secret_key)


def _send_reset_email(to_email, reset_url):
    """パスワードリセットメール送信（Resend API / HTTPS）"""
    resend_key = os.getenv('RESEND_API_KEY', '')
    # Resendドメイン未認証の場合はonboarding@resend.devを使用
    # 独自ドメイン取得後は MAIL_FROM 環境変数で変更可能
    from_email = os.getenv('MAIL_FROM', 'onboarding@resend.dev')
    from_name  = 'ミエルーム'

    if not resend_key:
        app.logger.warning('RESEND_API_KEY が未設定')
        return False

    try:
        import urllib.request, json as _json
        body_html = f"""
<div style="font-family:sans-serif;max-width:480px;margin:0 auto;padding:32px 24px;">
  <div style="text-align:center;margin-bottom:24px;">
    <h2 style="color:#0D9488;margin:0;">ミエルーム</h2>
    <p style="color:#6b7280;font-size:14px;margin:4px 0 0;">不動産賃貸仲介業務管理システム</p>
  </div>
  <div style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:32px;">
    <h3 style="color:#111827;margin:0 0 12px;">パスワードリセットのご案内</h3>
    <p style="color:#374151;line-height:1.7;">以下のボタンからパスワードをリセットしてください。<br>このリンクの有効期限は<strong>1時間</strong>です。</p>
    <div style="text-align:center;margin:28px 0;">
      <a href="{reset_url}" style="background:#0D9488;color:#fff;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:700;font-size:15px;">パスワードをリセットする</a>
    </div>
    <p style="color:#6b7280;font-size:13px;">ボタンが押せない場合は以下のURLをコピーしてください：<br>
    <a href="{reset_url}" style="color:#0D9488;word-break:break-all;">{reset_url}</a></p>
    <hr style="border:none;border-top:1px solid #e5e7eb;margin:20px 0;">
    <p style="color:#9ca3af;font-size:12px;text-align:center;">このメールに心当たりがない場合は無視してください。</p>
  </div>
</div>"""

        payload = _json.dumps({
            'from': f'{from_name} <{from_email}>',
            'to':   [to_email],
            'subject': 'パスワードリセットのご案内 - ミエルーム',
            'html': body_html,
        }).encode('utf-8')

        import ssl
        ctx = ssl.create_default_context()
        req = urllib.request.Request(
            'https://api.resend.com/emails',
            data=payload,
            headers={
                'Authorization': f'Bearer {resend_key}',
                'Content-Type': 'application/json',
                'User-Agent': 'mieroom-app/1.0',
                'Accept': 'application/json',
            },
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=15, context=ctx) as resp:
            result = _json.loads(resp.read())
            app.logger.info(f'Resend送信成功: {to_email} id={result.get("id")}')
            return True
    except urllib.error.HTTPError as he:
        err_body = he.read().decode('utf-8', errors='replace')
        app.logger.error(f'Resend HTTPエラー {he.code}: {err_body}')
        return False
    except Exception as e:
        app.logger.error(f'Resend送信エラー: {type(e).__name__}: {e}')
        return False



@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    """パスワードリセット要求"""
    message = None
    error = None
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        user = AppUser.query.filter(
            db.func.lower(AppUser.email) == email, AppUser.is_active == True
        ).first()
        # セキュリティのため、ユーザーが存在しなくても同じメッセージを返す
        if user and user.email:
            token = secrets.token_urlsafe(32)
            expires = datetime.utcnow() + timedelta(hours=1)
            rt = PasswordResetToken(user_id=user.id, token=token, expires_at=expires)
            db.session.add(rt)
            db.session.commit()
            reset_url = url_for('reset_password', token=token, _external=True)
            # メール送信を別スレッドで実行（画面がフリーズしないように）
            import threading
            def _send_bg(email=user.email, url=reset_url):
                try:
                    result = _send_reset_email(email, url)
                    if result:
                        app.logger.info(f'パスワードリセットメール送信成功: {email}')
                    else:
                        app.logger.error(f'パスワードリセットメール送信失敗: {email}')
                except Exception as e:
                    app.logger.error(f'パスワードリセットメール例外: {email} - {e}')
            threading.Thread(target=_send_bg, daemon=True).start()
            sent = True  # 非同期なので常にTrue扱い
            message = "パスワードリセット用のメールを送信しました。"
        else:
            message = "入力されたメールアドレスに一致するアカウントが見つかりませんでした。"
    return render_template("forgot_password.html", message=message, error=error)


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    """パスワードリセット実行"""
    rt = PasswordResetToken.query.filter_by(token=token, used=False).first()
    error = None
    if not rt or rt.expires_at < datetime.utcnow():
        return render_template("reset_password.html", error="リセットリンクが無効または期限切れです。", token=token, expired=True)

    if request.method == "POST":
        new_password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")
        if len(new_password) < 8:
            error = "パスワードは8文字以上にしてください。"
        elif new_password != confirm:
            error = "パスワードが一致しません。"
        else:
            user = AppUser.query.get(rt.user_id)
            user.password_hash = generate_password_hash(new_password)
            rt.used = True
            db.session.commit()
            return redirect(url_for('app_login') + '?reset=1')

    return render_template("reset_password.html", token=token, error=error, expired=False)


# ── ユーザー管理（オーナー専用） ──────────────────────────────

@app.route("/settings/users")
@login_required
def settings_users():
    """ユーザー管理ページ（super_admin: 管理者権限管理 / owner等: テナント内ユーザー管理）"""
    app_user = AppUser.query.get(session.get('app_user_id'))
    if not app_user or app_user.role not in ('owner', 'store_manager', 'super_admin', 'sys_admin'):
        return redirect(url_for('executive_dashboard'))
    if app_user.role in ('owner', 'store_manager'):
        return redirect(url_for('settings'))
    if app_user.role in ('super_admin', 'sys_admin'):
        # super_admin / sys_admin: クライアント管理画面の権限管理
        sys_admins = AppUser.query.filter(
            AppUser.role == 'sys_admin', AppUser.is_active == True
        ).order_by(AppUser.created_at).all()
        # super_admin 自身を先頭に追加してテンプレートに渡す
        all_admins = [app_user] + [u for u in sys_admins if u.id != app_user.id]
        return render_template("settings_admin_perms.html",
                               sys_admins=sys_admins,
                               all_admins=all_admins,
                               my_id=app_user.id,
                               is_super_admin=(app_user.role == 'super_admin'))
    stores      = get_allowed_stores(ignore_active=True)
    allowed_ids = [s.id for s in stores]
    users = AppUser.query.filter_by(is_active=True, tenant_id=app_user.tenant_id).order_by(AppUser.created_at).all()
    staff_list = Staff.query.filter(Staff.store_id.in_(allowed_ids), Staff.is_active == True).all()
    return render_template("settings_users.html", stores=stores, users=users, staff_list=staff_list)


@app.route("/api/users", methods=["POST"])
@login_required
def api_user_create():
    """ユーザー作成（オーナーのみ）"""
    app_user = AppUser.query.get(session.get('app_user_id'))
    if not app_user or app_user.role not in ('owner', 'super_admin'):
        return jsonify({'error': '権限がありません'}), 403
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    email    = (data.get('email') or '').strip().lower()
    password = data.get('password', '')
    role     = data.get('role', 'staff')
    staff_id = data.get('staff_id') or None
    if not username or not password:
        return jsonify({'error': 'ユーザー名とパスワードは必須です'}), 400
    if AppUser.query.filter_by(username=username).first():
        return jsonify({'error': 'そのユーザー名は既に使用されています'}), 400
    u = AppUser(
        username=username, email=email or None,
        password_hash=generate_password_hash(password),
        role=role, staff_id=staff_id,
        tenant_id=app_user.tenant_id,  # 作成者と同じテナントに所属
        can_view_accounting=bool(data.get('can_view_accounting', True)),
        can_view_all_staff=bool(data.get('can_view_all_staff', True)),
        can_edit_kpi=bool(data.get('can_edit_kpi', True)),
        can_manage_uncollected=bool(data.get('can_manage_uncollected', True)),
        can_view_executive=bool(data.get('can_view_executive', True)),
        can_view_leads_page=bool(data.get('can_view_leads_page', True)),
        can_view_daily_report=bool(data.get('can_view_daily_report', True)),
        can_view_leave=bool(data.get('can_view_leave', True)),
    )
    db.session.add(u)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': u.id})


@app.route("/api/users/<int:uid>", methods=["PUT"])
@login_required
def api_user_update(uid):
    """ユーザー更新（オーナー・店長・super_admin）"""
    app_user = AppUser.query.get(session.get('app_user_id'))
    if not app_user or app_user.role not in ('owner', 'store_manager', 'super_admin'):
        return jsonify({'error': '権限がありません'}), 403
    u = AppUser.query.get_or_404(uid)
    data = request.get_json() or {}
    if 'email'    in data: u.email    = (data['email'] or '').strip().lower() or None
    if 'role'     in data: u.role     = data['role']
    if 'staff_id' in data: u.staff_id = data['staff_id'] or None
    if 'password' in data and data['password']:
        u.password_hash = generate_password_hash(data['password'])
    if 'can_view_accounting'    in data: u.can_view_accounting    = bool(data['can_view_accounting'])
    if 'can_view_all_staff'     in data: u.can_view_all_staff     = bool(data['can_view_all_staff'])
    if 'can_edit_kpi'           in data: u.can_edit_kpi           = bool(data['can_edit_kpi'])
    if 'can_manage_uncollected' in data: u.can_manage_uncollected = bool(data['can_manage_uncollected'])
    if 'can_view_executive'     in data: u.can_view_executive     = bool(data['can_view_executive'])
    if 'can_view_leads_page'    in data: u.can_view_leads_page    = bool(data['can_view_leads_page'])
    if 'can_view_daily_report'  in data: u.can_view_daily_report  = bool(data['can_view_daily_report'])
    if 'can_view_leave'         in data: u.can_view_leave         = bool(data['can_view_leave'])
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/users/<int:uid>", methods=["DELETE"])
@login_required
def api_user_delete(uid):
    """ユーザー無効化（オーナー・店長・super_admin）"""
    app_user = AppUser.query.get(session.get('app_user_id'))
    if not app_user or app_user.role not in ('owner', 'store_manager', 'super_admin'):
        return jsonify({'error': '権限がありません'}), 403
    if uid == app_user.id:
        return jsonify({'error': '自分自身は削除できません'}), 400
    u = AppUser.query.get_or_404(uid)
    u.is_active = False
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/users")
@login_required
def api_user_list():
    """ユーザー一覧（オーナー：自テナントのみ / super_admin：全体）"""
    app_user = AppUser.query.get(session.get('app_user_id'))
    if not app_user or app_user.role not in ('owner', 'super_admin'):
        return jsonify({'error': '権限がありません'}), 403
    q = AppUser.query.filter_by(is_active=True)
    if app_user.role == 'owner':
        q = q.filter_by(tenant_id=app_user.tenant_id)
    users = q.order_by(AppUser.created_at).all()
    return jsonify([{
        'id': u.id,
        'username': u.username,
        'email': u.email or '',
        'role': u.role,
        'staff_id': u.staff_id,
        'tenant_id': u.tenant_id,
        'can_view_accounting': u.can_view_accounting,
        'can_view_all_staff': u.can_view_all_staff,
        'can_edit_kpi': u.can_edit_kpi,
        'can_manage_uncollected': u.can_manage_uncollected,
        'last_login': u.last_login.isoformat() if u.last_login else None,
    } for u in users])


# ── テナント管理（super_admin専用） ────────────────────────

@app.route("/api/admin/store-data-check")
@super_admin_required
def api_admin_store_data_check():
    """全店舗のデータ件数を返す（デバッグ用）"""
    stores = Store.query.filter_by(is_active=True).all()
    result = []
    for s in stores:
        tenant = Tenant.query.get(s.tenant_id) if s.tenant_id else None
        kpi_count  = SalesKPI.query.filter_by(store_id=s.id).count()
        staff_count = Staff.query.filter_by(store_id=s.id, is_active=True).count()
        app_count  = ApplicationRecord.query.filter_by(store_id=s.id).count()
        result.append({
            'store_id': s.id,
            'store_name': s.name,
            'tenant_name': tenant.name if tenant else 'なし',
            'kpi_records': kpi_count,
            'staff_count': staff_count,
            'application_count': app_count,
        })
    return jsonify(result)


@app.route("/api/admin/clear-store-data/<int:sid>", methods=["POST"])
@super_admin_required
def api_admin_clear_store_data(sid):
    """指定店舗のデータを強制クリア"""
    store = Store.query.get_or_404(sid)
    kpi_del  = SalesKPI.query.filter_by(store_id=sid).delete()
    pl_del   = PLRecord.query.filter_by(store_id=sid).delete()
    app_del  = ApplicationRecord.query.filter_by(store_id=sid).delete()
    Staff.query.filter_by(store_id=sid).update({'is_active': False})
    db.session.commit()
    return jsonify({'status': 'ok', 'store_name': store.name,
                    'deleted_kpi': kpi_del, 'deleted_app': app_del})


@app.route("/admin/tenants")
@super_admin_required
def admin_tenants():
    """テナント管理ページ（super_adminのみ）"""
    tenants = Tenant.query.order_by(Tenant.created_at.desc()).all()
    cur_user = AppUser.query.get(session.get('app_user_id'))
    is_super_admin = cur_user and cur_user.role == 'super_admin'
    return render_template("admin_tenants.html", tenants=tenants, now=datetime.now(),
                           is_super_admin=is_super_admin, plan_options=PLAN_OPTION_DEFS)


@app.route("/admin/applications")
@super_admin_required
def admin_applications():
    """トライアル申込管理ページ（super_adminのみ）"""
    apps = TrialApplication.query.order_by(TrialApplication.created_at.desc()).all()
    total   = len(apps)
    new_cnt = sum(1 for a in apps if a.status == 'new')
    contacted_cnt  = sum(1 for a in apps if a.status == 'contacted')
    contracted_cnt = sum(1 for a in apps if a.status == 'contracted')
    rejected_cnt   = sum(1 for a in apps if a.status == 'rejected')
    return render_template("admin_applications.html", apps=apps,
                           total=total, new_cnt=new_cnt,
                           contacted_cnt=contacted_cnt,
                           contracted_cnt=contracted_cnt,
                           rejected_cnt=rejected_cnt)


@app.route("/api/admin/applications/<int:app_id>", methods=["DELETE"])
@super_admin_required
def api_admin_application_delete(app_id):
    """申込削除"""
    rec = TrialApplication.query.get_or_404(app_id)
    db.session.delete(rec)
    db.session.commit()
    return jsonify({'ok': True})


@app.route("/api/admin/applications/<int:app_id>", methods=["PATCH"])
@super_admin_required
def api_admin_application_update(app_id):
    """申込ステータス・メモ更新"""
    rec  = TrialApplication.query.get_or_404(app_id)
    data = request.get_json() or {}
    if 'status' in data:
        rec.status = data['status']
    if 'memo' in data:
        rec.memo = data['memo']
    db.session.commit()
    return jsonify({'ok': True})


@app.route("/api/tenants", methods=["GET"])
@super_admin_required
def api_tenants_get():
    """テナント一覧"""
    tenants = Tenant.query.order_by(Tenant.created_at.desc()).all()
    result = []
    for t in tenants:
        stores = Store.query.filter_by(tenant_id=t.id, is_active=True).all()
        owner = AppUser.query.filter_by(tenant_id=t.id, role='owner', is_active=True).first()
        result.append({
            'id': t.id,
            'name': t.name,
            'plan': t.plan,
            'options': sorted(tenant_option_keys(t.id)),
            'is_active': t.is_active,
            'subscription_status': t.subscription_status or 'trial',
            'trial_ends_at': t.trial_ends_at.strftime('%Y-%m-%dT%H:%M:%S') if t.trial_ends_at else None,
            'contract_start_date': t.contract_start_date.strftime('%Y-%m-%d') if t.contract_start_date else None,
            'store_count': len(stores),
            'owner_username': owner.username if owner else None,
            'owner_email': owner.email if owner else None,
            'created_at': t.created_at.strftime('%Y-%m-%d') if t.created_at else None,
        })
    return jsonify(result)


@app.route("/api/tenants", methods=["POST"])
@super_admin_required
def api_tenants_post():
    """テナント新規作成（デフォルト店舗+オーナーアカウントも作成）"""
    err = _check_admin_perm('admin_can_add_tenant')
    if err: return err
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': '会社名は必須です'}), 400

    owner_username = (data.get('owner_username') or '').strip()
    owner_password = (data.get('owner_password') or '').strip()
    owner_email    = (data.get('owner_email') or '').strip()

    if not owner_username or not owner_password:
        return jsonify({'error': 'オーナーのユーザー名とパスワードは必須です'}), 400

    if AppUser.query.filter_by(username=owner_username).first():
        return jsonify({'error': 'そのユーザー名は既に使われています'}), 400

    trial_days = int(data.get('trial_days', 14))
    from datetime import timedelta
    trial_ends_at = datetime.utcnow() + timedelta(days=trial_days)

    tenant = Tenant(name=name, plan=data.get('plan', 'standard'), is_active=True,
                    trial_ends_at=trial_ends_at, subscription_status='trial')
    db.session.add(tenant)
    db.session.flush()

    store = Store(name=name, is_active=True, tenant_id=tenant.id)
    db.session.add(store)
    db.session.flush()

    owner = AppUser(
        username=owner_username,
        email=owner_email or None,
        password_hash=generate_password_hash(owner_password),
        role='owner',
        tenant_id=tenant.id,
        is_active=True,
    )
    db.session.add(owner)
    db.session.flush()
    set_tenant_options(tenant.id, data.get('options') or [])
    db.session.commit()
    return jsonify({'status': 'ok', 'id': tenant.id})


@app.route("/api/tenants/<int:tid>", methods=["PUT"])
@super_admin_required
def api_tenant_update(tid):
    """テナント更新"""
    tenant = Tenant.query.get_or_404(tid)
    data = request.get_json() or {}
    if 'name' in data and data['name'].strip():
        tenant.name = data['name'].strip()
    if 'plan' in data:
        tenant.plan = data['plan']
    if 'options' in data:
        set_tenant_options(tid, data['options'])
    if 'is_active' in data:
        tenant.is_active = bool(data['is_active'])
    if 'contract_start_date' in data:
        from datetime import datetime as _dt
        try:
            tenant.contract_start_date = _dt.strptime(data['contract_start_date'], '%Y-%m-%d').date() if data['contract_start_date'] else None
        except Exception:
            pass
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/tenants/<int:tid>/reset-all-data", methods=["POST"])
@super_admin_required
def api_tenant_reset_all_data(tid):
    """テナントの全業務データを削除（アカウント・店舗・テナントは残す）"""
    tenant = Tenant.query.get_or_404(tid)
    store_ids = [s.id for s in Store.query.filter_by(tenant_id=tid).all()]
    staff_ids  = [s.id for s in Staff.query.filter(Staff.store_id.in_(store_ids)).all()] if store_ids else []
    counts = {}

    if store_ids:
        counts['SalesKPI']             = SalesKPI.query.filter(SalesKPI.store_id.in_(store_ids)).delete(synchronize_session=False)
        counts['ApplicationRecord']    = ApplicationRecord.query.filter(ApplicationRecord.store_id.in_(store_ids)).delete(synchronize_session=False)
        counts['EchoRecord']           = EchoRecord.query.filter(EchoRecord.store_id.in_(store_ids)).delete(synchronize_session=False)
        counts['CustomerServiceRecord']= CustomerServiceRecord.query.filter(CustomerServiceRecord.store_id.in_(store_ids)).delete(synchronize_session=False)
        counts['PLRecord']             = PLRecord.query.filter(PLRecord.store_id.in_(store_ids)).delete(synchronize_session=False)

        # 日報関連
        report_ids = [r.id for r in DailyReport.query.filter(DailyReport.store_id.in_(store_ids)).all()]
        if report_ids:
            DailyReportCustomer.query.filter(DailyReportCustomer.report_id.in_(report_ids)).delete(synchronize_session=False)
            DailyTaskCheck.query.filter(DailyTaskCheck.report_id.in_(report_ids)).delete(synchronize_session=False)
            counts['DailyReport'] = DailyReport.query.filter(DailyReport.store_id.in_(store_ids)).delete(synchronize_session=False)

    if staff_ids:
        counts['LeaveRecord'] = LeaveRecord.query.filter(LeaveRecord.staff_id.in_(staff_ids)).delete(synchronize_session=False)
        try:
            counts['LeaveAllocation'] = LeaveAllocation.query.filter(LeaveAllocation.staff_id.in_(staff_ids)).delete(synchronize_session=False)
        except Exception:
            pass

    # UncollectedPayment
    if store_ids:
        try:
            counts['UncollectedPayment'] = UncollectedPayment.query.filter(UncollectedPayment.store_id.in_(store_ids)).delete(synchronize_session=False)
        except Exception:
            pass

    db.session.commit()
    return jsonify({'status': 'ok', 'tenant_name': tenant.name, 'deleted': counts})


@app.route("/api/tenants/<int:tid>/activate", methods=["POST"])
@super_admin_required
def api_tenant_activate(tid):
    """トライアル → 契約開始（super_admin / sys_admin 共通可）"""
    from datetime import date as _date
    tenant = Tenant.query.get_or_404(tid)
    tenant.subscription_status = 'active'
    tenant.is_active = True
    if not tenant.contract_start_date:
        tenant.contract_start_date = _date.today()
    db.session.commit()
    return jsonify({'status': 'ok', 'contract_start_date': tenant.contract_start_date.strftime('%Y-%m-%d')})


@app.route("/api/admin-users", methods=["GET"])
@super_admin_required
def api_admin_users_list():
    """sys_admin ユーザー一覧（super_admin のみ全表示）"""
    users = AppUser.query.filter(
        AppUser.role.in_(['sys_admin']),
        AppUser.is_active == True
    ).all()
    return jsonify([{
        'id': u.id, 'username': u.username,
        'email': u.email or '',
        'created_at': u.created_at.strftime('%Y-%m-%d') if u.created_at else None,
        'admin_can_add_tenant':    bool(getattr(u, 'admin_can_add_tenant',    False)),
        'admin_can_manage_stores': bool(getattr(u, 'admin_can_manage_stores', False)),
        'admin_can_delete_tenant': bool(getattr(u, 'admin_can_delete_tenant', False)),
        'admin_can_lock_tenant':   bool(getattr(u, 'admin_can_lock_tenant',   False)),
    } for u in users])


@app.route("/api/admin-users", methods=["POST"])
@super_admin_required
def api_admin_users_create():
    """sys_admin ユーザー作成（super_admin のみ）"""
    if session.get('app_user_role') != 'super_admin':
        return jsonify({'error': 'スーパー管理者のみ作成できます'}), 403
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    email    = (data.get('email') or '').strip()
    if not username or not password:
        return jsonify({'error': 'ユーザー名とパスワードは必須です'}), 400
    if AppUser.query.filter_by(username=username).first():
        return jsonify({'error': 'そのユーザー名は既に使われています'}), 400
    user = AppUser(
        username=username,
        email=email or None,
        password_hash=generate_password_hash(password),
        role='sys_admin',
        tenant_id=None,
        is_active=True,
        admin_can_add_tenant=bool(data.get('admin_can_add_tenant', False)),
        admin_can_manage_stores=bool(data.get('admin_can_manage_stores', False)),
        admin_can_delete_tenant=bool(data.get('admin_can_delete_tenant', False)),
        admin_can_lock_tenant=bool(data.get('admin_can_lock_tenant', False)),
    )
    db.session.add(user)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': user.id})


@app.route("/api/admin-users/<int:uid>", methods=["PUT"])
@super_admin_required
def api_admin_users_update(uid):
    """sys_admin の権限更新（super_admin のみ）"""
    if session.get('app_user_role') != 'super_admin':
        return jsonify({'error': 'スーパー管理者のみ変更できます'}), 403
    user = AppUser.query.get_or_404(uid)
    data = request.get_json() or {}
    for field in ['admin_can_add_tenant','admin_can_manage_stores','admin_can_delete_tenant','admin_can_lock_tenant']:
        if field in data:
            setattr(user, field, bool(data[field]))
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/my-account", methods=["PUT"])
@login_required
def api_my_account_update():
    """ログイン中のアカウント情報を更新（メール・パスワード）"""
    user = AppUser.query.get(session.get('app_user_id'))
    if not user:
        return jsonify({'error': '認証エラー'}), 401
    data = request.get_json() or {}
    if 'email' in data and data['email']:
        user.email = data['email'].strip().lower()
    if 'password' in data and data['password']:
        if len(data['password']) < 8:
            return jsonify({'error': 'パスワードは8文字以上にしてください'}), 400
        user.password_hash = generate_password_hash(data['password'])
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/admin-users/<int:uid>", methods=["DELETE"])
@super_admin_required
def api_admin_users_delete(uid):
    """sys_admin ユーザー削除（super_admin のみ）"""
    if session.get('app_user_role') != 'super_admin':
        return jsonify({'error': 'スーパー管理者のみ削除できます'}), 403
    user = AppUser.query.get_or_404(uid)
    if user.role not in ('sys_admin',):
        return jsonify({'error': '削除できないアカウントです'}), 400
    user.is_active = False
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/tenants/<int:tid>/lock", methods=["POST"])
@super_admin_only
def api_tenant_lock(tid):
    """テナントをロック（アクセス不可）"""
    err = _check_admin_perm("admin_can_lock_tenant")
    if err: return err
    tenant = Tenant.query.get_or_404(tid)
    tenant.subscription_status = 'locked'
    tenant.is_active = False
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/tenants/<int:tid>/unlock", methods=["POST"])
@super_admin_only
def api_tenant_unlock(tid):
    """テナントを有効化（ロック解除）"""
    err = _check_admin_perm("admin_can_lock_tenant")
    if err: return err
    from datetime import date as _date
    tenant = Tenant.query.get_or_404(tid)
    tenant.subscription_status = 'active'
    tenant.is_active = True
    # 契約開始日が未設定なら今日をセット
    if not tenant.contract_start_date:
        tenant.contract_start_date = _date.today()
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/tenants/<int:tid>/extend-trial", methods=["POST"])
@super_admin_only
def api_tenant_extend_trial(tid):
    """トライアル期間を延長"""
    from datetime import timedelta
    tenant = Tenant.query.get_or_404(tid)
    data = request.get_json() or {}
    days = int(data.get('days', 7))
    base = tenant.trial_ends_at if (tenant.trial_ends_at and tenant.trial_ends_at > datetime.utcnow()) else datetime.utcnow()
    tenant.trial_ends_at = base + timedelta(days=days)
    tenant.subscription_status = 'trial'
    tenant.is_active = True
    db.session.commit()
    return jsonify({'status': 'ok', 'trial_ends_at': tenant.trial_ends_at.strftime('%Y-%m-%d')})


@app.route("/api/admin/test-email", methods=["POST"])
@super_admin_required
def api_test_email():
    """メール送信テスト（super_adminのみ）"""
    data = request.get_json() or {}
    to_email = data.get("email", "").strip()
    if not to_email:
        return jsonify({"error": "emailが必要です"}), 400
    try:
        # Resendを直接呼んでエラー詳細を取得
        import urllib.request, urllib.error, json as _json
        resend_key = os.getenv('RESEND_API_KEY', '')
        if not resend_key:
            return jsonify({"status": "error", "message": "RESEND_API_KEY が Railway に未設定"}), 500
        from_addr = os.getenv('MAIL_FROM', 'onboarding@resend.dev')
        payload = _json.dumps({
            'from': f'ミエルーム <{from_addr}>',
            'to': [to_email],
            'subject': '【ミエルーム】メール送信テスト',
            'html': '<p>ミエルームのテストメールです。</p>'
        }).encode('utf-8')
        import ssl
        ctx = ssl.create_default_context()
        req = urllib.request.Request(
            'https://api.resend.com/emails',
            data=payload,
            headers={
                'Authorization': f'Bearer {resend_key}',
                'Content-Type': 'application/json',
                'User-Agent': 'mieroom-app/1.0',
                'Accept': 'application/json',
            },
            method='POST'
        )
        try:
            with urllib.request.urlopen(req, timeout=15, context=ctx) as resp:
                result = _json.loads(resp.read())
                return jsonify({"status": "ok", "message": f"{to_email} に送信成功", "id": result.get("id")})
        except urllib.error.HTTPError as he:
            err_body = he.read().decode('utf-8', errors='replace')
            return jsonify({"status": "error", "http_status": he.code, "message": err_body[:500]}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/law")
def law_page():
    """特定商取引法に基づく表示ページ"""
    return render_template("law.html")


@app.route("/apply", methods=["GET", "POST"])
def apply_page():
    """無料トライアル申し込みフォーム"""
    success = False
    error   = None
    form    = {}

    if request.method == "POST":
        company = request.form.get("company", "").strip()
        name    = request.form.get("name", "").strip()
        email   = request.form.get("email", "").strip()
        phone   = request.form.get("phone", "").strip()
        stores  = request.form.get("stores", "").strip()
        message = request.form.get("message", "").strip()
        form    = {"company": company, "name": name, "email": email,
                   "phone": phone, "stores": stores, "message": message}

        if not all([company, name, email, phone, stores]):
            error = "必須項目をすべて入力してください。"
        else:
            # DBに保存
            try:
                app_record = TrialApplication(
                    company=company, name=name, email=email,
                    phone=phone, stores=stores, message=message
                )
                db.session.add(app_record)
                db.session.commit()
            except Exception as e:
                app.logger.error(f'TrialApplication DB保存エラー: {e}')
                db.session.rollback()

            # 管理者への通知メール
            import threading
            def _notify():
                try:
                    import urllib.request, json as _json, ssl
                    resend_key = os.getenv('RESEND_API_KEY', '')
                    if not resend_key:
                        return
                    body_html = f"""
<h2>【ミエルーム】新規トライアル申し込み</h2>
<table border="1" cellpadding="8" style="border-collapse:collapse;">
  <tr><th>会社名</th><td>{company}</td></tr>
  <tr><th>担当者名</th><td>{name}</td></tr>
  <tr><th>メール</th><td>{email}</td></tr>
  <tr><th>電話番号</th><td>{phone}</td></tr>
  <tr><th>店舗数</th><td>{stores}</td></tr>
  <tr><th>メッセージ</th><td>{message or "なし"}</td></tr>
</table>"""
                    payload = _json.dumps({
                        'from': 'ミエルーム申込通知 <onboarding@resend.dev>',
                        'to':   ['mieroom.cloud@gmail.com'],
                        'reply_to': email,
                        'subject': f'【申込】{company} - ミエルーム トライアル申請',
                        'html': body_html,
                    }).encode('utf-8')
                    req = urllib.request.Request(
                        'https://api.resend.com/emails', data=payload,
                        headers={'Authorization': f'Bearer {resend_key}',
                                 'Content-Type': 'application/json',
                                 'User-Agent': 'mieroom-app/1.0'},
                        method='POST')
                    ctx = ssl.create_default_context()
                    with urllib.request.urlopen(req, timeout=15, context=ctx) as r:
                        app.logger.info(f'管理者通知メール送信成功: {company} → mieroom.cloud@gmail.com')

                    # サンクスメール（申込者へ）
                    thanks_html = f"""
<div style="font-family:sans-serif;max-width:520px;margin:0 auto;padding:32px 24px;">
  <div style="text-align:center;margin-bottom:24px;">
    <h2 style="color:#0D9488;margin:0;">ミエルーム</h2>
    <p style="color:#6b7280;font-size:13px;margin:4px 0 0;">不動産賃貸仲介業務管理システム</p>
  </div>
  <div style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:32px;">
    <h3 style="color:#111827;margin:0 0 16px;">お申し込みありがとうございます</h3>
    <p style="color:#374151;line-height:1.8;">{name} 様<br><br>
    この度は、ミエルームの無料トライアルにお申し込みいただきありがとうございます。<br><br>
    担当者より <strong>1〜2営業日以内</strong> にご連絡いたします。<br>
    今しばらくお待ちください。</p>
    <div style="background:#F0FDFA;border-radius:8px;padding:16px;margin-top:20px;">
      <p style="margin:0;font-size:13px;color:#0F766E;font-weight:600;">お申し込み内容</p>
      <table style="margin-top:8px;font-size:13px;color:#374151;width:100%;">
        <tr><td style="padding:3px 0;color:#6b7280;">会社名</td><td>{company}</td></tr>
        <tr><td style="padding:3px 0;color:#6b7280;">担当者名</td><td>{name}</td></tr>
        <tr><td style="padding:3px 0;color:#6b7280;">電話番号</td><td>{phone}</td></tr>
        <tr><td style="padding:3px 0;color:#6b7280;">店舗数</td><td>{stores}</td></tr>
      </table>
    </div>
    <hr style="border:none;border-top:1px solid #e5e7eb;margin:24px 0;">
    <p style="color:#9ca3af;font-size:12px;margin:0;">
      ご不明な点は <a href="mailto:mieroom.cloud@gmail.com" style="color:#0D9488;">mieroom.cloud@gmail.com</a> までご連絡ください。
    </p>
  </div>
</div>"""
                    thanks_payload = _json.dumps({
                        'from': 'ミエルーム <onboarding@resend.dev>',
                        'to': [email],
                        'subject': '【ミエルーム】お申し込みありがとうございます',
                        'html': thanks_html,
                    }).encode('utf-8')
                    thanks_req = urllib.request.Request(
                        'https://api.resend.com/emails', data=thanks_payload,
                        headers={'Authorization': f'Bearer {resend_key}',
                                 'Content-Type': 'application/json',
                                 'User-Agent': 'mieroom-app/1.0'},
                        method='POST')
                    with urllib.request.urlopen(thanks_req, timeout=15, context=ctx) as r:
                        app.logger.info(f'サンクスメール送信成功: {email}')

                except Exception as e:
                    app.logger.error(f'申込通知メールエラー: {e}')
            threading.Thread(target=_notify, daemon=True).start()
            success = True

    return render_template("apply.html", success=success, error=error,
                           form=form, email=form.get("email",""))


@app.route("/terms")
def terms_page():
    """利用規約ページ"""
    return render_template("terms.html")


@app.route("/privacy")
def privacy_page():
    """プライバシーポリシーページ"""
    return render_template("privacy.html")


@app.route("/trial-expired")
def trial_expired_page():
    """トライアル期限切れ・ロック時のページ"""
    return render_template("trial_expired.html")


@app.route("/api/tenants/<int:tid>", methods=["DELETE"])
@super_admin_only
def api_tenant_delete(tid):
    """テナント物理削除（関連データを全てカスケード削除）"""
    err = _check_admin_perm("admin_can_delete_tenant")
    if err: return err
    tenant = Tenant.query.get_or_404(tid)

    # 対象の店舗・スタッフIDを収集
    store_ids = [s.id for s in Store.query.filter_by(tenant_id=tid).all()]

    if store_ids:
        # 日報の子テーブルを先に削除
        report_ids = [r.id for r in DailyReport.query.filter(
            DailyReport.store_id.in_(store_ids)).all()]
        if report_ids:
            DailyTaskCheck.query.filter(
                DailyTaskCheck.report_id.in_(report_ids)).delete(synchronize_session=False)
            DailyReportCustomer.query.filter(
                DailyReportCustomer.report_id.in_(report_ids)).delete(synchronize_session=False)
            DailyReport.query.filter(
                DailyReport.id.in_(report_ids)).delete(synchronize_session=False)

        DailyTaskTemplate.query.filter(
            DailyTaskTemplate.store_id.in_(store_ids)).delete(synchronize_session=False)
        LeaveBalance.query.filter(
            LeaveBalance.store_id.in_(store_ids)).delete(synchronize_session=False)
        LeaveRecord.query.filter(
            LeaveRecord.store_id.in_(store_ids)).delete(synchronize_session=False)
        SalesKPI.query.filter(
            SalesKPI.store_id.in_(store_ids)).delete(synchronize_session=False)
        UncollectedPayment.query.filter(
            UncollectedPayment.store_id.in_(store_ids)).delete(synchronize_session=False)
        Lead.query.filter(
            Lead.store_id.in_(store_ids)).delete(synchronize_session=False)
        ApplicationRecord.query.filter(
            ApplicationRecord.store_id.in_(store_ids)).delete(synchronize_session=False)
        ContractRecord.query.filter(
            ContractRecord.store_id.in_(store_ids)).delete(synchronize_session=False)
        LeadMediaStat.query.filter(
            LeadMediaStat.store_id.in_(store_ids)).delete(synchronize_session=False)
        AdCost.query.filter(
            AdCost.store_id.in_(store_ids)).delete(synchronize_session=False)
        PLCustomValue.query.filter(
            PLCustomValue.store_id.in_(store_ids)).delete(synchronize_session=False)
        PLCustomItem.query.filter(
            PLCustomItem.store_id.in_(store_ids)).delete(synchronize_session=False)
        PLRecord.query.filter(
            PLRecord.store_id.in_(store_ids)).delete(synchronize_session=False)
        MediaType.query.filter(
            MediaType.store_id.in_(store_ids)).delete(synchronize_session=False)
        StatusColor.query.filter(
            StatusColor.store_id.in_(store_ids)).delete(synchronize_session=False)
        EchoRecord.query.filter(
            EchoRecord.store_id.in_(store_ids)).delete(synchronize_session=False)
        CustomerServiceRecord.query.filter(
            CustomerServiceRecord.store_id.in_(store_ids)).delete(synchronize_session=False)
        Staff.query.filter(
            Staff.store_id.in_(store_ids)).delete(synchronize_session=False)
        Store.query.filter(
            Store.id.in_(store_ids)).delete(synchronize_session=False)

    # AppUser とそのリセットトークンを削除
    user_ids = [u.id for u in AppUser.query.filter_by(tenant_id=tid).all()]
    if user_ids:
        PasswordResetToken.query.filter(
            PasswordResetToken.user_id.in_(user_ids)).delete(synchronize_session=False)
        AppUser.query.filter(
            AppUser.id.in_(user_ids)).delete(synchronize_session=False)

    db.session.delete(tenant)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/tenants/<int:tid>/stores", methods=["GET"])
@super_admin_required
def api_tenant_stores(tid):
    """テナント内の店舗一覧（有効店舗＋ロック中店舗。論理削除のみ除外）"""
    stores = (Store.query.filter(
                Store.tenant_id == tid,
                db.or_(Store.is_active == True, Store.is_locked == True))
              .order_by(Store.created_at.asc()).all())
    return jsonify([{
        'id': s.id,
        'name': s.name,
        'is_locked': bool(getattr(s, 'is_locked', False)),
        'options': sorted(store_option_keys(s.id)),
        'contract_start_date': s.contract_start_date.strftime('%Y-%m-%d') if getattr(s, 'contract_start_date', None) else None,
        'created_at': s.created_at.strftime('%Y-%m-%d') if s.created_at else None
    } for s in stores])


@app.route("/api/tenants/<int:tid>/stores", methods=["POST"])
@super_admin_only
def api_tenant_store_add(tid):
    """テナントに店舗を追加"""
    err = _check_admin_perm("admin_can_manage_stores")
    if err: return err
    Tenant.query.get_or_404(tid)
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': '店舗名は必須です'}), 400
    store = Store(name=name, is_active=True, tenant_id=tid)
    db.session.add(store)
    db.session.flush()  # IDを確定させる
    # 新店舗IDに万が一紐づく既存データを全削除（まっさら保証）
    sid = store.id
    SalesKPI.query.filter_by(store_id=sid).delete()
    PLRecord.query.filter_by(store_id=sid).delete()
    ApplicationRecord.query.filter_by(store_id=sid).delete()
    Staff.query.filter_by(store_id=sid).delete()
    UncollectedPayment.query.filter_by(store_id=sid).delete()
    try:
        EchoRecord.query.filter_by(store_id=sid).delete()
        CustomerServiceRecord.query.filter_by(store_id=sid).delete()
    except Exception: pass
    db.session.commit()
    return jsonify({'status': 'ok', 'id': store.id})


@app.route("/api/tenants/<int:tid>/stores/<int:sid>", methods=["PUT"])
@super_admin_required
def api_tenant_store_update(tid, sid):
    """店舗情報を更新（名前・ロック・契約開始日）"""
    store = Store.query.filter_by(id=sid, tenant_id=tid).first_or_404()
    data = request.get_json() or {}
    if 'name' in data:
        name = (data['name'] or '').strip()
        if name: store.name = name
    if 'is_locked' in data:
        store.is_locked = bool(data['is_locked'])
        store.is_active = not bool(data['is_locked'])
    if 'contract_start_date' in data:
        from datetime import datetime as _dt
        try:
            store.contract_start_date = _dt.strptime(data['contract_start_date'], '%Y-%m-%d').date() if data['contract_start_date'] else None
        except Exception:
            pass
    if 'options' in data:
        set_store_options(sid, data.get('options') or [])
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/tenants/<int:tid>/owner", methods=["PUT"])
@super_admin_required
def api_tenant_owner_update(tid):
    """テナントのオーナーアカウント情報を更新（ユーザー名/パスワード/メール）"""
    owner = AppUser.query.filter_by(tenant_id=tid, role='owner', is_active=True).first()
    if not owner:
        return jsonify({'error': 'オーナーアカウントが見つかりません'}), 404
    data = request.get_json() or {}
    if 'username' in data and data['username'].strip():
        new_name = data['username'].strip()
        if AppUser.query.filter(AppUser.username == new_name, AppUser.id != owner.id).first():
            return jsonify({'error': 'そのユーザー名は既に使われています'}), 400
        owner.username = new_name
    if 'email' in data:
        owner.email = (data['email'] or '').strip() or None
    if 'password' in data and data['password']:
        if len(data['password']) < 8:
            return jsonify({'error': 'パスワードは8文字以上にしてください'}), 400
        owner.password_hash = generate_password_hash(data['password'])
    db.session.commit()
    return jsonify({'status': 'ok', 'username': owner.username, 'email': owner.email or ''})


@app.route("/api/tenants/<int:tid>/stores/<int:sid>/reset-data", methods=["POST"])
@super_admin_required
def api_tenant_store_reset_data(tid, sid):
    """店舗の全データをリセット（完全にまっさらな状態に戻す）"""
    store = Store.query.filter_by(id=sid, tenant_id=tid).first_or_404()
    # 1. KPI・売上データ
    SalesKPI.query.filter_by(store_id=sid).delete()
    PLRecord.query.filter_by(store_id=sid).delete()
    # 2. 申込データ
    ApplicationRecord.query.filter_by(store_id=sid).delete()
    # 3. 反響・接客データ
    try:
        EchoRecord.query.filter_by(store_id=sid).delete()
    except Exception: pass
    try:
        CustomerServiceRecord.query.filter_by(store_id=sid).delete()
    except Exception: pass
    # 4. スタッフと有給
    staff_ids = [s.id for s in Staff.query.filter_by(store_id=sid).all()]
    if staff_ids:
        LeaveRecord.query.filter(LeaveRecord.staff_id.in_(staff_ids)).delete(synchronize_session=False)
        LeaveBalance.query.filter(LeaveBalance.staff_id.in_(staff_ids)).delete(synchronize_session=False)
    Staff.query.filter_by(store_id=sid).delete()
    # 5. 未入金データ
    UncollectedPayment.query.filter_by(store_id=sid).delete()
    db.session.commit()
    return jsonify({'status': 'ok', 'store_name': store.name})


@app.route("/api/tenants/<int:tid>/stores/<int:sid>", methods=["DELETE"])
@super_admin_only
def api_tenant_store_delete(tid, sid):
    """店舗を論理削除"""
    err = _check_admin_perm("admin_can_manage_stores")
    if err: return err
    store = Store.query.filter_by(id=sid, tenant_id=tid).first_or_404()
    store.is_active = False
    store.is_locked = False   # 論理削除はロック扱いにしない（一覧から消す）
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/tenants/<int:tid>/staff", methods=["POST"])
@super_admin_required
def api_tenant_staff_add(tid):
    """テナントにスタッフを追加"""
    Tenant.query.get_or_404(tid)
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    store_id = int(data.get('store_id') or 0) or None
    role = (data.get('role') or '営業').strip()
    if not name:
        return jsonify({'error': 'スタッフ名は必須です'}), 400
    if store_id:
        store = Store.query.filter_by(id=store_id, tenant_id=tid, is_active=True).first()
        if not store:
            return jsonify({'error': '指定した店舗が見つかりません'}), 400
    staff = Staff(name=name, store_id=store_id, role=role, is_active=True)
    db.session.add(staff)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': staff.id, 'name': staff.name})


# ── テナントのログインアカウント管理 (super_admin 専用) ─────────────────────

@app.route("/api/tenants/<int:tid>/users", methods=["GET"])
@super_admin_required
def api_tenant_users_list(tid):
    Tenant.query.get_or_404(tid)
    users = AppUser.query.filter_by(tenant_id=tid, is_active=True).order_by(AppUser.created_at).all()
    return jsonify([{
        'id': u.id,
        'username': u.username,
        'email': u.email or '',
        'role': u.role,
        'last_login': u.last_login.strftime('%Y-%m-%d %H:%M') if u.last_login else None,
    } for u in users])


@app.route("/api/tenants/<int:tid>/users", methods=["POST"])
@super_admin_required
def api_tenant_users_add(tid):
    Tenant.query.get_or_404(tid)
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    role = (data.get('role') or 'staff').strip()
    email = (data.get('email') or '').strip() or None
    if not username or not password:
        return jsonify({'error': 'ユーザー名とパスワードは必須です'}), 400
    if AppUser.query.filter_by(username=username).first():
        return jsonify({'error': 'そのユーザー名は既に使用されています'}), 400
    user = AppUser(
        tenant_id=tid,
        username=username,
        password_hash=generate_password_hash(password),
        role=role,
        email=email,
        is_active=True,
    )
    db.session.add(user)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': user.id, 'username': user.username, 'role': user.role})


@app.route("/api/tenants/<int:tid>/users/<int:uid>/password", methods=["PUT"])
@super_admin_required
def api_tenant_user_password(tid, uid):
    user = AppUser.query.filter_by(id=uid, tenant_id=tid).first_or_404()
    data = request.get_json() or {}
    password = (data.get('password') or '').strip()
    if not password:
        return jsonify({'error': 'パスワードを入力してください'}), 400
    user.password_hash = generate_password_hash(password)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/tenants/<int:tid>/users/<int:uid>", methods=["DELETE"])
@super_admin_required
def api_tenant_user_delete(tid, uid):
    user = AppUser.query.filter_by(id=uid, tenant_id=tid).first_or_404()
    if user.role == 'owner':
        return jsonify({'error': 'オーナーアカウントは削除できません'}), 400
    db.session.delete(user)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/me")
@login_required
def api_me():
    """ログインユーザーの診断情報を返す（テナント・店舗・権限）"""
    uid = session.get('app_user_id')
    user = AppUser.query.get(uid)
    if not user:
        return jsonify({'error': 'user not found'}), 404
    allowed_ids = get_allowed_store_ids()
    allowed_ids_ignore = get_allowed_store_ids(ignore_active=True)
    stores_info = []
    for sid in allowed_ids_ignore:
        s = Store.query.get(sid)
        stores_info.append({'id': sid, 'name': s.name if s else '?', 'is_active_selected': sid in allowed_ids})
    tenant = Tenant.query.get(user.tenant_id) if user.tenant_id else None
    return jsonify({
        'user_id':    user.id,
        'username':   user.username,
        'role':       user.role,
        'tenant_id':  user.tenant_id,
        'tenant_name': tenant.name if tenant else None,
        'tenant_plan': tenant.plan if tenant else None,
        'store_id':   user.store_id,
        'active_store_id': session.get('active_store_id'),
        'allowed_store_ids': allowed_ids,
        'all_store_ids': allowed_ids_ignore,
        'stores': stores_info,
        'STORE_ID_template': allowed_ids[0] if allowed_ids else None,
    })


@app.route("/api/admin/data-audit")
@super_admin_required
def api_admin_data_audit():
    """テナント別データ件数を確認（テナント分離診断ツール）"""
    result = {}
    stores = Store.query.all()
    for s in stores:
        tenant = Tenant.query.get(s.tenant_id) if s.tenant_id else None
        result[s.id] = {
            'store_name':  s.name,
            'tenant_name': tenant.name if tenant else 'なし',
            'pl_records':     PLRecord.query.filter_by(store_id=s.id).count(),
            'lead_stats':     LeadMediaStat.query.filter_by(store_id=s.id).count(),
            'leads':          Lead.query.filter_by(store_id=s.id).count(),
            'kpis':           SalesKPI.query.filter_by(store_id=s.id).count(),
            'staff':          Staff.query.filter_by(store_id=s.id, is_active=True).count(),
        }
    return jsonify(result)


# ── 申込一覧 API ──────────────────────────────────────────

def _parse_date(s):
    if not s: return None
    try: return datetime.strptime(s, '%Y-%m-%d').date()
    except: return None


def _app_record_to_dict(r, staff_map):
    return {
        'id': r.id,
        'staff_id': r.staff_id,
        'staff_name': staff_map.get(r.staff_id, '-') if r.staff_id else '-',
        'application_date': r.application_date.isoformat() if r.application_date else None,
        'media': r.media or '',
        'property_name': r.property_name or '',
        'room_number': r.room_number or '',
        'customer_name': r.customer_name or '',
        'rent': r.rent or 0,
        'management_company': r.management_company or '',
        'review_ng': bool(r.review_ng),
        'review_status': r.review_status or None,
        'past_customer': bool(r.past_customer),
        'contract_start_date': r.contract_start_date.isoformat() if r.contract_start_date else None,
        'ad_payment_date': r.ad_payment_date.isoformat() if r.ad_payment_date else None,
        'brokerage_fee': r.brokerage_fee or 0,
        'option_amount': r.option_amount or 0,
        'ad_type': r.ad_type or 'amount',
        'ad_amount': r.ad_amount or 0,
        'ad_amount_yen': round((r.rent or 0) * (r.ad_amount or 0) / 100) if (r.ad_type or 'amount') == 'percent' else (r.ad_amount or 0),
        'lifeline': bool(r.lifeline),
        'moving': bool(r.moving),
        'fire_insurance': bool(r.fire_insurance),
        'status': r.status or '申込',
        'ad_settled': bool(r.ad_settled),
        'ad_approved': bool(r.ad_approved),
        'brokerage_payment_date': r.brokerage_payment_date.strftime('%Y-%m-%d') if r.brokerage_payment_date else None,
        'brokerage_settled': bool(r.brokerage_settled),
        'brokerage_approved': bool(r.brokerage_approved),
        'option_settled': bool(r.option_settled),
        'option_approved': bool(r.option_approved),
        'option_payment_date': r.option_payment_date.strftime('%Y-%m-%d') if r.option_payment_date else None,
        'created_at': r.created_at.isoformat() if r.created_at else None,
    }


def _approved_in_month(rec, field, year, month):
    """指定方向(field)が、指定年月に入金（承認）済みかを返す。
    入金日(payment_date)の年月で判定する。"""
    if field == 'brokerage':
        ok, dt, amt = rec.brokerage_approved, rec.brokerage_payment_date, (rec.brokerage_fee or 0)
    elif field == 'option':
        ok, dt, amt = rec.option_approved, rec.option_payment_date, (rec.option_amount or 0)
    elif field == 'ad':
        ok, dt, amt = rec.ad_approved, rec.ad_payment_date, _ad_yen(rec)
    else:
        return False
    return bool(ok and amt != 0 and dt and dt.year == year and dt.month == month)


@app.route("/api/applications/settled")
@login_required
def api_applications_settled():
    """入金済み一覧：その月に入金（承認）された方向がある案件を返す。
    入金日(payment_date)の月で判定（6月申込/7月AD入金→7月の一覧に表示）。"""
    allowed_ids = get_allowed_store_ids()
    store_id = request.args.get('store_id', type=int)
    staff_id = request.args.get('staff_id', type=int)
    year  = request.args.get('year',  type=int)
    month = request.args.get('month', type=int)

    q = ApplicationRecord.query.filter(
        ApplicationRecord.store_id.in_(allowed_ids),
        ~ApplicationRecord.status.in_(['キャンセル', 'キャンセル振替']),
    )
    if store_id and store_id in allowed_ids:
        q = q.filter(ApplicationRecord.store_id == store_id)
    if staff_id:
        q = q.filter(ApplicationRecord.staff_id == staff_id)

    if year and month:
        # その月に入金日がある方向を持つ案件のみ
        q = q.filter(db.or_(
            db.and_(ApplicationRecord.brokerage_approved == True, ApplicationRecord.brokerage_fee != 0,
                    db.extract('year', ApplicationRecord.brokerage_payment_date) == year,
                    db.extract('month', ApplicationRecord.brokerage_payment_date) == month),
            db.and_(ApplicationRecord.option_approved == True, ApplicationRecord.option_amount != 0,
                    db.extract('year', ApplicationRecord.option_payment_date) == year,
                    db.extract('month', ApplicationRecord.option_payment_date) == month),
            db.and_(ApplicationRecord.ad_approved == True, ApplicationRecord.ad_amount != 0,
                    db.extract('year', ApplicationRecord.ad_payment_date) == year,
                    db.extract('month', ApplicationRecord.ad_payment_date) == month),
        ))
    else:
        q = q.filter(db.or_(
            db.and_(ApplicationRecord.brokerage_fee != 0, ApplicationRecord.brokerage_approved == True),
            db.and_(ApplicationRecord.ad_amount != 0, ApplicationRecord.ad_approved == True),
            db.and_(ApplicationRecord.option_amount != 0, ApplicationRecord.option_approved == True),
        ))

    recs = q.order_by(ApplicationRecord.application_date.asc(), ApplicationRecord.id.asc()).all()
    staff_map = {s.id: s.name for s in Staff.query.filter(Staff.id.in_({r.staff_id for r in recs if r.staff_id})).all()}
    # フロントが当月入金分を判定できるよう、対象年月も返す
    out = []
    for r in recs:
        d = _app_record_to_dict(r, staff_map)
        if year and month:
            d['brokerage_paid_this_month'] = _approved_in_month(r, 'brokerage', year, month)
            d['option_paid_this_month']    = _approved_in_month(r, 'option', year, month)
            d['ad_paid_this_month']        = _approved_in_month(r, 'ad', year, month)
        out.append(d)
    return jsonify(out)


@app.route("/api/applications/unpaid")
@login_required
def api_applications_unpaid():
    """未入金一覧：仲介またはADが未承認の申込レコードを返す（月フィルタ）"""
    allowed_ids = get_allowed_store_ids()
    store_id = request.args.get('store_id', type=int)
    staff_id = request.args.get('staff_id', type=int)
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]

    from datetime import date as _date
    # その月以前に申し込まれた未承認レコードを表示（入力した月から記載）
    month_end = _date(year, month, 28)  # その月末日以前
    try:
        import calendar
        last_day = calendar.monthrange(year, month)[1]
        month_end = _date(year, month, last_day)
    except Exception:
        pass

    q = ApplicationRecord.query.filter(
        ApplicationRecord.store_id.in_(allowed_ids),
        ~ApplicationRecord.status.in_(['キャンセル', 'キャンセル振替']),
        ApplicationRecord.application_date <= month_end,  # 選択月以前の案件のみ
        db.or_(
            db.and_(ApplicationRecord.brokerage_fee != 0, ApplicationRecord.brokerage_approved == False),
            db.and_(ApplicationRecord.ad_amount != 0,     ApplicationRecord.ad_approved == False),
            db.and_(ApplicationRecord.option_amount != 0, ApplicationRecord.option_approved == False),
        )
    )
    if store_id and store_id in allowed_ids:
        q = q.filter(ApplicationRecord.store_id == store_id)
    if staff_id:
        q = q.filter(ApplicationRecord.staff_id == staff_id)

    # 古いものが上（申込日昇順）
    recs = q.order_by(ApplicationRecord.application_date.asc(), ApplicationRecord.id.asc()).all()
    staff_ids = list({r.staff_id for r in recs if r.staff_id})
    staff_map = {s.id: s.name for s in Staff.query.filter(Staff.id.in_(staff_ids)).all()} if staff_ids else {}
    return jsonify([_app_record_to_dict(r, staff_map) for r in recs])


@app.route("/api/applications/approved-sum")
@login_required
def api_applications_approved_sum():
    """入金済み一覧の売上合計と件数を返す（年月フィルタ付き）。
    入金日(payment_date)が当月の承認済み方向を持つ案件で判定する（入金済み一覧と同じ基準）。"""
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]
    store_id = request.args.get('store_id', type=int)
    allowed = get_allowed_store_ids()

    q = ApplicationRecord.query.filter(
        ApplicationRecord.store_id.in_(allowed),
        ~ApplicationRecord.status.in_(['キャンセル', 'キャンセル振替']),
    )
    if store_id and store_id in allowed:
        q = q.filter(ApplicationRecord.store_id == store_id)

    total = 0
    count = 0
    for r in q.all():
        amt = 0
        if _approved_in_month(r, 'brokerage', year, month): amt += (r.brokerage_fee or 0)
        if _approved_in_month(r, 'option', year, month):    amt += (r.option_amount or 0)
        if _approved_in_month(r, 'ad', year, month):         amt += _ad_yen(r)
        if amt:
            total += amt
            count += 1   # 入金済み一覧に出る案件数（= 契約数）
    return jsonify({'total': total, 'count': count})


def _ad_yen(r):
    if (r.ad_type or 'amount') == 'percent':
        return round((r.rent or 0) * (r.ad_amount or 0) / 100)
    return r.ad_amount or 0


def _record_total_amount(r):
    """案件の満額（仲手＋その他費用＋AD）"""
    return (r.brokerage_fee or 0) + (r.option_amount or 0) + _ad_yen(r)


def _record_approved_amount(r):
    """入金済み（承認済み）金額（方向ごとに独立加算）"""
    amt = 0
    if (r.brokerage_fee or 0) != 0 and r.brokerage_approved:
        amt += r.brokerage_fee or 0
    if (r.option_amount or 0) != 0 and r.option_approved:
        amt += r.option_amount or 0
    ady = _ad_yen(r)
    if ady != 0 and r.ad_approved:
        amt += ady
    return amt


@app.route("/api/applications/summary")
@login_required
def api_applications_summary():
    """顧客管理表の集計サマリー（年月・スタッフ別）

    定義:
      売上      = その月に入金済み一覧に入った金額（入金日がその月の承認済み金額）
      見込み売上 = 申込一覧表（その月の申込）の入金予定金額の合計（満額）
      未入金    = 未入金一覧（その月以前の申込で未承認の方向）の金額合計
      申込数    = 当月の申込件数（キャンセル除く）
      契約数    = ステータス「契約」件数
      ライフライン/火災保険/引越し = 各付帯件数と申込数に対する割合(%)
    """
    allowed = get_allowed_store_ids()
    store_id = request.args.get('store_id', type=int)
    staff_id = request.args.get('staff_id', type=int)
    year  = request.args.get('year',  type=int) or current_ym()[0]
    month = request.args.get('month', type=int) or current_ym()[1]

    import calendar as _cal
    month_end = date(year, month, _cal.monthrange(year, month)[1])

    base = ApplicationRecord.query.filter(
        ApplicationRecord.store_id.in_(allowed),
        ~ApplicationRecord.status.in_(['キャンセル', 'キャンセル振替']),
    )
    if store_id and store_id in allowed:
        base = base.filter(ApplicationRecord.store_id == store_id)
    if staff_id:
        base = base.filter(ApplicationRecord.staff_id == staff_id)

    # その月の申込（件数・付帯・見込み売上）
    month_recs = base.filter(
        db.extract('year',  ApplicationRecord.application_date) == year,
        db.extract('month', ApplicationRecord.application_date) == month,
    ).all()

    # 売上：入金日がその月の承認済み金額（申込月は問わない）
    paid_recs = base.all()
    sales = 0
    for r in paid_recs:
        if _approved_in_month(r, 'brokerage', year, month): sales += (r.brokerage_fee or 0)
        if _approved_in_month(r, 'option', year, month):    sales += (r.option_amount or 0)
        if _approved_in_month(r, 'ad', year, month):         sales += _ad_yen(r)

    # 未入金：その月以前の申込で未承認の方向の金額合計
    uncollected = 0
    for r in base.filter(ApplicationRecord.application_date <= month_end).all():
        if (r.brokerage_fee or 0) != 0 and not r.brokerage_approved: uncollected += r.brokerage_fee or 0
        if (r.option_amount or 0) != 0 and not r.option_approved:    uncollected += r.option_amount or 0
        ady = _ad_yen(r)
        if ady != 0 and not r.ad_approved:                            uncollected += ady

    expected_sales = sum(_record_total_amount(r) for r in month_recs)
    app_count    = len(month_recs)
    contract_count = sum(1 for r in month_recs if r.status == '契約')
    lifeline_count = sum(1 for r in month_recs if r.lifeline)
    moving_count   = sum(1 for r in month_recs if r.moving)
    fire_count     = sum(1 for r in month_recs if r.fire_insurance)
    rate = lambda c: round(c / app_count * 100, 1) if app_count else 0

    return jsonify({
        'sales': sales,
        'expected_sales': expected_sales,
        'uncollected': uncollected,
        'application_count': app_count,
        'contract_count': contract_count,
        'lifeline_count': lifeline_count,
        'moving_count': moving_count,
        'fire_insurance_count': fire_count,
        'lifeline_rate': rate(lifeline_count),
        'moving_rate': rate(moving_count),
        'fire_insurance_rate': rate(fire_count),
    })


@app.route("/api/applications/search")
@login_required
def api_applications_search():
    """全期間を横断して申込レコードをキーワード検索"""
    q = request.args.get('q', '').strip()
    if not q or len(q) < 1:
        return jsonify([])

    allowed_ids = get_allowed_store_ids()
    cur_user = AppUser.query.get(session.get('app_user_id'))
    like = f'%{q}%'

    # スタッフ名検索のためのID収集
    matching_staff_ids = [s.id for s in Staff.query.filter(Staff.name.ilike(like)).all()]

    conditions = [
        ApplicationRecord.property_name.ilike(like),
        ApplicationRecord.customer_name.ilike(like),
        ApplicationRecord.room_number.ilike(like),
        ApplicationRecord.media.ilike(like),
    ]
    if matching_staff_ids:
        conditions.append(ApplicationRecord.staff_id.in_(matching_staff_ids))

    query = ApplicationRecord.query.filter(
        ApplicationRecord.store_id.in_(allowed_ids),
        db.or_(*conditions)
    )
    # スタッフも他スタッフの情報を横断検索可（編集は本人のみ）

    records = query.order_by(ApplicationRecord.application_date.desc(), ApplicationRecord.id.desc()).limit(200).all()
    staff_map = {s.id: s.name for s in Staff.query.all()}
    return jsonify([_app_record_to_dict(r, staff_map) for r in records])


@app.route("/api/applications/management-companies")
@login_required
def api_management_companies():
    """過去に入力された管理会社名の一覧（入力候補＝記憶用）"""
    allowed_ids = get_allowed_store_ids()
    rows = db.session.query(ApplicationRecord.management_company).filter(
        ApplicationRecord.store_id.in_(allowed_ids),
        ApplicationRecord.management_company.isnot(None),
        ApplicationRecord.management_company != '',
    ).distinct().all()
    names = sorted({(r[0] or '').strip() for r in rows if (r[0] or '').strip()})
    return jsonify(names)


@app.route("/api/applications", methods=["GET"])
@login_required
def api_applications_list():
    allowed_ids = get_allowed_store_ids()
    cur_user = AppUser.query.get(session.get('app_user_id'))
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    staff_id_filter = request.args.get('staff_id', type=int)
    status_filter = request.args.get('status')

    q = ApplicationRecord.query.filter(ApplicationRecord.store_id.in_(allowed_ids))

    # スタッフも他スタッフの情報を閲覧可（編集は本人のみ＝フロント/更新APIで制御）
    if staff_id_filter:
        q = q.filter(ApplicationRecord.staff_id == staff_id_filter)

    if year and month:
        from sqlalchemy import extract
        q = q.filter(
            extract('year', ApplicationRecord.application_date) == year,
            extract('month', ApplicationRecord.application_date) == month
        )

    if status_filter:
        q = q.filter(ApplicationRecord.status == status_filter)

    # 申込一覧は全件表示（全部入金済みになっても申込一覧からは消さない）
    records = q.order_by(ApplicationRecord.application_date.asc(), ApplicationRecord.id.asc()).all()
    staff_map = {s.id: s.name for s in Staff.query.all()}
    return jsonify([_app_record_to_dict(r, staff_map) for r in records])


@app.route("/api/applications", methods=["POST"])
@login_required
def api_applications_create():
    cur_user = AppUser.query.get(session.get('app_user_id'))
    allowed_ids = get_allowed_store_ids()
    data = request.get_json() or {}

    req_sid = data.get('store_id')
    if req_sid and int(req_sid) in allowed_ids:
        store_id = int(req_sid)
    elif allowed_ids:
        store_id = allowed_ids[0]
    else:
        return jsonify({'error': '利用可能な店舗がありません'}), 403

    staff_id = data.get('staff_id') or None
    if cur_user and cur_user.role == 'staff':
        staff_id = cur_user.staff_id  # staff は自分のレコードのみ

    rec = ApplicationRecord(
        store_id=store_id, staff_id=staff_id,
        application_date=_parse_date(data.get('application_date')) or date.today(),
        media=data.get('media') or None,
        property_name=data.get('property_name') or None,
        room_number=data.get('room_number') or None,
        customer_name=data.get('customer_name') or None,
        rent=float(data.get('rent') or 0),
        management_company=data.get('management_company') or None,
        review_ng=bool(data.get('review_ng')),
        contract_start_date=_parse_date(data.get('contract_start_date')),
        ad_payment_date=_parse_date(data.get('ad_payment_date')),
        brokerage_fee=float(data.get('brokerage_fee') or 0),
        option_amount=float(data.get('option_amount') or 0),
        ad_type=data.get('ad_type') or 'percent',
        ad_amount=float(data.get('ad_amount') or 0),
        lifeline=bool(data.get('lifeline')),
        moving=bool(data.get('moving')),
        fire_insurance=bool(data.get('fire_insurance')),
        status=data.get('status') or '申込',
    )
    db.session.add(rec)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': rec.id})


@app.route("/api/applications/<int:rid>", methods=["GET"])
@login_required
def api_applications_get(rid):
    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403
    staff_map = {s.id: s.name for s in Staff.query.all()}
    return jsonify(_app_record_to_dict(rec, staff_map))


@app.route("/api/applications/<int:rid>", methods=["PUT"])
@login_required
def api_applications_update(rid):
    cur_user = AppUser.query.get(session.get('app_user_id'))
    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403
    if cur_user and cur_user.role == 'staff' and rec.staff_id != cur_user.staff_id:
        return jsonify({'error': '権限がありません'}), 403

    data = request.get_json() or {}
    is_manager = cur_user and cur_user.role in ('owner', 'store_manager', 'super_admin')

    for fld in ['media', 'property_name', 'room_number', 'customer_name', 'status', 'ad_type', 'management_company']:
        if fld in data: setattr(rec, fld, data[fld] or None)
    if 'staff_id' in data and is_manager:
        rec.staff_id = data['staff_id'] or None
    for fld in ['rent', 'brokerage_fee', 'ad_amount', 'option_amount']:
        if fld in data: setattr(rec, fld, float(data[fld] or 0))
    for fld in ['lifeline', 'moving', 'fire_insurance']:
        if fld in data: setattr(rec, fld, bool(data[fld]))
    for fld in ['application_date', 'contract_start_date', 'ad_payment_date', 'brokerage_payment_date', 'option_payment_date']:
        if fld in data: setattr(rec, fld, _parse_date(data[fld]))

    # 審査状態: None=— / 'ok'=○→契約 / 'ng'=×→キャンセル
    if 'review_status' in data:
        rs = data['review_status'] or None
        rec.review_status = rs
        rec.review_ng = (rs == 'ng')
        if rs == 'ok':
            rec.status = '契約'
        elif rs == 'ng':
            rec.status = 'キャンセル'
        else:
            # リセット: 審査由来のステータスを申込に戻す
            if rec.status in ('契約', 'キャンセル'):
                rec.status = '申込'

    rec.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/applications/<int:rid>", methods=["DELETE"])
@login_required
def api_applications_delete(rid):
    cur_user = AppUser.query.get(session.get('app_user_id'))
    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403
    if cur_user and cur_user.role == 'staff':
        return jsonify({'error': '削除権限がありません'}), 403
    db.session.delete(rec)
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/applications/<int:rid>/settle", methods=["POST"])
@login_required
def api_applications_settle(rid):
    """営業マンが入金報告（決済フラグ）"""
    cur_user = AppUser.query.get(session.get('app_user_id'))
    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403
    if cur_user and cur_user.role == 'staff' and rec.staff_id != cur_user.staff_id:
        return jsonify({'error': '権限がありません'}), 403

    data = request.get_json() or {}
    field = data.get('field')  # 'ad' / 'brokerage' / 'option'
    pay_date = _parse_date(data.get('date'))  # 報告時に入力された入金日
    if field == 'ad':
        rec.ad_settled = True
        if pay_date: rec.ad_payment_date = pay_date
    elif field == 'brokerage':
        rec.brokerage_settled = True
        if pay_date: rec.brokerage_payment_date = pay_date
    elif field == 'option':
        rec.option_settled = True
        if pay_date: rec.option_payment_date = pay_date
    else:
        return jsonify({'error': 'invalid field'}), 400
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/applications/<int:rid>/approve", methods=["POST"])
@login_required
def api_applications_approve(rid):
    """店長が入金を承認 → 全項目承認済みになったらSalesKPIの売上に反映"""
    cur_user = AppUser.query.get(session.get('app_user_id'))
    if not cur_user or cur_user.role not in ('owner', 'store_manager', 'super_admin'):
        return jsonify({'error': '権限がありません'}), 403

    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403

    data = request.get_json() or {}
    field = data.get('field')  # 'ad' / 'brokerage' / 'option'

    # 承認日（入金日）を指定可能。未指定なら当日
    approve_date = _parse_date(data.get('date')) or date.today()

    if field == 'ad' and not rec.ad_approved:
        rec.ad_approved = True
        rec.ad_payment_date = rec.ad_payment_date or approve_date
    elif field == 'brokerage' and not rec.brokerage_approved:
        rec.brokerage_approved = True
        rec.brokerage_payment_date = rec.brokerage_payment_date or approve_date
    elif field == 'option' and not rec.option_approved:
        rec.option_approved = True
        rec.option_payment_date = rec.option_payment_date or approve_date
    else:
        return jsonify({'error': 'invalid field or already approved'}), 400

    # 承認されたフィールド分だけ即座に売上に反映（部分承認対応）
    # 仲手・その他費用・ADはそれぞれ独立した入金方向として加算する
    ad_yen = round((rec.rent or 0) * (rec.ad_amount or 0) / 100) if (rec.ad_type or 'amount') == 'percent' else (rec.ad_amount or 0)

    approved_amount = 0
    if field == 'brokerage':
        approved_amount += (rec.brokerage_fee or 0)
    elif field == 'option':
        approved_amount += (rec.option_amount or 0)
    elif field == 'ad' and ad_yen > 0:
        approved_amount += ad_yen

    if approved_amount > 0 and rec.staff_id:
        # staff_idが未設定の場合はKPI反映をスキップ（IntegrityError防止）
        ref_date = rec.application_date or date.today()
        ref_year, ref_month = ref_date.year, ref_date.month
        kpi = SalesKPI.query.filter_by(
            staff_id=rec.staff_id, store_id=rec.store_id,
            year=ref_year, month=ref_month
        ).first()
        if not kpi:
            kpi = SalesKPI(staff_id=rec.staff_id, store_id=rec.store_id,
                           year=ref_year, month=ref_month)
            db.session.add(kpi)
        kpi.sales_amount = (kpi.sales_amount or 0) + approved_amount

    db.session.commit()
    return jsonify({'status': 'ok', 'approved_amount': approved_amount})


@app.route("/api/applications/<int:rid>/unapprove", methods=["POST"])
@login_required
def api_applications_unapprove(rid):
    """店長が入金承認を取消 → その方向を未報告に戻し、売上(KPI)から差し引く"""
    cur_user = AppUser.query.get(session.get('app_user_id'))
    if not cur_user or cur_user.role not in ('owner', 'store_manager', 'super_admin'):
        return jsonify({'error': '権限がありません'}), 403

    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403

    data = request.get_json() or {}
    field = data.get('field')  # 'ad' / 'brokerage' / 'option'
    ad_yen = round((rec.rent or 0) * (rec.ad_amount or 0) / 100) if (rec.ad_type or 'amount') == 'percent' else (rec.ad_amount or 0)

    removed_amount = 0
    if field == 'brokerage' and rec.brokerage_approved:
        rec.brokerage_approved = False; rec.brokerage_settled = False
        rec.brokerage_payment_date = None
        removed_amount = rec.brokerage_fee or 0
    elif field == 'option' and rec.option_approved:
        rec.option_approved = False; rec.option_settled = False
        rec.option_payment_date = None
        removed_amount = rec.option_amount or 0
    elif field == 'ad' and rec.ad_approved:
        rec.ad_approved = False; rec.ad_settled = False
        rec.ad_payment_date = None
        removed_amount = ad_yen
    else:
        return jsonify({'error': 'invalid field or not approved'}), 400

    # 承認時に加算したKPI売上を差し引く
    if removed_amount and rec.staff_id:
        ref_date = rec.application_date or date.today()
        kpi = SalesKPI.query.filter_by(
            staff_id=rec.staff_id, store_id=rec.store_id,
            year=ref_date.year, month=ref_date.month
        ).first()
        if kpi:
            kpi.sales_amount = (kpi.sales_amount or 0) - removed_amount

    db.session.commit()
    return jsonify({'status': 'ok', 'removed_amount': removed_amount})


@app.route("/api/applications/<int:rid>/unsettle", methods=["POST"])
@login_required
def api_applications_unsettle(rid):
    """否認：入金報告を取消して未報告に戻す（承認前のみ／スタッフは自分の案件のみ可）"""
    cur_user = AppUser.query.get(session.get('app_user_id'))
    allowed_ids = get_allowed_store_ids()
    rec = ApplicationRecord.query.get_or_404(rid)
    if rec.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403
    if cur_user and cur_user.role == 'staff' and rec.staff_id != cur_user.staff_id:
        return jsonify({'error': '権限がありません'}), 403

    data = request.get_json() or {}
    field = data.get('field')  # 'ad' / 'brokerage' / 'option'
    if field == 'brokerage':
        if rec.brokerage_approved:
            return jsonify({'error': 'already approved'}), 400
        rec.brokerage_settled = False; rec.brokerage_payment_date = None
    elif field == 'option':
        if rec.option_approved:
            return jsonify({'error': 'already approved'}), 400
        rec.option_settled = False; rec.option_payment_date = None
    elif field == 'ad':
        if rec.ad_approved:
            return jsonify({'error': 'already approved'}), 400
        rec.ad_settled = False; rec.ad_payment_date = None
    else:
        return jsonify({'error': 'invalid field'}), 400
    db.session.commit()
    return jsonify({'status': 'ok'})


@app.route("/api/pending-approvals")
@login_required
def api_pending_approvals():
    """通知ベル件数。
    super_admin / sys_admin: 新規問合せ（トライアル申込）件数 → 問合せ管理へ
    owner / store_manager: 入金承認待ち件数 → 営業分析へ"""
    cur_user = AppUser.query.get(session.get('app_user_id'))
    if not cur_user:
        return jsonify({'count': 0, 'url': '/sales'})
    # 管理アカウント：新規問合せ件数
    if cur_user.role in ('super_admin', 'sys_admin'):
        try:
            count = TrialApplication.query.filter_by(status='new').count()
        except Exception:
            count = 0
        return jsonify({'count': count, 'url': '/admin/applications'})
    if cur_user.role not in ('owner', 'store_manager'):
        return jsonify({'count': 0, 'url': '/sales'})
    allowed_ids = get_allowed_store_ids()
    count = ApplicationRecord.query.filter(
        ApplicationRecord.store_id.in_(allowed_ids),
        db.or_(
            db.and_(ApplicationRecord.ad_settled == True, ApplicationRecord.ad_approved == False),
            db.and_(ApplicationRecord.brokerage_settled == True, ApplicationRecord.brokerage_approved == False),
            db.and_(ApplicationRecord.option_settled == True, ApplicationRecord.option_approved == False)
        )
    ).count()
    return jsonify({'count': count, 'url': '/sales'})


@app.route("/api/notifications")
@login_required
def api_notifications():
    """通知ベルの中身（お知らせを文章で一覧）。ロール別。"""
    cur = AppUser.query.get(session.get('app_user_id'))
    items = []
    if not cur:
        return jsonify({'items': [], 'total': 0})

    if cur.role in ('super_admin', 'sys_admin'):
        try:
            n = TrialApplication.query.filter_by(status='new').count()
        except Exception:
            n = 0
        if n:
            items.append({'icon': '📨', 'title': f'新着のお問い合わせ {n}件',
                          'body': 'トライアル申込フォームから新しいお問い合わせが届いています。クリックで問合せ管理を開きます。',
                          'url': '/admin/applications', 'count': n})
    else:
        allowed = get_allowed_store_ids()
        # 入金承認待ち（オーナー・店長）
        if cur.role in ('owner', 'store_manager') and allowed:
            try:
                ap = ApplicationRecord.query.filter(
                    ApplicationRecord.store_id.in_(allowed),
                    db.or_(
                        db.and_(ApplicationRecord.ad_settled == True, ApplicationRecord.ad_approved == False),
                        db.and_(ApplicationRecord.brokerage_settled == True, ApplicationRecord.brokerage_approved == False),
                        db.and_(ApplicationRecord.option_settled == True, ApplicationRecord.option_approved == False),
                    )).count()
            except Exception:
                ap = 0
            if ap:
                items.append({'icon': '💰', 'title': f'入金承認待ち {ap}件',
                              'body': '営業から入金報告があり、店長の承認待ちの案件があります。営業分析の入金承認からご確認ください。',
                              'url': '/sales', 'count': ap})
        # 未返信メール（全スタッフ）
        if allowed:
            try:
                ur = EchoRecord.query.filter(EchoRecord.store_id.in_(allowed),
                                             EchoRecord.has_unread_reply == True).count()
            except Exception:
                ur = 0
            if ur:
                items.append({'icon': '✉️', 'title': f'未返信のメール {ur}件',
                              'body': 'お客様から返信が届いていて、まだ返信していないメールがあります。反響管理表のチャットからご返信ください。',
                              'url': '/echo-management', 'count': ur})

    total = sum(i['count'] for i in items)
    return jsonify({'items': items, 'total': total})


# ── 使い方マニュアル ──────────────────────────────────────
MANUAL_SECTIONS = [
    {'cat': 'はじめに', 'title': 'ログインと基本画面', 'body':
     'ユーザー名（またはメール）とパスワードでログインします。左のサイドバーから各機能へ移動できます。'
     'サイドバーの項目をクリックで画面が切り替わり、「事務作業」「各種設定」はクリックすると下にメニューが開きます。'
     '上部のベルでお知らせ、その隣の「ヘルプ」でこのマニュアルが開きます。'},
    {'cat': '売上分析', 'title': '売上分析ダッシュボード', 'body':
     '「売上分析」では、当月の売上・見込・各種KPIをカードで確認できます。総売上は顧客管理の入金額（入金月ベース）を集計したものです。'
     '上部の月切替で対象月を変更できます。'},
    {'cat': '顧客管理表', 'title': '顧客管理表（入金・売上）', 'body':
     '申込ごとの金額（仲介手数料・AD・その他費用など）と入金状況を管理します。入金報告→店長承認のワークフローがあり、'
     '承認時に入金日が自動で付きます。入金済み一覧は入金月ベースで表示され、金額は「円」表記です。'},
    {'cat': '反響管理表', 'title': '反響管理表の基本', 'body':
     'ポータル（SUUMO・HOME\'S等）からの反響を一覧で管理します。担当者・反響日・状況・顧客名・媒体・手段・番号有無・初回対応日・追客日程・メモなどをセルクリックで直接編集できます。'
     '上部の検索で顧客名・媒体・担当で絞り込めます。サマリーカードで総反響数や電話番号あり件数などを確認できます。'},
    {'cat': '反響管理表', 'title': '状況タグと行の色', 'body':
     '「状況」列でタグ（追客中・申込・終了など）を選べます。状況見出しの⚙からタグの追加・削除と、各タグの行の色を設定できます。'
     '行の背景色は状況タグの色で変わります。お客様から返信があって未返信の行は、左端に赤いラインが付いて目立ちます。'},
    {'cat': 'メール', 'title': '反響からメール送受信（チャット）', 'body':
     '各行の「✉️メール」ボタンでチャット画面が開き、お客様とメールでやり取りできます。テンプレ挿入（📋）・画像/PDF添付（📎）が可能です。'
     '送信は Shift+Enter（Enterは改行）または右下の➤ボタン。お客様がメールを開くと「既読」が表示されます。返信が来ると未返信としてアラート表示されます。'},
    {'cat': 'メール', 'title': '反響メールの自動取込（受信設定）', 'body':
     '「各種設定 → メール自動取込設定」で、反響メールが届くGmailを連携します。アプリパスワード方式またはGoogle連携で接続し、「自動取り込みを有効にする」をONにすると、'
     '届いた反響がリアルタイムで反響管理表に入ります。「ポータル登録」で差出人アドレス→媒体名を登録すると確実に振り分けられます。同じお客様の再反響は1件にまとまり、追加分はメモに記録されます。'},
    {'cat': 'メール', 'title': '自動返信（媒体ごとに変える）', 'body':
     '「各種設定 → メール自動返信設定」で「新着反響に自動返信する」をONにすると、反響が届いた瞬間に自動でテンプレを送信します（お客様メールがある反響のみ）。'
     '「ポータル登録」で媒体ごとに使うテンプレを選べるので、SUUMO用・HOME\'S用など媒体別に文面を変えられます。媒体未指定のときは既定テンプレが使われます。'},
    {'cat': 'メール', 'title': 'メールテンプレートの作り方（HTML・差し込み文字）', 'body':
     '「各種設定 → メールテンプレート設定」の「＋新しいテンプレートを作成」でエディタが開きます。HTML（太字・色・リンク・リスト）で装飾でき、'
     '差し込み文字（#name#＝お客様氏名、#お問い合わせ物件名#、#会社名#、#会社電話番号#、#公式LINE# など）を入れると送信時に実データへ自動で置き換わります。'},
    {'cat': 'メール', 'title': '会社情報の登録（差し込み用）', 'body':
     '「各種設定 → メールテンプレート設定」の会社情報カードに、会社名・電話・メール・住所・営業時間・定休日・公式LINEを登録します。'
     'ここに登録した内容が、テンプレ内の差し込み文字（#会社名# 等）に自動で入ります。'},
    {'cat': '接客管理表', 'title': '接客管理表', 'body':
     '来店・電話・メール等の接客対応を記録します。日付・担当・お客様名・対応種別・接客数・状況などを管理できます。'},
    {'cat': '営業分析', 'title': '営業分析・スタッフKPI', 'body':
     '店舗・スタッフ別の成約率や売上KPIを確認できます。成約率は接客数→申込数で算出。LL・火災保険・引越しの成約率カードや、スタッフ別の昨年同月比（YoY）も表示されます。'},
    {'cat': '事務作業', 'title': '事務作業メニュー', 'body':
     'サイドバーの「事務作業」をクリックすると、契約管理・顧客管理・日報・間取り作成・有給管理・経理（PL）が開きます。'},
    {'cat': 'チャット', 'title': '社内チャット', 'body':
     'サイドバーの「チャット」で社内メンバーとやり取りできます。全社チャンネル・店舗別チャンネル・グループ（オーナー/店長が作成）があり、'
     'チャットProプラン（オプション）では画像・PDFの添付と2年間の保存が可能です（通常はテキストのみ・60日保存）。'},
    {'cat': '通知', 'title': '通知ベルの見方', 'body':
     '上部のベルを押すと、未承認の入金報告・未返信メール・新着お問い合わせなどのお知らせが文章で一覧表示されます。各項目をクリックすると該当画面へ移動します。'},
    {'cat': '設定', 'title': '各種設定の場所', 'body':
     'メール関連の設定はサイドバー「各種設定」にまとまっています。メールテンプレート設定（文面）・メール自動取込設定（受信連携）・メール自動返信設定（自動返信）の3つです。'},
    {'cat': 'メール', 'title': 'アプリパスワードの作り方', 'body':
     '「メール自動取込設定」の下部「🔑アプリパスワードの作り方」を開くと手順が出ます。Googleで2段階認証をON→アプリパスワードを発行（16桁）→メールアドレスと一緒に貼り付け→接続テスト→保存、の流れです。'},
    {'cat': 'よくある質問', 'title': 'メールが送れない・受信できないとき', 'body':
     'まず「メール自動取込設定」で接続テストを行ってください。失敗する場合はメールアドレスの打ち間違い、または通常のログインパスワードを入れている（必ず16桁のアプリパスワードを使う）ことが多いです。'
     'Google連携（OAuth）を使う方法もあります。'},
]


@app.route("/api/manual")
@login_required
def api_manual():
    return jsonify(MANUAL_SECTIONS)


@app.route("/api/admin/net-diagnose")
@login_required
def api_net_diagnose():
    """本番サーバーからGmail等への到達性を計測する診断（送受信不可の原因切り分け用）。"""
    import socket as _sock, time as _t, ssl as _ssl
    out = {}

    # 1) DNS解決（A / AAAA）
    dns = {}
    for h in ['smtp.gmail.com', 'imap.gmail.com', 'www.google.com', 'api.resend.com']:
        e = {'A': [], 'AAAA': []}
        for fam, label in [(_sock.AF_INET, 'A'), (_sock.AF_INET6, 'AAAA')]:
            try:
                infos = _sock.getaddrinfo(h, None, fam, _sock.SOCK_STREAM)
                e[label] = sorted({i[4][0] for i in infos})
            except Exception as ex:
                e[label] = f'ERR {type(ex).__name__}: {ex}'
        dns[h] = e
    out['dns'] = dns

    # 2) TCP接続プローブ（family指定）
    def probe(host, port, family):
        t0 = _t.time()
        try:
            infos = _sock.getaddrinfo(host, port, family, _sock.SOCK_STREAM)
            if not infos:
                return {'ok': False, 'err': 'no address'}
            af, _st, _pr, _cn, addr = infos[0]
            s = _sock.socket(af, _sock.SOCK_STREAM)
            s.settimeout(8)
            s.connect(addr)
            s.close()
            return {'ok': True, 'ms': int((_t.time() - t0) * 1000), 'addr': addr[0]}
        except Exception as ex:
            return {'ok': False, 'err': f'{type(ex).__name__}: {ex}', 'ms': int((_t.time() - t0) * 1000)}

    probes = {}
    for host, port in [('smtp.gmail.com', 465), ('smtp.gmail.com', 587), ('smtp.gmail.com', 25),
                       ('imap.gmail.com', 993), ('www.google.com', 443), ('api.resend.com', 443)]:
        probes[f'{host}:{port}/v4'] = probe(host, port, _sock.AF_INET)
        probes[f'{host}:{port}/v6'] = probe(host, port, _sock.AF_INET6)
    out['probes'] = probes

    # 3) create_connection（smtplib/imaplibの既定動作と同じ AF_UNSPEC）
    def probe_unspec(host, port):
        t0 = _t.time()
        try:
            s = _sock.create_connection((host, port), timeout=8)
            peer = s.getpeername()[0]
            s.close()
            return {'ok': True, 'fam': ('v6' if ':' in peer else 'v4'), 'peer': peer, 'ms': int((_t.time() - t0) * 1000)}
        except Exception as ex:
            return {'ok': False, 'err': f'{type(ex).__name__}: {ex}', 'ms': int((_t.time() - t0) * 1000)}
    out['create_connection'] = {
        'smtp.gmail.com:465': probe_unspec('smtp.gmail.com', 465),
        'smtp.gmail.com:587': probe_unspec('smtp.gmail.com', 587),
        'imap.gmail.com:993': probe_unspec('imap.gmail.com', 993),
    }

    # 4) フルSMTPハンドシェイク（IPv4、ログイン手前まで）
    smtp_hs = {}
    for port in [465, 587]:
        t0 = _t.time()
        try:
            if port == 465:
                s = _SMTPSSLIPv4('smtp.gmail.com', 465, local_hostname='mieroom.cloud', timeout=12,
                                 context=_ssl.create_default_context())
            else:
                s = _SMTPIPv4('smtp.gmail.com', 587, local_hostname='mieroom.cloud', timeout=12)
                s.ehlo(); s.starttls(context=_ssl.create_default_context()); s.ehlo()
            code = s.noop()[0]
            try: s.quit()
            except Exception: pass
            smtp_hs[f'{port}'] = {'ok': True, 'noop': code, 'ms': int((_t.time() - t0) * 1000)}
        except Exception as ex:
            smtp_hs[f'{port}'] = {'ok': False, 'err': f'{type(ex).__name__}: {ex}', 'ms': int((_t.time() - t0) * 1000)}
    out['smtp_handshake_ipv4'] = smtp_hs

    return jsonify(out)


# ── 媒体マスター API ──────────────────────────────────────

@app.route("/api/media-types", methods=["GET"])
@login_required
def api_media_types_list():
    allowed_ids = get_allowed_store_ids()
    store_id = allowed_ids[0] if allowed_ids else 1
    items = MediaType.query.filter_by(store_id=store_id, is_active=True)\
        .order_by(MediaType.sort_order, MediaType.name).all()
    # 重複した媒体名を除外（同名は最初の1件だけ返す）
    seen = set()
    result = []
    for m in items:
        if m.name in seen:
            continue
        seen.add(m.name)
        result.append({'id': m.id, 'name': m.name})
    return jsonify(result)


@app.route("/api/media-types", methods=["POST"])
@login_required
def api_media_types_create():
    allowed_ids = get_allowed_store_ids()
    store_id = allowed_ids[0] if allowed_ids else 1
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': '媒体名を入力してください'}), 400
    if MediaType.query.filter_by(store_id=store_id, name=name, is_active=True).first():
        return jsonify({'error': 'すでに存在します'}), 400
    m = MediaType(store_id=store_id, name=name)
    db.session.add(m)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': m.id, 'name': m.name})


@app.route("/api/media-types/<int:mid>", methods=["DELETE"])
@login_required
def api_media_types_delete(mid):
    cur_user = AppUser.query.get(session.get('app_user_id'))
    if not cur_user or cur_user.role not in ('owner', 'store_manager', 'super_admin'):
        return jsonify({'error': '権限がありません'}), 403
    m = MediaType.query.get_or_404(mid)
    m.is_active = False
    db.session.commit()
    return jsonify({'status': 'ok'})


# ── ステータスカラー API ──────────────────────────────────

STATUS_COLOR_DEFAULTS = {
    '申込':          {'bg': '#ffffff', 'text': '#111827', 'row_bg': '#ffffff'},
    '契約':          {'bg': '#fef9c3', 'text': '#92400e', 'row_bg': '#fef9c3'},
    'キャンセル':     {'bg': '#fee2e2', 'text': '#b91c1c', 'row_bg': '#e5e7eb'},
    'キャンセル振替': {'bg': '#dcfce7', 'text': '#15803d', 'row_bg': '#e5e7eb'},
}


@app.route("/api/status-colors", methods=["GET"])
@login_required
def api_status_colors_get():
    allowed_ids = get_allowed_store_ids()
    store_id = allowed_ids[0] if allowed_ids else 1
    result = {}
    for key, default in STATUS_COLOR_DEFAULTS.items():
        sc = StatusColor.query.filter_by(store_id=store_id, status_key=key).first()
        result[key] = {
            'bg':     sc.bg_color if sc else default['bg'],
            'text':   sc.text_color if sc else default['text'],
            'row_bg': (sc.row_bg_color if (sc and sc.row_bg_color) else default.get('row_bg', '#ffffff')),
        }
    return jsonify(result)


@app.route("/api/status-colors", methods=["PUT"])
@login_required
def api_status_colors_update():
    cur_user = AppUser.query.get(session.get('app_user_id'))
    if not cur_user or cur_user.role not in ('owner', 'store_manager', 'super_admin'):
        return jsonify({'error': '権限がありません'}), 403
    allowed_ids = get_allowed_store_ids()
    data = request.get_json() or {}
    # フロントは {store_id, colors:{...}} の形で送る。colors を取り出す（後方互換で直接形も許容）
    req_sid = data.get('store_id')
    store_id = int(req_sid) if (req_sid and int(req_sid) in allowed_ids) else (allowed_ids[0] if allowed_ids else 1)
    colors_data = data.get('colors') if isinstance(data.get('colors'), dict) else data
    for status_key, colors in colors_data.items():
        if not isinstance(colors, dict):
            continue   # store_id 等の非カラー項目はスキップ
        sc = StatusColor.query.filter_by(store_id=store_id, status_key=status_key).first()
        if not sc:
            sc = StatusColor(store_id=store_id, status_key=status_key)
            db.session.add(sc)
        sc.bg_color = colors.get('bg', '#ffffff')
        sc.text_color = colors.get('text', '#111827')
        if 'row_bg' in colors:
            sc.row_bg_color = colors.get('row_bg') or '#ffffff'
    db.session.commit()
    return jsonify({'status': 'ok'})


# ── 目標売上 API ─────────────────────────────────────────

@app.route("/api/sales-kpi/target", methods=["POST"])
@login_required
def api_sales_kpi_target():
    """目標売上を設定（staff は自分のみ、manager は全員）"""
    cur_user = AppUser.query.get(session.get('app_user_id'))
    allowed_ids = get_allowed_store_ids()
    data = request.get_json() or {}

    staff_id = data.get('staff_id')
    if cur_user and cur_user.role == 'staff':
        staff_id = cur_user.staff_id

    year = data.get('year')
    month = data.get('month')
    target = float(data.get('target_sales') or 0)

    if not staff_id or not year or not month:
        return jsonify({'error': 'パラメータ不足'}), 400

    staff = Staff.query.get(staff_id)
    if not staff or staff.store_id not in allowed_ids:
        return jsonify({'error': '権限がありません'}), 403

    kpi = SalesKPI.query.filter_by(staff_id=staff_id, year=year, month=month,
                                    store_id=staff.store_id).first()
    if not kpi:
        kpi = SalesKPI(staff_id=staff_id, store_id=staff.store_id, year=year, month=month)
        db.session.add(kpi)
    kpi.target_sales = target
    db.session.commit()
    return jsonify({'status': 'ok'})


# ── プレミアム：本部ダッシュボード ────────────────────────

@app.route("/headquarters")
@login_required
@block_super_admin
def headquarters_dashboard():
    """本部ダッシュボード（プレミアプラン専用）"""
    cur_user = AppUser.query.get(session.get('app_user_id'))
    if not cur_user:
        return redirect(url_for('app_login'))
    if cur_user.role == 'staff':
        return redirect(url_for('sales_management'))
    # プレミアプランのみ許可
    if not is_premium_user():
        return redirect(url_for('executive_dashboard'))
    stores = get_allowed_stores(ignore_active=True)   # HQは全店舗（active_store_id無視）
    year, month = current_ym()
    return render_template("headquarters_dashboard.html",
                           stores=stores, year=year, month=month,
                           now=datetime.now())


@app.route("/api/hq/summary")
@login_required
@block_super_admin
def api_hq_summary():
    """全店舗KPIサマリー（本部ダッシュボード用・プレミアム限定）"""
    if not is_premium_user():
        return jsonify({'error': 'premium required'}), 403
    year  = request.args.get('year',  type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    store_ids = get_allowed_store_ids(ignore_active=True)

    today = date.today()
    days_in_month = (date(year, month % 12 + 1, 1) - timedelta(days=1)).day if month < 12 else 31
    elapsed_days  = min(today.day, days_in_month) if (today.year == year and today.month == month) else days_in_month

    result = []
    total_sales = 0
    total_target = 0
    total_apps   = 0
    total_contracts = 0

    for sid in store_ids:
        store = Store.query.get(sid)
        if not store:
            continue
        kpis = SalesKPI.query.filter_by(store_id=sid, year=year, month=month).all()
        sales    = sum(k.sales_amount  or 0 for k in kpis)
        target   = sum(k.target_sales  or 0 for k in kpis)
        apps     = sum(k.applications  or 0 for k in kpis)
        contracts= sum(k.contracts     or 0 for k in kpis)
        inquiries= sum(k.inquiries     or 0 for k in kpis)

        # 前月売上
        pm = month - 1 if month > 1 else 12
        py = year if month > 1 else year - 1
        prev_kpis = SalesKPI.query.filter_by(store_id=sid, year=py, month=pm).all()
        prev_sales = sum(k.sales_amount or 0 for k in prev_kpis)

        # 着地予測（現ペース × 月日数）
        forecast = int(sales / elapsed_days * days_in_month) if elapsed_days > 0 else 0

        close_rate = round(contracts / apps * 100, 1) if apps > 0 else 0
        vs_prev    = round((sales - prev_sales) / prev_sales * 100, 1) if prev_sales > 0 else None

        total_sales    += sales
        total_target   += target
        total_apps     += apps
        total_contracts += contracts

        # 利益・経費はPLRecordから取得
        pl = PLRecord.query.filter_by(store_id=sid, year=year, month=month).first()
        profit   = pl.net_profit if pl else None
        expenses = max(0, sales - (profit or 0)) if profit is not None else None

        result.append({
            'store_id':   sid,
            'store_name': store.name,
            'sales':      sales,
            'target':     target,
            'forecast':   forecast,
            'apps':       apps,
            'contracts':  contracts,
            'inquiries':  inquiries,
            'close_rate': close_rate,
            'vs_prev':    vs_prev,
            'prev_sales': prev_sales,
            'profit':     profit,
            'expenses':   expenses,
            'is_danger':  (target > 0 and sales < target * 0.5),
            'is_drop':    (vs_prev is not None and vs_prev <= -20),
        })

    total_forecast  = int(total_sales / elapsed_days * days_in_month) if elapsed_days > 0 else 0
    total_close     = round(total_contracts / total_apps * 100, 1) if total_apps > 0 else 0
    total_inquiries = sum(r['inquiries'] for r in result)
    total_profit    = sum(r['profit']    for r in result if r['profit'] is not None)
    total_expenses  = sum(r['expenses']  for r in result if r['expenses'] is not None)
    return jsonify({
        'stores':           result,
        'total_sales':      total_sales,
        'total_target':     total_target,
        'total_forecast':   total_forecast,
        'total_apps':       total_apps,
        'total_contracts':  total_contracts,
        'total_inquiries':  total_inquiries,
        'total_profit':     total_profit,
        'total_expenses':   total_expenses,
        'total_close_rate': total_close,
        'year': year, 'month': month,
        'elapsed_days': elapsed_days, 'days_in_month': days_in_month,
    })


@app.route("/api/hq/rankings")
@login_required
@block_super_admin
def api_hq_rankings():
    """店舗別ランキング（metric: sales / apps / close_rate・プレミアム限定）"""
    if not is_premium_user():
        return jsonify({'error': 'premium required'}), 403
    year   = request.args.get('year',   type=int) or date.today().year
    month  = request.args.get('month',  type=int) or date.today().month
    metric = request.args.get('metric', 'sales')
    store_ids = get_allowed_store_ids(ignore_active=True)

    rows = []
    for sid in store_ids:
        store = Store.query.get(sid)
        if not store:
            continue
        kpis = SalesKPI.query.filter_by(store_id=sid, year=year, month=month).all()
        sales     = sum(k.sales_amount or 0 for k in kpis)
        apps      = sum(k.applications or 0 for k in kpis)
        contracts = sum(k.contracts    or 0 for k in kpis)
        inquiries = sum(k.inquiries    or 0 for k in kpis)
        close_rate= round(contracts / apps * 100, 1) if apps > 0 else 0
        pl = PLRecord.query.filter_by(store_id=sid, year=year, month=month).first()
        profit   = pl.net_profit if pl else 0
        expenses = max(0, sales - profit) if pl else 0
        rows.append({'store_id': sid, 'store_name': store.name,
                     'sales': sales, 'apps': apps, 'contracts': contracts,
                     'inquiries': inquiries, 'close_rate': close_rate,
                     'profit': profit, 'expenses': expenses})

    key_map = {'sales':'sales','apps':'apps','contracts':'contracts',
               'inquiries':'inquiries','close_rate':'close_rate',
               'profit':'profit','expenses':'expenses'}
    sort_key = key_map.get(metric, 'sales')
    rows.sort(key=lambda r: r.get(sort_key, 0), reverse=True)
    for i, r in enumerate(rows):
        r['rank'] = i + 1
    return jsonify(rows)


@app.route("/api/hq/store-comparison")
@login_required
@block_super_admin
def api_hq_store_comparison():
    """店舗別月次推移（グラフ用・プレミアム限定）"""
    if not is_premium_user():
        return jsonify({'error': 'premium required'}), 403
    from_param = request.args.get('from')
    to_param   = request.args.get('to')
    metric     = request.args.get('metric', 'sales')
    store_ids  = get_allowed_store_ids(ignore_active=True)

    today = date.today()
    if from_param and to_param:
        try:
            fy, fm = int(from_param[:4]), int(from_param[5:7])
            ty, tm = int(to_param[:4]),   int(to_param[5:7])
        except Exception:
            fy, fm = today.year, today.month - 5
            ty, tm = today.year, today.month
    else:
        base = today.year * 12 + today.month - 1
        fy, fm = divmod(base - 5, 12)
        fm += 1
        ty, tm = today.year, today.month

    base_s = fy * 12 + fm - 1
    base_e = ty * 12 + tm - 1
    periods = [(t // 12, t % 12 + 1) for t in range(base_s, base_e + 1)]
    labels  = [f'{y}/{m:02d}' for y, m in periods]

    datasets = []
    colors_list = ['#16a34a','#2563eb','#dc2626','#d97706','#7c3aed','#db2777','#0891b2']
    for ci, sid in enumerate(store_ids):
        store = Store.query.get(sid)
        if not store:
            continue
        values = []
        for y, m in periods:
            kpis = SalesKPI.query.filter_by(store_id=sid, year=y, month=m).all()
            if metric == 'sales':
                v = sum(k.sales_amount or 0 for k in kpis)
            elif metric == 'apps':
                v = sum(k.applications or 0 for k in kpis)
            elif metric == 'contracts':
                v = sum(k.contracts    or 0 for k in kpis)
            elif metric == 'inquiries':
                v = sum(k.inquiries    or 0 for k in kpis)
            elif metric == 'close_rate':
                apps = sum(k.applications or 0 for k in kpis)
                ctrs = sum(k.contracts   or 0 for k in kpis)
                v = round(ctrs / apps * 100, 1) if apps > 0 else 0
            elif metric in ('profit', 'expenses'):
                pl = PLRecord.query.filter_by(store_id=sid, year=y, month=m).first()
                sales_v = sum(k.sales_amount or 0 for k in kpis)
                profit_v = pl.net_profit if pl else 0
                v = profit_v if metric == 'profit' else max(0, sales_v - profit_v)
            else:
                v = 0
            values.append(v)
        datasets.append({
            'store_id':   sid,
            'store_name': store.name,
            'color':      colors_list[ci % len(colors_list)],
            'data':       values,
        })
    return jsonify({'labels': labels, 'datasets': datasets})


@app.route("/api/hq/ai-summary", methods=["POST"])
@login_required
def api_hq_ai_summary():
    """Claude APIで全店舗データのAIサマリーを生成"""
    if not is_premium_user():
        return jsonify({'error': 'premium required'}), 403
    year  = request.json.get('year',  date.today().year)
    month = request.json.get('month', date.today().month)
    store_ids = get_allowed_store_ids(ignore_active=True)

    store_lines = []
    for sid in store_ids:
        store = Store.query.get(sid)
        if not store:
            continue
        kpis = SalesKPI.query.filter_by(store_id=sid, year=year, month=month).all()
        sales     = sum(k.sales_amount or 0 for k in kpis)
        target    = sum(k.target_sales or 0 for k in kpis)
        apps      = sum(k.applications or 0 for k in kpis)
        contracts = sum(k.contracts    or 0 for k in kpis)
        close     = round(contracts / apps * 100, 1) if apps > 0 else 0
        store_lines.append(
            f"・{store.name}: 売上¥{sales:,.0f} (目標¥{target:,.0f}), 申込{apps}件, 成約{contracts}件, 成約率{close}%"
        )

    prompt = f"""あなたは不動産会社の経営コンサルタントです。
以下は{year}年{month}月の全店舗KPIデータです。

{chr(10).join(store_lines) if store_lines else '（データなし）'}

以下の観点で簡潔に分析してください（400字以内）：
1. 全体の状況と注目すべき店舗
2. 課題・リスク
3. 来月に向けたアドバイス"""

    try:
        msg = anthropic_client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=600,
            messages=[{"role": "user", "content": prompt}],
        )
        return jsonify({'summary': msg.content[0].text})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 反響メール自動取込サービス（IMAP IDLE）を起動（複数ワーカーでも1つのみ）
try:
    start_mail_service()
except Exception as _e:
    print(f"start_mail_service error: {_e}")


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    host = "0.0.0.0" if _IS_POSTGRES else "127.0.0.1"
    app.run(debug=False, port=port, host=host, use_reloader=False)
